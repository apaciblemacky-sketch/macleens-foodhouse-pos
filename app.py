import io
import json
import base64
import calendar
import csv
import hashlib
import hmac
import ipaddress
import logging
import math
import os
import re
import secrets
from datetime import datetime, date, timedelta, time, timezone
from functools import wraps
from urllib.parse import urljoin, urlparse
from zoneinfo import ZoneInfo

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_from_directory, Response, has_request_context
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text, UniqueConstraint, and_, or_
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

import qrcode
import qrcode.image.svg
import requests
from PIL import Image, ImageDraw, ImageFont, ImageOps, UnidentifiedImageError

from marketing_agent import (
    analyze_marketing_insights, extract_peso_amounts, generate_ai_marketing_decision,
    gemini_configured, openai_configured,
)

logging.basicConfig(
    level=os.environ.get('LOG_LEVEL', 'INFO').upper(),
    format='%(asctime)s %(levelname)s %(name)s: %(message)s'
)

app = Flask(__name__)
# Trust Render's single reverse-proxy hop so generated QR URLs use the public HTTPS host.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
IS_PRODUCTION = bool(os.environ.get('RENDER') or os.environ.get('FLASK_ENV') == 'production')
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    _secret_key = secrets.token_hex(32)
    app.logger.warning('SECRET_KEY is not configured; generated a temporary key. Set SECRET_KEY in Render for stable sessions.')
app.config['SECRET_KEY'] = _secret_key

database_url = os.environ.get('DATABASE_URL')
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url or 'sqlite:///foodhouse_pos.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
# A non-JavaScript bulk-catalog submission contains about fourteen controls per
# product. Keep the guarded fallback usable for catalogs larger than 70 items;
# the modern editor sends only changed rows and normally stays far below this.
app.config['MAX_FORM_PARTS'] = 5000
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = IS_PRODUCTION

db = SQLAlchemy(app)

APP_RELEASE = '2026.09.04-community-lite-scale-safety-v6'
MANILA_TZ = ZoneInfo('Asia/Manila')
STAFF_SESSION_TIMEOUT = timedelta(hours=8)
_DB_INITIALIZED = False

def utc_now():
    """UTC-naive timestamp for backwards-compatible database storage."""
    return datetime.now(timezone.utc).replace(tzinfo=None)

def ph_now():
    return datetime.now(MANILA_TZ)

def ph_today():
    return ph_now().date()

def ph_day_utc_bounds(day=None):
    """Return UTC-naive [start, end) bounds for one Philippine calendar day."""
    day = day or ph_today()
    start_local = datetime.combine(day, time.min, tzinfo=MANILA_TZ)
    next_local = start_local + timedelta(days=1)
    return (
        start_local.astimezone(timezone.utc).replace(tzinfo=None),
        next_local.astimezone(timezone.utc).replace(tzinfo=None),
    )

def utc_naive_to_ph(value):
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(MANILA_TZ)

# ==================== DATA MODELS ====================

class StoreSetting(db.Model):
    __tablename__ = 'store_setting'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Text, nullable=False)

class InvestorInterest(db.Model):
    """Private proposal requests submitted by selected invitees."""
    __tablename__ = 'investor_interest'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    contact = db.Column(db.String(160), nullable=False)
    preferred_contact = db.Column(db.String(20), nullable=False, default='PHONE')
    business_area = db.Column(db.String(30), nullable=False, default='WHOLE_BUSINESS')
    funding_range = db.Column(db.String(40), nullable=False, default='DISCUSS_PRIVATELY')
    offer_code = db.Column(db.String(40), nullable=False, default='GENERAL')
    payout_option = db.Column(db.String(30), nullable=True)
    proposed_amount = db.Column(db.Float, nullable=True)
    monthly_rate_percent = db.Column(db.Float, nullable=True)
    term_months = db.Column(db.Integer, nullable=True)
    monthly_interest_amount = db.Column(db.Float, nullable=True)
    total_interest_amount = db.Column(db.Float, nullable=True)
    maturity_payment_amount = db.Column(db.Float, nullable=True)
    total_contract_amount = db.Column(db.Float, nullable=True)
    is_counter_offer = db.Column(db.Boolean, nullable=False, default=False)
    message = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), nullable=False, default='NEW')
    cashier_notes = db.Column(db.Text, nullable=True)
    cashier_reviewed_by = db.Column(db.String(50), nullable=True)
    cashier_reviewed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False, index=True)

class Staff(db.Model):
    __tablename__ = 'staff'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    pin_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    active = db.Column(db.Boolean, default=True)

class Customer(db.Model):
    __tablename__ = 'customer'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), nullable=True)
    contact = db.Column(db.String(30), unique=True, nullable=False)
    fb_messenger = db.Column(db.String(150), nullable=True)
    pin_hash = db.Column(db.String(255), nullable=False)
    profile_image = db.Column(db.Text, nullable=True)
    default_address = db.Column(db.Text, nullable=True)
    default_landmark = db.Column(db.String(150), nullable=True)
    points_balance = db.Column(db.Float, default=0.0)
    is_credit_eligible = db.Column(db.Boolean, default=False)
    credit_limit = db.Column(db.Float, default=0.0)
    outstanding_ar = db.Column(db.Float, default=0.0)
    accumulated_spend = db.Column(db.Float, default=0.0)
    card_number = db.Column(db.String(20), unique=True, nullable=True)
    card_status = db.Column(db.String(20), default="ACTIVE")
    card_expires_at = db.Column(db.Date, nullable=True)
    card_theme = db.Column(db.String(30), default="pink-classic")
    card_logo_scale = db.Column(db.Float, default=1.0)
    card_photo_scale = db.Column(db.Float, default=1.0)
    card_qr_scale = db.Column(db.Float, default=1.0)
    card_text_scale = db.Column(db.Float, default=1.0)
    card_info_scale = db.Column(db.Float, default=1.0)
    referred_by = db.Column(db.String(50), nullable=True)
    last_daily_login = db.Column(db.Date, nullable=True)
    login_streak = db.Column(db.Integer, default=1)
    wifi_voucher_code = db.Column(db.String(20), nullable=True)
    wifi_minutes_left = db.Column(db.Integer, default=0)
    last_active_at = db.Column(db.DateTime, default=utc_now)
    campus_name = db.Column(db.String(120), nullable=True)
    break_start = db.Column(db.String(5), nullable=True)
    break_end = db.Column(db.String(5), nullable=True)
    favorite_alerts = db.Column(db.Boolean, default=False)
    community_student_preapproved = db.Column(db.Boolean, default=False, nullable=False)
    community_student_preapproved_at = db.Column(db.DateTime, nullable=True)
    community_student_preapproved_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class DeliveryZone(db.Model):
    __tablename__ = 'delivery_zone'
    id = db.Column(db.Integer, primary_key=True)
    place_name = db.Column(db.String(100), nullable=False)
    barangay = db.Column(db.String(100), nullable=False)
    rate = db.Column(db.Float, nullable=False)
    distance = db.Column(db.String(50), nullable=True)
    note = db.Column(db.String(150), nullable=True)
    is_active = db.Column(db.Boolean, default=True)

class Category(db.Model):
    __tablename__ = 'category'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)

class Product(db.Model):
    __tablename__ = 'product'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    category_name = db.Column(db.String(80), nullable=False)
    price = db.Column(db.Float, nullable=False)
    cost = db.Column(db.Float, default=0.0, nullable=True)
    allow_custom_amount = db.Column(db.Boolean, default=False)
    minimum_order_amount = db.Column(db.Float, nullable=True)
    option_schema = db.Column(db.Text, nullable=True)
    size_schema = db.Column(db.Text, nullable=True)
    stock = db.Column(db.Integer, default=100)
    image_url = db.Column(db.Text, nullable=True)
    is_featured = db.Column(db.Boolean, default=False)
    is_top_seller = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    available_start_time = db.Column(db.String(10), nullable=True)
    available_end_time = db.Column(db.String(10), nullable=True)
    prep_minutes = db.Column(db.Integer, default=10)
    total_likes = db.Column(db.Integer, default=0)
    comments = db.relationship('ProductComment', backref='product_rel', cascade="all, delete-orphan", lazy=True)

class ProductLike(db.Model):
    __tablename__ = 'product_like'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='CASCADE'), nullable=False)
    ip_address = db.Column(db.String(50), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='SET NULL'), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class ProductComment(db.Model):
    __tablename__ = 'product_comment'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='CASCADE'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='SET NULL'), nullable=True)
    author_name = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(50), nullable=False)
    comment_text = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class Order(db.Model):
    __tablename__ = 'order'
    id = db.Column(db.Integer, primary_key=True)
    order_type = db.Column(db.String(30), nullable=False)
    dining_option = db.Column(db.String(20), default='DINE-IN')
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    customer_name = db.Column(db.String(100), default='Customer')
    contact_number = db.Column(db.String(50), default='N/A')
    fb_messenger = db.Column(db.String(150), nullable=True)
    delivery_address = db.Column(db.Text, nullable=True)
    landmark = db.Column(db.String(150), nullable=True)
    pickup_time = db.Column(db.String(50), nullable=True)
    target_time = db.Column(db.String(50), nullable=True)
    change_for = db.Column(db.Float, nullable=True)
    gcash_ref = db.Column(db.String(10), nullable=True)
    subtotal = db.Column(db.Float, nullable=False)
    delivery_fee = db.Column(db.Float, default=0.0)
    total_amount = db.Column(db.Float, nullable=False)
    points_redeemed = db.Column(db.Float, default=0.0)
    points_discount = db.Column(db.Float, default=0.0)
    payment_method = db.Column(db.String(20), nullable=False)
    payment_verified = db.Column(db.Boolean, default=False)
    status = db.Column(db.String(30), default="VERIFICATION")
    is_unpaid = db.Column(db.Boolean, default=False)
    collection_notes = db.Column(db.String(255), nullable=True)
    notes = db.Column(db.Text, nullable=False, default="None")
    public_token = db.Column(db.String(64), unique=True, nullable=True, default=lambda: secrets.token_urlsafe(24))
    fulfillment_status = db.Column(db.String(30), default='SUBMITTED')
    created_at = db.Column(db.DateTime, default=utc_now)
    customer = db.relationship('Customer', backref='orders', lazy=True)
    items = db.relationship('OrderItem', backref='order_rel', cascade="all, delete-orphan", lazy=True)

class OrderItem(db.Model):
    __tablename__ = 'order_item'
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=True)
    product_name = db.Column(db.String(120), nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    cost_price = db.Column(db.Float, default=0.0, nullable=True)
    quantity = db.Column(db.Integer, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    selected_options = db.Column(db.Text, nullable=True)

class Expense(db.Model):
    __tablename__ = 'expense'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(50), default='General')
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class CashFlowPlan(db.Model):
    __tablename__ = 'cash_flow_plan'
    id = db.Column(db.Integer, primary_key=True)
    entry_type = db.Column(db.String(20), nullable=False)  # EXPENSE or INCOME
    title = db.Column(db.String(150), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    frequency = db.Column(db.String(20), nullable=False)  # DAILY, WEEKLY, BIWEEKLY, MONTHLY
    start_date = db.Column(db.Date, nullable=False)
    duration_count = db.Column(db.Integer, nullable=False, default=1)  # 0 means indefinite recurrence
    category = db.Column(db.String(80), nullable=True)
    notes = db.Column(db.String(255), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class CashFlowExpensePayment(db.Model):
    __tablename__ = 'cash_flow_expense_payment'
    __table_args__ = (UniqueConstraint('plan_id', 'occurrence_date', name='uq_cashflow_expense_occurrence'),)
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('cash_flow_plan.id', ondelete='CASCADE'), nullable=False)
    occurrence_date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='PAYABLE', nullable=False)
    paid_at = db.Column(db.DateTime, nullable=True)
    paid_by = db.Column(db.String(50), nullable=True)
    payment_method = db.Column(db.String(20), nullable=True)
    reference = db.Column(db.String(100), nullable=True)
    notes = db.Column(db.String(255), nullable=True)
    updated_at = db.Column(db.DateTime, default=utc_now, onupdate=utc_now)
    plan = db.relationship('CashFlowPlan', lazy=True)

class VaultDrop(db.Model):
    __tablename__ = 'vault_drop'
    id = db.Column(db.Integer, primary_key=True)
    drop_number = db.Column(db.Integer, nullable=False)
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.String(255), nullable=True)
    cash_breakdown = db.Column(db.Text, nullable=True)
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class ChangeFund(db.Model):
    __tablename__ = 'change_fund'
    id = db.Column(db.Integer, primary_key=True)
    fund_title = db.Column(db.String(150), nullable=False, default="Cashier Opening Change Fund")
    amount = db.Column(db.Float, nullable=False)
    notes = db.Column(db.String(255), nullable=True)
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class RewardLedger(db.Model):
    __tablename__ = 'reward_ledger'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    points_change = db.Column(db.Float, nullable=False)
    reason = db.Column(db.String(150), nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class SiteVisitor(db.Model):
    __tablename__ = 'site_visitor'
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(50), unique=True, nullable=False)
    visit_count = db.Column(db.Integer, default=1)
    visited_at = db.Column(db.DateTime, default=utc_now)

class PromotionTracker(db.Model):
    __tablename__ = 'promotion_tracker'
    id = db.Column(db.Integer, primary_key=True)
    promo_code = db.Column(db.String(50), unique=True, nullable=False)
    title = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text, nullable=True)
    promo_price = db.Column(db.Float, nullable=False)
    promo_cost = db.Column(db.Float, default=0.0, nullable=True)
    page_views = db.Column(db.Integer, default=0)
    claims_count = db.Column(db.Integer, default=0)
    total_revenue = db.Column(db.Float, default=0.0)
    is_active = db.Column(db.Boolean, default=True)
    is_visible = db.Column(db.Boolean, default=True)
    portal_only = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=utc_now)

# Legacy historical tables retained for non-destructive compatibility with older deployments.
class CustomerWishlist(db.Model):
    __tablename__ = 'customer_wishlist'
    __table_args__ = (UniqueConstraint('customer_id', 'product_id', name='uq_customer_wishlist_product'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='CASCADE'), nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class MenuVote(db.Model):
    __tablename__ = 'menu_vote'
    __table_args__ = (UniqueConstraint('customer_id', 'product_id', 'period_key', name='uq_menu_vote_period'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='CASCADE'), nullable=False)
    period_key = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class MenuVoteCandidate(db.Model):
    __tablename__ = 'menu_vote_candidate'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    normalized_name = db.Column(db.String(140), unique=True, nullable=False)
    category_name = db.Column(db.String(80), nullable=False, default='Customer Requests')
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='SET NULL'), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class MenuPreferenceVote(db.Model):
    __tablename__ = 'menu_preference_vote'
    __table_args__ = (UniqueConstraint('customer_id', 'candidate_id', 'period_key', name='uq_menu_preference_vote_period'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('menu_vote_candidate.id', ondelete='CASCADE'), nullable=False)
    period_key = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)
    candidate = db.relationship('MenuVoteCandidate', backref=db.backref('preference_votes', lazy=True, cascade='all, delete-orphan'))

class EngagementClaim(db.Model):
    __tablename__ = 'engagement_claim'
    __table_args__ = (UniqueConstraint('customer_id', 'action_code', 'period_key', name='uq_engagement_claim_period'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    action_code = db.Column(db.String(50), nullable=False)
    period_key = db.Column(db.String(30), nullable=False)
    points_awarded = db.Column(db.Float, default=0.0, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class BonusCampaign(db.Model):
    __tablename__ = 'bonus_campaign'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text, nullable=True)
    bonus_points = db.Column(db.Float, nullable=False, default=0.0)
    points_multiplier = db.Column(db.Float, nullable=False, default=1.0)
    min_spend = db.Column(db.Float, nullable=False, default=0.0)
    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    start_time = db.Column(db.String(5), nullable=True)
    end_time = db.Column(db.String(5), nullable=True)
    weekdays = db.Column(db.String(30), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class BonusCampaignClaim(db.Model):
    __tablename__ = 'bonus_campaign_claim'
    __table_args__ = (UniqueConstraint('campaign_id', 'order_id', name='uq_bonus_campaign_order'),)
    id = db.Column(db.Integer, primary_key=True)
    campaign_id = db.Column(db.Integer, db.ForeignKey('bonus_campaign.id', ondelete='CASCADE'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='CASCADE'), nullable=False)
    points_awarded = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class ReferralReward(db.Model):
    __tablename__ = 'referral_reward'
    __table_args__ = (UniqueConstraint('referred_customer_id', name='uq_referral_referred_customer'),)
    id = db.Column(db.Integer, primary_key=True)
    referrer_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    referred_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    first_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True)
    referrer_points = db.Column(db.Float, default=0.0, nullable=False)
    referred_points = db.Column(db.Float, default=0.0, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class PortalEvent(db.Model):
    __tablename__ = 'portal_event'
    id = db.Column(db.Integer, primary_key=True)
    source = db.Column(db.String(30), nullable=True)
    event_type = db.Column(db.String(50), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='SET NULL'), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class ProductSuggestion(db.Model):
    __tablename__ = 'product_suggestion'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='SET NULL'), nullable=True)
    customer_name = db.Column(db.String(100), nullable=False)
    suggestion_text = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), nullable=False, default='NEW')
    created_at = db.Column(db.DateTime, default=utc_now)

SUGGESTION_STATUSES = ('NEW', 'CONSIDERING', 'PLANNED', 'AVAILABLE', 'ARCHIVED')

# ==================== MACLEEN'S COMMUNITY ====================

class CommunityProfile(db.Model):
    """Optional public identity layered on top of a private loyalty account."""
    __tablename__ = 'community_profile'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), unique=True, nullable=False)
    handle = db.Column(db.String(32), unique=True, nullable=False, index=True)
    role = db.Column(db.String(20), nullable=False)  # STUDENT or RESIDENT
    campus_name = db.Column(db.String(120), nullable=True)
    department = db.Column(db.String(80), nullable=True)
    graduating_year = db.Column(db.Integer, nullable=True)
    vibe_status = db.Column(db.String(40), nullable=True)
    barangay = db.Column(db.String(80), nullable=True)
    resident_since_year = db.Column(db.Integer, nullable=True)
    verification_status = db.Column(db.String(20), nullable=False, default='PENDING')
    verification_method = db.Column(db.String(30), nullable=False, default='IN_PERSON')
    verification_note = db.Column(db.String(255), nullable=True)
    is_community_admin = db.Column(db.Boolean, nullable=False, default=False)
    public_bio = db.Column(db.String(160), nullable=True)
    is_profile_locked = db.Column(db.Boolean, nullable=False, default=False)
    student_id_image_data = db.Column(db.Text, nullable=True)
    student_id_uploaded_at = db.Column(db.DateTime, nullable=True)
    student_id_deleted_at = db.Column(db.DateTime, nullable=True)
    student_application_status = db.Column(db.String(20), nullable=True)
    student_application_campus = db.Column(db.String(120), nullable=True)
    student_application_department = db.Column(db.String(80), nullable=True)
    student_application_graduating_year = db.Column(db.Integer, nullable=True)
    first_post_approved = db.Column(db.Boolean, nullable=False, default=False)
    community_score = db.Column(db.Float, nullable=False, default=0.0)
    community_streak = db.Column(db.Integer, nullable=False, default=0)
    last_checkin_date = db.Column(db.Date, nullable=True)
    push_opt_in = db.Column(db.Boolean, nullable=False, default=False)
    privacy_consent_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    terms_accepted_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    customer = db.relationship('Customer', backref=db.backref('community_profile', uselist=False, cascade='all, delete-orphan'))

class CommunityPost(db.Model):
    __tablename__ = 'community_post'
    id = db.Column(db.Integer, primary_key=True)
    author_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    reshared_post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='SET NULL'), nullable=True, index=True)
    channel = db.Column(db.String(20), nullable=False)  # CAMPUS or TOWN
    module = db.Column(db.String(40), nullable=False)
    post_type = db.Column(db.String(20), nullable=False, default='TEXT')
    body = db.Column(db.String(280), nullable=False)
    image_data = db.Column(db.Text, nullable=True)
    link_url = db.Column(db.String(500), nullable=True)
    status = db.Column(db.String(24), nullable=False, default='PENDING')
    flags_count = db.Column(db.Integer, nullable=False, default=0)
    moderation_hits = db.Column(db.Text, nullable=True)
    score_awarded = db.Column(db.Boolean, nullable=False, default=False)
    is_flash_poll = db.Column(db.Boolean, nullable=False, default=False)
    publish_date = db.Column(db.Date, nullable=True)
    expires_at = db.Column(db.DateTime, nullable=True)
    published_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now, index=True)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    author = db.relationship('CommunityProfile', lazy=True)
    reshared_post = db.relationship('CommunityPost', remote_side=[id], foreign_keys=[reshared_post_id], lazy=True)
    poll_options = db.relationship('CommunityPollOption', backref='post', cascade='all, delete-orphan', lazy=True, order_by='CommunityPollOption.sort_order')

class CommunityPollOption(db.Model):
    __tablename__ = 'community_poll_option'
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    option_text = db.Column(db.String(80), nullable=False)
    sort_order = db.Column(db.Integer, nullable=False, default=0)

class CommunityPollVote(db.Model):
    __tablename__ = 'community_poll_vote'
    __table_args__ = (UniqueConstraint('post_id', 'customer_id', name='uq_community_poll_customer'),)
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    option_id = db.Column(db.Integer, db.ForeignKey('community_poll_option.id', ondelete='CASCADE'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    role_snapshot = db.Column(db.String(20), nullable=False)
    score_awarded = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityReaction(db.Model):
    __tablename__ = 'community_reaction'
    __table_args__ = (UniqueConstraint('post_id', 'customer_id', name='uq_community_reaction_customer'),)
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    reaction_type = db.Column(db.String(20), nullable=False, default='LIKE')
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityFollow(db.Model):
    __tablename__ = 'community_follow'
    __table_args__ = (UniqueConstraint('follower_profile_id', 'followed_profile_id', name='uq_community_follow_pair'),)
    id = db.Column(db.Integer, primary_key=True)
    follower_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    followed_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityEngagementReward(db.Model):
    """One-time loyalty award receipt; unfollowing/unliking never resets eligibility."""
    __tablename__ = 'community_engagement_reward'
    __table_args__ = (UniqueConstraint('customer_id', 'event_type', 'target_key', name='uq_community_engagement_reward'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False, index=True)
    event_type = db.Column(db.String(30), nullable=False)
    target_key = db.Column(db.String(80), nullable=False)
    points_awarded = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityMention(db.Model):
    __tablename__ = 'community_mention'
    __table_args__ = (UniqueConstraint('post_id', 'mentioned_profile_id', name='uq_community_post_mention'),)
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    mentioned_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityNotification(db.Model):
    __tablename__ = 'community_notification'
    __table_args__ = (UniqueConstraint('recipient_profile_id', 'kind', 'target_key', name='uq_community_notification_target'),)
    id = db.Column(db.Integer, primary_key=True)
    recipient_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    actor_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    kind = db.Column(db.String(30), nullable=False)
    target_key = db.Column(db.String(80), nullable=False)
    message = db.Column(db.String(180), nullable=False)
    is_read = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityAdminNotice(db.Model):
    """Private Community Admin inbox event; never rendered to members."""
    __tablename__ = 'community_admin_notice'
    __table_args__ = (UniqueConstraint('kind', 'target_key', name='uq_community_admin_notice_target'),)
    id = db.Column(db.Integer, primary_key=True)
    profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True, index=True)
    kind = db.Column(db.String(40), nullable=False)
    target_key = db.Column(db.String(80), nullable=False)
    message = db.Column(db.String(200), nullable=False)
    is_read = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now, index=True)
    profile = db.relationship('CommunityProfile', lazy=True)

class CommunityGroup(db.Model):
    __tablename__ = 'community_group'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    channel = db.Column(db.String(20), nullable=False)
    created_by_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True, index=True)
    external_chat_url = db.Column(db.String(500), nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    creator = db.relationship('CommunityProfile', foreign_keys=[created_by_profile_id], lazy=True)

class CommunityGroupMember(db.Model):
    __tablename__ = 'community_group_member'
    __table_args__ = (UniqueConstraint('group_id', 'profile_id', name='uq_community_group_profile'),)
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('community_group.id', ondelete='CASCADE'), nullable=False, index=True)
    profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    invited_by_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    member_role = db.Column(db.String(20), nullable=False, default='MEMBER')
    status = db.Column(db.String(20), nullable=False, default='INVITED')
    joined_at = db.Column(db.DateTime, nullable=True)
    last_read_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    group = db.relationship('CommunityGroup', backref=db.backref('memberships', cascade='all, delete-orphan', lazy=True))
    profile = db.relationship('CommunityProfile', foreign_keys=[profile_id], lazy=True)
    invited_by = db.relationship('CommunityProfile', foreign_keys=[invited_by_profile_id], lazy=True)

class CommunityGroupMessage(db.Model):
    __tablename__ = 'community_group_message'
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('community_group.id', ondelete='CASCADE'), nullable=False, index=True)
    author_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True, index=True)
    body = db.Column(db.String(1000), nullable=False)
    status = db.Column(db.String(24), nullable=False, default='PUBLISHED')
    moderation_hits = db.Column(db.Text, nullable=True)
    flags_count = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now, index=True)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    group = db.relationship('CommunityGroup', lazy=True)
    author = db.relationship('CommunityProfile', lazy=True)

class CommunityGroupMessageReport(db.Model):
    __tablename__ = 'community_group_message_report'
    __table_args__ = (UniqueConstraint('message_id', 'reporter_profile_id', name='uq_community_group_message_reporter'),)
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.Integer, db.ForeignKey('community_group_message.id', ondelete='CASCADE'), nullable=False, index=True)
    reporter_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    reason = db.Column(db.String(40), nullable=False)
    details = db.Column(db.String(240), nullable=True)
    status = db.Column(db.String(20), nullable=False, default='OPEN')
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityGroupTask(db.Model):
    __tablename__ = 'community_group_task'
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('community_group.id', ondelete='CASCADE'), nullable=False, index=True)
    created_by_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    assigned_to_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True, index=True)
    title = db.Column(db.String(120), nullable=False)
    details = db.Column(db.String(500), nullable=True)
    priority = db.Column(db.String(16), nullable=False, default='NORMAL')
    status = db.Column(db.String(16), nullable=False, default='TODO')
    due_date = db.Column(db.Date, nullable=True)
    completed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    creator = db.relationship('CommunityProfile', foreign_keys=[created_by_profile_id], lazy=True)
    assignee = db.relationship('CommunityProfile', foreign_keys=[assigned_to_profile_id], lazy=True)

class CommunityGroupNote(db.Model):
    __tablename__ = 'community_group_note'
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('community_group.id', ondelete='CASCADE'), nullable=False, index=True)
    author_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    title = db.Column(db.String(120), nullable=False)
    body = db.Column(db.String(2000), nullable=False)
    is_pinned = db.Column(db.Boolean, nullable=False, default=False)
    status = db.Column(db.String(16), nullable=False, default='ACTIVE')
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)
    author = db.relationship('CommunityProfile', lazy=True)

class CommunityGroupPoll(db.Model):
    __tablename__ = 'community_group_poll'
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('community_group.id', ondelete='CASCADE'), nullable=False, index=True)
    author_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    question = db.Column(db.String(240), nullable=False)
    status = db.Column(db.String(16), nullable=False, default='OPEN')
    closes_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    author = db.relationship('CommunityProfile', lazy=True)
    options = db.relationship('CommunityGroupPollOption', backref='poll', cascade='all, delete-orphan', lazy=True, order_by='CommunityGroupPollOption.sort_order')

class CommunityGroupPollOption(db.Model):
    __tablename__ = 'community_group_poll_option'
    id = db.Column(db.Integer, primary_key=True)
    poll_id = db.Column(db.Integer, db.ForeignKey('community_group_poll.id', ondelete='CASCADE'), nullable=False, index=True)
    option_text = db.Column(db.String(100), nullable=False)
    sort_order = db.Column(db.Integer, nullable=False, default=0)

class CommunityGroupPollVote(db.Model):
    __tablename__ = 'community_group_poll_vote'
    __table_args__ = (UniqueConstraint('poll_id', 'profile_id', name='uq_community_group_poll_profile'),)
    id = db.Column(db.Integer, primary_key=True)
    poll_id = db.Column(db.Integer, db.ForeignKey('community_group_poll.id', ondelete='CASCADE'), nullable=False, index=True)
    option_id = db.Column(db.Integer, db.ForeignKey('community_group_poll_option.id', ondelete='CASCADE'), nullable=False)
    profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityConnection(db.Model):
    """Mutual opt-in connection used to protect friend-only vibe status."""
    __tablename__ = 'community_connection'
    __table_args__ = (UniqueConstraint('profile_a_id', 'profile_b_id', name='uq_community_connection_pair'),)
    id = db.Column(db.Integer, primary_key=True)
    profile_a_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    profile_b_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False, index=True)
    requested_by_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='CASCADE'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='PENDING')
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    responded_at = db.Column(db.DateTime, nullable=True)

class CommunityComment(db.Model):
    __tablename__ = 'community_comment'
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    author_profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    body = db.Column(db.String(280), nullable=False)
    status = db.Column(db.String(24), nullable=False, default='PUBLISHED')
    moderation_hits = db.Column(db.Text, nullable=True)
    score_awarded = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    author = db.relationship('CommunityProfile', lazy=True)

class CommunityReport(db.Model):
    __tablename__ = 'community_report'
    __table_args__ = (UniqueConstraint('post_id', 'reporter_customer_id', name='uq_community_reporter_post'),)
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='CASCADE'), nullable=False, index=True)
    reporter_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    reason = db.Column(db.String(40), nullable=False)
    details = db.Column(db.String(240), nullable=True)
    status = db.Column(db.String(20), nullable=False, default='OPEN')
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityKeyword(db.Model):
    __tablename__ = 'community_keyword'
    id = db.Column(db.Integer, primary_key=True)
    phrase = db.Column(db.String(100), unique=True, nullable=False)
    category = db.Column(db.String(40), nullable=False, default='ABUSE')
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityAd(db.Model):
    __tablename__ = 'community_ad'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(120), nullable=False)
    body = db.Column(db.String(240), nullable=False)
    image_url = db.Column(db.Text, nullable=True)
    target_role = db.Column(db.String(20), nullable=False, default='ALL')
    channel = db.Column(db.String(20), nullable=False, default='ALL')
    cta_label = db.Column(db.String(40), nullable=False, default='View menu')
    cta_url = db.Column(db.String(500), nullable=False, default='/')
    start_at = db.Column(db.DateTime, nullable=True)
    end_at = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    impression_count = db.Column(db.Integer, nullable=False, default=0)
    click_count = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityAlert(db.Model):
    __tablename__ = 'community_alert'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    body = db.Column(db.String(240), nullable=False)
    target_role = db.Column(db.String(20), nullable=False, default='ALL')
    cta_url = db.Column(db.String(500), nullable=True)
    starts_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    ends_at = db.Column(db.DateTime, nullable=False)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityCheckin(db.Model):
    __tablename__ = 'community_checkin'
    __table_args__ = (UniqueConstraint('customer_id', 'checkin_date', name='uq_community_daily_checkin'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    checkin_date = db.Column(db.Date, nullable=False)
    streak_day = db.Column(db.Integer, nullable=False)
    score_awarded = db.Column(db.Float, nullable=False, default=1.0)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityStoreCheckin(db.Model):
    """A verified visit created only from one completed paid member order."""
    __tablename__ = 'community_store_checkin'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False, index=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='CASCADE'), unique=True, nullable=False)
    checkin_date = db.Column(db.Date, nullable=False, index=True)
    recorded_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class CommunityDrop(db.Model):
    __tablename__ = 'community_drop'
    __table_args__ = (UniqueConstraint('customer_id', 'milestone_day', 'period_key', name='uq_community_drop_milestone'),)
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    milestone_day = db.Column(db.Integer, nullable=False)
    period_key = db.Column(db.String(30), nullable=False)
    reward_type = db.Column(db.String(30), nullable=False)
    reward_title = db.Column(db.String(120), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='SET NULL'), nullable=True)
    status = db.Column(db.String(24), nullable=False, default='ACTIVE')
    approved_by = db.Column(db.String(50), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    redeemed_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True)
    points_awarded = db.Column(db.Float, nullable=False, default=0.0)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    redeemed_at = db.Column(db.DateTime, nullable=True)
    customer = db.relationship('Customer', lazy=True)
    product = db.relationship('Product', lazy=True)

class CommunityGift(db.Model):
    __tablename__ = 'community_gift'
    id = db.Column(db.Integer, primary_key=True)
    sender_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    recipient_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    gift_type = db.Column(db.String(20), nullable=False)
    points_amount = db.Column(db.Float, nullable=False, default=0.0)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='SET NULL'), nullable=True)
    product_name = db.Column(db.String(120), nullable=True)
    note = db.Column(db.String(120), nullable=True)
    claim_code = db.Column(db.String(24), unique=True, nullable=False)
    status = db.Column(db.String(20), nullable=False, default='SENT')
    claimed_by = db.Column(db.String(50), nullable=True)
    claimed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now, index=True)
    sender = db.relationship('Customer', foreign_keys=[sender_customer_id], lazy=True)
    recipient = db.relationship('Customer', foreign_keys=[recipient_customer_id], lazy=True)
    product = db.relationship('Product', lazy=True)

class CommunityPushSubscription(db.Model):
    __tablename__ = 'community_push_subscription'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    endpoint = db.Column(db.Text, unique=True, nullable=False)
    p256dh = db.Column(db.Text, nullable=False)
    auth = db.Column(db.Text, nullable=False)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)

class CommunityModerationAction(db.Model):
    __tablename__ = 'community_moderation_action'
    id = db.Column(db.Integer, primary_key=True)
    post_id = db.Column(db.Integer, db.ForeignKey('community_post.id', ondelete='SET NULL'), nullable=True)
    profile_id = db.Column(db.Integer, db.ForeignKey('community_profile.id', ondelete='SET NULL'), nullable=True)
    admin_username = db.Column(db.String(50), nullable=False)
    action = db.Column(db.String(40), nullable=False)
    note = db.Column(db.String(240), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=utc_now)

class GroupOrder(db.Model):
    __tablename__ = 'group_order'
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(64), unique=True, nullable=False, default=lambda: secrets.token_urlsafe(20))
    organizer_customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='CASCADE'), nullable=False)
    title = db.Column(db.String(120), nullable=False, default='Barkada Order')
    status = db.Column(db.String(20), nullable=False, default='OPEN')
    submitted_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)
    submitted_at = db.Column(db.DateTime, nullable=True)
    organizer = db.relationship('Customer', lazy=True)
    lines = db.relationship('GroupOrderLine', backref='group_order', cascade='all, delete-orphan', lazy=True)

class GroupOrderLine(db.Model):
    __tablename__ = 'group_order_line'
    id = db.Column(db.Integer, primary_key=True)
    group_order_id = db.Column(db.Integer, db.ForeignKey('group_order.id', ondelete='CASCADE'), nullable=False)
    participant_name = db.Column(db.String(100), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id', ondelete='CASCADE'), nullable=False)
    product_name = db.Column(db.String(120), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=1)
    unit_price = db.Column(db.Float, nullable=False)
    selected_options = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)
    product = db.relationship('Product', lazy=True)


# ==================== INTEGRATED MACLEEN'S CRAFT SHOP ====================

class CraftCategory(db.Model):
    __tablename__ = 'craft_category'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)
    image_url = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)

class CraftItem(db.Model):
    __tablename__ = 'craft_item'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text, nullable=False, default='')
    category_name = db.Column(db.String(80), nullable=False, default='General')
    price = db.Column(db.Float, nullable=False)
    cost = db.Column(db.Float, nullable=False, default=0.0)
    image_url = db.Column(db.Text, nullable=True)
    availability_type = db.Column(db.String(20), default='IN_STOCK', nullable=False)  # IN_STOCK / PREORDER
    stock_quantity = db.Column(db.Integer, default=0, nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    is_top_seller = db.Column(db.Boolean, default=False, nullable=False)
    is_featured = db.Column(db.Boolean, default=False, nullable=False)
    likes = db.Column(db.Integer, default=0, nullable=False)
    views = db.Column(db.Integer, default=0, nullable=False)
    orders_count = db.Column(db.Integer, default=0, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)
    comments = db.relationship('CraftComment', backref='craft_item', cascade='all, delete-orphan', lazy=True)
    craft_orders = db.relationship('CraftOrder', backref='craft_item', lazy=True)

class CraftSiteVisitor(db.Model):
    __tablename__ = 'craft_site_visitor'
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(64), unique=True, nullable=False, index=True)
    visit_count = db.Column(db.Integer, default=1, nullable=False)
    first_seen_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    last_seen_at = db.Column(db.DateTime, default=utc_now, nullable=False)

class CraftItemView(db.Model):
    __tablename__ = 'craft_item_view'
    __table_args__ = (UniqueConstraint('item_id', 'ip_address', name='uq_craft_item_view_ip'),)
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('craft_item.id', ondelete='CASCADE'), nullable=False, index=True)
    ip_address = db.Column(db.String(64), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)

class CraftItemLike(db.Model):
    __tablename__ = 'craft_item_like'
    __table_args__ = (UniqueConstraint('item_id', 'ip_address', name='uq_craft_item_like_ip'),)
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('craft_item.id', ondelete='CASCADE'), nullable=False, index=True)
    ip_address = db.Column(db.String(64), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)

class CraftComment(db.Model):
    __tablename__ = 'craft_comment'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('craft_item.id', ondelete='CASCADE'), nullable=False)
    author = db.Column(db.String(80), nullable=False)
    content = db.Column(db.Text, nullable=False)
    ip_address = db.Column(db.String(64), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class CraftOrder(db.Model):
    __tablename__ = 'craft_order'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('craft_item.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id', ondelete='SET NULL'), nullable=True)
    customer_name = db.Column(db.String(100), nullable=False)
    contact_number = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120), nullable=True)
    fb_account = db.Column(db.String(150), nullable=True)
    quantity = db.Column(db.Integer, default=1, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    unit_cost = db.Column(db.Float, nullable=False, default=0.0)
    total_price = db.Column(db.Float, nullable=False)
    total_cost = db.Column(db.Float, nullable=False, default=0.0)
    payment_method = db.Column(db.String(20), nullable=False, default='CASH')
    payment_status = db.Column(db.String(20), nullable=False, default='PENDING')
    gcash_ref = db.Column(db.String(20), nullable=True)
    status = db.Column(db.String(30), nullable=False, default='PENDING')
    pickup_location = db.Column(db.String(150), default="Macleen's Food House")
    notes = db.Column(db.Text, nullable=True)
    main_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True, unique=True)
    stock_restored = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=utc_now)
    completed_at = db.Column(db.DateTime, nullable=True)
    customer = db.relationship('Customer', lazy=True)
    main_order = db.relationship('Order', lazy=True, foreign_keys=[main_order_id])

class CraftLedger(db.Model):
    __tablename__ = 'craft_ledger'
    id = db.Column(db.Integer, primary_key=True)
    event_type = db.Column(db.String(30), nullable=False)  # SALE / EXPENSE / OTHER_INCOME / REFUND
    title = db.Column(db.String(150), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), nullable=True)
    craft_order_id = db.Column(db.Integer, db.ForeignKey('craft_order.id', ondelete='SET NULL'), nullable=True)
    main_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True)
    expense_id = db.Column(db.Integer, db.ForeignKey('expense.id', ondelete='SET NULL'), nullable=True)
    notes = db.Column(db.String(255), nullable=True)
    created_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

class DigitalCategory(db.Model):
    __tablename__ = 'digital_category'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)

class DigitalItem(db.Model):
    __tablename__ = 'digital_item'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text, nullable=True)
    category_name = db.Column(db.String(80), default='General', nullable=False)
    product_type = db.Column(db.String(30), default='DOWNLOAD', nullable=False)
    price = db.Column(db.Float, nullable=False)
    cost = db.Column(db.Float, default=0.0)
    image_url = db.Column(db.Text, nullable=True)
    sample_url = db.Column(db.Text, nullable=True)
    file_format = db.Column(db.String(80), nullable=True)
    license_terms = db.Column(db.Text, nullable=True)
    turnaround_days = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    is_featured = db.Column(db.Boolean, default=False)
    views = db.Column(db.Integer, default=0)
    orders_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=utc_now)

class DigitalOrder(db.Model):
    __tablename__ = 'digital_order'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('digital_item.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    customer_name = db.Column(db.String(100), nullable=False)
    contact_number = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    quantity = db.Column(db.Integer, default=1, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    unit_cost = db.Column(db.Float, default=0.0)
    total_price = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(20), nullable=False)
    payment_status = db.Column(db.String(20), default='PENDING')
    gcash_ref = db.Column(db.String(30), nullable=True)
    status = db.Column(db.String(30), default='PENDING_PAYMENT', nullable=False)
    requirements = db.Column(db.Text, nullable=True)
    fulfillment_url = db.Column(db.Text, nullable=True)
    license_key = db.Column(db.String(255), nullable=True)
    fulfillment_notes = db.Column(db.Text, nullable=True)
    tracking_token = db.Column(db.String(64), unique=True, nullable=False, default=lambda: secrets.token_urlsafe(24))
    main_order_id = db.Column(db.Integer, db.ForeignKey('order.id', ondelete='SET NULL'), nullable=True, unique=True)
    created_at = db.Column(db.DateTime, default=utc_now)
    completed_at = db.Column(db.DateTime, nullable=True)
    item = db.relationship('DigitalItem', lazy=True)
    main_order = db.relationship('Order', lazy=True, foreign_keys=[main_order_id])

class MarketingInsightImport(db.Model):
    __tablename__ = 'marketing_insight_import'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    row_count = db.Column(db.Integer, default=0)
    summary_json = db.Column(db.Text, nullable=False)
    analysis = db.Column(db.Text, nullable=False)
    uploaded_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)


class MarketingFacebookPage(db.Model):
    __tablename__ = 'marketing_facebook_page'
    id = db.Column(db.Integer, primary_key=True)
    page_id = db.Column(db.String(80), unique=True, nullable=False)
    page_name = db.Column(db.String(150), nullable=False)
    access_token_encrypted = db.Column(db.Text, nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    connected_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    token_updated_at = db.Column(db.DateTime, default=utc_now, nullable=False)

class MarketingGroup(db.Model):
    __tablename__ = 'marketing_group'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    group_url = db.Column(db.Text, nullable=False)
    business_scope = db.Column(db.String(20), default='BOTH', nullable=False)
    post_types = db.Column(db.String(255), nullable=True)
    cooldown_days = db.Column(db.Integer, default=7, nullable=False)
    notes = db.Column(db.Text, nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    last_posted_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)

class MarketingPost(db.Model):
    __tablename__ = 'marketing_post'
    id = db.Column(db.Integer, primary_key=True)
    target_type = db.Column(db.String(30), nullable=False, default='FACEBOOK_PAGE')
    business = db.Column(db.String(20), nullable=False)
    post_type = db.Column(db.String(50), nullable=False)
    source_kind = db.Column(db.String(30), nullable=False, default='PAGE')
    source_id = db.Column(db.Integer, nullable=True)
    caption = db.Column(db.Text, nullable=False)
    reason = db.Column(db.Text, nullable=True)
    link_url = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(30), nullable=False, default='DRAFT')
    ai_model = db.Column(db.String(80), nullable=True)
    facebook_post_id = db.Column(db.String(150), nullable=True)
    group_id = db.Column(db.Integer, db.ForeignKey('marketing_group.id', ondelete='SET NULL'), nullable=True)
    error_message = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    approved_at = db.Column(db.DateTime, nullable=True)
    published_at = db.Column(db.DateTime, nullable=True)
    insight_reach = db.Column(db.Integer, default=0)
    insight_impressions = db.Column(db.Integer, default=0)
    insight_reactions = db.Column(db.Integer, default=0)
    insight_comments = db.Column(db.Integer, default=0)
    insight_shares = db.Column(db.Integer, default=0)
    insight_saves = db.Column(db.Integer, default=0)
    insight_link_clicks = db.Column(db.Integer, default=0)
    insight_new_followers = db.Column(db.Integer, default=0)
    insight_spend = db.Column(db.Float, default=0.0)
    insight_notes = db.Column(db.Text, nullable=True)
    insight_analysis = db.Column(db.Text, nullable=True)
    insight_ai_model = db.Column(db.String(100), nullable=True)
    insights_updated_at = db.Column(db.DateTime, nullable=True)
    insights_analyzed_at = db.Column(db.DateTime, nullable=True)
    group = db.relationship('MarketingGroup', lazy=True)


class FacebookMenuRun(db.Model):
    """One idempotent daily-menu action per channel and Philippine calendar day."""
    __tablename__ = 'facebook_menu_run'
    __table_args__ = (UniqueConstraint('menu_date', 'channel', name='uq_facebook_menu_run_day_channel'),)
    id = db.Column(db.Integer, primary_key=True)
    menu_date = db.Column(db.Date, nullable=False)
    channel = db.Column(db.String(30), nullable=False, default='PAGE_POST')
    status = db.Column(db.String(20), nullable=False, default='DRAFT')
    caption = db.Column(db.Text, nullable=False)
    external_id = db.Column(db.String(180), nullable=True)
    error_message = db.Column(db.Text, nullable=True)
    triggered_by = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    sent_at = db.Column(db.DateTime, nullable=True)


class MessengerContact(db.Model):
    """A Page-scoped Messenger contact learned only after that person messages the Page."""
    __tablename__ = 'messenger_contact'
    id = db.Column(db.Integer, primary_key=True)
    psid = db.Column(db.String(180), unique=True, nullable=False)
    source = db.Column(db.String(40), default='INBOUND', nullable=False)
    opted_out = db.Column(db.Boolean, default=False, nullable=False)
    first_seen_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    last_interaction_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    last_menu_sent_at = db.Column(db.DateTime, nullable=True)


class MessengerDelivery(db.Model):
    """Auditable Messenger reply delivery/read/click status without storing message text."""
    __tablename__ = 'messenger_delivery'
    id = db.Column(db.Integer, primary_key=True)
    contact_id = db.Column(db.Integer, db.ForeignKey('messenger_contact.id', ondelete='CASCADE'), nullable=False)
    message_kind = db.Column(db.String(30), nullable=False, default='MENU_REPLY')
    message_id = db.Column(db.String(180), nullable=True)
    tracking_token = db.Column(db.String(64), unique=True, nullable=False, default=lambda: secrets.token_urlsafe(24))
    status = db.Column(db.String(20), nullable=False, default='QUEUED')
    error_message = db.Column(db.Text, nullable=True)
    sent_at = db.Column(db.DateTime, nullable=True)
    delivered_at = db.Column(db.DateTime, nullable=True)
    read_at = db.Column(db.DateTime, nullable=True)
    clicked_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    contact = db.relationship('MessengerContact', lazy=True)

CRAFT_ORDER_STATUSES = ('PENDING', 'READY', 'COMPLETED', 'CANCELLED')
CRAFT_PAYMENT_METHODS = ('CASH', 'GCASH')
CRAFT_DEFAULT_IMAGE = '/static/craft/default-craft.png'

# Catalog imported from the previous public Macleen's Crafts storefront.
# This is intentionally an idempotent seed: existing items are not overwritten.
LEGACY_CRAFT_SOURCE_URL = 'https://macleens-crafts.onrender.com/'
LEGACY_CRAFT_CATALOG = (
    {
        'name': 'Flower Keychain - Yellow',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 2,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/744917655_122094499527400948_430343626372547796_n.jpg',
    },
    {
        'name': 'Flower Keychain - Violet',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/782512923_122112516585400948_7338110726271509685_n.jpg',
    },
    {
        'name': 'Flower Keychain - Blue',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/783256072_122112516717400948_1064296111081603629_n.jpg',
    },
    {
        'name': 'Octopus Keychain - Yellow',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 50.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/745123653_122094499209400948_8154867140069997677_n.jpg',
    },
    {
        'name': 'Octopus Keychain - Blue',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 50.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/744955113_122094499077400948_6511059889199350008_n.jpg',
    },
    {
        'name': 'Strawberry Keychain - Red',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 3,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/782148124_122112516669400948_7137228521426484421_n.jpg',
    },
    {
        'name': 'Strawberry Keychain - Yellow',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 2,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/783591753_122112516627400948_7432897152173266573_n.jpg',
    },
    {
        'name': '1L Tumbler Holder - Violet',
        'description': 'Handknitted',
        'category_name': 'Tumbler Holder',
        'price': 200.0,
        'availability_type': 'PREORDER',
        'stock_quantity': 0,
        'likes': 0,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/743588778_122094498951400948_8539128599543056976_n_1.jpg',
    },
    {
        'name': 'Strawhat by Luffy Keychain',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 50.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 2,
        'likes': 22,
        'is_featured': True,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/745477083_122094499455400948_7413700848929646075_n.jpg',
    },
    {
        'name': 'Flower Keychain - Pink',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 40.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 30,
        'is_featured': False,
        'is_top_seller': True,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/745503684_122094499635400948_5860449721293184690_n.jpg',
    },
    {
        'name': 'Octopus Keychain - Pink',
        'description': 'Handknitted',
        'category_name': 'Handknitted Crochet Keychains',
        'price': 50.0,
        'availability_type': 'IN_STOCK',
        'stock_quantity': 1,
        'likes': 4,
        'is_featured': False,
        'is_top_seller': False,
        'image_url': 'https://macleens-crafts.onrender.com/static/uploads/745556386_122094499143400948_8005875496749238542_n.jpg',
    },
)


def ensure_legacy_craft_catalog():
    """Seed the current Craft Shop with the catalog visible on the former standalone storefront.

    Existing items are preserved. Newly seeded costs are 0 because the public source does not expose COGS.
    """
    category_names = sorted({entry['category_name'] for entry in LEGACY_CRAFT_CATALOG})
    for category_name in category_names:
        existing_category = CraftCategory.query.filter(db.func.lower(CraftCategory.name) == category_name.lower()).first()
        if not existing_category:
            first_image = next((x['image_url'] for x in LEGACY_CRAFT_CATALOG if x['category_name'] == category_name), CRAFT_DEFAULT_IMAGE)
            db.session.add(CraftCategory(name=category_name, image_url=first_image, is_active=True))

    imported = 0
    for entry in LEGACY_CRAFT_CATALOG:
        existing = CraftItem.query.filter(db.func.lower(CraftItem.name) == entry['name'].lower()).first()
        if existing:
            # Preserve live/current admin edits. Only repair obviously blank/default metadata and retain higher engagement counts.
            if not (existing.image_url or '').strip() or existing.image_url == CRAFT_DEFAULT_IMAGE:
                existing.image_url = entry['image_url']
            if not (existing.description or '').strip():
                existing.description = entry['description']
            if not (existing.category_name or '').strip() or existing.category_name == 'General':
                existing.category_name = entry['category_name']
            existing.likes = max(parse_int(existing.likes, 0), parse_int(entry.get('likes'), 0))
            existing.is_featured = bool(existing.is_featured or entry.get('is_featured'))
            existing.is_top_seller = bool(existing.is_top_seller or entry.get('is_top_seller'))
            continue

        db.session.add(CraftItem(
            name=entry['name'],
            description=entry['description'],
            category_name=entry['category_name'],
            price=entry['price'],
            cost=0.0,
            image_url=entry['image_url'],
            availability_type=entry['availability_type'],
            stock_quantity=entry['stock_quantity'],
            is_active=True,
            is_top_seller=entry.get('is_top_seller', False),
            is_featured=entry.get('is_featured', False),
            likes=entry.get('likes', 0),
            views=0,
            orders_count=0,
        ))
        imported += 1

    db.session.commit()
    if imported:
        app.logger.info('Imported %s legacy Craft Shop catalog item(s) into the unified system.', imported)

def ensure_community_defaults():
    """Add conservative moderation phrases and role-targeted house banners once."""
    if CommunityKeyword.query.count() == 0:
        defaults = (
            ('kill yourself', 'THREAT_OR_HARASSMENT'),
            ('putang ina', 'PROFANITY'),
            ('gago', 'PROFANITY_OR_HARASSMENT'),
            ('bobo', 'HARASSMENT'),
            ('ulol', 'HARASSMENT'),
        )
        for phrase, category in defaults:
            db.session.add(CommunityKeyword(phrase=phrase, category=category, is_active=True))
    if CommunityAd.query.count() == 0:
        db.session.add(CommunityAd(
            title='Study break, sorted.',
            body='Budget-friendly meals, coffee, and printing support for busy campus days.',
            target_role='STUDENT',
            channel='CAMPUS',
            cta_label='See today’s menu',
            cta_url='/',
            is_active=True,
        ))
        db.session.add(CommunityAd(
            title='Comfort food for the neighborhood.',
            body='Morning brews, affordable ulam, snacks, and advance food inquiries in one local place.',
            target_role='RESIDENT',
            channel='TOWN',
            cta_label='View Macleen’s',
            cta_url='/',
            is_active=True,
        ))
    # The reserved handle can never be claimed by a new account. If it already
    # existed before this release, promote it safely and revoke any old main-admin flag.
    main_profile = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == 'uzu.macky').first()
    if main_profile:
        CommunityProfile.query.filter(CommunityProfile.id != main_profile.id).update({'is_community_admin': False})
        main_profile.is_community_admin = True
        main_profile.verification_status = 'VERIFIED'
        main_profile.first_post_approved = True
    db.session.commit()


# ==================== PWA ROOT ROUTES ====================

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'manifest.json', mimetype='application/manifest+json')

@app.route('/sw.js')
def serve_sw():
    response = send_from_directory(os.path.join(app.root_path, 'static'), 'sw.js', mimetype='application/javascript')
    response.headers['Service-Worker-Allowed'] = '/'
    return response

# ==================== SAFE MIGRATION & RUN HOOKS ====================

def _ensure_column(table_name, column_name, ddl_type):
    """Idempotently add one missing column on SQLite or PostgreSQL."""
    inspector = inspect(db.engine)
    existing = {c['name'] for c in inspector.get_columns(table_name)}
    if column_name in existing:
        return False

    prep = db.engine.dialect.identifier_preparer
    q_table = prep.quote(table_name)
    q_col = prep.quote(column_name)
    try:
        with db.engine.begin() as conn:
            conn.execute(text(f"ALTER TABLE {q_table} ADD COLUMN {q_col} {ddl_type}"))
        app.logger.info('Database migration: added %s.%s', table_name, column_name)
        return True
    except Exception:
        # Another process may have applied it after our inspection. Re-check before failing.
        inspector = inspect(db.engine)
        existing = {c['name'] for c in inspector.get_columns(table_name)}
        if column_name in existing:
            app.logger.info('Database migration: %s.%s was added concurrently', table_name, column_name)
            return False
        app.logger.exception('Database migration failed while adding %s.%s', table_name, column_name)
        raise

def run_schema_migrations():
    """Apply the known additive schema upgrades needed by the current application."""
    migrations = {
        'customer': [
            ('referred_by', 'VARCHAR(50)'),
            ('last_daily_login', 'DATE'),
            ('login_streak', 'INTEGER DEFAULT 1'),
            ('last_active_at', 'TIMESTAMP'),
            ('card_theme', "VARCHAR(30) DEFAULT 'pink-classic'"),
            ('card_logo_scale', 'FLOAT DEFAULT 1.0'),
            ('card_photo_scale', 'FLOAT DEFAULT 1.0'),
            ('card_qr_scale', 'FLOAT DEFAULT 1.0'),
            ('card_text_scale', 'FLOAT DEFAULT 1.0'),
            ('card_info_scale', 'FLOAT DEFAULT 1.0'),
            ('campus_name', 'VARCHAR(120)'),
            ('break_start', 'VARCHAR(5)'),
            ('break_end', 'VARCHAR(5)'),
            ('favorite_alerts', 'BOOLEAN DEFAULT FALSE'),
            ('community_student_preapproved', 'BOOLEAN DEFAULT FALSE'),
            ('community_student_preapproved_at', 'TIMESTAMP'),
            ('community_student_preapproved_by', 'VARCHAR(50)'),
        ],
        'order': [
            ('dining_option', "VARCHAR(20) DEFAULT 'DINE-IN'"),
            ('points_redeemed', 'FLOAT DEFAULT 0.0'),
            ('points_discount', 'FLOAT DEFAULT 0.0'),
            ('public_token', 'VARCHAR(64)'),
            ('fulfillment_status', "VARCHAR(30) DEFAULT 'SUBMITTED'"),
        ],
        'promotion_tracker': [
            ('promo_cost', 'FLOAT DEFAULT 0.0'),
            ('is_visible', 'BOOLEAN DEFAULT TRUE'),
            ('portal_only', 'BOOLEAN DEFAULT FALSE'),
            ('description', 'TEXT'),
        ],
        'marketing_post': [
            ('insight_reach', 'INTEGER DEFAULT 0'),
            ('insight_impressions', 'INTEGER DEFAULT 0'),
            ('insight_reactions', 'INTEGER DEFAULT 0'),
            ('insight_comments', 'INTEGER DEFAULT 0'),
            ('insight_shares', 'INTEGER DEFAULT 0'),
            ('insight_saves', 'INTEGER DEFAULT 0'),
            ('insight_link_clicks', 'INTEGER DEFAULT 0'),
            ('insight_new_followers', 'INTEGER DEFAULT 0'),
            ('insight_spend', 'FLOAT DEFAULT 0.0'),
            ('insight_notes', 'TEXT'),
            ('insight_analysis', 'TEXT'),
            ('insight_ai_model', 'VARCHAR(100)'),
            ('insights_updated_at', 'TIMESTAMP'),
            ('insights_analyzed_at', 'TIMESTAMP'),
        ],
        'product': [
            ('cost', 'FLOAT DEFAULT 0.0'),
            ('allow_custom_amount', 'BOOLEAN DEFAULT FALSE'),
            ('minimum_order_amount', 'FLOAT'),
            ('option_schema', 'TEXT'),
            ('size_schema', 'TEXT'),
            ('available_start_time', 'VARCHAR(10)'),
            ('available_end_time', 'VARCHAR(10)'),
            ('prep_minutes', 'INTEGER DEFAULT 10'),
        ],
        'order_item': [
            ('cost_price', 'FLOAT DEFAULT 0.0'),
            ('selected_options', 'TEXT'),
        ],
        'vault_drop': [
            ('cash_breakdown', 'TEXT'),
        ],
        'craft_comment': [
            ('ip_address', 'VARCHAR(64)'),
        ],
        'investor_interest': [
            ('offer_code', "VARCHAR(40) DEFAULT 'GENERAL'"),
            ('payout_option', 'VARCHAR(30)'),
            ('proposed_amount', 'FLOAT'),
            ('monthly_rate_percent', 'FLOAT'),
            ('term_months', 'INTEGER'),
            ('monthly_interest_amount', 'FLOAT'),
            ('total_interest_amount', 'FLOAT'),
            ('maturity_payment_amount', 'FLOAT'),
            ('total_contract_amount', 'FLOAT'),
            ('is_counter_offer', 'BOOLEAN DEFAULT FALSE'),
            ('cashier_notes', 'TEXT'),
            ('cashier_reviewed_by', 'VARCHAR(50)'),
            ('cashier_reviewed_at', 'TIMESTAMP'),
        ],
        'community_profile': [
            ('is_community_admin', 'BOOLEAN DEFAULT FALSE'),
            ('public_bio', 'VARCHAR(160)'),
            ('is_profile_locked', 'BOOLEAN DEFAULT FALSE'),
            ('student_id_image_data', 'TEXT'),
            ('student_id_uploaded_at', 'TIMESTAMP'),
            ('student_id_deleted_at', 'TIMESTAMP'),
            ('student_application_status', 'VARCHAR(20)'),
            ('student_application_campus', 'VARCHAR(120)'),
            ('student_application_department', 'VARCHAR(80)'),
            ('student_application_graduating_year', 'INTEGER'),
        ],
        'community_post': [
            ('reshared_post_id', 'INTEGER'),
        ],
        'community_group': [
            ('external_chat_url', 'VARCHAR(500)'),
        ],
    }

    inspector = inspect(db.engine)
    tables = set(inspector.get_table_names())
    for table_name, columns in migrations.items():
        if table_name not in tables:
            continue
        for column_name, ddl_type in columns:
            _ensure_column(table_name, column_name, ddl_type)

    # Normalize data for columns that were introduced after initial deployments.
    with db.engine.begin() as conn:
        if 'customer' in tables:
            conn.execute(text("UPDATE customer SET login_streak = 1 WHERE login_streak IS NULL"))
            conn.execute(text("UPDATE customer SET card_theme = 'pink-classic' WHERE card_theme IS NULL OR card_theme = ''"))
            conn.execute(text("UPDATE customer SET card_logo_scale = 1.0 WHERE card_logo_scale IS NULL"))
            conn.execute(text("UPDATE customer SET card_photo_scale = 1.0 WHERE card_photo_scale IS NULL"))
            conn.execute(text("UPDATE customer SET card_qr_scale = 1.0 WHERE card_qr_scale IS NULL"))
            conn.execute(text("UPDATE customer SET card_text_scale = 1.0 WHERE card_text_scale IS NULL"))
            conn.execute(text("UPDATE customer SET card_info_scale = 1.0 WHERE card_info_scale IS NULL"))
            conn.execute(text("UPDATE customer SET favorite_alerts = FALSE WHERE favorite_alerts IS NULL"))
            conn.execute(text("UPDATE customer SET community_student_preapproved = FALSE WHERE community_student_preapproved IS NULL"))
        if 'community_profile' in tables:
            conn.execute(text("UPDATE community_profile SET is_community_admin = FALSE WHERE is_community_admin IS NULL"))
            conn.execute(text("UPDATE community_profile SET is_profile_locked = FALSE WHERE is_profile_locked IS NULL"))
            # Community Lite never retains student-ID images. Existing images are
            # erased during migration; verification continues through staff review.
            conn.execute(text("UPDATE community_profile SET student_id_image_data = NULL, student_id_deleted_at = CURRENT_TIMESTAMP WHERE student_id_image_data IS NOT NULL"))
        if 'promotion_tracker' in tables:
            conn.execute(text("UPDATE promotion_tracker SET is_visible = TRUE WHERE is_visible IS NULL"))
            conn.execute(text("UPDATE promotion_tracker SET portal_only = FALSE WHERE portal_only IS NULL"))
        if 'product' in tables:
            conn.execute(text("UPDATE product SET allow_custom_amount = FALSE WHERE allow_custom_amount IS NULL"))
            conn.execute(text("UPDATE product SET prep_minutes = 10 WHERE prep_minutes IS NULL OR prep_minutes < 1"))
        if 'order' in tables:
            conn.execute(text("UPDATE \"order\" SET fulfillment_status = CASE WHEN status = 'COMPLETED' THEN 'FULFILLED' WHEN status = 'CANCELLED' THEN 'CANCELLED' ELSE 'SUBMITTED' END WHERE fulfillment_status IS NULL OR fulfillment_status = ''"))
        if 'investor_interest' in tables:
            conn.execute(text("UPDATE investor_interest SET offer_code = 'GENERAL' WHERE offer_code IS NULL OR offer_code = ''"))
            conn.execute(text("UPDATE investor_interest SET is_counter_offer = FALSE WHERE is_counter_offer IS NULL"))

def run_db_setup():
    global _DB_INITIALIZED
    if _DB_INITIALIZED:
        return
    try:
        db.create_all()
        run_schema_migrations()

        # Create bootstrap accounts only when they do not already exist. Never reset an existing PIN on restart.
        default_roles = [
            ('admin', os.environ.get('DEFAULT_ADMIN_PIN') or '1234', 'ADMIN'),
            ('cashier1', os.environ.get('DEFAULT_CASHIER_PIN') or '1111', 'CASHIER'),
        ]
        for username, pin, role in default_roles:
            existing = Staff.query.filter(db.func.lower(Staff.username) == username.lower()).first()
            if not existing:
                db.session.add(Staff(
                    username=username,
                    pin_hash=generate_password_hash(pin),
                    role=role,
                    active=True,
                ))
                if not os.environ.get('DEFAULT_ADMIN_PIN' if role == 'ADMIN' else 'DEFAULT_CASHIER_PIN'):
                    app.logger.warning('Created bootstrap %s account %r using the built-in first-run PIN. Change it in Admin immediately.', role, username)
        if not CraftCategory.query.filter(db.func.lower(CraftCategory.name) == 'general').first():
            db.session.add(CraftCategory(name='General', image_url=CRAFT_DEFAULT_IMAGE, is_active=True))
        # Categories are useful immediately but never invent products, prices, or sales.
        for category_name in ('School', 'Internship & Work', 'Personal Finance', 'Small Business', 'Productivity', 'General'):
            if not DigitalCategory.query.filter(db.func.lower(DigitalCategory.name) == category_name.lower()).first():
                db.session.add(DigitalCategory(name=category_name, is_active=True))
        db.session.commit()

        # Existing orders pre-date private tracking links. Backfill them once with
        # unguessable tokens; no customer information is placed in the URL.
        tokenless_orders = Order.query.filter(
            (Order.public_token.is_(None)) | (Order.public_token == '')
        ).all()
        for existing_order in tokenless_orders:
            existing_order.public_token = secrets.token_urlsafe(24)
        if tokenless_orders:
            db.session.commit()

        ensure_default_promos()
        ensure_legacy_craft_catalog()
        ensure_community_defaults()
        disable_legacy_meta_connection()
        _DB_INITIALIZED = True
        app.logger.info('Database setup completed successfully.')
    except Exception:
        db.session.rollback()
        app.logger.exception('Database setup failed. Deployment/request should fail visibly instead of hiding schema errors.')
        raise

@app.before_request
def app_startup_and_session_handler():
    if not _DB_INITIALIZED:
        run_db_setup()

    # Customer logins may persist for up to 30 days. Staff sessions remain browser-session cookies
    # and are additionally protected by the explicit 8-hour inactivity check in the staff guards.
    session.permanent = bool(session.get('customer_id')) and not bool(session.get('admin_user') or session.get('cashier_user'))

    if 'customer_id' in session:
        try:
            cust = db.session.get(Customer, session['customer_id'])
            if cust and hasattr(cust, 'last_active_at'):
                cust.last_active_at = utc_now()
                db.session.commit()
        except Exception:
            db.session.rollback()
            app.logger.exception('Could not update customer last_active_at for customer_id=%s', session.get('customer_id'))

@app.after_request
def add_live_ui_progressive_enhancement(response):
    """Load one safe, shared AJAX-navigation layer on every rendered HTML page."""
    content_type = (response.headers.get('Content-Type') or '').lower()
    if response.direct_passthrough or 'text/html' not in content_type:
        return response
    try:
        html = response.get_data(as_text=True)
        marker = 'data-macleens-live'
        if marker not in html and '</body>' in html.lower():
            script = '<script data-macleens-live src="/static/live-ui.js?v=5"></script>'
            closing = html.lower().rfind('</body>')
            html = html[:closing] + script + html[closing:]
            response.set_data(html)
    except Exception:
        app.logger.exception('Could not attach progressive live UI script')
    return response

# ==================== HELPERS & GUARDS ====================

def get_client_ip():
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr or '127.0.0.1'

class OrderValidationError(ValueError):
    pass

def parse_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)

def parse_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)

MIN_REDEMPTION_POINTS = 20.0
POINT_VALUE_PHP = 1.0

def calculate_points_redemption(cust, requested_points, merchandise_total):
    """Validate redemption using the system rule: minimum 20 points, 1 point = ₱1."""
    points = round(parse_float(requested_points, 0.0), 2)
    balance = round(max(0.0, parse_float(cust.points_balance, 0.0)), 2)
    total = round(max(0.0, parse_float(merchandise_total, 0.0)), 2)
    if points <= 0:
        return 0.0, 0.0
    if points < MIN_REDEMPTION_POINTS:
        raise OrderValidationError(f'Minimum redemption is {MIN_REDEMPTION_POINTS:g} points.')
    if points > balance + 1e-9:
        raise OrderValidationError(f'Only {balance:,.2f} points are available.')
    max_points = round(total / POINT_VALUE_PHP, 2)
    if points > max_points + 1e-9:
        raise OrderValidationError(f'Only {max_points:,.2f} points can be used on this purchase.')
    return points, round(points * POINT_VALUE_PHP, 2)

def record_points_redemption(cust, points, order_id=None, reason='Purchase discount'):
    if points <= 0:
        return
    cust.points_balance = round(max(0.0, parse_float(cust.points_balance, 0.0) - points), 2)
    suffix = f' / Order #{order_id}' if order_id else ''
    db.session.add(RewardLedger(customer_id=cust.id, points_change=-points, reason=f'{reason}{suffix}'))

CASH_FLOW_FREQUENCIES = ('DAILY', 'WEEKLY', 'BIWEEKLY', 'MONTHLY')
CASH_FLOW_ENTRY_TYPES = ('EXPENSE', 'INCOME')
CASH_FLOW_DEFAULT_HORIZON_YEARS = 2
CASH_FLOW_MAX_HORIZON_YEARS = 20
CASH_FLOW_MAX_DURATION = {'DAILY': 7320, 'WEEKLY': 1060, 'BIWEEKLY': 540, 'MONTHLY': 240}
CASH_FLOW_COGS_RATE = 0.60


def cashflow_month_start(value):
    return date(value.year, value.month, 1)


def cashflow_add_months(value, months):
    """Add whole calendar months, clamping to the last valid day."""
    month_index = (value.year * 12 + (value.month - 1)) + int(months)
    year, month_zero = divmod(month_index, 12)
    month = month_zero + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def cashflow_plan_end_date(plan):
    count = parse_int(getattr(plan, 'duration_count', 1), 1)
    if count == 0:
        return None
    count = max(1, count)
    start = plan.start_date
    if plan.frequency == 'DAILY':
        return start + timedelta(days=count - 1)
    if plan.frequency == 'WEEKLY':
        return start + timedelta(weeks=count - 1)
    if plan.frequency == 'BIWEEKLY':
        return start + timedelta(weeks=2 * (count - 1))
    return cashflow_add_months(start, count - 1)


def cashflow_occurrence_dates(plan, window_start, window_end_exclusive):
    """Yield scheduled dates inside a window. duration_count=0 means indefinite."""
    count = parse_int(getattr(plan, 'duration_count', 0), 0)
    if count < 0 or plan.frequency not in CASH_FLOW_FREQUENCIES:
        return

    indefinite = count == 0
    index = 0
    while indefinite or index < count:
        if plan.frequency == 'DAILY':
            occurrence = plan.start_date + timedelta(days=index)
        elif plan.frequency == 'WEEKLY':
            occurrence = plan.start_date + timedelta(weeks=index)
        elif plan.frequency == 'BIWEEKLY':
            occurrence = plan.start_date + timedelta(weeks=2 * index)
        else:
            occurrence = cashflow_add_months(plan.start_date, index)

        if occurrence >= window_end_exclusive:
            break
        if occurrence >= window_start:
            yield occurrence
        index += 1


def cashflow_utc_bounds(start_day, end_day_exclusive):
    start_local = datetime.combine(start_day, time.min, tzinfo=MANILA_TZ)
    end_local = datetime.combine(end_day_exclusive, time.min, tzinfo=MANILA_TZ)
    return (
        start_local.astimezone(timezone.utc).replace(tzinfo=None),
        end_local.astimezone(timezone.utc).replace(tzinfo=None),
    )

def parse_product_option_schema(schema, strict=False):
    """Parse a compact product-option definition.

    Admin syntax examples:
        Sauce[]: Hot|Sweet; Flavor: Ube|Chocolate|Vanilla
        Sauce: Hot|Sweet

    A group name ending in [] is a checkbox group and allows one or more choices.
    A normal group is a single-choice group and is shown as radio buttons.
    Every configured group is required when the product is ordered. Choices do not
    change price; they only describe the selected preparation/flavor.
    """
    raw = str(schema or '').strip()
    if not raw:
        return []

    groups = []
    seen_groups = set()
    segments = [x.strip() for x in re.split(r'[;\n]+', raw) if x.strip()]
    if len(segments) > 6:
        raise OrderValidationError('A product can have at most 6 option groups.')

    for segment in segments:
        if ':' in segment:
            label, choices_raw = segment.split(':', 1)
        elif '=' in segment:
            label, choices_raw = segment.split('=', 1)
        else:
            if strict:
                raise OrderValidationError(
                    f"Invalid product option '{segment}'. Use Group: Choice1|Choice2."
                )
            continue

        label = re.sub(r'\s+', ' ', label.strip())
        multiple = False
        if label.endswith('[]'):
            multiple = True
            label = label[:-2].strip()
        label = label[:40]
        if not label:
            if strict:
                raise OrderValidationError('Every product option group needs a name.')
            continue
        key = label.casefold()
        if key in seen_groups:
            raise OrderValidationError(f"Duplicate product option group: {label}.")
        seen_groups.add(key)

        # Prefer | as the separator; allow commas as a convenient fallback.
        splitter = '|' if '|' in choices_raw else ','
        choices = []
        seen_choices = set()
        for value in choices_raw.split(splitter):
            value = re.sub(r'\s+', ' ', value.strip())[:60]
            if not value:
                continue
            vkey = value.casefold()
            if vkey not in seen_choices:
                seen_choices.add(vkey)
                choices.append(value)
        if len(choices) < 2:
            if strict:
                raise OrderValidationError(f"{label} needs at least 2 choices.")
            continue
        if len(choices) > 20:
            raise OrderValidationError(f"{label} can have at most 20 choices.")
        groups.append({'name': label, 'choices': choices, 'multiple': multiple})

    if strict and raw and not groups:
        raise OrderValidationError('Product options could not be read. Example: Sauce[]: Hot|Sweet; Flavor: Ube|Chocolate')
    return groups

def normalize_product_option_schema(schema):
    groups = parse_product_option_schema(schema, strict=bool(str(schema or '').strip()))
    return '; '.join(
        f"{g['name']}{'[]' if g.get('multiple') else ''}: {'|'.join(g['choices'])}"
        for g in groups
    ) or None

def parse_product_size_schema(schema, strict=False):
    """Parse per-size selling prices from a compact admin definition.

    Examples:
        Small=35|Medium=45|Large=55
        12oz=50 | 16oz=65 | 22oz=80

    A configured size is a required single-choice option and its selling price
    replaces the product's base price for that cart/order line.
    """
    raw = str(schema or '').strip()
    if not raw:
        return []

    entries = []
    seen = set()
    parts = [x.strip() for x in re.split(r'[|;\n]+', raw) if x.strip()]
    if len(parts) > 12:
        raise OrderValidationError('A product can have at most 12 priced sizes.')

    for part in parts:
        match = re.match(r'^(.+?)\s*(?:=|:|@)\s*₱?\s*([0-9]+(?:\.[0-9]{1,2})?)$', part)
        if not match:
            if strict:
                raise OrderValidationError(
                    f"Invalid size '{part}'. Use Size=Price, for example 12oz=50|16oz=65."
                )
            continue
        name = re.sub(r'\s+', ' ', match.group(1).strip())[:40]
        price = round(parse_float(match.group(2), 0.0), 2)
        if not name or price <= 0:
            if strict:
                raise OrderValidationError('Every size needs a name and a price greater than ₱0.00.')
            continue
        key = name.casefold()
        if key in seen:
            raise OrderValidationError(f'Duplicate product size: {name}.')
        seen.add(key)
        entries.append({'name': name, 'price': price})

    if strict and raw and not entries:
        raise OrderValidationError('Product sizes could not be read. Example: Small=35|Medium=45|Large=55')
    return entries


def normalize_product_size_schema(schema):
    sizes = parse_product_size_schema(schema, strict=bool(str(schema or '').strip()))
    return '|'.join(f"{row['name']}={row['price']:.2f}" for row in sizes) or None


def product_choice_groups(option_schema=None, size_schema=None):
    """Return all selectable groups, with a priced Size group first when configured."""
    groups = []
    sizes = parse_product_size_schema(size_schema, strict=False)
    if sizes:
        groups.append({
            'name': 'Size',
            'choices': [row['name'] for row in sizes],
            'multiple': False,
            'priced': True,
            'prices': {row['name']: row['price'] for row in sizes},
        })
    groups.extend(parse_product_option_schema(option_schema, strict=False))
    return groups


def normalize_product_configuration(option_schema, size_schema):
    normalized_options = normalize_product_option_schema(option_schema)
    normalized_sizes = normalize_product_size_schema(size_schema)
    if normalized_sizes:
        if any(g['name'].casefold() == 'size' for g in parse_product_option_schema(normalized_options, strict=False)):
            raise OrderValidationError(
                "Do not add a normal 'Size' sub-option when Priced Sizes are configured. Use the Priced Sizes field instead."
            )
    return normalized_options, normalized_sizes


def product_starting_price(prod):
    sizes = parse_product_size_schema(getattr(prod, 'size_schema', None), strict=False)
    if sizes:
        return min(row['price'] for row in sizes)
    return max(0.0, parse_float(getattr(prod, 'price', 0.0), 0.0))


PRODUCT_SHARE_IMAGE_SIZE = (1200, 630)
PRODUCT_SHARE_MAX_SOURCE_BYTES = 8_000_000
PRODUCT_SHARE_CACHE_LIMIT = 128
PRODUCT_SHARE_STYLE_VERSION = 'original-photo-v3'
_PRODUCT_SHARE_CACHE = {}
Image.MAX_IMAGE_PIXELS = 40_000_000


def product_share_version(prod):
    """Change the shared URL whenever visible product details change."""
    visible_state = json.dumps([
        PRODUCT_SHARE_STYLE_VERSION,
        getattr(prod, 'id', None),
        getattr(prod, 'name', ''),
        getattr(prod, 'category_name', ''),
        round(parse_float(getattr(prod, 'price', 0.0), 0.0), 2),
        getattr(prod, 'size_schema', '') or '',
        getattr(prod, 'image_url', '') or '',
        bool(getattr(prod, 'is_active', True)),
    ], ensure_ascii=False, separators=(',', ':'))
    return hashlib.sha256(visible_state.encode('utf-8')).hexdigest()[:12]


def _preview_font(size):
    candidates = (
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf',
    )
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except OSError:
            continue
    try:
        return ImageFont.load_default(size=size)
    except TypeError:
        return ImageFont.load_default()


def _preview_url_is_public(url):
    parsed = urlparse(url)
    if parsed.scheme not in ('http', 'https') or not parsed.hostname:
        return False
    host = parsed.hostname.casefold().rstrip('.')
    if host in ('localhost', 'localhost.localdomain') or host.endswith('.local'):
        return False
    try:
        address = ipaddress.ip_address(host)
    except ValueError:
        return True
    return not any((
        address.is_private, address.is_loopback, address.is_link_local,
        address.is_multicast, address.is_reserved, address.is_unspecified,
    ))


def _download_preview_source(url):
    current = url
    for _ in range(4):
        if not _preview_url_is_public(current):
            raise ValueError('Product image URL is not a public HTTP image.')
        with requests.get(
            current,
            stream=True,
            allow_redirects=False,
            # Facebook's crawler will abandon slow image responses. Keep this
            # upstream fetch short, then use the bright branded fallback.
            timeout=(2.0, 4.0),
            headers={
                'User-Agent': 'Mozilla/5.0 (compatible; MacleensFoodHouse-SocialPreview/2.0)',
                'Accept': 'image/jpeg,image/png,image/webp,image/*;q=0.8',
            },
        ) as response:
            if response.status_code in (301, 302, 303, 307, 308):
                location = response.headers.get('Location')
                if not location:
                    raise ValueError('Product image redirect has no destination.')
                current = urljoin(current, location)
                continue
            response.raise_for_status()
            content_type = (response.headers.get('Content-Type') or '').lower()
            if content_type and not content_type.startswith('image/'):
                raise ValueError('Product image URL did not return an image.')
            content_length = parse_int(response.headers.get('Content-Length'), 0)
            if content_length > PRODUCT_SHARE_MAX_SOURCE_BYTES:
                raise ValueError('Product image is too large for a social preview.')
            payload = bytearray()
            for chunk in response.iter_content(64 * 1024):
                if not chunk:
                    continue
                payload.extend(chunk)
                if len(payload) > PRODUCT_SHARE_MAX_SOURCE_BYTES:
                    raise ValueError('Product image is too large for a social preview.')
            return bytes(payload)
    raise ValueError('Product image redirected too many times.')


def _open_preview_image(payload):
    with Image.open(io.BytesIO(payload)) as source:
        source.load()
        return ImageOps.exif_transpose(source).convert('RGB')


def _product_preview_fallback(prod):
    """Create a colorful food/drink illustration when a remote photo is unusable."""
    width, height = PRODUCT_SHARE_IMAGE_SIZE
    canvas = Image.new('RGB', (width, height), '#fb7185')
    draw = ImageDraw.Draw(canvas)
    for y in range(height):
        blend = y / max(1, height - 1)
        color = (
            int(251 * (1 - blend) + 190 * blend),
            int(113 * (1 - blend) + 24 * blend),
            int(133 * (1 - blend) + 93 * blend),
        )
        draw.line((0, y, width, y), fill=color)

    draw.ellipse((-180, -240, 470, 410), fill='#fdba74')
    draw.ellipse((870, 330, 1390, 850), fill='#0f766e')
    draw.ellipse((60, 70, 1140, 900), fill='#ffe4e6', outline='#ffffff', width=16)

    keywords = f"{getattr(prod, 'name', '')} {getattr(prod, 'category_name', '')}".casefold()
    is_drink = any(word in keywords for word in ('shake', 'coffee', 'drink', 'juice', 'tea', 'float', 'beverage'))
    if is_drink:
        # Layered iced drink with cream, chocolate drizzle, straw, and glass shine.
        draw.rounded_rectangle((390, 145, 810, 555), radius=58, fill='#f8fafc', outline='#ffffff', width=14)
        draw.rounded_rectangle((420, 250, 780, 530), radius=38, fill='#713f12')
        draw.rectangle((420, 330, 780, 530), fill='#451a03')
        draw.rectangle((420, 430, 780, 530), fill='#d97706')
        draw.rounded_rectangle((515, 40, 565, 285), radius=20, fill='#0f766e')
        for cx, cy, radius in ((455, 245, 72), (535, 210, 88), (625, 205, 92), (715, 245, 74)):
            draw.ellipse((cx-radius, cy-radius, cx+radius, cy+radius), fill='#fff7ed', outline='#ffffff', width=5)
        draw.arc((450, 135, 745, 340), start=195, end=345, fill='#7c2d12', width=18)
        draw.arc((480, 150, 720, 315), start=195, end=345, fill='#be123c', width=10)
        for cx, cy in ((470, 365), (555, 410), (690, 340), (735, 455), (620, 485)):
            draw.ellipse((cx-22, cy-15, cx+22, cy+15), fill='#fde68a')
        draw.rounded_rectangle((446, 280, 480, 500), radius=17, fill='#ffffff')
    else:
        # Warm plated meal illustration with rice, viand, vegetables, and steam.
        draw.ellipse((265, 185, 935, 585), fill='#f8fafc', outline='#ffffff', width=18)
        draw.ellipse((330, 230, 870, 535), fill='#fecdd3')
        draw.ellipse((420, 245, 650, 455), fill='#fff7ed', outline='#ffffff', width=8)
        draw.ellipse((590, 290, 820, 475), fill='#92400e')
        draw.ellipse((625, 325, 760, 420), fill='#ef4444')
        for cx, cy, color in ((355, 360, '#16a34a'), (395, 410, '#65a30d'), (790, 265, '#facc15'), (820, 380, '#22c55e')):
            draw.ellipse((cx-38, cy-24, cx+38, cy+24), fill=color)
        for x in (485, 580, 680):
            draw.arc((x, 95, x+90, 255), start=120, end=245, fill='#ffffff', width=13)

    return canvas


def _product_preview_photo(prod):
    source_url = (getattr(prod, 'image_url', '') or '').strip()
    try:
        if source_url.startswith('data:image/') and ';base64,' in source_url:
            encoded = source_url.split(';base64,', 1)[1]
            if len(encoded) > PRODUCT_SHARE_MAX_SOURCE_BYTES * 2:
                raise ValueError('Embedded product image is too large.')
            return _open_preview_image(base64.b64decode(encoded, validate=True))
        if source_url.startswith('/static/'):
            relative = source_url[len('/static/'):].replace('\\', '/')
            candidate = os.path.realpath(os.path.join(app.static_folder, relative))
            static_root = os.path.realpath(app.static_folder)
            if os.path.commonpath((static_root, candidate)) != static_root:
                raise ValueError('Product image path is outside the static folder.')
            with open(candidate, 'rb') as image_file:
                payload = image_file.read(PRODUCT_SHARE_MAX_SOURCE_BYTES + 1)
            if len(payload) > PRODUCT_SHARE_MAX_SOURCE_BYTES:
                raise ValueError('Product image is too large for a social preview.')
            return _open_preview_image(payload)
        if source_url.startswith(('http://', 'https://')):
            return _open_preview_image(_download_preview_source(source_url))
    except (
        OSError, ValueError, UnidentifiedImageError, Image.DecompressionBombError,
        requests.exceptions.RequestException,
    ) as exc:
        app.logger.info('Using branded fallback for product preview %s: %s', getattr(prod, 'id', '?'), exc)

    return _product_preview_fallback(prod)


def _original_product_photo(photo):
    """Preserve the uploaded product photo's original color and brightness."""
    return photo.convert('RGB').copy()


def _wrap_preview_text(draw, text_value, font, max_width, max_lines=3):
    words = str(text_value or '').split() or ['Menu Pick']
    lines = []
    current = ''
    for word in words:
        candidate = f'{current} {word}'.strip()
        if not current or draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        remaining = lines[-1]
        while remaining and draw.textbbox((0, 0), remaining + '...', font=font)[2] > max_width:
            remaining = remaining[:-1]
        lines[-1] = remaining.rstrip() + '...'
    return lines


def render_product_social_preview(prod):
    width, height = PRODUCT_SHARE_IMAGE_SIZE
    photo = _original_product_photo(_product_preview_photo(prod))

    # Keep the original photo untouched and fully visible. The surrounding
    # layout supplies the brightness and branding instead of filtering food.
    background = Image.new('RGBA', (width, height), '#fff7fb')
    draw = ImageDraw.Draw(background)
    for x in range(width):
        blend = x / max(1, width - 1)
        color = (
            int(255 * (1 - blend) + 240 * blend),
            int(247 * (1 - blend) + 253 * blend),
            int(251 * (1 - blend) + 250 * blend),
            255,
        )
        draw.line((x, 0, x, height), fill=color)
    draw.ellipse((-170, -240, 440, 350), fill=(251, 207, 232, 150))
    draw.ellipse((960, 390, 1340, 780), fill=(153, 246, 228, 145))

    # Soft shadows and bright cards; nothing is placed over the product photo.
    draw.rounded_rectangle((48, 48, 724, 594), radius=34, fill=(131, 24, 67, 38))
    draw.rounded_rectangle((750, 48, 1178, 594), radius=34, fill=(131, 24, 67, 38))
    draw.rounded_rectangle((38, 38, 714, 584), radius=34, fill='white', outline='#f9a8d4', width=4)
    draw.rounded_rectangle((740, 38, 1168, 584), radius=34, fill='white', outline='#f9a8d4', width=4)

    photo_area = (626, 494)
    contained = ImageOps.contain(photo, photo_area, method=Image.Resampling.LANCZOS)
    photo_x = 62 + (photo_area[0] - contained.width) // 2
    photo_y = 64 + (photo_area[1] - contained.height) // 2
    background.paste(contained.convert('RGBA'), (photo_x, photo_y))

    try:
        with Image.open(os.path.join(app.static_folder, 'logo.png')) as logo_source:
            logo = ImageOps.fit(ImageOps.exif_transpose(logo_source).convert('RGBA'), (70, 70), method=Image.Resampling.LANCZOS)
        logo_mask = Image.new('L', (70, 70), 0)
        ImageDraw.Draw(logo_mask).ellipse((0, 0, 69, 69), fill=255)
        background.paste(logo, (772, 68), logo_mask)
        draw.ellipse((771, 67, 843, 139), outline=(236, 72, 153, 255), width=3)
    except (OSError, UnidentifiedImageError):
        pass

    draw.text((858, 72), "MACLEEN'S", font=_preview_font(28), fill='#831843')
    draw.text((858, 108), 'FOOD HOUSE', font=_preview_font(21), fill='#ec4899')

    category = (getattr(prod, 'category_name', '') or 'MENU PICK').upper()
    category_font = _preview_font(18)
    category_width = min(345, draw.textbbox((0, 0), category, font=category_font)[2] + 34)
    draw.rounded_rectangle((772, 165, 772 + category_width, 203), radius=19, fill='#0f766e')
    draw.text((789, 173), category, font=category_font, fill='white')

    name = (getattr(prod, 'name', '') or 'Fresh Menu Pick').strip()
    name_font = _preview_font(52)
    name_lines = _wrap_preview_text(draw, name, name_font, 340, max_lines=3)
    if len(name_lines) > 2:
        name_font = _preview_font(44)
        name_lines = _wrap_preview_text(draw, name, name_font, 340, max_lines=3)
    draw.multiline_text((772, 220), '\n'.join(name_lines), font=name_font, fill='#172033', spacing=3)

    price_prefix = 'From ' if parse_product_size_schema(getattr(prod, 'size_schema', None), strict=False) else ''
    price_label = f'{price_prefix}₱{product_starting_price(prod):,.2f}'
    draw.text((772, 414), price_label, font=_preview_font(48), fill='#db2777')
    draw.text((774, 468), 'Fresh • Affordable • Made for you', font=_preview_font(17), fill='#475569')
    draw.rounded_rectangle((772, 512, 1135, 560), radius=24, fill='#0f766e')
    draw.text((796, 524), 'CLICK TO VIEW & ORDER', font=_preview_font(20), fill='white')

    output = io.BytesIO()
    background.convert('RGB').save(
        output, format='JPEG', quality=90, optimize=True, progressive=True, subsampling=0
    )
    return output.getvalue()


def cached_product_social_preview(prod):
    """Return an immutable JPEG without repeatedly re-fetching the source photo."""
    cache_key = f"{getattr(prod, 'id', 'new')}:{product_share_version(prod)}"
    payload = _PRODUCT_SHARE_CACHE.get(cache_key)
    if payload is None:
        payload = render_product_social_preview(prod)
        if len(_PRODUCT_SHARE_CACHE) >= PRODUCT_SHARE_CACHE_LIMIT:
            _PRODUCT_SHARE_CACHE.pop(next(iter(_PRODUCT_SHARE_CACHE)), None)
        _PRODUCT_SHARE_CACHE[cache_key] = payload
    return payload


def product_price_for_options(prod, selected_options):
    """Resolve the authoritative regular selling price after a priced Size choice."""
    base = round(max(0.0, parse_float(getattr(prod, 'price', 0.0), 0.0)), 2)
    sizes = parse_product_size_schema(getattr(prod, 'size_schema', None), strict=False)
    if not sizes:
        return base
    selected_size = str((selected_options or {}).get('Size') or '').strip()
    for row in sizes:
        if row['name'].casefold() == selected_size.casefold():
            return round(row['price'], 2)
    raise OrderValidationError(f'Please choose a valid Size for {prod.name}.')


def validate_product_options(prod, raw_options):
    groups = product_choice_groups(getattr(prod, 'option_schema', None), getattr(prod, 'size_schema', None))
    if not groups:
        if raw_options not in (None, '', {}, []):
            raise OrderValidationError(f'{prod.name} does not have selectable options.')
        return {}

    if isinstance(raw_options, str):
        raw_options = raw_options.strip()
        if not raw_options:
            raw_options = {}
        else:
            try:
                raw_options = json.loads(raw_options)
            except (TypeError, ValueError):
                raise OrderValidationError(f'Invalid option selection for {prod.name}.')
    if not isinstance(raw_options, dict):
        raise OrderValidationError(f'Please choose the required options for {prod.name}.')

    provided = {str(k).strip().casefold(): v for k, v in raw_options.items()}
    selected = {}
    allowed_group_keys = {g['name'].casefold() for g in groups}
    extras = set(provided) - allowed_group_keys
    if extras:
        raise OrderValidationError(f'Invalid option group submitted for {prod.name}.')

    for group in groups:
        label = group['name']
        value = provided.get(label.casefold())
        is_multiple = bool(group.get('multiple'))

        if is_multiple:
            values = value if isinstance(value, list) else ([value] if value not in (None, '') else [])
            canonical_values = []
            seen = set()
            for raw_value in values:
                value_text = re.sub(r'\s+', ' ', str(raw_value or '').strip())
                if not value_text:
                    continue
                canonical = next((c for c in group['choices'] if c.casefold() == value_text.casefold()), None)
                if canonical is None:
                    raise OrderValidationError(f'Invalid {label} choice for {prod.name}.')
                key = canonical.casefold()
                if key not in seen:
                    seen.add(key)
                    canonical_values.append(canonical)
            if not canonical_values:
                raise OrderValidationError(f'Please choose at least one {label} for {prod.name}.')
            selected[label] = canonical_values
        else:
            if isinstance(value, list):
                if len(value) != 1:
                    raise OrderValidationError(f'Please choose exactly one {label} for {prod.name}.')
                value = value[0]
            value_text = re.sub(r'\s+', ' ', str(value or '').strip())
            if not value_text:
                raise OrderValidationError(f'Please choose {label} for {prod.name}.')
            canonical = next((c for c in group['choices'] if c.casefold() == value_text.casefold()), None)
            if canonical is None:
                raise OrderValidationError(f'Invalid {label} choice for {prod.name}.')
            selected[label] = canonical
    return selected

def serialize_selected_options(options):
    return json.dumps(options or {}, ensure_ascii=False, separators=(',', ':')) if options else None

def option_summary(raw):
    if not raw:
        return ''
    data = raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
        except (TypeError, ValueError):
            return raw
    if isinstance(data, dict):
        def fmt(v):
            if isinstance(v, list):
                return ' + '.join(str(x) for x in v)
            return str(v)
        return ' • '.join(f'{k}: {fmt(v)}' for k, v in data.items())
    return str(data)

app.jinja_env.filters['option_summary'] = option_summary

def is_valid_customer_pin(pin):
    return bool(pin) and len(pin) == 4 and pin.isdigit()

def mask_card_number(card_number):
    value = (card_number or '').strip()
    if not value:
        return 'Member'
    if len(value) <= 4:
        return '*' * len(value)
    return f"{value[:4]}****{value[-2:]}"

def customer_access_issue(cust):
    if not cust:
        return 'Customer account was not found.'
    status = (cust.card_status or 'ACTIVE').upper()
    if cust.card_expires_at and ph_today() > cust.card_expires_at:
        return 'Your rewards card has expired. Please ask staff to renew it.'
    if status != 'ACTIVE':
        if status == 'LOCKED':
            return 'Your rewards account is locked. Please ask staff for assistance.'
        if status == 'EXPIRED':
            return 'Your rewards card has expired. Please ask staff to renew it.'
        return f'Your rewards account is currently {status.lower()}.'
    return None

def get_customer_by_identifier(identifier):
    value = (identifier or '').strip()
    if not value:
        return None
    return Customer.query.filter(
        (Customer.contact == value) | (db.func.lower(Customer.card_number) == value.lower())
    ).first()

def generate_unique_card_number(customer_id=None):
    """Continue the MFH sequence without random collisions."""
    highest = 0
    for (card,) in db.session.query(Customer.card_number).filter(Customer.card_number.isnot(None)).all():
        match = re.fullmatch(r'MFH-(\d+)', card or '', flags=re.IGNORECASE)
        if match:
            highest = max(highest, int(match.group(1)))
    number = max(highest + 1, parse_int(customer_id, 0) or 1)
    while True:
        width = 4 if number <= 9999 else 6
        candidate = f"MFH-{number:0{width}d}"
        if not Customer.query.filter(db.func.lower(Customer.card_number) == candidate.lower()).first():
            return candidate
        number += 1

COMMUNITY_ROLES = ('STUDENT', 'RESIDENT')
COMMUNITY_CHANNELS = ('CAMPUS', 'TOWN')
COMMUNITY_CHANNEL_MODULES = {
    'CAMPUS': (
        ('STUDY_COLLAB', 'Study & Collab'),
        ('CAMPUS_BOARD', 'Campus Board'),
        ('BUY_SELL_SWAP', 'Textbook Buy / Sell / Swap'),
        ('LOST_FOUND', 'Lost & Found'),
    ),
    'TOWN': (
        ('MUNICIPAL_BULLETIN', 'Municipal Bulletin'),
        ('LOCAL_CLASSIFIEDS', 'Local Classifieds'),
        ('HELP_WANTED', 'Neighborhood Help Wanted'),
        ('LOST_FOUND', 'Lost Pets / Belongings'),
    ),
}
COMMUNITY_CAMPUSES = (
    'Binalbagan Catholic College',
    'CHMSU Binalbagan Campus',
    'Other Binalbagan college or campus',
)
COMMUNITY_DEPARTMENTS = (
    'Business', 'Criminology', 'Education', 'Information Technology',
    'Engineering / Technology', 'Arts & Sciences', 'Other',
)
COMMUNITY_VIBES = (
    'Quiet study mode', 'Group cramming', 'Free to chat', 'On a break', 'Offline',
)
BINALBAGAN_BARANGAYS = (
    'Amontay', 'Bagroy', 'Bi-ao', 'Canmoros', 'Enclaro', 'Marina', 'Paglaum',
    'Payao', 'Progreso', 'San Jose', 'San Juan', 'San Pedro', 'San Teodoro',
    'San Vicente', 'Santo Rosario', 'Santol',
)
COMMUNITY_REACTIONS = ('LIKE', 'HELPFUL', 'INTERESTED')
COMMUNITY_REPORT_REASONS = ('HARASSMENT', 'HATE_OR_ABUSE', 'MISINFORMATION', 'SCAM', 'PRIVACY', 'SPAM', 'OTHER')
COMMUNITY_GIFT_DAILY_CAP = 50.0
COMMUNITY_GROUP_PRIVACY_MINIMUM = 3
COMMUNITY_MAIN_ADMIN_HANDLE = 'uzu.macky'
COMMUNITY_GROUP_MAX_MEMBERS = 25
COMMUNITY_TRUSTED_POST_THRESHOLD = max(1, int(os.environ.get('COMMUNITY_TRUSTED_POST_THRESHOLD', '3')))
COMMUNITY_MAX_OWNED_GROUPS = max(1, int(os.environ.get('COMMUNITY_MAX_OWNED_GROUPS', '2')))
COMMUNITY_GROUP_TASK_STATUSES = ('TODO', 'DOING', 'DONE')
COMMUNITY_GROUP_TASK_PRIORITIES = ('LOW', 'NORMAL', 'HIGH')
COMMUNITY_MENTION_PATTERN = re.compile(r'(?<![A-Za-z0-9._])@([A-Za-z0-9][A-Za-z0-9._]{2,23})', re.IGNORECASE)

def community_env_enabled(name, default=True):
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() not in {'0', 'false', 'no', 'off', 'disabled'}

COMMUNITY_REGISTRATION_OPEN = community_env_enabled('COMMUNITY_REGISTRATION_OPEN', True)
COMMUNITY_POSTING_OPEN = community_env_enabled('COMMUNITY_POSTING_OPEN', True)
COMMUNITY_GROUP_WORKSPACES_OPEN = community_env_enabled('COMMUNITY_GROUP_WORKSPACES_OPEN', True)
# Expensive or abuse-prone v5 features are intentionally paused in Community Lite.
COMMUNITY_INTERNAL_CHAT_OPEN = community_env_enabled('COMMUNITY_INTERNAL_CHAT_OPEN', False)
COMMUNITY_SOCIAL_REWARDS_OPEN = community_env_enabled('COMMUNITY_SOCIAL_REWARDS_OPEN', False)
COMMUNITY_GIFTING_OPEN = community_env_enabled('COMMUNITY_GIFTING_OPEN', False)
COMMUNITY_RESHARING_OPEN = community_env_enabled('COMMUNITY_RESHARING_OPEN', False)

def normalize_community_handle(value, allow_main_admin=False):
    handle = (value or '').strip().lower().lstrip('@')
    if not re.fullmatch(r'[a-z0-9][a-z0-9._]{2,23}', handle):
        raise OrderValidationError('Handle must be 3–24 characters using letters, numbers, dots, or underscores.')
    if handle in {'admin', 'administrator', 'cashier', 'staff', 'macleens', 'macleensfoodhouse', 'support', COMMUNITY_MAIN_ADMIN_HANDLE} and not (allow_main_admin and handle == COMMUNITY_MAIN_ADMIN_HANDLE):
        raise OrderValidationError('That handle is reserved. Please choose another one.')
    return handle

def community_channel_for_role(role):
    return 'CAMPUS' if (role or '').upper() == 'STUDENT' else 'TOWN'

def community_can_interact(profile, channel):
    normalized_channel = (channel or '').upper()
    if not profile:
        return False
    if profile.is_community_admin:
        return normalized_channel in {'GLOBAL', 'CAMPUS', 'TOWN'}
    if profile.role == 'STUDENT' and profile.verification_status != 'VERIFIED':
        return False
    return normalized_channel == 'GLOBAL' or community_channel_for_role(profile.role) == normalized_channel

def community_connection_pair(first_profile_id, second_profile_id):
    left, right = sorted((parse_int(first_profile_id, 0), parse_int(second_profile_id, 0)))
    return left, right

def community_friend_profile_ids(profile_id):
    rows = CommunityConnection.query.filter(
        CommunityConnection.status == 'ACCEPTED',
        or_(CommunityConnection.profile_a_id == profile_id, CommunityConnection.profile_b_id == profile_id),
    ).all()
    return {
        row.profile_b_id if row.profile_a_id == profile_id else row.profile_a_id
        for row in rows
    }

def community_can_view_profile(viewer, target):
    """Keep student/resident membership walls separated at the query boundary."""
    if not viewer or not target:
        return False
    if viewer.id == target.id or viewer.is_community_admin or target.is_community_admin:
        return True
    if target.role == 'STUDENT' and target.verification_status != 'VERIFIED':
        return False
    return viewer.role == target.role

def community_add_admin_notice(kind, target_key, message, profile=None):
    """Queue an idempotent private notice for the Community Admin page."""
    # New-member activity is summarized from profiles on the Admin page. Avoid
    # one database notification per join; retain urgent verification/safety notices.
    if kind == 'NEW_MEMBER':
        return
    if not CommunityAdminNotice.query.filter_by(kind=kind, target_key=target_key).first():
        db.session.add(CommunityAdminNotice(
            profile_id=profile.id if profile else None,
            kind=kind,
            target_key=target_key,
            message=message[:200],
        ))
    main_admin = CommunityProfile.query.filter_by(is_community_admin=True).first()
    member_kind = {
        'STUDENT_APPLICATION': 'STUDENT_REVIEW',
        'STUDENT_ID_RESUBMITTED': 'STUDENT_REVIEW',
    }.get(kind)
    if member_kind and main_admin and (not profile or main_admin.id != profile.id):
        if not CommunityNotification.query.filter_by(
            recipient_profile_id=main_admin.id, kind=member_kind, target_key=target_key,
        ).first():
            db.session.add(CommunityNotification(
                recipient_profile_id=main_admin.id,
                actor_profile_id=profile.id if profile else None,
                kind=member_kind,
                target_key=target_key,
                message=message[:180],
            ))

def community_suggested_profiles(viewer, followed_ids=None, limit=18):
    """Return role-safe follow suggestions without exposing private account data."""
    followed_ids = set(followed_ids or ())
    query = CommunityProfile.query.filter(
        CommunityProfile.id != viewer.id,
        CommunityProfile.verification_status != 'REJECTED',
        or_(
            CommunityProfile.role != 'STUDENT',
            CommunityProfile.verification_status == 'VERIFIED',
            CommunityProfile.is_community_admin.is_(True),
        ),
    )
    if not viewer.is_community_admin:
        query = query.filter(or_(
            CommunityProfile.role == viewer.role,
            CommunityProfile.is_community_admin.is_(True),
        ))
    candidates = [row for row in query.all() if row.id not in followed_ids]

    def suggestion_rank(target):
        same_group = False
        if viewer.role == 'STUDENT' and target.role == 'STUDENT':
            same_group = bool(
                (viewer.department and viewer.department == target.department)
                or (viewer.campus_name and viewer.campus_name == target.campus_name)
            )
        elif viewer.role == 'RESIDENT' and target.role == 'RESIDENT':
            same_group = bool(viewer.barangay and viewer.barangay == target.barangay)
        return (
            0 if target.is_community_admin else 1,
            0 if same_group else 1,
            0 if not target.is_profile_locked else 1,
            -parse_float(target.community_score, 0.0),
            target.handle.casefold(),
        )

    return sorted(candidates, key=suggestion_rank)[:limit]

def community_group_membership(profile_id, group_id, status='ACTIVE'):
    query = CommunityGroupMember.query.filter_by(profile_id=profile_id, group_id=group_id)
    return query.filter_by(status=status).first() if status else query.first()

def community_group_actor(group_id):
    cust, profile, error = community_api_actor()
    if error:
        return cust, profile, None, None, error
    group = db.session.get(CommunityGroup, group_id)
    membership = community_group_membership(profile.id, group_id, status='ACTIVE') if group else None
    if not group or not group.is_active or not membership:
        return cust, profile, group, membership, (jsonify({'success': False, 'message': 'This group workspace is unavailable or you are not an active member.'}), 403)
    if not profile.is_community_admin and group.channel != community_channel_for_role(profile.role):
        return cust, profile, group, membership, (jsonify({'success': False, 'message': 'This workspace is outside your role-locked community.'}), 403)
    if profile.role == 'STUDENT' and profile.verification_status != 'VERIFIED' and not profile.is_community_admin:
        return cust, profile, group, membership, (jsonify({'success': False, 'message': 'Student verification is required before using group workspaces.'}), 403)
    return cust, profile, group, membership, None

def community_group_target_allowed(group, target):
    if not target or target.verification_status == 'REJECTED':
        return False
    if target.role == 'STUDENT' and target.verification_status != 'VERIFIED' and not target.is_community_admin:
        return False
    return target.is_community_admin or community_channel_for_role(target.role) == group.channel

def community_group_message_payload(message):
    author = message.author
    return {
        'id': message.id,
        'body': message.body,
        'handle': f'@{author.handle}' if author else '@member',
        'profile_url': url_for('community_member_profile', handle=author.handle) if author else None,
        'avatar': author.customer.profile_image if author and author.customer else None,
        'created_at': ph_datetime_filter(message.created_at),
        'flags_count': message.flags_count or 0,
    }

def community_group_poll_payload(poll, viewer_profile_id=None):
    votes = CommunityGroupPollVote.query.filter_by(poll_id=poll.id).all()
    counts = {}
    selected = None
    for vote in votes:
        counts[vote.option_id] = counts.get(vote.option_id, 0) + 1
        if vote.profile_id == viewer_profile_id:
            selected = vote.option_id
    total = len(votes)
    return {
        'id': poll.id,
        'author_profile_id': poll.author_profile_id,
        'question': poll.question,
        'status': poll.status,
        'closes_at': ph_datetime_filter(poll.closes_at) if poll.closes_at else None,
        'selected_option_id': selected,
        'total_votes': total,
        'options': [
            {
                'id': option.id,
                'text': option.option_text,
                'votes': counts.get(option.id, 0),
                'percent': round(counts.get(option.id, 0) / total * 100) if total else 0,
            }
            for option in poll.options
        ],
    }

def community_safe_link(value, required=False):
    link = (value or '').strip()
    if not link:
        if required:
            raise OrderValidationError('Please enter a link.')
        return None
    if len(link) > 500:
        raise OrderValidationError('Link must be 500 characters or less.')
    parsed = urlparse(link)
    if parsed.scheme not in {'http', 'https'} or not parsed.netloc or parsed.username or parsed.password:
        raise OrderValidationError('Use a complete public http:// or https:// link without embedded login details.')
    return link

def community_safe_cta(value, default='/community'):
    link = (value or '').strip() or default
    if link.startswith('/') and not link.startswith('//') and len(link) <= 500:
        return link
    return community_safe_link(link, required=True)

def community_local_datetime(value, default=None):
    raw = (value or '').strip()
    if not raw:
        return default
    try:
        local_value = datetime.fromisoformat(raw)
    except ValueError:
        raise OrderValidationError('Please enter a valid date and time.')
    if local_value.tzinfo is None:
        local_value = local_value.replace(tzinfo=MANILA_TZ)
    return local_value.astimezone(timezone.utc).replace(tzinfo=None)

def community_moderation_hits(text_value):
    normalized = re.sub(r'\s+', ' ', (text_value or '').casefold()).strip()
    if not normalized:
        return []
    hits = []
    for keyword in CommunityKeyword.query.filter_by(is_active=True).all():
        phrase = re.sub(r'\s+', ' ', (keyword.phrase or '').casefold()).strip()
        if not phrase:
            continue
        pattern = r'(?<!\w)' + re.escape(phrase) + r'(?!\w)'
        if re.search(pattern, normalized):
            hits.append({'phrase': keyword.phrase, 'category': keyword.category})
    return hits

def community_image_from_request(file_key='image'):
    upload = request.files.get(file_key)
    if not upload or not upload.filename:
        return None
    declared = (upload.mimetype or '').lower()
    if declared not in {'image/jpeg', 'image/png', 'image/webp'}:
        raise OrderValidationError('Community images must be JPG, PNG, or WEBP.')
    raw = upload.read(3_000_001)
    if len(raw) > 3_000_000:
        raise OrderValidationError('Community images must be 3 MB or smaller.')
    try:
        with Image.open(io.BytesIO(raw)) as source:
            source.verify()
        with Image.open(io.BytesIO(raw)) as source:
            image = ImageOps.exif_transpose(source).convert('RGB')
            image.thumbnail((1280, 1280), Image.Resampling.LANCZOS)
            canvas = io.BytesIO()
            image.save(canvas, format='WEBP', quality=82, method=6)
    except (UnidentifiedImageError, OSError, ValueError):
        raise OrderValidationError('The uploaded community image could not be read safely.')
    encoded = base64.b64encode(canvas.getvalue()).decode('ascii')
    return f'data:image/webp;base64,{encoded}'

def community_student_id_from_request(file_key='student_id'):
    """Read a private student-ID image; never expose it through public serializers."""
    upload = request.files.get(file_key)
    if not upload or not upload.filename:
        return None
    declared = (upload.mimetype or '').lower()
    if declared not in {'image/jpeg', 'image/png', 'image/webp'}:
        raise OrderValidationError('Student ID must be a JPG, PNG, or WEBP image.')
    raw = upload.read(3_000_001)
    if len(raw) > 3_000_000:
        raise OrderValidationError('Student ID image must be 3 MB or smaller.')
    try:
        with Image.open(io.BytesIO(raw)) as source:
            source.verify()
        with Image.open(io.BytesIO(raw)) as source:
            image = ImageOps.exif_transpose(source).convert('RGB')
            image.thumbnail((1600, 1600), Image.Resampling.LANCZOS)
            canvas = io.BytesIO()
            image.save(canvas, format='WEBP', quality=84, method=6)
    except (UnidentifiedImageError, OSError, ValueError):
        raise OrderValidationError('The student ID image could not be read safely.')
    return 'data:image/webp;base64,' + base64.b64encode(canvas.getvalue()).decode('ascii')

def community_award_once(cust, event_type, target_key, amount, reason):
    """Award loyalty points once per immutable event target to prevent toggle farming."""
    if not COMMUNITY_SOCIAL_REWARDS_OPEN:
        return 0.0
    # Serialize reward writes for one loyalty account on production databases.
    cust = Customer.query.filter_by(id=cust.id).with_for_update().first() or cust
    existing = CommunityEngagementReward.query.filter_by(
        customer_id=cust.id, event_type=event_type, target_key=str(target_key),
    ).first()
    if existing:
        return 0.0
    amount = round(parse_float(amount, 0.0), 2)
    db.session.add(CommunityEngagementReward(
        customer_id=cust.id, event_type=event_type, target_key=str(target_key), points_awarded=amount,
    ))
    cust.points_balance = round(parse_float(cust.points_balance, 0.0) + amount, 2)
    db.session.add(RewardLedger(customer_id=cust.id, points_change=amount, reason=reason[:150]))
    return amount

def community_mentioned_handles(body):
    handles = []
    seen = set()
    for match in COMMUNITY_MENTION_PATTERN.findall(body or ''):
        handle = match.lower()
        if handle not in seen:
            seen.add(handle)
            handles.append(handle)
    if len(handles) > 10:
        raise OrderValidationError('A post can tag up to 10 people.')
    return handles

def community_validate_mentions(author_profile, channel, body):
    handles = community_mentioned_handles(body)
    if not handles:
        return []
    profiles = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle).in_(handles)).all()
    found = {row.handle.lower(): row for row in profiles}
    missing = [handle for handle in handles if handle not in found]
    if missing:
        raise OrderValidationError('Unknown community handle: @' + missing[0])
    allowed = []
    for handle in handles:
        target = found[handle]
        same_feed = community_channel_for_role(target.role) == channel
        if not author_profile.is_community_admin and not target.is_community_admin and not same_feed:
            raise OrderValidationError(f'@{target.handle} is not available in your community feed.')
        if target.role == 'STUDENT' and target.verification_status != 'VERIFIED' and not target.is_community_admin:
            raise OrderValidationError(f'@{target.handle} is not yet a verified student member.')
        if target.id != author_profile.id:
            allowed.append(target)
    return allowed

def community_sync_post_mentions(post, mentioned_profiles=None, notify=True):
    existing_rows = CommunityMention.query.filter_by(post_id=post.id).all()
    existing_ids = {row.mentioned_profile_id for row in existing_rows}
    if mentioned_profiles is None:
        if existing_ids:
            mentioned_profiles = CommunityProfile.query.filter(CommunityProfile.id.in_(existing_ids)).all()
        else:
            mentioned_profiles = community_validate_mentions(post.author, post.channel, post.body)
    for target in mentioned_profiles:
        if target.id not in existing_ids:
            db.session.add(CommunityMention(post_id=post.id, mentioned_profile_id=target.id))
        if notify and post.status == 'PUBLISHED':
            target_key = f'post:{post.id}'
            if not CommunityNotification.query.filter_by(recipient_profile_id=target.id, kind='MENTION', target_key=target_key).first():
                db.session.add(CommunityNotification(
                    recipient_profile_id=target.id,
                    actor_profile_id=post.author_profile_id,
                    kind='MENTION',
                    target_key=target_key,
                    message=f'@{post.author.handle} tagged you in a post.',
                ))

def get_current_community_customer():
    customer_id = parse_int(session.get('customer_id'), 0)
    if customer_id <= 0:
        return None
    cust = db.session.get(Customer, customer_id)
    if not cust or customer_access_issue(cust):
        return None
    return cust

def community_rate_limit(customer_id, model, minutes, maximum, author_field='author_profile_id', profile_id=None):
    threshold = utc_now() - timedelta(minutes=minutes)
    query = model.query.filter(model.created_at >= threshold)
    actor_value = customer_id if author_field.endswith('customer_id') else profile_id
    query = query.filter(getattr(model, author_field) == actor_value)
    if query.count() >= maximum:
        raise OrderValidationError(f'Please slow down and try again later. Limit: {maximum} actions per {minutes} minutes.')

def community_check_in(profile):
    today = ph_today()
    existing = CommunityCheckin.query.filter_by(customer_id=profile.customer_id, checkin_date=today).first()
    if existing:
        return existing, None
    yesterday = today - timedelta(days=1)
    streak = (profile.community_streak or 0) + 1 if profile.last_checkin_date == yesterday else 1
    profile.community_streak = streak
    profile.last_checkin_date = today
    profile.community_score = round(parse_float(profile.community_score, 0.0) + 1.0, 2)
    checkin = CommunityCheckin(
        customer_id=profile.customer_id,
        checkin_date=today,
        streak_day=streak,
        score_awarded=1.0,
    )
    db.session.add(checkin)
    drop = None
    if streak in {7, 30}:
        reward_type = 'NEXT_ORDER_1_5X' if streak == 7 else 'STAFF_FREEBIE'
        status = 'ACTIVE' if streak == 7 else 'PENDING_APPROVAL'
        title = '1.5× points on your next paid order' if streak == 7 else 'Staff-approved Mystery Freebie'
        period_key = f'{today.isoformat()}-{streak}'
        drop = CommunityDrop(
            customer_id=profile.customer_id,
            milestone_day=streak,
            period_key=period_key,
            reward_type=reward_type,
            reward_title=title,
            status=status,
        )
        db.session.add(drop)
    return checkin, drop

def community_poll_results(post):
    votes = CommunityPollVote.query.filter_by(post_id=post.id).all()
    option_rows = []
    role_totals = {'STUDENT': 0, 'RESIDENT': 0}
    counts = {}
    for vote in votes:
        role = vote.role_snapshot if vote.role_snapshot in role_totals else 'RESIDENT'
        role_totals[role] += 1
        key = (vote.option_id, role)
        counts[key] = counts.get(key, 0) + 1
    for option in post.poll_options:
        student = counts.get((option.id, 'STUDENT'), 0)
        resident = counts.get((option.id, 'RESIDENT'), 0)
        option_rows.append({
            'id': option.id,
            'text': option.option_text,
            'total': student + resident,
            'student': student if role_totals['STUDENT'] >= COMMUNITY_GROUP_PRIVACY_MINIMUM else None,
            'resident': resident if role_totals['RESIDENT'] >= COMMUNITY_GROUP_PRIVACY_MINIMUM else None,
        })
    return {
        'total_votes': len(votes),
        'options': option_rows,
        'student_total_visible': role_totals['STUDENT'] >= COMMUNITY_GROUP_PRIVACY_MINIMUM,
        'resident_total_visible': role_totals['RESIDENT'] >= COMMUNITY_GROUP_PRIVACY_MINIMUM,
    }

def serialize_community_post(post, viewer_customer_id=None):
    author = post.author
    original = post.reshared_post if post.reshared_post and post.reshared_post.status == 'PUBLISHED' else None
    original_author = original.author if original else None
    root_id = original.id if original else post.id
    reaction = CommunityReaction.query.filter_by(post_id=post.id, customer_id=viewer_customer_id).first() if viewer_customer_id else None
    return {
        'id': post.id,
        'channel': post.channel,
        'module': post.module,
        'post_type': post.post_type,
        'body': post.body,
        'image_data': post.image_data,
        'link_url': post.link_url,
        'status': post.status,
        'is_flash_poll': bool(post.is_flash_poll),
        'created_at': ph_datetime_filter(post.created_at),
        'author_handle': f'@{author.handle}' if author else '@macleens',
        'author_role': author.role if author else 'ADMIN',
        'author_badge': (author.department if author and author.role == 'STUDENT' else author.barangay if author else 'MacLeen’s'),
        'author_avatar': author.customer.profile_image if author and author.customer else None,
        'reaction_count': CommunityReaction.query.filter_by(post_id=post.id).count(),
        'comment_count': CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED').count(),
        'reshare_count': CommunityPost.query.filter_by(reshared_post_id=root_id, status='PUBLISHED').count(),
        'reshare_root_id': root_id,
        'original': ({
            'id': original.id,
            'body': original.body,
            'author_handle': f'@{original_author.handle}' if original_author else '@macleens',
        } if original else None),
        'viewer_reaction': reaction.reaction_type if reaction else None,
        'poll': community_poll_results(post) if post.post_type == 'POLL' else None,
    }

def community_group_leaderboards():
    profiles = CommunityProfile.query.filter(CommunityProfile.verification_status != 'REJECTED').all()
    visit_counts = dict(
        db.session.query(CommunityStoreCheckin.customer_id, db.func.count(db.distinct(CommunityStoreCheckin.checkin_date)))
        .group_by(CommunityStoreCheckin.customer_id).all()
    )
    campus = {}
    barangays = {}
    for profile in profiles:
        if profile.role == 'STUDENT' and profile.verification_status != 'VERIFIED':
            continue
        label = profile.department if profile.role == 'STUDENT' else profile.barangay
        target = campus if profile.role == 'STUDENT' else barangays
        if not label:
            continue
        row = target.setdefault(label, {'label': label, 'members': 0, 'score': 0.0, 'visits': 0})
        row['members'] += 1
        visits = parse_int(visit_counts.get(profile.customer_id), 0)
        row['visits'] += visits
        row['score'] += parse_float(profile.community_score, 0.0) + (visits * 2.0)
    def visible(rows):
        return sorted(
            (row for row in rows.values() if row['members'] >= COMMUNITY_GROUP_PRIVACY_MINIMUM),
            key=lambda row: (-row['score'], row['label'].casefold()),
        )[:8]
    return visible(campus), visible(barangays)

def active_community_ads(role, channel):
    now = utc_now()
    return CommunityAd.query.filter(
        CommunityAd.is_active.is_(True),
        CommunityAd.target_role.in_(('ALL', role)),
        CommunityAd.channel.in_(('ALL', channel)),
        ((CommunityAd.start_at.is_(None)) | (CommunityAd.start_at <= now)),
        ((CommunityAd.end_at.is_(None)) | (CommunityAd.end_at >= now)),
    ).order_by(CommunityAd.created_at.desc()).all()

def active_community_alerts(role, include_all_roles=False):
    now = utc_now()
    query = CommunityAlert.query.filter(
        CommunityAlert.is_active.is_(True),
        CommunityAlert.starts_at <= now,
        CommunityAlert.ends_at >= now,
    )
    if not include_all_roles:
        query = query.filter(CommunityAlert.target_role.in_(('ALL', role)))
    return query.order_by(CommunityAlert.created_at.desc()).all()

def send_community_alert_push(alert):
    """Send consent-based Web Push when VAPID is configured; keep in-app alerts otherwise."""
    private_key = (os.environ.get('WEBPUSH_VAPID_PRIVATE_KEY') or '').strip()
    subject = (os.environ.get('WEBPUSH_VAPID_SUBJECT') or '').strip()
    if not private_key or not subject:
        return {'configured': False, 'sent': 0, 'failed': 0}
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        app.logger.warning('Web Push is configured but pywebpush is not installed')
        return {'configured': False, 'sent': 0, 'failed': 0}

    query = CommunityPushSubscription.query.filter_by(is_active=True)
    if alert.target_role in COMMUNITY_ROLES:
        customer_ids = [row.customer_id for row in CommunityProfile.query.filter_by(role=alert.target_role).all()]
        if not customer_ids:
            return {'configured': True, 'sent': 0, 'failed': 0}
        query = query.filter(CommunityPushSubscription.customer_id.in_(customer_ids))
    payload = json.dumps({
        'title': alert.title,
        'body': alert.body,
        'url': alert.cta_url or '/community',
        'tag': f'community-alert-{alert.id}',
    })
    sent = 0
    failed = 0
    for subscription in query.all():
        try:
            webpush(
                subscription_info={
                    'endpoint': subscription.endpoint,
                    'keys': {'p256dh': subscription.p256dh, 'auth': subscription.auth},
                },
                data=payload,
                vapid_private_key=private_key,
                vapid_claims={'sub': subject},
                timeout=10,
            )
            sent += 1
        except WebPushException as exc:
            failed += 1
            status_code = getattr(getattr(exc, 'response', None), 'status_code', None)
            if status_code in {404, 410}:
                subscription.is_active = False
            app.logger.warning('Community Web Push failed for subscription %s: HTTP %s', subscription.id, status_code)
        except Exception:
            failed += 1
            app.logger.exception('Unexpected Community Web Push failure for subscription %s', subscription.id)
    db.session.commit()
    return {'configured': True, 'sent': sent, 'failed': failed}

LOYALTY_CARD_THEMES = {
    'pink-classic': 'Classic Pink',
    'cafe-cream': 'Café Cream',
    'midnight-gold': 'Midnight Gold',
    'mint-fresh': 'Mint Fresh',
    'purple-craft': 'Purple Craft',
}

def loyalty_card_scale(value, default=1.0):
    """Clamp card element scaling to a print-safe range."""
    return max(0.80, min(1.45, parse_float(value, default)))

def normalize_menu_vote_name(value):
    value = re.sub(r'[^a-z0-9]+', ' ', (value or '').lower()).strip()
    return re.sub(r'\s+', ' ', value)

def menu_vote_period_key(day=None):
    day = day or ph_today()
    return day.strftime('%Y-%m')

def ensure_menu_vote_candidate(name, category_name='Customer Requests', product_id=None, commit=False):
    clean_name = re.sub(r'\s+', ' ', (name or '').strip())[:120]
    normalized = normalize_menu_vote_name(clean_name)
    if not clean_name or not normalized:
        return None
    candidate = MenuVoteCandidate.query.filter_by(normalized_name=normalized).first()
    if candidate:
        if product_id and not candidate.product_id:
            candidate.product_id = product_id
        if category_name and (not candidate.category_name or candidate.category_name == 'Customer Requests'):
            candidate.category_name = category_name[:80]
    else:
        candidate = MenuVoteCandidate(
            name=clean_name,
            normalized_name=normalized,
            category_name=(category_name or 'Customer Requests')[:80],
            product_id=product_id,
            is_active=True,
        )
        db.session.add(candidate)
    if commit:
        db.session.commit()
    return candidate

def ensure_menu_vote_candidates():
    """Keep a non-destructive voting catalog, including past favorites like Palabok."""
    changed = False
    if not MenuVoteCandidate.query.filter_by(normalized_name='palabok').first():
        db.session.add(MenuVoteCandidate(
            name='Palabok', normalized_name='palabok', category_name='Past Favorites', is_active=True
        ))
        changed = True
    for product in Product.query.order_by(Product.id.asc()).all():
        normalized = normalize_menu_vote_name(product.name)
        if not normalized:
            continue
        candidate = MenuVoteCandidate.query.filter_by(normalized_name=normalized).first()
        if not candidate:
            db.session.add(MenuVoteCandidate(
                name=product.name[:120],
                normalized_name=normalized,
                category_name=(product.category_name or 'Menu')[:80],
                product_id=product.id,
                is_active=True,
            ))
            changed = True
        else:
            if not candidate.product_id:
                candidate.product_id = product.id
                changed = True
            if not candidate.category_name:
                candidate.category_name = (product.category_name or 'Menu')[:80]
                changed = True

    # Preserve historical votes from the older product-only voting feature once.
    db.session.flush()
    legacy_marker = StoreSetting.query.filter_by(key='menu_vote_legacy_migrated_v1').first()
    if not legacy_marker:
        for legacy_vote in MenuVote.query.order_by(MenuVote.id.asc()).all():
            product = Product.query.get(legacy_vote.product_id)
            if not product:
                continue
            candidate = MenuVoteCandidate.query.filter_by(normalized_name=normalize_menu_vote_name(product.name)).first()
            if not candidate:
                continue
            existing_vote = MenuPreferenceVote.query.filter_by(
                customer_id=legacy_vote.customer_id,
                candidate_id=candidate.id,
                period_key=legacy_vote.period_key,
            ).first()
            if not existing_vote:
                db.session.add(MenuPreferenceVote(
                    customer_id=legacy_vote.customer_id,
                    candidate_id=candidate.id,
                    period_key=legacy_vote.period_key,
                    created_at=legacy_vote.created_at or utc_now(),
                ))
        db.session.add(StoreSetting(key='menu_vote_legacy_migrated_v1', value='1'))
        changed = True
    if changed:
        db.session.commit()

def menu_vote_rankings(period_key=None, include_inactive=False):
    period_key = period_key or menu_vote_period_key()
    q = db.session.query(
        MenuVoteCandidate,
        db.func.count(MenuPreferenceVote.id).label('vote_count')
    ).outerjoin(
        MenuPreferenceVote,
        and_(
            MenuPreferenceVote.candidate_id == MenuVoteCandidate.id,
            MenuPreferenceVote.period_key == period_key,
        )
    )
    if not include_inactive:
        q = q.filter(MenuVoteCandidate.is_active.is_(True))
    return q.group_by(MenuVoteCandidate.id).order_by(
        db.func.count(MenuPreferenceVote.id).desc(), MenuVoteCandidate.name.asc()
    ).all()

def customer_profile_image_from_request(file_key='profile_photo', url_key='profile_image', existing=None):
    image_url = (request.form.get(url_key) or '').strip()
    upload = request.files.get(file_key)
    if upload and upload.filename:
        content_type = (upload.mimetype or '').lower()
        if content_type not in ('image/png', 'image/jpeg', 'image/webp', 'image/gif'):
            raise OrderValidationError('Profile photo must be PNG, JPG, WEBP, or GIF.')
        raw = upload.read(2_500_001)
        if len(raw) > 2_500_000:
            raise OrderValidationError('Profile photo must be 2.5 MB or smaller.')
        encoded = base64.b64encode(raw).decode('ascii')
        return f'data:{content_type};base64,{encoded}'
    return image_url or existing

def qr_svg_data_url(target):
    image = qrcode.make(target, image_factory=qrcode.image.svg.SvgPathImage)
    buffer = io.BytesIO()
    image.save(buffer)
    encoded = base64.b64encode(buffer.getvalue()).decode('ascii')
    return f'data:image/svg+xml;base64,{encoded}'

def customer_available_credit(cust, include_pending=True):
    limit = max(0.0, parse_float(cust.credit_limit, 0.0))
    used = max(0.0, parse_float(cust.outstanding_ar, 0.0))
    if include_pending:
        pending = db.session.query(db.func.coalesce(db.func.sum(Order.total_amount), 0.0)).filter(
            Order.customer_id == cust.id,
            Order.payment_method == 'CREDIT',
            Order.status == 'VERIFICATION',
        ).scalar() or 0.0
        used += float(pending)
    return max(0.0, limit - used)

def validate_and_lock_cart(
    raw_items,
    require_available=True,
    allow_cashier_custom_amount=False,
    allow_storefront_custom_amount=False,
):
    """Validate quantities, options, products and stock using server-side truth.

    Products may define required sub-options such as Sauce (Hot/Sweet) or Flavor.
    The same base product can appear as multiple cart lines when option choices differ.
    """
    if not isinstance(raw_items, list) or not raw_items:
        raise OrderValidationError('No items were selected.')

    raw_entries = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            raise OrderValidationError('Invalid cart item.')
        product_id = parse_int(raw.get('product_id'), 0)
        quantity = parse_int(raw.get('quantity'), 0)
        if product_id <= 0:
            raise OrderValidationError('A cart item has an invalid product ID.')
        if quantity < 1:
            raise OrderValidationError('Item quantity must be at least 1.')
        if quantity > 999:
            raise OrderValidationError('Item quantity is too large.')

        requested_price = None
        custom_amount_channel = allow_cashier_custom_amount or allow_storefront_custom_amount
        if custom_amount_channel and raw.get('unit_price') not in (None, ''):
            requested_price = round(parse_float(raw.get('unit_price'), -1.0), 2)
            if requested_price <= 0:
                raise OrderValidationError('Specific amount must be greater than ₱0.00.')

        raw_entries.append({
            'product_id': product_id,
            'quantity': quantity,
            'requested_price': requested_price,
            'raw_options': raw.get('options', {}),
        })

    product_cache = {}
    combined = {}
    aggregate_qty = {}

    for entry in raw_entries:
        product_id = entry['product_id']
        prod = product_cache.get(product_id)
        if prod is None:
            stmt = db.select(Product).where(Product.id == product_id)
            if db.engine.dialect.name != 'sqlite':
                stmt = stmt.with_for_update()
            prod = db.session.execute(stmt).scalar_one_or_none()
            if not prod:
                raise OrderValidationError(f'Product #{product_id} no longer exists.')
            product_cache[product_id] = prod

        if not prod.is_active:
            raise OrderValidationError(f'{prod.name} is currently inactive.')
        if require_available and not is_product_available_now(prod):
            raise OrderValidationError(f'{prod.name} is not available at this time.')

        base_price = round(max(0.0, parse_float(prod.price, 0.0)), 2)
        if base_price <= 0:
            raise OrderValidationError(f'{prod.name} has an invalid selling price.')
        selected_options = validate_product_options(prod, entry['raw_options'])
        unit_price = product_price_for_options(prod, selected_options)

        requested_price = entry['requested_price']
        if requested_price is not None:
            if not (allow_cashier_custom_amount or allow_storefront_custom_amount):
                raise OrderValidationError('Specific amounts are not available for this checkout channel.')
            if not bool(getattr(prod, 'allow_custom_amount', False)):
                raise OrderValidationError(f'{prod.name} does not allow a specific selling amount.')
            configured_min = getattr(prod, 'minimum_order_amount', None)
            min_amount = round(parse_float(configured_min, unit_price), 2)
            if min_amount <= 0:
                min_amount = unit_price
            if requested_price + 0.004 < min_amount:
                raise OrderValidationError(
                    f'{prod.name} cannot be sold below its ₱{min_amount:,.2f} minimum order amount.'
                )
            unit_price = requested_price

        selected_json = serialize_selected_options(selected_options)
        identity = (product_id, round(unit_price, 2), selected_json or '')
        if identity in combined:
            combined[identity]['quantity'] += entry['quantity']
            combined[identity]['subtotal'] = combined[identity]['unit_price'] * combined[identity]['quantity']
        else:
            cost_price = max(0.0, parse_float(prod.cost, 0.0))
            combined[identity] = {
                'product': prod,
                'quantity': entry['quantity'],
                'unit_price': unit_price,
                'cost_price': cost_price,
                'subtotal': unit_price * entry['quantity'],
                'selected_options': selected_options,
                'selected_options_json': selected_json,
            }
        aggregate_qty[product_id] = aggregate_qty.get(product_id, 0) + entry['quantity']

    for product_id, total_qty in aggregate_qty.items():
        prod = product_cache[product_id]
        stock = max(0, parse_int(prod.stock, 0))
        if stock < total_qty:
            raise OrderValidationError(f'Not enough stock for {prod.name}. Available: {stock}.')

    return list(combined.values())

def reserve_cart_stock(lines):
    for line in lines:
        line['product'].stock = parse_int(line['product'].stock, 0) - line['quantity']

def cart_subtotal(lines):
    return sum(line['subtotal'] for line in lines)

def staff_session_valid():
    raw = session.get('_staff_last_activity')
    if not raw:
        return False
    try:
        last = datetime.fromisoformat(raw)
    except (TypeError, ValueError):
        return False
    now = utc_now()
    if now - last > STAFF_SESSION_TIMEOUT:
        return False
    session['_staff_last_activity'] = now.isoformat()
    return True

def clear_staff_session():
    for key in ('admin_id', 'admin_user', 'cashier_id', 'cashier_user', '_staff_last_activity'):
        session.pop(key, None)

def get_store_settings():
    try:
        settings = {row.key: row.value for row in StoreSetting.query.all()}
    except Exception:
        app.logger.exception('Store settings query failed; using safe defaults for this request')
        settings = {}
    defaults = {
        'store_open_time': '08:00',
        'store_close_time': '21:00',
        'delivery_open_time': '08:00',
        'delivery_close_time': '20:00',
        'is_store_open': 'true',
        'is_delivery_enabled': 'true'
    }
    for k, v in defaults.items():
        settings.setdefault(k, v)
    return settings

def check_operating_status():
    s = get_store_settings()
    now_ph = ph_now().time()

    def parse_t(val, fallback):
        try:
            return datetime.strptime(val, '%H:%M').time()
        except Exception:
            return fallback

    store_open = parse_t(s.get('store_open_time', '08:00'), time(8, 0))
    store_close = parse_t(s.get('store_close_time', '21:00'), time(21, 0))
    del_open = parse_t(s.get('delivery_open_time', '08:00'), time(8, 0))
    del_close = parse_t(s.get('delivery_close_time', '20:00'), time(20, 0))

    is_store_active = (s.get('is_store_open') == 'true') and (store_open <= now_ph <= store_close)
    is_delivery_active = (s.get('is_delivery_enabled') == 'true') and (del_open <= now_ph <= del_close) and is_store_active

    return {
        'store_open': is_store_active,
        'delivery_open': is_delivery_active,
        'settings': s,
        'current_time': now_ph.strftime('%I:%M %p')
    }

def is_product_available_now(prod):
    if not prod.is_active:
        return False
    if not getattr(prod, 'available_start_time', None) or not getattr(prod, 'available_end_time', None):
        return True
    try:
        now_ph = ph_now().time()
        start = datetime.strptime(prod.available_start_time, '%H:%M').time()
        end = datetime.strptime(prod.available_end_time, '%H:%M').time()
        if start <= end:
            return start <= now_ph <= end
        return now_ph >= start or now_ph <= end
    except Exception:
        return True

def ensure_default_promos():
    try:
        deal1 = PromotionTracker.query.filter_by(promo_code='BURGER_FRIES_50').first()
        if not deal1:
            db.session.add(PromotionTracker(
                promo_code='BURGER_FRIES_50',
                title='1 Regular Burger + Crispy Fries',
                promo_price=50.0,
                page_views=0,
                claims_count=0,
                total_revenue=0.0,
                is_active=True,
                is_visible=True
            ))

        deal2 = PromotionTracker.query.filter_by(promo_code='BEEFY_NACHOS_75').first()
        if not deal2:
            db.session.add(PromotionTracker(
                promo_code='BEEFY_NACHOS_75',
                title='All-New Loaded Beefy Nachos Supreme',
                promo_price=75.0,
                page_views=0,
                claims_count=0,
                total_revenue=0.0,
                is_active=True,
                is_visible=True
            ))

        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Could not ensure default promotions')
        raise

@app.template_filter('ph_datetime')
def ph_datetime_filter(value, fmt='%b %d, %Y - %I:%M %p'):
    local_value = utc_naive_to_ph(value)
    return local_value.strftime(fmt) if local_value else ''

def track_portal_event(event_type, source=None, customer_id=None):
    """Best-effort portal/QR analytics. Never block a customer action if tracking fails."""
    try:
        db.session.add(PortalEvent(
            source=(source or session.get('portal_source') or 'direct')[:30],
            event_type=str(event_type)[:50],
            customer_id=customer_id,
        ))
    except Exception:
        app.logger.exception('Could not queue portal event %s', event_type)

def _parse_campaign_time(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%H:%M').time()
    except (TypeError, ValueError):
        return None

def bonus_campaign_applies(campaign, order=None, moment=None):
    if not campaign.is_active:
        return False
    if order and order.created_at:
        local_dt = utc_naive_to_ph(order.created_at)
    else:
        local_dt = moment or ph_now()
    day = local_dt.date()
    local_time = local_dt.time().replace(tzinfo=None)

    if campaign.start_date and day < campaign.start_date:
        return False
    if campaign.end_date and day > campaign.end_date:
        return False

    if campaign.weekdays:
        allowed = {x.strip() for x in campaign.weekdays.split(',') if x.strip()}
        if allowed and str(local_dt.weekday()) not in allowed:
            return False

    start = _parse_campaign_time(campaign.start_time)
    end = _parse_campaign_time(campaign.end_time)
    if start and end:
        if start <= end:
            if not (start <= local_time <= end):
                return False
        elif not (local_time >= start or local_time <= end):
            return False
    elif start and local_time < start:
        return False
    elif end and local_time > end:
        return False

    if order and (order.total_amount or 0.0) + 1e-9 < (campaign.min_spend or 0.0):
        return False
    return True

def get_active_bonus_campaigns(moment=None):
    campaigns = BonusCampaign.query.filter_by(is_active=True).order_by(BonusCampaign.created_at.desc()).all()
    return [c for c in campaigns if bonus_campaign_applies(c, moment=moment or ph_now())]

def award_active_bonus_campaigns(cust, order):
    if not cust or not order or order.id is None or order.status != 'COMPLETED':
        return 0.0
    total_bonus = 0.0
    campaigns = BonusCampaign.query.filter_by(is_active=True).all()
    for campaign in campaigns:
        if not bonus_campaign_applies(campaign, order=order):
            continue
        existing = BonusCampaignClaim.query.filter_by(campaign_id=campaign.id, order_id=order.id).first()
        if existing:
            continue
        fixed_bonus = max(0.0, float(campaign.bonus_points or 0.0))
        multiplier = max(1.0, float(campaign.points_multiplier or 1.0))
        base_points = max(0, int((order.total_amount or 0.0) // 30))
        multiplier_bonus = max(0.0, (multiplier - 1.0) * base_points)
        bonus = fixed_bonus + multiplier_bonus
        if bonus <= 0:
            continue
        cust.points_balance = (cust.points_balance or 0.0) + bonus
        db.session.add(BonusCampaignClaim(
            campaign_id=campaign.id,
            customer_id=cust.id,
            order_id=order.id,
            points_awarded=bonus,
        ))
        db.session.add(RewardLedger(
            customer_id=cust.id,
            points_change=bonus,
            reason=f"Bonus Campaign: {campaign.title} / Order #{order.id}",
        ))
        total_bonus += bonus
    return total_bonus

def award_referral_first_purchase(cust, order):
    """Reward a valid referral only after the referred member completes a paid purchase."""
    if not cust or not order or not cust.referred_by or order.status != 'COMPLETED' or not order.payment_verified:
        return 0.0, 0.0
    if ReferralReward.query.filter_by(referred_customer_id=cust.id).first():
        return 0.0, 0.0

    referrer = get_customer_by_identifier(cust.referred_by)
    if not referrer or referrer.id == cust.id or customer_access_issue(referrer):
        return 0.0, 0.0

    # Older versions awarded the referrer +2 immediately on signup. Detect that ledger entry
    # so existing accounts are never rewarded twice after this upgrade.
    legacy_reason = f"Referral Bonus: Invited {cust.name}"
    already_paid_referrer = RewardLedger.query.filter_by(customer_id=referrer.id, reason=legacy_reason).first() is not None
    referrer_points = 0.0 if already_paid_referrer else 2.0
    referred_points = 2.0

    if referrer_points:
        referrer.points_balance = (referrer.points_balance or 0.0) + referrer_points
        db.session.add(RewardLedger(
            customer_id=referrer.id,
            points_change=referrer_points,
            reason=f"Referral First-Purchase Bonus: {cust.name}",
        ))
    cust.points_balance = (cust.points_balance or 0.0) + referred_points
    db.session.add(RewardLedger(
        customer_id=cust.id,
        points_change=referred_points,
        reason=f"Referral Welcome Purchase Bonus / Order #{order.id}",
    ))
    db.session.add(ReferralReward(
        referrer_customer_id=referrer.id,
        referred_customer_id=cust.id,
        first_order_id=order.id,
        referrer_points=referrer_points,
        referred_points=referred_points,
    ))
    return referrer_points, referred_points

def award_community_drop_reward(cust, order):
    """Consume one earned 1.5x drop only on an eligible completed, paid order."""
    if not cust or not order or order.status != 'COMPLETED' or not order.payment_verified:
        return 0.0
    base_points = max(0, int(parse_float(order.total_amount, 0.0) // 30))
    if base_points <= 0:
        return 0.0
    drop = CommunityDrop.query.filter_by(
        customer_id=cust.id,
        reward_type='NEXT_ORDER_1_5X',
        status='ACTIVE',
    ).order_by(CommunityDrop.created_at.asc()).first()
    if not drop:
        return 0.0
    bonus = round(base_points * 0.5, 2)
    if bonus <= 0:
        return 0.0
    cust.points_balance = round(parse_float(cust.points_balance, 0.0) + bonus, 2)
    drop.status = 'REDEEMED'
    drop.redeemed_order_id = order.id
    drop.points_awarded = bonus
    drop.redeemed_at = utc_now()
    db.session.add(RewardLedger(
        customer_id=cust.id,
        points_change=bonus,
        reason=f'Community Mystery Drop 1.5x Bonus / Order #{order.id}',
    ))
    return bonus

def record_community_store_checkin(cust, order):
    """Count a real visit once, from an eligible completed paid member order."""
    if not cust or not order or not order.id or order.status != 'COMPLETED' or not order.payment_verified:
        return False
    if parse_float(order.total_amount, 0.0) <= 0 or (order.payment_method or '').upper() == 'REWARD':
        return False
    if (order.dining_option or '').upper() in {'DELIVERY', 'DIGITAL'} or (order.order_type or '').upper() in {'DIGITAL', 'CRAFT_ONLINE'}:
        return False
    if CommunityStoreCheckin.query.filter_by(order_id=order.id).first():
        return False
    db.session.add(CommunityStoreCheckin(
        customer_id=cust.id,
        order_id=order.id,
        checkin_date=ph_today(),
        recorded_by=((session.get('cashier_user') or session.get('admin_user')) if has_request_context() else None) or 'verified-order',
    ))
    return True

def apply_member_marketing_rewards(cust, order):
    """Apply non-base marketing rewards to a completed, paid member transaction."""
    if not cust or not order or order.status != 'COMPLETED' or not order.payment_verified:
        return {'bonus_points': 0.0, 'referral_member_points': 0.0, 'referrer_points': 0.0, 'community_drop_points': 0.0}
    bonus = award_active_bonus_campaigns(cust, order)
    referrer_pts, referred_pts = award_referral_first_purchase(cust, order)
    community_drop_points = award_community_drop_reward(cust, order)
    record_community_store_checkin(cust, order)
    return {
        'bonus_points': bonus,
        'referral_member_points': referred_pts,
        'referrer_points': referrer_pts,
        'community_drop_points': community_drop_points,
    }

def reverse_member_marketing_rewards_for_order(order):
    """Reverse bonus/referral points tied to a completed order before reassign/delete."""
    if not order or order.id is None:
        return

    CommunityStoreCheckin.query.filter_by(order_id=order.id).delete(synchronize_session=False)

    for claim in BonusCampaignClaim.query.filter_by(order_id=order.id).all():
        cust = db.session.get(Customer, claim.customer_id)
        if cust and claim.points_awarded:
            cust.points_balance = max(0.0, (cust.points_balance or 0.0) - claim.points_awarded)
            db.session.add(RewardLedger(
                customer_id=cust.id,
                points_change=-claim.points_awarded,
                reason=f"Reversed Bonus Campaign / Order #{order.id}",
            ))
        db.session.delete(claim)

    for drop in CommunityDrop.query.filter_by(redeemed_order_id=order.id, status='REDEEMED').all():
        cust = db.session.get(Customer, drop.customer_id)
        if cust and drop.points_awarded:
            cust.points_balance = max(0.0, parse_float(cust.points_balance, 0.0) - drop.points_awarded)
            db.session.add(RewardLedger(
                customer_id=cust.id,
                points_change=-drop.points_awarded,
                reason=f'Reversed Community Mystery Drop / Order #{order.id}',
            ))
        drop.status = 'ACTIVE'
        drop.redeemed_order_id = None
        drop.points_awarded = 0.0
        drop.redeemed_at = None

    referral = ReferralReward.query.filter_by(first_order_id=order.id).first()
    if referral:
        referred = db.session.get(Customer, referral.referred_customer_id)
        referrer = db.session.get(Customer, referral.referrer_customer_id)
        if referred and referral.referred_points:
            referred.points_balance = max(0.0, (referred.points_balance or 0.0) - referral.referred_points)
            db.session.add(RewardLedger(
                customer_id=referred.id,
                points_change=-referral.referred_points,
                reason=f"Reversed Referral Purchase Bonus / Order #{order.id}",
            ))
        if referrer and referral.referrer_points:
            referrer.points_balance = max(0.0, (referrer.points_balance or 0.0) - referral.referrer_points)
            db.session.add(RewardLedger(
                customer_id=referrer.id,
                points_change=-referral.referrer_points,
                reason=f"Reversed Referral Bonus / Order #{order.id}",
            ))
        db.session.delete(referral)

@app.context_processor
def inject_globals():
    try:
        setting = StoreSetting.query.filter_by(key='logo_url').first()
        logo = setting.value if setting else '/static/logo.png'
    except Exception:
        app.logger.exception('Logo setting query failed; using /static/logo.png fallback')
        logo = '/static/logo.png'
    status = check_operating_status()
    return dict(store_logo=logo, status=status, app_release=APP_RELEASE, mask_card_number=mask_card_number, product_option_groups=parse_product_option_schema, product_size_options=parse_product_size_schema, product_choice_groups=product_choice_groups, product_starting_price=product_starting_price, product_share_version=product_share_version, marketing_post_public_link=marketing_post_public_link)

def _staff_auth_failure(target, message):
    """Return JSON for fetch/API calls and a normal login redirect for browser forms."""
    clear_staff_session()
    if request.is_json or request.path.startswith('/api/'):
        return jsonify({'success': False, 'message': message}), 401
    flash(message, 'info')
    return redirect(url_for('staff_login', target=target))

def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_user') or not staff_session_valid():
            return _staff_auth_failure('admin', 'Admin session expired. Please log in again.')
        return f(*args, **kwargs)
    return decorated

def require_cashier(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not (session.get('cashier_user') or session.get('admin_user')) or not staff_session_valid():
            return _staff_auth_failure('cashier', 'Staff session expired. Please log in again.')
        return f(*args, **kwargs)
    return decorated


# ==================== CRAFT SHOP HELPERS ====================

def craft_image_from_request(file_key='image', url_key='image_url', existing=None):
    """Persist small uploaded images in the database as data URLs so Render deploys do not erase them."""
    image_url = (request.form.get(url_key) or '').strip()
    upload = request.files.get(file_key)
    if upload and upload.filename:
        content_type = (upload.mimetype or '').lower()
        if content_type not in ('image/png', 'image/jpeg', 'image/webp', 'image/gif'):
            raise OrderValidationError('Craft image must be PNG, JPG, WEBP, or GIF.')
        raw = upload.read(2_500_001)
        if len(raw) > 2_500_000:
            raise OrderValidationError('Craft image must be 2.5 MB or smaller.')
        encoded = base64.b64encode(raw).decode('ascii')
        return f'data:{content_type};base64,{encoded}'
    return image_url or existing or CRAFT_DEFAULT_IMAGE


def craft_restore_stock(craft_order):
    if not craft_order or craft_order.stock_restored:
        return
    item = db.session.get(CraftItem, craft_order.item_id)
    if item and item.availability_type == 'IN_STOCK':
        item.stock_quantity = parse_int(item.stock_quantity, 0) + parse_int(craft_order.quantity, 0)
    craft_order.stock_restored = True


def craft_add_sale_ledger(craft_order):
    if not craft_order:
        return
    existing = CraftLedger.query.filter_by(event_type='SALE', craft_order_id=craft_order.id).first()
    if existing:
        return
    db.session.add(CraftLedger(
        event_type='SALE',
        title=f'Craft Sale - {craft_order.craft_item.name if craft_order.craft_item else "Craft Order"}',
        amount=max(0.0, parse_float(craft_order.total_price, 0.0)),
        payment_method=craft_order.payment_method,
        craft_order_id=craft_order.id,
        main_order_id=craft_order.main_order_id,
        notes=f'Completed Craft Order #{craft_order.id}',
        created_by=session.get('cashier_user') or session.get('admin_user') or 'system',
    ))


def sync_craft_order_after_main_verification(main_order, accepted):
    if not main_order or not str(main_order.order_type or '').upper().startswith('CRAFT'):
        return
    craft_order = CraftOrder.query.filter_by(main_order_id=main_order.id).first()
    if not craft_order:
        return
    if accepted:
        craft_order.status = 'COMPLETED'
        craft_order.payment_status = 'PAID' if main_order.payment_method != 'CREDIT' else 'CREDIT'
        craft_order.total_price = max(0.0, parse_float(main_order.total_amount, craft_order.total_price))
        craft_order.completed_at = utc_now()
        craft_order.stock_restored = False
        craft_add_sale_ledger(craft_order)
    else:
        craft_order.status = 'CANCELLED'
        craft_order.payment_status = 'CANCELLED'
        craft_restore_stock(craft_order)


def create_main_craft_order(craft_order):
    member = db.session.get(Customer, craft_order.customer_id) if craft_order.customer_id else None
    main_order = Order(
        order_type='CRAFT',
        dining_option='TAKEOUT',
        customer_id=member.id if member else None,
        customer_name=craft_order.customer_name,
        contact_number=craft_order.contact_number,
        fb_messenger=craft_order.fb_account,
        pickup_time='Craft Shop Pickup',
        target_time='Craft Shop Pickup',
        gcash_ref=craft_order.gcash_ref,
        subtotal=craft_order.total_price,
        delivery_fee=0.0,
        total_amount=craft_order.total_price,
        payment_method=craft_order.payment_method,
        payment_verified=False,
        status='VERIFICATION',
        notes=f'[CRAFT SHOP] Craft Order #{craft_order.id}: {craft_order.notes or "No special note"}',
    )
    db.session.add(main_order)
    db.session.flush()
    db.session.add(OrderItem(
        order_id=main_order.id,
        product_id=None,
        product_name=f'[Craft] {craft_order.craft_item.name}',
        unit_price=craft_order.unit_price,
        cost_price=craft_order.unit_cost,
        quantity=craft_order.quantity,
        subtotal=craft_order.total_price,
        selected_options=None,
    ))
    craft_order.main_order_id = main_order.id
    return main_order

def create_main_digital_order(digital_order):
    main_order = Order(
        order_type='DIGITAL', dining_option='TAKEOUT', customer_id=digital_order.customer_id,
        customer_name=digital_order.customer_name, contact_number=digital_order.contact_number,
        subtotal=digital_order.total_price, delivery_fee=0.0, total_amount=digital_order.total_price,
        payment_method=digital_order.payment_method, payment_verified=False, status='VERIFICATION',
        gcash_ref=digital_order.gcash_ref,
        notes=f'[DIGITAL BUSINESS] Digital Order #{digital_order.id}: {digital_order.requirements or "No customization requirements"}',
    )
    db.session.add(main_order); db.session.flush()
    db.session.add(OrderItem(
        order_id=main_order.id, product_id=None, product_name=f'[Digital] {digital_order.item.name}',
        unit_price=digital_order.unit_price, cost_price=digital_order.unit_cost,
        quantity=digital_order.quantity, subtotal=digital_order.total_price,
    ))
    digital_order.main_order_id = main_order.id
    return main_order

def sync_digital_order_after_main_verification(main_order, accepted):
    if not main_order or str(main_order.order_type or '').upper() != 'DIGITAL':
        return
    digital_order = DigitalOrder.query.filter_by(main_order_id=main_order.id).first()
    if not digital_order:
        return
    if accepted:
        digital_order.payment_status = 'PAID'
        digital_order.status = 'IN_PROGRESS' if digital_order.item.product_type == 'CUSTOM_SERVICE' else 'PAID'
    else:
        digital_order.payment_status = 'CANCELLED'
        digital_order.status = 'CANCELLED'

def _meta_number(value):
    try:
        return float(str(value or '0').replace(',', '').replace('%', '').strip() or 0)
    except (TypeError, ValueError):
        return 0.0

def parse_meta_insight_upload(upload):
    """Read aggregate CSV/XLSX exports in memory; raw Meta files are never retained."""
    filename = (upload.filename or '').strip()
    raw = upload.read(8_000_001)
    if not filename or len(raw) > 8_000_000:
        raise OrderValidationError('Choose a Meta CSV/XLSX export no larger than 8 MB.')
    ext = os.path.splitext(filename.lower())[1]
    if ext == '.csv':
        text_data = raw.decode('utf-8-sig', errors='replace')
        rows = list(csv.DictReader(io.StringIO(text_data)))
    elif ext == '.xlsx':
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
            values = list(sheet.iter_rows(values_only=True))
            headers = [str(x or '').strip() for x in (values[0] if values else [])]
            rows = [dict(zip(headers, row)) for row in values[1:]]
        except Exception as exc:
            raise OrderValidationError(f'Could not read the XLSX export: {exc}')
    else:
        raise OrderValidationError('Meta insight upload must be a .csv or .xlsx file.')
    if not rows:
        raise OrderValidationError('The Meta export did not contain any insight rows.')
    products = Product.query.filter_by(is_active=True).all()
    aliases = {
        'reach': ('reach', 'people reached'), 'impressions': ('impressions', 'views'),
        'reactions': ('reactions', 'likes'), 'comments': ('comments',), 'shares': ('shares',),
        'clicks': ('link clicks', 'clicks'), 'date': ('publish time', 'date', 'created time'),
        'text': ('description', 'post message', 'title', 'caption'),
    }
    totals = {k: 0 for k in ('reach','impressions','reactions','comments','shares','clicks')}
    by_product, by_hour = {}, {}
    for row in rows[:5000]:
        normalized = {str(k or '').strip().casefold(): v for k, v in row.items()}
        def pick(field):
            for candidate in aliases[field]:
                for key, value in normalized.items():
                    if candidate in key:
                        return value
            return ''
        metrics = {k: _meta_number(pick(k)) for k in totals}
        for key, value in metrics.items(): totals[key] += value
        message = str(pick('text') or '').casefold()
        matched = [p.name for p in products if p.name.casefold() in message]
        for name in matched:
            bucket = by_product.setdefault(name, {'posts': 0, 'reach': 0, 'engagement': 0, 'clicks': 0})
            bucket['posts'] += 1; bucket['reach'] += metrics['reach']; bucket['clicks'] += metrics['clicks']
            bucket['engagement'] += metrics['reactions'] + metrics['comments'] + metrics['shares']
        raw_date = str(pick('date') or '')
        hour_match = re.search(r'(?:\s|T)([01]?\d|2[0-3]):', raw_date)
        if hour_match:
            hour = f'{int(hour_match.group(1)):02d}:00'
            bucket = by_hour.setdefault(hour, {'posts': 0, 'reach': 0, 'engagement': 0})
            bucket['posts'] += 1; bucket['reach'] += metrics['reach']
            bucket['engagement'] += metrics['reactions'] + metrics['comments'] + metrics['shares']
    top_products = sorted(by_product.items(), key=lambda x: (x[1]['reach'] + x[1]['engagement'] * 5), reverse=True)[:8]
    top_hours = sorted(by_hour.items(), key=lambda x: (x[1]['reach'] / max(1, x[1]['posts'])), reverse=True)[:5]
    summary = {'rows': len(rows), 'totals': totals, 'products': dict(top_products), 'best_hours': dict(top_hours)}
    if top_products:
        product_text = ', '.join(f'{name} ({int(stats["reach"])} reach, {int(stats["engagement"])} engagements)' for name, stats in top_products[:3])
    else:
        product_text = 'No active product names were reliably matched in the exported captions.'
    time_text = ', '.join(f'{hour} ({int(stats["reach"] / max(1, stats["posts"]))} average reach)' for hour, stats in top_hours[:3]) or 'No publish-time column was detected.'
    analysis = f'Product interest: {product_text}\n\nBest visibility times: {time_text}\n\nSuggestion: feature the strongest matched product during the best-performing time, then compare reach and engagement after the next post.'
    return filename[:255], summary, analysis


# ==================== AI MARKETING HELPERS ====================

MARKETING_POST_TYPES = (
    'PRODUCT_SPOTLIGHT', 'SLOW_SELLER', 'TOP_SELLER', 'NEW_OR_FEATURED',
    'LOYALTY', 'ENGAGEMENT', 'BRAND_AWARENESS', 'RESTOCK_OR_AVAILABILITY',
    'CRAFT_STORY', 'VALUE_REMINDER',
)

def marketing_setting(key, default=''):
    row = StoreSetting.query.filter_by(key=key).first()
    return row.value if row else default

def save_marketing_setting(key, value):
    value = str(value)
    row = StoreSetting.query.filter_by(key=key).first()
    if row:
        row.value = value
    else:
        db.session.add(StoreSetting(key=key, value=value))

def investor_setting(key, default=''):
    row = StoreSetting.query.filter_by(key=key).first()
    return row.value if row else default

def save_investor_setting(key, value):
    value = str(value)
    row = StoreSetting.query.filter_by(key=key).first()
    if row:
        row.value = value
    else:
        db.session.add(StoreSetting(key=key, value=value))

def marketing_settings():
    return {
        'enabled': marketing_setting('marketing_enabled', 'false') == 'true',
        # General AI posts remain approval-first/manual. Daily menu publishing has
        # its own explicit controls and environment-only Meta credentials.
        'mode': 'MANUAL_POSTING',
        'ai_provider': marketing_setting('marketing_ai_provider', 'GEMINI').upper(),
        'business_scope': marketing_setting('marketing_business_scope', 'BOTH'),
        'posts_per_week': max(1, min(7, parse_int(marketing_setting('marketing_posts_per_week', '4'), 4))),
        'start_hour': marketing_setting('marketing_start_hour', '08:00'),
        'end_hour': marketing_setting('marketing_end_hour', '19:00'),
        'repeat_cooldown_days': max(0, min(90, parse_int(marketing_setting('marketing_repeat_cooldown_days', '10'), 10))),
        'max_posts_per_day': max(1, min(3, parse_int(marketing_setting('marketing_max_posts_per_day', '1'), 1))),
        'facebook_page_name': marketing_setting('marketing_facebook_page_name', 'Macleen\'s Facebook Page'),
        'facebook_page_url': marketing_setting('marketing_facebook_page_url', ''),
        'daily_menu_enabled': marketing_setting('fb_daily_menu_enabled', 'false') == 'true',
        'daily_menu_time': marketing_setting('fb_daily_menu_time', '07:00'),
        'daily_menu_require_approval': marketing_setting('fb_daily_menu_require_approval', 'true') == 'true',
        'daily_menu_auto_page_post': marketing_setting('fb_daily_menu_auto_page_post', 'false') == 'true',
        'daily_menu_messenger_reply': marketing_setting('fb_daily_menu_messenger_reply', 'true') == 'true',
        'daily_menu_provider_enabled': marketing_setting('fb_daily_menu_provider_enabled', 'false') == 'true',
        'daily_menu_intro': marketing_setting('fb_daily_menu_intro', 'Affordable meals and snacks available today!').strip(),
        'daily_menu_page_username': marketing_setting('fb_daily_menu_page_username', '').strip(),
    }

def _marketing_public_base_url():
    configured = os.environ.get('PUBLIC_BASE_URL', '').strip().rstrip('/')
    if configured:
        return configured
    if has_request_context():
        return request.url_root.rstrip('/')
    return 'https://macleens-foodhouse-pos.onrender.com'


def _meta_graph_version():
    value = os.environ.get('META_GRAPH_VERSION', 'v24.0').strip()
    return value if re.fullmatch(r'v\d+\.\d+', value) else 'v24.0'


def facebook_page_api_ready():
    return bool(os.environ.get('META_PAGE_ID', '').strip() and os.environ.get('META_PAGE_ACCESS_TOKEN', '').strip())


def messenger_webhook_ready():
    return bool(
        facebook_page_api_ready()
        and os.environ.get('META_WEBHOOK_VERIFY_TOKEN', '').strip()
        and os.environ.get('META_APP_SECRET', '').strip()
    )


def marketing_provider_ready():
    url = os.environ.get('META_MARKETING_PROVIDER_WEBHOOK_URL', '').strip()
    return bool(url and url.startswith('https://'))


def messenger_menu_start_url():
    username = marketing_settings().get('daily_menu_page_username', '').strip().lstrip('@')
    if not username or not re.fullmatch(r'[A-Za-z0-9._-]{3,100}', username):
        return ''
    return f'https://m.me/{username}?ref=daily_menu'


def _daily_menu_products():
    """Return active, stocked Food House products without inventing availability."""
    return [
        product for product in Product.query.filter_by(is_active=True).order_by(
            Product.category_name.asc(), Product.price.desc(), Product.name.asc()
        ).all()
        if parse_int(product.stock, 0) > 0
    ]


def build_daily_menu_caption(menu_date=None, compact=False):
    """Build a deterministic menu from current product records; AI cannot invent items or prices."""
    menu_date = menu_date or ph_today()
    products = _daily_menu_products()
    if not products:
        raise OrderValidationError('No active in-stock Food House products are available for the daily menu.')

    cfg = marketing_settings()
    grouped = {}
    for product in products:
        grouped.setdefault(product.category_name or 'Menu', []).append(product)

    lines = [f"🍽️ MACLEEN'S MENU TODAY — {menu_date.strftime('%B %d, %Y')}"]
    if cfg['daily_menu_intro']:
        lines.extend(['', cfg['daily_menu_intro'][:240]])
    for category, items in grouped.items():
        lines.extend(['', f'【{category}】'])
        for product in items:
            starting_price = product_starting_price(product)
            prefix = 'From ' if product.size_schema or product.allow_custom_amount else ''
            schedule = ''
            if product.available_start_time and product.available_end_time:
                schedule = f' • {product.available_start_time}–{product.available_end_time}'
            lines.append(f'• {product.name} — {prefix}₱{starting_price:,.2f}{schedule}')

    order_url = _marketing_public_base_url() + '/'
    lines.extend(['', f'View the live menu and order: {order_url}', 'Prices and stock shown in the portal are the latest available.'])
    caption = '\n'.join(lines)
    if compact and len(caption) > 1700:
        short_lines = [lines[0], '', cfg['daily_menu_intro'][:160]]
        for category, items in grouped.items():
            short_lines.extend(['', f'【{category}】'])
            for product in items[:8]:
                price = product_starting_price(product)
                prefix = 'From ' if product.size_schema or product.allow_custom_amount else ''
                short_lines.append(f'• {product.name} — {prefix}₱{price:,.2f}')
        short_lines.extend(['', f'Complete live menu: {order_url}'])
        caption = '\n'.join(short_lines)
    return caption


def _daily_menu_run(channel, caption, triggered_by):
    run = FacebookMenuRun.query.filter_by(menu_date=ph_today(), channel=channel).first()
    if not run:
        run = FacebookMenuRun(
            menu_date=ph_today(), channel=channel, status='DRAFT', caption=caption,
            triggered_by=(triggered_by or 'system')[:50],
        )
        db.session.add(run)
    elif run.status != 'SENT':
        run.caption = caption
        run.triggered_by = (triggered_by or run.triggered_by or 'system')[:50]
        run.error_message = None
        run.status = 'DRAFT'
    db.session.commit()
    return run


def publish_daily_menu_page(caption=None, triggered_by='admin'):
    caption = (caption or '').strip() or build_daily_menu_caption()
    run = _daily_menu_run('PAGE_POST', caption[:12000], triggered_by)
    if run.status == 'SENT':
        return {'sent': False, 'status': 'already_sent', 'message': "Today's Facebook Page menu was already published.", 'run_id': run.id}

    page_id = os.environ.get('META_PAGE_ID', '').strip()
    access_token = os.environ.get('META_PAGE_ACCESS_TOKEN', '').strip()
    if not page_id or not access_token:
        run.status = 'BLOCKED'
        run.error_message = 'META_PAGE_ID and META_PAGE_ACCESS_TOKEN are not configured in Render.'
        db.session.commit()
        raise OrderValidationError(run.error_message)

    try:
        response = requests.post(
            f'https://graph.facebook.com/{_meta_graph_version()}/{page_id}/feed',
            data={'message': run.caption, 'link': _marketing_public_base_url() + '/'},
            params={'access_token': access_token},
            timeout=20,
        )
        payload = response.json() if response.content else {}
        if not response.ok or not payload.get('id'):
            message = str((payload.get('error') or {}).get('message') or f'Meta returned HTTP {response.status_code}')
            raise RuntimeError(message[:500])
        run.status = 'SENT'
        run.external_id = str(payload['id'])[:180]
        run.error_message = None
        run.sent_at = utc_now()
        db.session.commit()
        return {'sent': True, 'status': 'sent', 'message': "Today's menu was published to the Facebook Page.", 'run_id': run.id, 'external_id': run.external_id}
    except Exception as exc:
        db.session.rollback()
        run = db.session.get(FacebookMenuRun, run.id)
        run.status = 'FAILED'
        run.error_message = str(exc)[:1000]
        db.session.commit()
        app.logger.exception('Daily Facebook Page menu publishing failed')
        raise OrderValidationError(f'Facebook could not publish the menu: {exc}')


def send_daily_menu_to_provider(caption=None, triggered_by='scheduler'):
    """Hand an approved, opted-in paid campaign to a configured Meta technology provider."""
    caption = (caption or '').strip() or build_daily_menu_caption(compact=True)
    run = _daily_menu_run('PROVIDER_CAMPAIGN', caption[:12000], triggered_by)
    if run.status == 'SENT':
        return {'sent': False, 'status': 'already_sent', 'message': "Today's opted-in Messenger campaign was already handed to the provider.", 'run_id': run.id}
    webhook_url = os.environ.get('META_MARKETING_PROVIDER_WEBHOOK_URL', '').strip()
    if not webhook_url.startswith('https://'):
        run.status = 'BLOCKED'
        run.error_message = 'An HTTPS META_MARKETING_PROVIDER_WEBHOOK_URL is not configured.'
        db.session.commit()
        raise OrderValidationError(run.error_message)
    headers = {'Content-Type': 'application/json'}
    provider_token = os.environ.get('META_MARKETING_PROVIDER_TOKEN', '').strip()
    if provider_token:
        headers['Authorization'] = f'Bearer {provider_token}'
    try:
        response = requests.post(webhook_url, json={
            'event': 'macleens_daily_menu',
            'menu_date': ph_today().isoformat(),
            'page_name': marketing_settings()['facebook_page_name'],
            'message': run.caption,
            'order_url': _marketing_public_base_url() + '/',
        }, headers=headers, timeout=20)
        if not response.ok:
            raise RuntimeError(f'Provider returned HTTP {response.status_code}')
        try:
            payload = response.json()
        except ValueError:
            payload = {}
        run.status = 'SENT'
        run.external_id = str(payload.get('campaign_id') or payload.get('id') or '')[:180] or None
        run.error_message = None
        run.sent_at = utc_now()
        db.session.commit()
        return {'sent': True, 'status': 'sent', 'message': 'The opted-in Messenger campaign was handed to the configured provider.', 'run_id': run.id}
    except Exception as exc:
        db.session.rollback()
        run = db.session.get(FacebookMenuRun, run.id)
        run.status = 'FAILED'
        run.error_message = str(exc)[:1000]
        db.session.commit()
        app.logger.exception('Daily Messenger provider campaign failed')
        raise OrderValidationError(f'Messenger provider could not accept the campaign: {exc}')


def create_daily_menu_drafts(triggered_by='admin'):
    caption = build_daily_menu_caption()
    channels = ['PAGE_POST']
    if marketing_settings()['daily_menu_provider_enabled']:
        channels.append('PROVIDER_CAMPAIGN')
    runs = [_daily_menu_run(channel, caption, triggered_by) for channel in channels]
    return runs


def run_daily_menu_automation_once(force_due=False, triggered_by='scheduler'):
    cfg = marketing_settings()
    if not cfg['daily_menu_enabled']:
        return {'ran': False, 'message': 'Daily menu automation is paused.', 'channels': {}}
    try:
        scheduled = datetime.strptime(cfg['daily_menu_time'], '%H:%M').time()
    except ValueError:
        scheduled = time(7, 0)
    if not force_due and ph_now().time() < scheduled:
        return {'ran': False, 'message': f"Daily menu is scheduled for {scheduled.strftime('%I:%M %p')} PH time.", 'channels': {}}

    caption = build_daily_menu_caption()
    results = {}
    if cfg['daily_menu_require_approval']:
        for run in create_daily_menu_drafts(triggered_by):
            results[run.channel] = {'status': run.status.lower(), 'run_id': run.id}
        return {'ran': True, 'message': "Today's menu draft is ready for approval.", 'channels': results}

    if cfg['daily_menu_auto_page_post']:
        try:
            results['PAGE_POST'] = publish_daily_menu_page(caption, triggered_by)
        except OrderValidationError as exc:
            results['PAGE_POST'] = {'status': 'failed', 'message': str(exc)}
    if cfg['daily_menu_provider_enabled']:
        try:
            results['PROVIDER_CAMPAIGN'] = send_daily_menu_to_provider(caption, triggered_by)
        except OrderValidationError as exc:
            results['PROVIDER_CAMPAIGN'] = {'status': 'failed', 'message': str(exc)}
    if not results:
        return {'ran': False, 'message': 'Daily menu is enabled, but no automatic publishing channel is enabled.', 'channels': {}}
    return {'ran': True, 'message': 'Daily menu automation check completed.', 'channels': results}


def _messenger_contact(psid, source='INBOUND'):
    contact = MessengerContact.query.filter_by(psid=psid).first()
    if not contact:
        contact = MessengerContact(psid=psid[:180], source=source[:40])
        db.session.add(contact)
    contact.last_interaction_at = utc_now()
    return contact


def send_messenger_text(psid, text, message_kind='MENU_REPLY', include_tracked_menu_link=False):
    if not facebook_page_api_ready():
        raise OrderValidationError('META_PAGE_ID and META_PAGE_ACCESS_TOKEN are not configured in Render.')
    contact = _messenger_contact(psid)
    delivery = MessengerDelivery(contact=contact, message_kind=message_kind, status='QUEUED')
    db.session.add(delivery)
    db.session.commit()
    if include_tracked_menu_link:
        tracked_url = f'{_marketing_public_base_url()}/m/menu/{delivery.tracking_token}'
        text = text[:max(200, 1850 - len(tracked_url))].rstrip() + f'\n\nView the complete live menu and order: {tracked_url}\nReply STOP to stop marketing updates.'
    try:
        response = requests.post(
            f"https://graph.facebook.com/{_meta_graph_version()}/{os.environ.get('META_PAGE_ID', '').strip()}/messages",
            params={'access_token': os.environ.get('META_PAGE_ACCESS_TOKEN', '').strip()},
            json={'recipient': {'id': psid}, 'messaging_type': 'RESPONSE', 'message': {'text': text[:2000]}},
            timeout=20,
        )
        payload = response.json() if response.content else {}
        if not response.ok or not payload.get('message_id'):
            message = str((payload.get('error') or {}).get('message') or f'Meta returned HTTP {response.status_code}')
            raise RuntimeError(message[:500])
        delivery.message_id = str(payload['message_id'])[:180]
        delivery.status = 'SENT'
        delivery.sent_at = utc_now()
        if message_kind == 'MENU_REPLY':
            contact.last_menu_sent_at = utc_now()
        db.session.commit()
        return delivery
    except Exception as exc:
        db.session.rollback()
        delivery = db.session.get(MessengerDelivery, delivery.id)
        delivery.status = 'FAILED'
        delivery.error_message = str(exc)[:1000]
        db.session.commit()
        app.logger.exception('Messenger reply failed for delivery_id=%s', delivery.id)
        raise OrderValidationError(f'Messenger could not send the reply: {exc}')


def send_messenger_menu_reply(psid):
    return send_messenger_text(psid, build_daily_menu_caption(compact=True), 'MENU_REPLY', True)

def _marketing_source_link(source_kind, source_id, business):
    base = _marketing_public_base_url()
    if source_kind == 'PRODUCT' and source_id:
        product_id = int(source_id)
        prod = db.session.get(Product, product_id)
        version = product_share_version(prod) if prod else APP_RELEASE.replace('.', '-')
        return f'{base}/product/{product_id}?pv={version}'
    if source_kind == 'CRAFT_ITEM' and source_id:
        return f'{base}/craft/item/{int(source_id)}'
    return f'{base}/craft' if business == 'CRAFT' else f'{base}/'


def marketing_post_public_link(post):
    """Refresh legacy product-draft links without mutating saved post history."""
    if (getattr(post, 'source_kind', '') or '').upper() == 'PRODUCT' and getattr(post, 'source_id', None):
        return _marketing_source_link('PRODUCT', post.source_id, 'FOODHOUSE')
    return (getattr(post, 'link_url', '') or '').strip()

def _product_marketing_rows():
    cutoff = utc_now() - timedelta(days=30)
    stats = {}
    rows = (db.session.query(
                OrderItem.product_id,
                db.func.coalesce(db.func.sum(OrderItem.quantity), 0),
                db.func.coalesce(db.func.sum(OrderItem.subtotal), 0.0),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .filter(Order.status == 'COMPLETED', Order.created_at >= cutoff, OrderItem.product_id.isnot(None))
            .group_by(OrderItem.product_id).all())
    for pid, qty, revenue in rows:
        stats[int(pid)] = {'qty_30d': int(qty or 0), 'revenue_30d': round(float(revenue or 0.0), 2)}
    result = []
    for prod in Product.query.filter_by(is_active=True).order_by(Product.name.asc()).all():
        stock = parse_int(prod.stock, 0)
        if stock <= 0:
            continue
        stat = stats.get(prod.id, {'qty_30d': 0, 'revenue_30d': 0.0})
        result.append({
            'id': prod.id, 'name': prod.name, 'category': prod.category_name,
            'price': round(float(prod.price or 0.0), 2), 'stock': stock,
            'featured': bool(prod.is_featured), 'top_seller': bool(prod.is_top_seller),
            'likes': parse_int(prod.total_likes, 0), **stat,
        })
    return result

def _craft_marketing_rows():
    result = []
    for item in CraftItem.query.filter_by(is_active=True).order_by(CraftItem.name.asc()).all():
        availability = (item.availability_type or 'IN_STOCK').upper()
        if availability == 'IN_STOCK' and parse_int(item.stock_quantity, 0) <= 0:
            continue
        result.append({
            'id': item.id, 'name': item.name, 'category': item.category_name,
            'price': round(float(item.price or 0.0), 2),
            'availability': availability, 'stock': parse_int(item.stock_quantity, 0),
            'featured': bool(item.is_featured), 'top_seller': bool(item.is_top_seller),
            'likes': parse_int(item.likes, 0), 'views': parse_int(item.views, 0),
            'orders': parse_int(item.orders_count, 0),
        })
    return result

def build_marketing_context():
    recent = MarketingPost.query.order_by(MarketingPost.created_at.desc()).limit(25).all()
    active_promos = PromotionTracker.query.filter_by(is_active=True, is_visible=True, portal_only=False).all()
    active_promos = [p for p in active_promos if not p.created_at or (utc_now() - p.created_at).days <= 3]
    return {
        'foodhouse_products': _product_marketing_rows(),
        'craft_items': _craft_marketing_rows(),
        'active_promotions': [
            {'title': p.title, 'price': round(float(p.promo_price or 0.0), 2), 'code': p.promo_code}
            for p in active_promos
        ],
        'recent_posts': [
            {
                'business': p.business, 'post_type': p.post_type, 'source_kind': p.source_kind,
                'source_id': p.source_id, 'status': p.status,
                'created_at': p.created_at.isoformat() if p.created_at else None,
                'meta_insights': ({
                    'reach': p.insight_reach or 0,
                    'impressions': p.insight_impressions or 0,
                    'reactions': p.insight_reactions or 0,
                    'comments': p.insight_comments or 0,
                    'shares': p.insight_shares or 0,
                    'saves': p.insight_saves or 0,
                    'link_clicks': p.insight_link_clicks or 0,
                    'new_followers': p.insight_new_followers or 0,
                    'analysis': (p.insight_analysis or '')[:1000],
                } if p.insights_updated_at else None),
            } for p in recent
        ],
        'settings': marketing_settings(),
    }

def validate_marketing_decision(decision):
    business = str(decision.get('business') or '').upper()
    source_kind = str(decision.get('source_kind') or 'PAGE').upper()
    source_id = decision.get('source_id')
    post_type = str(decision.get('post_type') or '').upper()
    caption = str(decision.get('caption') or '').strip()
    if business not in ('FOODHOUSE', 'CRAFT'):
        raise OrderValidationError('AI returned an invalid business target.')
    if post_type not in MARKETING_POST_TYPES:
        raise OrderValidationError('AI returned an invalid marketing post type.')
    if not caption:
        raise OrderValidationError('AI returned an empty caption.')

    public_promos = PromotionTracker.query.filter_by(is_active=True, is_visible=True, portal_only=False).all()
    public_promos = [p for p in public_promos if not p.created_at or (utc_now() - p.created_at).days <= 3]
    allowed_amounts = {round(float(p.promo_price or 0.0), 2) for p in public_promos}
    if source_kind == 'PRODUCT':
        prod = db.session.get(Product, parse_int(source_id, 0))
        if not prod or not prod.is_active or parse_int(prod.stock, 0) <= 0:
            raise OrderValidationError('AI selected an unavailable Food House product.')
        if business != 'FOODHOUSE':
            raise OrderValidationError('AI mixed a Food House product with the Craft business.')
        source_id = prod.id
        allowed_amounts.add(round(float(prod.price or 0.0), 2))
    elif source_kind == 'CRAFT_ITEM':
        item = db.session.get(CraftItem, parse_int(source_id, 0))
        if not item or not item.is_active:
            raise OrderValidationError('AI selected an unavailable Craft item.')
        if (item.availability_type or 'IN_STOCK').upper() == 'IN_STOCK' and parse_int(item.stock_quantity, 0) <= 0:
            raise OrderValidationError('AI selected an out-of-stock Craft item.')
        if business != 'CRAFT':
            raise OrderValidationError('AI mixed a Craft item with the Food House business.')
        source_id = item.id
        allowed_amounts.add(round(float(item.price or 0.0), 2))
    elif source_kind == 'PAGE':
        source_id = None
    else:
        raise OrderValidationError('AI returned an invalid source type.')

    for amount in extract_peso_amounts(caption):
        if amount not in allowed_amounts:
            raise OrderValidationError(f'AI caption contained an unapproved price: ₱{amount:,.2f}.')

    cooldown = marketing_settings()['repeat_cooldown_days']
    if source_id and cooldown > 0:
        cutoff = utc_now() - timedelta(days=cooldown)
        repeated = MarketingPost.query.filter(
            MarketingPost.source_kind == source_kind,
            MarketingPost.source_id == source_id,
            MarketingPost.created_at >= cutoff,
            MarketingPost.status.in_(['APPROVED', 'POSTED']),
        ).first()
        if repeated:
            raise OrderValidationError(f'This item was already promoted within the {cooldown}-day repeat cooldown.')

    return {
        'business': business, 'post_type': post_type, 'source_kind': source_kind,
        'source_id': source_id, 'caption': caption,
        'reason': str(decision.get('reason') or '').strip(),
        'model': str(decision.get('model') or '').strip(),
    }

def create_ai_marketing_post(business_hint='AUTO', post_type_hint='AUTO', group=None, product_id=None):
    context = build_marketing_context()
    selected_product = None
    if product_id:
        selected_product = db.session.get(Product, parse_int(product_id, 0))
        if not selected_product or not selected_product.is_active:
            raise OrderValidationError('The selected product is no longer active.')
        if parse_int(selected_product.stock, 0) <= 0:
            raise OrderValidationError('The selected active product is currently out of stock.')
        context['foodhouse_products'] = [row for row in context['foodhouse_products'] if row['id'] == selected_product.id]
        context['craft_items'] = []
        business_hint = 'FOODHOUSE'
        if post_type_hint == 'AUTO':
            post_type_hint = 'PRODUCT_SPOTLIGHT'
    target_type = 'FACEBOOK_PAGE'
    group_context = None
    if group:
        target_type = 'GROUP_ASSIST'
        group_context = {
            'name': group.name,
            'url': group.group_url,
            'business_scope': group.business_scope,
            'allowed_post_types': group.post_types,
            'notes_or_rules': group.notes,
        }
        if group.business_scope in ('FOODHOUSE', 'CRAFT'):
            business_hint = group.business_scope
    decision = generate_ai_marketing_decision(
        context, business_hint, post_type_hint, group_context,
        provider=marketing_settings().get('ai_provider', 'GEMINI'),
    )
    if selected_product and decision.get('should_post'):
        decision['business'] = 'FOODHOUSE'
        decision['source_kind'] = 'PRODUCT'
        decision['source_id'] = selected_product.id
    if not decision.get('should_post'):
        post = MarketingPost(
            target_type=target_type,
            business=str(decision.get('business') or 'FOODHOUSE').upper(),
            post_type=str(decision.get('post_type') or 'BRAND_AWARENESS').upper(),
            source_kind='PAGE',
            source_id=None,
            caption=str(decision.get('caption') or 'AI chose not to post today.'),
            reason=str(decision.get('reason') or ''),
            status='SKIPPED',
            ai_model=decision.get('model'),
            group_id=group.id if group else None,
        )
        db.session.add(post)
        db.session.commit()
        return post
    safe = validate_marketing_decision(decision)
    post = MarketingPost(
        target_type=target_type,
        business=safe['business'],
        post_type=safe['post_type'],
        source_kind=safe['source_kind'],
        source_id=safe['source_id'],
        caption=safe['caption'],
        reason=safe['reason'],
        link_url=_marketing_source_link(safe['source_kind'], safe['source_id'], safe['business']),
        status='DRAFT',
        ai_model=safe['model'],
        group_id=group.id if group else None,
    )
    db.session.add(post)
    db.session.commit()
    return post

def disable_legacy_meta_connection():
    """Permanently disable/clear any old Meta API tokens; manual posting no longer needs them."""
    changed = False
    for row in MarketingFacebookPage.query.all():
        if row.is_active or row.access_token_encrypted:
            row.is_active = False
            row.access_token_encrypted = ''
            changed = True
    pending = StoreSetting.query.filter_by(key='meta_pending_user_token').first()
    if pending and pending.value:
        pending.value = ''
        changed = True
    legacy_mode = StoreSetting.query.filter_by(key='marketing_mode').first()
    if legacy_mode and legacy_mode.value != 'MANUAL_POSTING':
        legacy_mode.value = 'MANUAL_POSTING'
        changed = True
    if changed:
        db.session.commit()


def manual_facebook_page_url():
    return marketing_settings().get('facebook_page_url', '').strip()

def marketing_agent_due_now():
    cfg = marketing_settings()
    if not cfg['enabled']:
        return False, 'AI Marketing is paused.'
    now = ph_now()
    try:
        start = datetime.strptime(cfg['start_hour'], '%H:%M').time()
        end = datetime.strptime(cfg['end_hour'], '%H:%M').time()
    except ValueError:
        start, end = time(8, 0), time(19, 0)
    if not (start <= now.time() <= end):
        return False, 'Outside the allowed posting window.'
    day_start, day_end = ph_day_utc_bounds(now.date())
    today_decisions = MarketingPost.query.filter(
        MarketingPost.target_type == 'FACEBOOK_PAGE',
        MarketingPost.status.in_(['DRAFT', 'APPROVED', 'POSTED', 'SKIPPED']),
        MarketingPost.created_at >= day_start,
        MarketingPost.created_at < day_end,
    ).count()
    if today_decisions >= cfg['max_posts_per_day']:
        return False, 'Daily AI marketing decision limit already reached.'
    # Use the last Page decision, not only the last published post. This prevents an
    # hourly UptimeRobot/cron check from generating duplicate approval drafts.
    last = MarketingPost.query.filter(
        MarketingPost.target_type == 'FACEBOOK_PAGE',
        MarketingPost.status.in_(['DRAFT', 'APPROVED', 'POSTED', 'SKIPPED']),
    ).order_by(MarketingPost.created_at.desc()).first()
    min_gap_hours = max(20.0, (7.0 * 24.0) / float(cfg['posts_per_week']))
    if last and last.created_at and (utc_now() - last.created_at).total_seconds() < min_gap_hours * 3600:
        return False, 'Posting cadence is not due yet.'
    return True, 'Due'

def run_marketing_agent_once():
    due, reason = marketing_agent_due_now()
    if not due:
        return {'ran': False, 'message': reason}
    cfg = marketing_settings()
    post = create_ai_marketing_post(business_hint=cfg['business_scope'])
    if post.status == 'SKIPPED':
        return {'ran': True, 'post_id': post.id, 'message': post.reason or 'AI chose to skip.'}
    return {
        'ran': True,
        'post_id': post.id,
        'message': 'AI draft created. Copy it from Marketing and post it manually to Facebook when ready.',
    }


# ==================== REAL-TIME POLLING API ====================

@app.route('/api/queue-counts')
def api_queue_counts():
    try:
        pending_cashier = Order.query.filter_by(status="VERIFICATION").count()
        return jsonify({'pending_cashier': pending_cashier})
    except Exception:
        app.logger.exception('Queue-count query failed')
        return jsonify({'error': 'Queue data is temporarily unavailable.'}), 500

# ==================== STOREFRONT ====================

@app.route('/')
def store_catalog():
    try:
        ip = get_client_ip()
        v = SiteVisitor.query.filter_by(ip_address=ip).first()
        if not v:
            db.session.add(SiteVisitor(ip_address=ip, visit_count=1))
        else:
            v.visit_count = (v.visit_count or 0) + 1
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Storefront visitor tracking update failed')

    try:
        unique_visitors = SiteVisitor.query.count()
        total_accumulated_visits = db.session.query(db.func.coalesce(db.func.sum(SiteVisitor.visit_count), 0)).scalar() or unique_visitors
    except Exception:
        app.logger.exception('Storefront visitor metrics query failed')
        unique_visitors = 0
        total_accumulated_visits = 0

    categories = Category.query.all()
    # Visibility and orderability are intentionally separate on the public storefront.
    # If Admin marks a product Active, customers should still be able to see it even
    # outside its optional selling-time window. The time window only disables ordering.
    all_active_products = Product.query.filter_by(is_active=True).all()

    featured = [p for p in all_active_products if p.is_featured]
    top_sellers = [p for p in all_active_products if p.is_top_seller]
    products = sorted(all_active_products, key=lambda x: (-(x.total_likes or 0), x.id))

    # "Popular now" is based on real completed sales from the last 30 days.
    # It remains empty when there is not enough history—no fake popularity labels.
    recent_cutoff = utc_now() - timedelta(days=30)
    trending_rows = db.session.query(
        OrderItem.product_id,
        db.func.coalesce(db.func.sum(OrderItem.quantity), 0),
    ).join(Order, Order.id == OrderItem.order_id).filter(
        Order.status == 'COMPLETED',
        Order.created_at >= recent_cutoff,
        OrderItem.product_id.isnot(None),
    ).group_by(OrderItem.product_id).order_by(db.func.sum(OrderItem.quantity).desc()).limit(8).all()
    trending_ids = {row[0] for row in trending_rows}
    student_picks = sorted(
        [p for p in all_active_products if product_starting_price(p) <= 50 and is_product_available_now(p)],
        key=lambda p: (p.id not in trending_ids, not p.is_top_seller, product_starting_price(p), p.name.casefold()),
    )[:10]

    liked_ids = {pl.product_id for pl in ProductLike.query.filter_by(ip_address=get_client_ip()).all()}
    delivery_zones = DeliveryZone.query.filter_by(is_active=True).all()
    status = check_operating_status()
    active_promos = PromotionTracker.query.filter_by(is_active=True, is_visible=True, portal_only=False).all()
    active_promos = [p for p in active_promos if not p.created_at or (utc_now() - p.created_at).days <= 3]

    cust = None
    credit_available = 0.0
    favorite_ids = set()
    if 'customer_id' in session:
        cust = Customer.query.get(session['customer_id'])
        if cust and cust.is_credit_eligible and not customer_access_issue(cust):
            credit_available = customer_available_credit(cust, include_pending=True)
        if cust:
            favorite_ids = {
                row.product_id for row in CustomerWishlist.query.filter_by(customer_id=cust.id).all()
            }

    reorder_cart = session.pop('reorder_cart', None)

    return render_template('store_catalog.html', 
                           categories=categories, 
                           featured=featured, 
                           top_sellers=top_sellers, 
                           products=products, 
                           liked_ids=liked_ids, 
                           delivery_zones=delivery_zones, 
                           active_promos=active_promos, 
                           cust=cust, 
                           status=status, 
                           unique_visitors=unique_visitors, 
                           total_accumulated_visits=total_accumulated_visits,
                           credit_available=credit_available,
                           favorite_ids=favorite_ids,
                           trending_ids=trending_ids,
                           student_picks=student_picks,
                           reorder_cart=reorder_cart,
                           messenger_menu_url=(messenger_menu_start_url() if marketing_settings()['daily_menu_messenger_reply'] else ''),
                           product_is_available_now=is_product_available_now)

@app.route('/promo/burger-deal')
def promo_burger_deal():
    promo = PromotionTracker.query.filter_by(promo_code='BURGER_FRIES_50').first()
    if promo:
        promo.page_views = (promo.page_views or 0) + 1
        db.session.commit()
    return render_template('promo_burger_deal.html')

@app.route('/promo/beefy-nachos')
def promo_beefy_nachos():
    promo = PromotionTracker.query.filter_by(promo_code='BEEFY_NACHOS_75').first()
    if promo:
        promo.page_views = (promo.page_views or 0) + 1
        db.session.commit()
    return render_template('promo_beefy_nachos.html')

@app.route('/product/<int:product_id>')
def product_detail(product_id):
    prod = Product.query.get_or_404(product_id)
    ip = get_client_ip()
    liked = bool(ProductLike.query.filter_by(product_id=product_id, ip_address=ip).first())
    share_version = product_share_version(prod)
    crawler_agent = (request.headers.get('User-Agent') or '').casefold()
    if any(bot in crawler_agent for bot in ('facebookexternalhit', 'facebot', 'twitterbot', 'linkedinbot')):
        # Social crawlers request the HTML first. Warm the matching image now so
        # the immediately-following image request cannot time out on a remote photo.
        try:
            cached_product_social_preview(prod)
        except Exception:
            app.logger.exception('Could not warm product preview %s for social crawler', prod.id)
    return render_template(
        'product_detail.html',
        prod=prod,
        liked=liked,
        product_share_image=url_for(
            'product_social_preview', product_id=prod.id, version=share_version, _external=True
        ),
        product_share_page_url=url_for(
            'product_detail', product_id=prod.id, pv=share_version, _external=True
        ),
    )


@app.route('/social/product/<int:product_id>/<version>.jpg')
def product_social_preview(product_id, version):
    prod = Product.query.get_or_404(product_id)
    current_version = product_share_version(prod)
    payload = cached_product_social_preview(prod)
    response = Response(payload, mimetype='image/jpeg')
    response.headers['Cache-Control'] = 'public, max-age=31536000, immutable' if version == current_version else 'public, max-age=3600'
    response.headers['Content-Disposition'] = f'inline; filename="macleens-product-{product_id}-{current_version}.jpg"'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Content-Length'] = str(len(payload))
    return response


@app.route('/social/product/<int:product_id>.png')
def product_social_preview_legacy(product_id):
    """Keep old Facebook links working while moving previews to reliable JPEGs."""
    prod = Product.query.get_or_404(product_id)
    return redirect(url_for(
        'product_social_preview', product_id=prod.id, version=product_share_version(prod)
    ), code=302)

@app.route('/api/toggle-like/<int:product_id>', methods=['POST'])
def api_toggle_like(product_id):
    ip = get_client_ip()
    cust_id = session.get('customer_id')
    prod = Product.query.get_or_404(product_id)

    existing = ProductLike.query.filter_by(product_id=product_id, ip_address=ip).first()
    if existing:
        db.session.delete(existing)
        prod.total_likes = max(0, (prod.total_likes or 0) - 1)
        db.session.commit()
        return jsonify({'liked': False, 'total_likes': prod.total_likes})
    else:
        db.session.add(ProductLike(product_id=product_id, ip_address=ip, customer_id=cust_id))
        prod.total_likes = (prod.total_likes or 0) + 1
        db.session.commit()
        return jsonify({'liked': True, 'total_likes': prod.total_likes})

@app.route('/api/favorite/<int:product_id>', methods=['POST'])
def api_toggle_favorite(product_id):
    if 'customer_id' not in session:
        return jsonify({'success': False, 'message': 'Log in to save favorites.'}), 401
    cust = db.session.get(Customer, session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        return jsonify({'success': False, 'message': issue}), 403
    product = Product.query.filter_by(id=product_id, is_active=True).first_or_404()
    saved = CustomerWishlist.query.filter_by(customer_id=cust.id, product_id=product.id).first()
    if saved:
        db.session.delete(saved)
        favorited = False
    else:
        db.session.add(CustomerWishlist(customer_id=cust.id, product_id=product.id))
        favorited = True
    track_portal_event('FAVORITE_ADD' if favorited else 'FAVORITE_REMOVE', customer_id=cust.id)
    db.session.commit()
    return jsonify({'success': True, 'favorited': favorited})

@app.route('/api/add-comment/<int:product_id>', methods=['POST'])
def api_add_comment(product_id):
    ip = get_client_ip()
    cust_id = session.get('customer_id')
    data = request.get_json() or {}
    text_content = data.get('comment', '').strip()
    if not text_content:
        return jsonify({'success': False, 'message': 'Comment cannot be empty.'}), 400

    name = f"Guest ({ip})"
    if cust_id:
        cust = Customer.query.get(cust_id)
        if cust:
            name = cust.name

    comment = ProductComment(product_id=product_id, customer_id=cust_id, author_name=name, ip_address=ip, comment_text=text_content)
    db.session.add(comment)
    db.session.commit()
    return jsonify({'success': True, 'author': name, 'comment': text_content, 'created_at': 'Just now'})

@app.route('/api/storefront-checkout', methods=['POST'])
def api_storefront_checkout():
    if 'customer_id' not in session:
        return jsonify({'success': False, 'message': 'Registration / Login is required. No guest checkout.'}), 403

    cust = db.session.get(Customer, session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        return jsonify({'success': False, 'message': issue}), 403

    data = request.get_json() or {}
    order_type = str(data.get('order_type', 'PICKUP')).upper()
    dining_opt = str(data.get('dining_option', 'TAKEOUT')).upper()
    pay_method = str(data.get('payment_method', 'CASH')).upper()
    notes = str(data.get('notes', '')).strip() or 'None'
    target_time = str(data.get('target_time', '')).strip()
    zone_id = data.get('delivery_zone_id')
    landmark = str(data.get('landmark', '')).strip()
    delivery_address = str(data.get('delivery_address', '')).strip()
    gcash_ref = str(data.get('gcash_ref', '')).strip()
    fb = str(data.get('fb_messenger', '')).strip()

    if order_type not in {'PICKUP', 'DELIVERY'}:
        return jsonify({'success': False, 'message': 'Invalid order type.'}), 400
    if pay_method not in {'CASH', 'GCASH', 'CREDIT'}:
        return jsonify({'success': False, 'message': 'Invalid payment method.'}), 400

    status = check_operating_status()
    if order_type == 'PICKUP' and not status['store_open']:
        return jsonify({'success': False, 'message': 'Store ordering is currently closed.'}), 400
    if order_type == 'DELIVERY' and not status['delivery_open']:
        return jsonify({'success': False, 'message': 'Barangay delivery is currently unavailable/closed.'}), 400
    if not target_time:
        return jsonify({'success': False, 'message': 'Please provide your target time.'}), 400

    delivery_fee = 0.0
    final_address = delivery_address
    final_landmark = landmark
    if order_type == 'DELIVERY':
        dining_opt = 'DELIVERY'
        if not zone_id and (not landmark or not delivery_address):
            return jsonify({'success': False, 'message': 'Please choose a Barangay Delivery Zone or provide address info.'}), 400
        if zone_id:
            zone = db.session.get(DeliveryZone, parse_int(zone_id, 0))
            if not zone or not zone.is_active:
                return jsonify({'success': False, 'message': 'The selected delivery zone is unavailable.'}), 400
            delivery_fee = max(0.0, parse_float(zone.rate, 0.0))
            final_address = f"Barangay: {zone.barangay} ({zone.place_name})"
            final_landmark = landmark or zone.note or 'Designated Delivery Spot'
    else:
        dining_opt = 'TAKEOUT' if dining_opt not in {'DINE-IN', 'TAKEOUT'} else dining_opt

    if pay_method == 'CREDIT' and not cust.is_credit_eligible:
        return jsonify({'success': False, 'message': 'Your account is not authorized for A/R Credit.'}), 403
    if pay_method in {'GCASH', 'CREDIT'} and not fb:
        return jsonify({'success': False, 'message': 'Facebook messenger link is required for evaluation.'}), 400
    if pay_method == 'GCASH' and (len(gcash_ref) != 6 or not gcash_ref.isdigit()):
        return jsonify({'success': False, 'message': 'Please input the 6-digit GCash Reference Number.'}), 400

    try:
        lines = validate_and_lock_cart(
            data.get('items', []),
            require_available=True,
            allow_storefront_custom_amount=True,
        )
        subtotal = cart_subtotal(lines)
        points_redeemed, points_discount = calculate_points_redemption(cust, data.get('redeem_points'), subtotal)
        total = max(0.0, subtotal + delivery_fee - points_discount)

        if pay_method == 'CREDIT':
            available_credit = customer_available_credit(cust, include_pending=True)
            if total > available_credit + 1e-9:
                return jsonify({
                    'success': False,
                    'message': f'Credit limit exceeded. Available credit: ₱{available_credit:,.2f}.'
                }), 400

        change_for = parse_float(data.get('change_for'), 0.0) if pay_method == 'CASH' else 0.0
        if pay_method == 'CASH' and change_for and change_for < total:
            return jsonify({'success': False, 'message': 'Cash bill cannot be less than the order total.'}), 400

        order = Order(
            order_type=order_type,
            dining_option=dining_opt,
            customer_id=cust.id,
            customer_name=cust.name,
            contact_number=cust.contact,
            fb_messenger=fb or cust.fb_messenger,
            delivery_address=final_address if order_type == 'DELIVERY' else None,
            landmark=final_landmark if order_type == 'DELIVERY' else None,
            pickup_time=target_time if order_type == 'PICKUP' else None,
            target_time=target_time,
            change_for=change_for if pay_method == 'CASH' and change_for > 0 else None,
            gcash_ref=gcash_ref if pay_method == 'GCASH' else None,
            subtotal=subtotal,
            delivery_fee=delivery_fee,
            total_amount=total,
            points_redeemed=points_redeemed,
            points_discount=points_discount,
            payment_method=pay_method,
            payment_verified=False,
            status='VERIFICATION',
            notes=notes,
        )
        db.session.add(order)
        db.session.flush()
        record_points_redemption(cust, points_redeemed, order.id, 'Storefront points discount')

        for line in lines:
            prod = line['product']
            db.session.add(OrderItem(
                order_id=order.id,
                product_id=prod.id,
                product_name=prod.name,
                unit_price=line['unit_price'],
                cost_price=line['cost_price'],
                quantity=line['quantity'],
                subtotal=line['subtotal'],
                selected_options=line.get('selected_options_json'),
            ))
        reserve_cart_stock(lines)
        db.session.commit()
        return jsonify({'success': True, 'order_id': order.id, 'total': total,
                        'tracking_url': url_for('order_tracking', token=order.public_token),
                        'points_redeemed': points_redeemed, 'points_discount': points_discount})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Storefront checkout failed for customer_id=%s', cust.id)
        return jsonify({'success': False, 'message': 'Checkout failed due to a server error. No order was recorded.'}), 500


@app.route('/order/track/<token>')
def order_tracking(token):
    """Privacy-safe public order tracker reached through an unguessable link."""
    order = Order.query.filter_by(public_token=token).first_or_404()
    stage = (order.fulfillment_status or 'SUBMITTED').upper()
    if order.status == 'CANCELLED':
        stage = 'CANCELLED'
    stages = [
        ('SUBMITTED', 'Order sent', 'Your order is waiting for cashier confirmation.'),
        ('PREPARING', 'Preparing', 'The kitchen or counter is working on your order.'),
        ('READY', 'Ready', 'Your order is ready for pickup or delivery handoff.'),
        ('FULFILLED', 'Completed', 'Your order has been handed over.'),
    ]
    stage_index = next((i for i, row in enumerate(stages) if row[0] == stage), -1)
    return render_template('order_tracking.html', order=order, stages=stages,
                           current_stage=stage, stage_index=stage_index)


@app.route('/portal/reorder/<int:order_id>', methods=['POST'])
def customer_reorder(order_id):
    if 'customer_id' not in session:
        return redirect(url_for('customer_login'))
    cust = db.session.get(Customer, session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('customer_login'))
    order = Order.query.filter_by(id=order_id, customer_id=cust.id).first_or_404()
    reorder_cart = {}
    skipped = []
    for index, item in enumerate(order.items[:50]):
        product = db.session.get(Product, item.product_id) if item.product_id else None
        if not product or not product.is_active or not is_product_available_now(product) or product.stock < 1:
            skipped.append(item.product_name)
            continue
        try:
            raw_options = json.loads(item.selected_options) if item.selected_options else {}
            if isinstance(raw_options, dict):
                raw_options.pop('For', None)
            options = validate_product_options(product, raw_options)
            regular_price = product_price_for_options(product, options)
        except (OrderValidationError, TypeError, ValueError):
            skipped.append(item.product_name)
            continue
        unit_price = regular_price
        custom_amount = False
        if product.allow_custom_amount:
            minimum = parse_float(product.minimum_order_amount, regular_price)
            old_price = round(parse_float(item.unit_price, regular_price), 2)
            if old_price + 0.004 >= minimum:
                unit_price = old_price
                custom_amount = abs(old_price - regular_price) > 0.004
        qty = min(max(1, parse_int(item.quantity, 1)), max(1, parse_int(product.stock, 1)))
        reorder_cart[f'{product.id}::reorder-{index}'] = {
            'productId': product.id,
            'name': product.name,
            'price': unit_price,
            'regularPrice': regular_price,
            'qty': qty,
            'options': options,
            'customAmount': custom_amount,
            'minimumAmount': product.minimum_order_amount,
        }
    if not reorder_cart:
        flash('None of the items in that order are available right now.', 'info')
        return redirect(url_for('customer_dashboard') + '#orders')
    session['reorder_cart'] = reorder_cart
    if skipped:
        flash('Available items were added. Some unavailable or changed items were skipped.', 'info')
    else:
        flash('Your previous order is ready in the basket. Review it before checkout.', 'success')
    return redirect(url_for('store_catalog'))


@app.route('/portal/preferences', methods=['POST'])
def customer_preferences():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login'))
    cust = Customer.query.get_or_404(session['customer_id'])
    campus = re.sub(r'\s+', ' ', request.form.get('campus_name', '').strip())[:120] or None
    break_start = request.form.get('break_start', '').strip() or None
    break_end = request.form.get('break_end', '').strip() or None
    parsed = []
    for label, value in (('start', break_start), ('end', break_end)):
        if value:
            try:
                parsed.append(datetime.strptime(value, '%H:%M').time())
            except ValueError:
                flash(f'Break {label} time is invalid.', 'error')
                return redirect(url_for('customer_dashboard') + '#preferences')
    if len(parsed) == 2 and parsed[1] <= parsed[0]:
        flash('Break end time must be later than the start time.', 'error')
        return redirect(url_for('customer_dashboard') + '#preferences')
    cust.campus_name = campus
    cust.break_start = break_start
    cust.break_end = break_end
    cust.favorite_alerts = bool(request.form.get('favorite_alerts'))
    db.session.commit()
    flash('Campus and pickup preferences saved.', 'success')
    return redirect(url_for('customer_dashboard') + '#preferences')


@app.route('/portal/group-order/create', methods=['POST'])
def group_order_create():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login'))
    cust = Customer.query.get_or_404(session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('customer_login'))
    title = re.sub(r'\s+', ' ', request.form.get('title', '').strip())[:120] or f"{cust.name.split()[0]}'s Barkada Order"
    group = GroupOrder(organizer_customer_id=cust.id, title=title)
    db.session.add(group)
    track_portal_event('GROUP_ORDER_CREATE', customer_id=cust.id)
    db.session.commit()
    return redirect(url_for('group_order_page', token=group.token))


def _open_group_order_or_404(token):
    group = GroupOrder.query.filter_by(token=token).first_or_404()
    if group.status == 'OPEN' and group.created_at and utc_now() - group.created_at > timedelta(hours=48):
        group.status = 'EXPIRED'
        db.session.commit()
    return group


@app.route('/group-order/<token>')
def group_order_page(token):
    group = _open_group_order_or_404(token)
    products = [
        product for product in Product.query.filter_by(is_active=True).order_by(Product.name).all()
        if is_product_available_now(product) and parse_int(product.stock, 0) > 0
    ]
    organizer = session.get('customer_id') == group.organizer_customer_id
    return render_template('group_order.html', group=group, products=products, organizer=organizer)


@app.route('/group-order/<token>/item', methods=['POST'])
def group_order_add_item(token):
    group = _open_group_order_or_404(token)
    is_json = request.is_json
    data = request.get_json(silent=True) or request.form
    if group.status != 'OPEN':
        message = 'This group order is already closed.'
        return (jsonify({'success': False, 'message': message}), 409) if is_json else (flash(message, 'info') or redirect(url_for('group_order_page', token=token)))
    if len(group.lines) >= 80:
        message = 'This group order has reached its item limit.'
        return (jsonify({'success': False, 'message': message}), 400) if is_json else (flash(message, 'error') or redirect(url_for('group_order_page', token=token)))
    participant = re.sub(r'\s+', ' ', str(data.get('participant_name', '')).strip())[:100]
    if len(participant) < 2:
        message = 'Enter your name so the organizer knows which item is yours.'
        return (jsonify({'success': False, 'message': message}), 400) if is_json else (flash(message, 'error') or redirect(url_for('group_order_page', token=token)))
    try:
        raw_options = data.get('options', {})
        if isinstance(raw_options, str):
            raw_options = json.loads(raw_options or '{}')
        lines = validate_and_lock_cart([{
            'product_id': data.get('product_id'),
            'quantity': data.get('quantity', 1),
            'options': raw_options,
        }], require_available=True)
        line = lines[0]
        product = line['product']
        saved_line = GroupOrderLine(
            group_order_id=group.id,
            participant_name=participant,
            product_id=product.id,
            product_name=product.name,
            quantity=line['quantity'],
            unit_price=line['unit_price'],
            selected_options=line['selected_options_json'],
        )
        db.session.add(saved_line)
        db.session.commit()
        if is_json:
            return jsonify({
                'success': True,
                'message': f'{product.name} added for {participant}.',
                'line': {
                    'id': saved_line.id,
                    'participant_name': saved_line.participant_name,
                    'product_name': saved_line.product_name,
                    'quantity': saved_line.quantity,
                    'total': round(saved_line.unit_price * saved_line.quantity, 2),
                    'options': option_summary(saved_line.selected_options),
                },
                'line_count': GroupOrderLine.query.filter_by(group_order_id=group.id).count(),
            })
        flash(f'{product.name} added for {participant}.', 'success')
    except (OrderValidationError, TypeError, ValueError, json.JSONDecodeError) as exc:
        db.session.rollback()
        if is_json:
            return jsonify({'success': False, 'message': str(exc)}), 400
        flash(str(exc), 'error')
    return redirect(url_for('group_order_page', token=token))


@app.route('/group-order/<token>/line/<int:line_id>/delete', methods=['POST'])
def group_order_delete_line(token, line_id):
    group = _open_group_order_or_404(token)
    if session.get('customer_id') != group.organizer_customer_id:
        return jsonify({'success': False, 'message': 'Only the organizer can remove items.'}), 403
    line = GroupOrderLine.query.filter_by(id=line_id, group_order_id=group.id).first_or_404()
    if group.status != 'OPEN':
        return jsonify({'success': False, 'message': 'This group order is closed.'}), 409
    db.session.delete(line)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/group-order/<token>/submit', methods=['POST'])
def group_order_submit(token):
    group = _open_group_order_or_404(token)
    if session.get('customer_id') != group.organizer_customer_id:
        flash('Only the organizer can submit this order.', 'error')
        return redirect(url_for('group_order_page', token=token))
    if group.status != 'OPEN':
        if group.submitted_order_id:
            order = db.session.get(Order, group.submitted_order_id)
            if order:
                return redirect(url_for('order_tracking', token=order.public_token))
        flash('This group order is already closed.', 'info')
        return redirect(url_for('group_order_page', token=token))
    if not group.lines:
        flash('Add at least one item before submitting.', 'error')
        return redirect(url_for('group_order_page', token=token))
    cust = db.session.get(Customer, group.organizer_customer_id)
    target_time = request.form.get('target_time', '').strip()
    if not target_time:
        flash('Choose a pickup time.', 'error')
        return redirect(url_for('group_order_page', token=token))
    try:
        # Claim the open cart atomically so a double-click or two browser tabs
        # cannot create duplicate orders.
        claimed = GroupOrder.query.filter_by(id=group.id, status='OPEN').update(
            {'status': 'SUBMITTING'}, synchronize_session=False
        )
        if claimed != 1:
            db.session.rollback()
            flash('This group order is already being submitted.', 'info')
            return redirect(url_for('group_order_page', token=token))
        raw_items = []
        for saved in group.lines:
            raw_items.append({
                'product_id': saved.product_id,
                'quantity': saved.quantity,
                'options': json.loads(saved.selected_options) if saved.selected_options else {},
            })
        lines = validate_and_lock_cart(raw_items, require_available=True)
        subtotal = cart_subtotal(lines)
        people = ', '.join(sorted({line.participant_name for line in group.lines}))[:500]
        order = Order(
            order_type='PICKUP', dining_option='TAKEOUT', customer_id=cust.id,
            customer_name=cust.name, contact_number=cust.contact, fb_messenger=cust.fb_messenger,
            pickup_time=target_time, target_time=target_time, subtotal=subtotal, delivery_fee=0,
            total_amount=subtotal, payment_method='CASH', payment_verified=False,
            status='VERIFICATION', fulfillment_status='SUBMITTED',
            notes=f'[BARKADA ORDER] {group.title} • For: {people}',
        )
        db.session.add(order)
        db.session.flush()
        # Keep each participant visible on the cashier slip/tracker even when two
        # friends chose the same product and options. The aggregate validation
        # above still protects combined stock and the authoritative total.
        for saved in group.lines:
            product = db.session.get(Product, saved.product_id)
            selected = validate_product_options(
                product, json.loads(saved.selected_options) if saved.selected_options else {}
            )
            unit_price = product_price_for_options(product, selected)
            display_options = dict(selected)
            display_options['For'] = saved.participant_name
            db.session.add(OrderItem(
                order_id=order.id, product_id=product.id, product_name=product.name,
                unit_price=unit_price, cost_price=max(0.0, parse_float(product.cost, 0.0)), quantity=saved.quantity,
                subtotal=unit_price * saved.quantity, selected_options=serialize_selected_options(display_options),
            ))
        reserve_cart_stock(lines)
        group.status = 'SUBMITTED'
        group.submitted_order_id = order.id
        group.submitted_at = utc_now()
        db.session.commit()
        return redirect(url_for('order_tracking', token=order.public_token))
    except (OrderValidationError, TypeError, ValueError, json.JSONDecodeError) as exc:
        db.session.rollback()
        flash(f'Please review the group cart: {exc}', 'error')
        return redirect(url_for('group_order_page', token=token))

# ==================== TABLET KIOSK ENDPOINTS ====================

@app.route('/tablet')
def tablet_kiosk():
    categories = Category.query.all()
    all_active_products = Product.query.filter_by(is_active=True).all()
    available_products = [p for p in all_active_products if is_product_available_now(p)]
    return render_template('tablet.html', categories=categories, products=available_products)

@app.route('/api/tablet-checkout', methods=['POST'])
def api_tablet_checkout():
    data = request.get_json() or {}
    dining_opt = str(data.get('dining_option', 'DINE-IN')).upper()
    if dining_opt not in {'DINE-IN', 'TAKEOUT'}:
        dining_opt = 'DINE-IN'
    pay_method = str(data.get('payment_method', 'CASH')).upper()
    if pay_method not in {'CASH', 'GCASH'}:
        return jsonify({'success': False, 'message': 'Invalid kiosk payment method.'}), 400
    notes = str(data.get('notes', 'Tablet Self-Order')).strip() or 'Tablet Self-Order'

    cust = None
    member_identifier = str(data.get('member_identifier', '')).strip()
    member_pin = str(data.get('member_pin', '')).strip()
    if member_identifier or member_pin:
        if not member_identifier or not member_pin:
            return jsonify({'success': False, 'message': 'Enter both member mobile/card ID and 4-digit PIN, or leave both blank for guest checkout.'}), 400
        if not is_valid_customer_pin(member_pin):
            return jsonify({'success': False, 'message': 'Member PIN must be exactly 4 digits.'}), 400
        cust = get_customer_by_identifier(member_identifier)
        if not cust or not check_password_hash(cust.pin_hash, member_pin):
            return jsonify({'success': False, 'message': 'Invalid member ID/mobile number or PIN.'}), 403
        issue = customer_access_issue(cust)
        if issue:
            return jsonify({'success': False, 'message': issue}), 403

    try:
        lines = validate_and_lock_cart(data.get('items', []), require_available=True)
        subtotal = cart_subtotal(lines)
        order = Order(
            order_type='TABLET',
            dining_option=dining_opt,
            customer_id=cust.id if cust else None,
            customer_name=cust.name if cust else 'Tablet Kiosk Guest',
            contact_number=cust.contact if cust else 'Kiosk',
            subtotal=subtotal,
            delivery_fee=0.0,
            total_amount=subtotal,
            payment_method=pay_method,
            payment_verified=False,
            status='VERIFICATION',
            notes=notes,
        )
        db.session.add(order)
        db.session.flush()

        for line in lines:
            prod = line['product']
            db.session.add(OrderItem(
                order_id=order.id,
                product_id=prod.id,
                product_name=prod.name,
                unit_price=line['unit_price'],
                cost_price=line['cost_price'],
                quantity=line['quantity'],
                subtotal=line['subtotal'],
                selected_options=line.get('selected_options_json'),
            ))
        reserve_cart_stock(lines)
        db.session.commit()
        return jsonify({
            'success': True,
            'order_id': order.id,
            'total': subtotal,
            'member_name': cust.name if cust else None,
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Tablet checkout failed')
        return jsonify({'success': False, 'message': 'Server error. No kiosk order was recorded.'}), 500

# ==================== CASHIER TERMINAL & CLAIM DISPATCH ====================

@app.route('/pos/cashier')
@require_cashier
def cashier_terminal():
    categories = Category.query.all()
    # Cashier is an internal staff terminal, so show every active product even
    # when its public/tablet availability window is currently closed.
    products = Product.query.filter_by(is_active=True).order_by(Product.id.asc()).all()

    # These are core cashier queries. Do not hide database/schema failures behind empty panels.
    pending_orders = Order.query.filter_by(status='VERIFICATION').order_by(Order.created_at.asc()).all()
    completed_orders = Order.query.filter_by(status='COMPLETED').order_by(Order.created_at.desc()).limit(15).all()
    unpaid_collections = Order.query.filter_by(is_unpaid=True).order_by(Order.created_at.desc()).all()

    staff_list = Staff.query.all()
    customers_list = Customer.query.order_by(Customer.name.asc()).all()
    two_mins_ago = utc_now() - timedelta(minutes=2)
    online_customers = Customer.query.filter(Customer.last_active_at >= two_mins_ago).order_by(Customer.last_active_at.desc()).all()
    credit_customers = Customer.query.filter(Customer.outstanding_ar > 0).order_by(Customer.outstanding_ar.desc()).all()

    start_today, next_day = ph_day_utc_bounds()
    today_expenses = Expense.query.filter(Expense.created_at >= start_today, Expense.created_at < next_day).order_by(Expense.created_at.desc()).all()
    today_drops = VaultDrop.query.filter(VaultDrop.created_at >= start_today, VaultDrop.created_at < next_day).order_by(VaultDrop.drop_number.asc()).all()
    today_change_funds = ChangeFund.query.filter(ChangeFund.created_at >= start_today, ChangeFund.created_at < next_day).order_by(ChangeFund.created_at.desc()).all()
    next_drop_num = len(today_drops) + 1
    active_bonus_campaigns = get_active_bonus_campaigns()
    community_gift_vouchers = CommunityGift.query.filter_by(gift_type='PRODUCT', status='AVAILABLE').order_by(CommunityGift.created_at.asc()).all()
    community_mystery_drops = CommunityDrop.query.filter_by(reward_type='STAFF_FREEBIE', status='ACTIVE').order_by(CommunityDrop.created_at.asc()).all()

    return render_template(
        'cashier_pos.html',
        categories=categories,
        products=products,
        pending_orders=pending_orders,
        completed_orders=completed_orders,
        unpaid_collections=unpaid_collections,
        credit_customers=credit_customers,
        online_customers=online_customers,
        staff_list=staff_list,
        customers_list=customers_list,
        today_expenses=today_expenses,
        today_drops=today_drops,
        today_change_funds=today_change_funds,
        next_drop_num=next_drop_num,
        active_bonus_campaigns=active_bonus_campaigns,
        community_gift_vouchers=community_gift_vouchers,
        community_mystery_drops=community_mystery_drops,
    )

@app.route('/pos/direct-sale', methods=['POST'])
@require_cashier
def cashier_direct_sale():
    data = request.get_json() or {}
    cust_type = str(data.get('customer_type', 'WALKIN')).upper()
    reg_id = data.get('registered_customer_id')
    dining_opt = str(data.get('dining_option', 'DINE-IN')).upper()
    pay_method = str(data.get('payment_method', 'CASH')).upper()
    cust_name = str(data.get('customer_name', 'Counter Walk-in')).strip() or 'Counter Walk-in'
    notes = str(data.get('notes', 'Cashier Counter POS Sale')).strip() or 'Cashier Counter POS Sale'
    change_for = parse_float(data.get('change_for'), 0.0)

    if dining_opt not in {'DINE-IN', 'TAKEOUT'}:
        return jsonify({'success': False, 'message': 'Invalid dining option.'}), 400
    if pay_method not in {'CASH', 'GCASH'}:
        return jsonify({'success': False, 'message': 'Direct paid sales support Cash or GCash.'}), 400

    try:
        lines = validate_and_lock_cart(data.get('items', []), require_available=False, allow_cashier_custom_amount=True)
        subtotal = cart_subtotal(lines)

        cust = None
        points_earned = 0
        points_redeemed = 0.0
        points_discount = 0.0
        contact = 'N/A'
        if cust_type == 'REGISTERED':
            cust = db.session.get(Customer, parse_int(reg_id, 0))
            if not cust:
                raise OrderValidationError('Please select a valid registered member.')
            issue = customer_access_issue(cust)
            if issue:
                raise OrderValidationError(issue)
            cust_name = cust.name
            contact = cust.contact
            points_redeemed, points_discount = calculate_points_redemption(cust, data.get('redeem_points'), subtotal)
        elif parse_float(data.get('redeem_points'), 0.0) > 0:
            raise OrderValidationError('Select a registered member before redeeming points.')

        total = max(0.0, subtotal - points_discount)
        if pay_method == 'CASH' and change_for and change_for < total:
            raise OrderValidationError('Cash bill cannot be less than the discounted sale total.')

        if cust:
            cust.accumulated_spend = (cust.accumulated_spend or 0.0) + total
            points_earned = int(total // 30)
            if points_earned > 0:
                cust.points_balance = (cust.points_balance or 0.0) + points_earned
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=points_earned,
                    reason=f'Counter POS Sale (₱{total:,.2f} paid)',
                ))

        order = Order(
            order_type='COUNTER_SALE',
            dining_option=dining_opt,
            customer_id=cust.id if cust else None,
            customer_name=cust_name,
            contact_number=contact,
            subtotal=subtotal,
            delivery_fee=0.0,
            total_amount=total,
            points_redeemed=points_redeemed,
            points_discount=points_discount,
            payment_method=pay_method,
            payment_verified=True,
            change_for=change_for if change_for > 0 else None,
            status='COMPLETED',
            fulfillment_status='FULFILLED',
            notes=notes,
        )
        db.session.add(order)
        db.session.flush()
        if cust:
            record_points_redemption(cust, points_redeemed, order.id, 'Counter POS points discount')

        for line in lines:
            prod = line['product']
            db.session.add(OrderItem(
                order_id=order.id,
                product_id=prod.id,
                product_name=prod.name,
                unit_price=line['unit_price'],
                cost_price=line['cost_price'],
                quantity=line['quantity'],
                subtotal=line['subtotal'],
                selected_options=line.get('selected_options_json'),
            ))
        reserve_cart_stock(lines)
        marketing = apply_member_marketing_rewards(cust, order) if cust else {'bonus_points': 0.0, 'referral_member_points': 0.0, 'referrer_points': 0.0}
        db.session.commit()
        return jsonify({
            'success': True,
            'order_id': order.id,
            'subtotal': subtotal,
            'total': total,
            'points_earned': points_earned,
            'points_redeemed': points_redeemed,
            'points_discount': points_discount,
            'bonus_points': marketing['bonus_points'],
            'referral_points': marketing['referral_member_points'],
            'member_balance': (cust.points_balance or 0.0) if cust else None,
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Cashier direct sale failed')
        return jsonify({'success': False, 'message': 'Sale failed due to a server error. Nothing was recorded.'}), 500

@app.route('/pos/claim-promo', methods=['POST'])
@require_cashier
def cashier_claim_promo():
    promo_code = request.form.get('promo_code', 'BURGER_FRIES_50')
    reg_id = request.form.get('registered_customer_id')
    dining_opt = request.form.get('dining_option', 'DINE-IN').upper()
    pay_method = request.form.get('payment_method', 'CASH').upper()
    
    promo = PromotionTracker.query.filter_by(promo_code=promo_code).first()
    if not promo or not promo.is_active:
        flash("Promotion deal is not active or could not be found.", "error")
        return redirect(url_for('cashier_terminal'))
    if dining_opt not in {'DINE-IN', 'TAKEOUT'}:
        flash('Invalid dining option.', 'error')
        return redirect(url_for('cashier_terminal'))
    if pay_method not in {'CASH', 'GCASH'}:
        flash('Promo sales support Cash or GCash only.', 'error')
        return redirect(url_for('cashier_terminal'))

    cust_id = None
    cust_name = 'Walk-in Member'
    contact = 'N/A'
    points_earned = 0

    if reg_id:
        cust = Customer.query.get(reg_id)
        if cust:
            issue = customer_access_issue(cust)
            if issue:
                flash(issue, 'error')
                return redirect(url_for('cashier_terminal'))
            cust_id = cust.id
            cust_name = cust.name
            contact = cust.contact
            cust.accumulated_spend = (cust.accumulated_spend or 0.0) + promo.promo_price
            points_earned = int(promo.promo_price // 30)
            if points_earned > 0:
                cust.points_balance = (cust.points_balance or 0.0) + points_earned
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=points_earned,
                    reason=f"Promo Deal: {promo.title} (₱{promo.promo_price:,.2f})"
                ))

    order = Order(
        order_type='PROMO_DEAL',
        dining_option=dining_opt,
        customer_id=cust_id,
        customer_name=cust_name,
        contact_number=contact,
        subtotal=promo.promo_price,
        delivery_fee=0.0,
        total_amount=promo.promo_price,
        payment_method=pay_method,
        payment_verified=True,
        status='COMPLETED',
        fulfillment_status='FULFILLED',
        notes=f"[PROMO:{promo.promo_code}] {promo.title}"
    )
    db.session.add(order)
    db.session.flush()

    db.session.add(OrderItem(
        order_id=order.id,
        product_id=None,
        product_name=f"[PROMO] {promo.title}",
        unit_price=promo.promo_price,
        cost_price=max(0.0, parse_float(promo.promo_cost, 0.0)),
        quantity=1,
        subtotal=promo.promo_price
    ))

    promo.claims_count = (promo.claims_count or 0) + 1
    promo.total_revenue = (promo.total_revenue or 0.0) + promo.promo_price
    if reg_id and cust_id:
        apply_member_marketing_rewards(cust, order)

    db.session.commit()
    flash(f"🎉 Promo Deal '{promo.title}' recorded ({dining_opt}) for {cust_name} (₱{promo.promo_price:,.2f})!", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/create-reservation', methods=['POST'])
@require_cashier
def cashier_create_reservation():
    cust_type = request.form.get('customer_type', 'REGISTERED')
    reg_id = request.form.get('registered_customer_id')
    product_id = parse_int(request.form.get('product_id'), 0)
    qty = parse_int(request.form.get('quantity'), 1)
    dining_opt = request.form.get('dining_option', 'TAKEOUT').upper()
    target_time = request.form.get('target_time', '').strip() or 'Today'
    pay_method = request.form.get('payment_method', 'CASH').upper()
    notes = request.form.get('notes', '').strip() or 'In-Store Reservation'

    try:
        lines = validate_and_lock_cart([{'product_id': product_id, 'quantity': qty, 'options': request.form.get('selected_options', '')}], require_available=True)
        line = lines[0]
        prod = line['product']
        subtotal = line['subtotal']

        cust = None
        cust_name = 'Walk-in Guest'
        contact = 'N/A'
        if cust_type == 'REGISTERED':
            cust = db.session.get(Customer, parse_int(reg_id, 0))
            if not cust:
                raise OrderValidationError('Please select a registered member.')
            issue = customer_access_issue(cust)
            if issue:
                raise OrderValidationError(issue)
            cust_name = cust.name
            contact = cust.contact
        else:
            cust_name = request.form.get('custom_customer_name', '').strip() or 'Walk-in Reservation'
            contact = request.form.get('custom_contact', '').strip() or 'N/A'

        order = Order(
            order_type='RESERVATION',
            dining_option=dining_opt if dining_opt in {'DINE-IN', 'TAKEOUT'} else 'TAKEOUT',
            customer_id=cust.id if cust else None,
            customer_name=f'{cust_name} (Reserved: {target_time})',
            contact_number=contact,
            pickup_time=target_time,
            target_time=target_time,
            subtotal=subtotal,
            delivery_fee=0.0,
            total_amount=subtotal,
            payment_method=pay_method if pay_method in {'CASH', 'GCASH'} else 'CASH',
            payment_verified=False,
            status='VERIFICATION',
            notes=f'[{dining_opt} - RESERVED for {target_time}] {notes}',
        )
        db.session.add(order)
        db.session.flush()
        db.session.add(OrderItem(
            order_id=order.id,
            product_id=prod.id,
            product_name=prod.name,
            unit_price=line['unit_price'],
            cost_price=line['cost_price'],
            quantity=qty,
            subtotal=subtotal,
            selected_options=line.get('selected_options_json'),
        ))
        reserve_cart_stock(lines)
        db.session.commit()
        flash(f'📌 Reserved {prod.name} x{qty} ({dining_opt}) for {cust_name} (Pickup: {target_time}) — ₱{subtotal:,.2f}', 'success')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    except Exception:
        db.session.rollback()
        app.logger.exception('Cashier reservation failed')
        flash('Reservation failed due to a server error. Nothing was recorded.', 'error')
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/misc-sale', methods=['POST'])
@require_cashier
def cashier_misc_sale():
    service_name = request.form.get('service_name', '').strip() or 'Printing / Custom Service'
    amount = parse_float(request.form.get('amount'), 0.0)
    pay_method = request.form.get('payment_method', 'CASH').upper()
    notes = request.form.get('notes', '').strip() or 'Over-the-counter Misc Service'
    reg_cust_id = request.form.get('registered_customer_id')
    custom_name = request.form.get('custom_customer_name', '').strip()

    if amount <= 0:
        flash("Amount must be greater than zero.", "error")
        return redirect(url_for('cashier_terminal'))
    if pay_method not in {'CASH', 'GCASH'}:
        flash('Misc sales support Cash or GCash only.', 'error')
        return redirect(url_for('cashier_terminal'))

    cust_id = None
    customer_name = 'Walk-in Customer'
    contact = 'N/A'

    if reg_cust_id:
        cust = Customer.query.get(reg_cust_id)
        if cust:
            issue = customer_access_issue(cust)
            if issue:
                flash(issue, 'error')
                return redirect(url_for('cashier_terminal'))
            cust_id = cust.id
            customer_name = cust.name
            contact = cust.contact
            cust.accumulated_spend = (cust.accumulated_spend or 0.0) + amount
            earned = int(amount // 30)
            if earned > 0:
                cust.points_balance = (cust.points_balance or 0.0) + earned
                db.session.add(RewardLedger(customer_id=cust.id, points_change=earned, reason=f"Service Purchase: {service_name}"))
    elif custom_name:
        customer_name = custom_name

    order = Order(
        order_type='SERVICE/MISC',
        dining_option='SERVICE',
        customer_id=cust_id,
        customer_name=customer_name,
        contact_number=contact,
        subtotal=amount,
        delivery_fee=0.0,
        total_amount=amount,
        payment_method=pay_method,
        payment_verified=True,
        status='COMPLETED',
        fulfillment_status='FULFILLED',
        notes=notes
    )
    db.session.add(order)
    db.session.flush()

    db.session.add(OrderItem(
        order_id=order.id,
        product_id=None,
        product_name=f"[Service] {service_name}",
        unit_price=amount,
        cost_price=0.0,
        quantity=1,
        subtotal=amount
    ))
    if reg_cust_id and cust_id:
        apply_member_marketing_rewards(cust, order)

    db.session.commit()
    flash(f"Misc Sale recorded: {service_name} for {customer_name} (₱{amount:,.2f})", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/record-change-fund', methods=['POST'])
@require_cashier
def cashier_record_change_fund():
    title = request.form.get('fund_title', 'Opening Petty/Change Fund').strip()
    amount = parse_float(request.form.get('amount'), 0.0)
    notes = request.form.get('notes', 'Ulam / Register Drawer Starting Cash').strip()
    staff_user = session.get('cashier_user') or session.get('admin_user') or 'Cashier'

    if amount <= 0:
        flash("Change fund amount must be greater than zero.", "error")
        return redirect(url_for('cashier_terminal'))

    db.session.add(ChangeFund(fund_title=title, amount=amount, notes=notes, created_by=staff_user))
    db.session.commit()
    flash(f"Change Fund (₱{amount:,.2f}) added to register drawer notes.", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/create-collection', methods=['POST'])
@require_cashier
def cashier_create_collection():
    cust_type = request.form.get('customer_type', 'REGISTERED')
    item_choice_type = request.form.get('item_choice_type', 'PRODUCT')
    product_id = parse_int(request.form.get('product_id'), 0)
    qty = parse_int(request.form.get('quantity'), 1)
    dining_opt = request.form.get('dining_option', 'TAKEOUT').upper()
    notes = request.form.get('notes', '').strip()

    try:
        lines = None
        prod = None
        if item_choice_type == 'PRODUCT':
            lines = validate_and_lock_cart([{'product_id': product_id, 'quantity': qty, 'options': request.form.get('selected_options', '')}], require_available=True)
            line = lines[0]
            prod = line['product']
            prod_name = prod.name
            unit_p = line['unit_price']
            cost_p = line['cost_price']
            amount = line['subtotal']
        else:
            if qty < 1:
                raise OrderValidationError('Quantity must be at least 1.')
            prod_name = request.form.get('custom_title', '').strip() or 'Custom Receivable Service'
            unit_p = parse_float(request.form.get('custom_amount'), 0.0)
            cost_p = 0.0
            amount = unit_p * qty
            if amount <= 0:
                raise OrderValidationError('Amount must be greater than zero.')

        cust = None
        cust_name = 'Walk-in Customer'
        contact = 'N/A'
        if cust_type == 'REGISTERED':
            cust = db.session.get(Customer, parse_int(request.form.get('registered_customer_id'), 0))
            if not cust:
                raise OrderValidationError('Please select a registered member.')
            issue = customer_access_issue(cust)
            if issue:
                raise OrderValidationError(issue)
            if not cust.is_credit_eligible:
                raise OrderValidationError('This member is not authorized for A/R credit.')
            available = customer_available_credit(cust, include_pending=True)
            if amount > available + 1e-9:
                raise OrderValidationError(f'Credit limit exceeded. Available credit: ₱{available:,.2f}.')
            cust_name = cust.name
            contact = cust.contact
            cust.outstanding_ar = (cust.outstanding_ar or 0.0) + amount
        else:
            cust_name = request.form.get('custom_customer_name', '').strip() or 'Custom Account'
            contact = request.form.get('custom_contact', '').strip() or 'N/A'

        order = Order(
            order_type='COLLECTION',
            dining_option=dining_opt if dining_opt in {'DINE-IN', 'TAKEOUT'} else 'TAKEOUT',
            customer_id=cust.id if cust else None,
            customer_name=cust_name,
            contact_number=contact,
            subtotal=amount,
            delivery_fee=0.0,
            total_amount=amount,
            payment_method='UNPAID',
            payment_verified=False,
            status='UNPAID_COLLECTION',
            is_unpaid=True,
            collection_notes=notes,
            notes=f'[{dining_opt}] Attributable Item: {prod_name} (x{qty})',
        )
        db.session.add(order)
        db.session.flush()
        db.session.add(OrderItem(
            order_id=order.id,
            product_id=prod.id if prod else None,
            product_name=prod_name,
            unit_price=unit_p,
            cost_price=cost_p,
            quantity=qty,
            subtotal=amount,
            selected_options=line.get('selected_options_json') if lines else None,
        ))
        if lines:
            reserve_cart_stock(lines)
        db.session.commit()
        flash(f'For Collection ({dining_opt}) recorded for {cust_name}: {prod_name} x{qty} (₱{amount:,.2f})', 'info')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    except Exception:
        db.session.rollback()
        app.logger.exception('For Collection entry failed')
        flash('For Collection entry failed due to a server error.', 'error')
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/settle-collection/<int:order_id>', methods=['POST'])
@require_cashier
def cashier_settle_collection(order_id):
    order = Order.query.get_or_404(order_id)
    pay_method = request.form.get('payment_method', 'CASH').upper()
    if pay_method not in {'CASH', 'GCASH'}:
        flash('Settlement payment must be Cash or GCash.', 'error')
        return redirect(url_for('cashier_terminal'))
    if not order.is_unpaid:
        flash(f'Order #{order.id} has already been settled.', 'info')
        return redirect(url_for('cashier_terminal'))
    
    order.is_unpaid = False
    order.status = 'COMPLETED'
    order.payment_method = pay_method
    order.payment_verified = True

    is_same_day = (utc_naive_to_ph(order.created_at).date() == ph_today())
    earned = 0

    if order.customer_id:
        cust = Customer.query.get(order.customer_id)
        if cust:
            cust.outstanding_ar = max(0.0, (cust.outstanding_ar or 0.0) - order.total_amount)
            cust.accumulated_spend = (cust.accumulated_spend or 0.0) + order.total_amount
            
            if is_same_day:
                earned = int(order.total_amount // 30)
                if earned > 0:
                    cust.points_balance = (cust.points_balance or 0.0) + earned
                    db.session.add(RewardLedger(customer_id=cust.id, points_change=earned, reason=f"Same-Day Settled Credit #{order.id}"))
            # Referral and scheduled campaign rewards are tied to the original purchase,
            # so they can be released once the credit order is actually paid.
            apply_member_marketing_rewards(cust, order)

    db.session.commit()
    bonus_msg = f" (+{earned} pts earned for Same-Day payment!)" if earned > 0 else " (No points: paid after order date)"
    flash(f"Collection #{order.id} for {order.customer_name} settled via {pay_method}!{bonus_msg}", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/settle-customer-credit/<int:cust_id>', methods=['POST'])
@require_cashier
def cashier_settle_customer_credit(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    amount = parse_float(request.form.get('amount'), cust.outstanding_ar or 0.0)
    pay_method = request.form.get('payment_method', 'CASH').upper()

    if pay_method not in {'CASH', 'GCASH'}:
        flash('Credit settlement payment must be Cash or GCash.', 'error')
        return redirect(url_for('cashier_terminal'))
    if (cust.outstanding_ar or 0.0) <= 0:
        flash(f'{cust.name} has no outstanding credit balance.', 'info')
        return redirect(url_for('cashier_terminal'))
    if amount <= 0:
        flash("Amount must be greater than zero.", "error")
        return redirect(url_for('cashier_terminal'))

    amount_to_pay = min(amount, cust.outstanding_ar)
    cust.outstanding_ar = max(0.0, cust.outstanding_ar - amount_to_pay)
    cust.accumulated_spend = (cust.accumulated_spend or 0.0) + amount_to_pay

    db.session.commit()
    flash(f"Settled ₱{amount_to_pay:,.2f} credit balance for {cust.name} via {pay_method}!", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/record-expense', methods=['POST'])
@require_cashier
def cashier_record_expense():
    title = request.form.get('title', '').strip()
    amount = parse_float(request.form.get('amount'), 0.0)
    category = request.form.get('category', 'Supplies')
    staff_user = session.get('cashier_user') or session.get('admin_user') or 'Cashier'

    if not title or amount <= 0:
        flash("Please provide a valid title and amount.", "error")
        return redirect(url_for('cashier_terminal'))

    db.session.add(Expense(title=title, amount=amount, category=category, created_by=staff_user))
    db.session.commit()
    flash(f"Expense recorded: {title} (₱{amount:,.2f})", "success")
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/record-vault-drop', methods=['POST'])
@require_cashier
def cashier_record_vault_drop():
    drop_num = parse_int(request.form.get('drop_number'), 1)
    amount = parse_float(request.form.get('amount'), 0.0)
    notes = request.form.get('notes', '').strip()
    breakdown = request.form.get('cash_breakdown', '').strip()
    staff_user = session.get('cashier_user') or session.get('admin_user') or 'Cashier'

    if drop_num < 1:
        flash('Vault drop number must be at least 1.', 'error')
        return redirect(url_for('cashier_terminal'))
    if amount <= 0:
        flash("Drop amount must be greater than zero.", "error")
        return redirect(url_for('cashier_terminal'))

    db.session.add(VaultDrop(drop_number=drop_num, amount=amount, notes=notes, cash_breakdown=breakdown, created_by=staff_user))
    db.session.commit()
    flash(f"Vault Cash Drop #{drop_num} recorded: ₱{amount:,.2f}", "success")
    return redirect(url_for('cashier_terminal'))

def _apply_pending_order_total_delta(order, delta):
    """Apply a cashier adjustment without disturbing existing delivery/discount math."""
    delta = round(parse_float(delta, 0.0), 2)
    order.subtotal = max(0.0, round(parse_float(order.subtotal, 0.0) + delta, 2))
    order.total_amount = max(0.0, round(parse_float(order.total_amount, 0.0) + delta, 2))


def _append_order_adjustment_audit(order, message):
    stamp = ph_now().strftime('%Y-%m-%d %I:%M %p')
    staff_user = session.get('cashier_user') or session.get('admin_user') or 'Cashier'
    entry = f'[{stamp} PH] {staff_user}: {message}'
    existing = (order.collection_notes or '').strip()
    order.collection_notes = (existing + (' | ' if existing else '') + entry)[-250:]


@app.route('/pos/adjust-order/<int:order_id>', methods=['POST'])
@require_cashier
def cashier_adjust_pending_order(order_id):
    """Add/remove cashier-approved extra charges before a queued order is accepted.

    This is intentionally restricted to VERIFICATION orders. Original product lines
    cannot be deleted here; only lines created by this adjustment route can be removed.
    """
    order = Order.query.get_or_404(order_id)
    if order.status != 'VERIFICATION':
        flash(f'Order #{order.id} can no longer be adjusted because it is {order.status}.', 'error')
        return redirect(url_for('cashier_terminal'))

    action = str(request.form.get('action', 'ADD_CHARGE')).upper()
    try:
        if action == 'ADD_CHARGE':
            description = re.sub(r'\s+', ' ', str(request.form.get('description', '')).strip())[:100]
            qty = parse_int(request.form.get('quantity'), 1)
            unit_fee = round(parse_float(request.form.get('unit_fee'), 0.0), 2)

            if not description:
                raise OrderValidationError('Please describe the additional item or service.')
            if qty < 1 or qty > 99:
                raise OrderValidationError('Additional charge quantity must be between 1 and 99.')
            if unit_fee <= 0 or unit_fee > 100000:
                raise OrderValidationError('Additional fee must be greater than ₱0.00.')

            line_total = round(unit_fee * qty, 2)
            item = OrderItem(
                order_id=order.id,
                product_id=None,
                product_name=f'[Cashier Add-on] {description}',
                unit_price=unit_fee,
                cost_price=0.0,
                quantity=qty,
                subtotal=line_total,
                selected_options=None,
            )
            db.session.add(item)
            db.session.flush()
            _apply_pending_order_total_delta(order, line_total)

            cash_received_raw = str(request.form.get('cash_received', '')).strip()
            if order.payment_method == 'CASH' and cash_received_raw:
                cash_received = round(parse_float(cash_received_raw, -1), 2)
                if cash_received < 0:
                    raise OrderValidationError('Cash received/change-for amount is invalid.')
                order.change_for = cash_received or None

            _append_order_adjustment_audit(
                order,
                f'Added {description} x{qty} @ ₱{unit_fee:,.2f} (+₱{line_total:,.2f}). New total ₱{order.total_amount:,.2f}.'
            )
            db.session.commit()
            extra = ''
            if order.payment_method == 'GCASH':
                extra = ' Confirm the additional GCash/payment amount before accepting.'
            elif order.payment_method == 'CASH' and order.change_for and order.change_for + 1e-9 < order.total_amount:
                extra = ' Cash received/change-for is below the new total; update it before accepting.'
            flash(f'Order #{order.id} updated: +₱{line_total:,.2f}. New total ₱{order.total_amount:,.2f}.{extra}', 'success')

        elif action == 'REMOVE_LINE':
            item_id = parse_int(request.form.get('item_id'), 0)
            item = db.session.get(OrderItem, item_id)
            if not item or item.order_id != order.id:
                raise OrderValidationError('Adjustment line was not found.')
            if item.product_id is not None or not str(item.product_name or '').startswith('[Cashier Add-on] '):
                raise OrderValidationError('Only cashier-added charge lines can be removed here.')

            removed_total = parse_float(item.subtotal, 0.0)
            removed_name = str(item.product_name).replace('[Cashier Add-on] ', '', 1)
            db.session.delete(item)
            _apply_pending_order_total_delta(order, -removed_total)
            _append_order_adjustment_audit(
                order,
                f'Removed cashier add-on {removed_name} (-₱{removed_total:,.2f}). New total ₱{order.total_amount:,.2f}.'
            )
            db.session.commit()
            flash(f'Cashier add-on removed from Order #{order.id}. New total ₱{order.total_amount:,.2f}.', 'info')

        elif action == 'UPDATE_CASH':
            if order.payment_method != 'CASH':
                raise OrderValidationError('Cash received can only be updated for CASH orders.')
            cash_received = round(parse_float(request.form.get('cash_received'), -1), 2)
            if cash_received < order.total_amount:
                raise OrderValidationError(f'Cash received cannot be below the order total of ₱{order.total_amount:,.2f}.')
            order.change_for = cash_received
            _append_order_adjustment_audit(order, f'Updated cash received/change-for to ₱{cash_received:,.2f}.')
            db.session.commit()
            flash(f'Cash received for Order #{order.id} updated to ₱{cash_received:,.2f}.', 'success')
        else:
            raise OrderValidationError('Invalid order adjustment action.')

    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    except Exception:
        db.session.rollback()
        app.logger.exception('Failed to adjust pending Order #%s', order_id)
        flash('Unable to adjust this order. Please check the server log.', 'error')

    return redirect(url_for('cashier_terminal'))


@app.route('/api/redeem-order-points', methods=['POST'])
def redeem_points_on_recent_order():
    """Apply points to a pending/recent completed order without permitting double redemption."""
    data = request.get_json() or {}
    order = db.session.get(Order, parse_int(data.get('order_id'), 0))
    if not order:
        return jsonify({'success': False, 'message': 'Purchase not found.'}), 404

    staff_access = bool(session.get('cashier_user') or session.get('admin_user'))
    customer_access = bool(session.get('customer_id') and order.customer_id == session.get('customer_id'))
    if not staff_access and not customer_access:
        return jsonify({'success': False, 'message': 'You cannot update this purchase.'}), 403
    if order.status not in {'VERIFICATION', 'COMPLETED'}:
        return jsonify({'success': False, 'message': 'Points can only be applied to a pending or completed purchase.'}), 400
    if not order.customer_id:
        return jsonify({'success': False, 'message': 'This purchase is not linked to a loyalty account.'}), 400
    if parse_float(order.points_redeemed, 0.0) > 0:
        return jsonify({'success': False, 'message': 'Points were already applied to this purchase.'}), 400
    if utc_now() - order.created_at > timedelta(hours=24):
        return jsonify({'success': False, 'message': 'Only purchases from the last 24 hours are eligible.'}), 400

    cust = db.session.get(Customer, order.customer_id)
    issue = customer_access_issue(cust)
    if issue:
        return jsonify({'success': False, 'message': issue}), 400

    try:
        old_total = round(parse_float(order.total_amount, 0.0), 2)
        points, discount = calculate_points_redemption(cust, data.get('redeem_points'), old_total)
        new_total = round(max(0.0, old_total - discount), 2)
        if order.status == 'COMPLETED' and order.payment_verified and not order.is_unpaid:
            old_earned = int(old_total // 30)
            new_earned = int(new_total // 30)
            earned_adjustment = new_earned - old_earned
            if parse_float(cust.points_balance, 0.0) + earned_adjustment < points - 1e-9:
                raise OrderValidationError('Not enough available points after recalculating this purchase.')
            if earned_adjustment:
                cust.points_balance = round(parse_float(cust.points_balance, 0.0) + earned_adjustment, 2)
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=earned_adjustment,
                    reason=f'Base points recalculated after discount / Order #{order.id}',
                ))
            cust.accumulated_spend = max(0.0, parse_float(cust.accumulated_spend, 0.0) - discount)

        order.points_redeemed = points
        order.points_discount = discount
        order.total_amount = new_total
        record_points_redemption(cust, points, order.id, 'Recent purchase points discount')
        db.session.commit()
        return jsonify({
            'success': True,
            'order_id': order.id,
            'old_total': old_total,
            'total': new_total,
            'points_redeemed': points,
            'discount': discount,
            'refund_due': discount if order.status == 'COMPLETED' and order.payment_verified else 0.0,
            'points_balance': round(parse_float(cust.points_balance, 0.0), 2),
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Recent purchase points redemption failed for order_id=%s', order.id)
        return jsonify({'success': False, 'message': 'Could not apply points. No changes were saved.'}), 500


@app.route('/pos/verify/<int:order_id>', methods=['POST'])
@require_cashier
def verify_order(order_id):
    order = Order.query.get_or_404(order_id)
    action = request.form.get('action')

    if order.status != 'VERIFICATION':
        flash(f'Order #{order.id} is already {order.status}.', 'info')
        return redirect(url_for('cashier_terminal'))

    if action == 'ACCEPT':
        if order.payment_method == 'CASH' and order.change_for and order.change_for + 1e-9 < order.total_amount:
            flash(
                f'Order #{order.id} total is ₱{order.total_amount:,.2f}, but cash received/change-for is only ₱{order.change_for:,.2f}. '
                'Use Adjust to update the cash amount before accepting.',
                'error'
            )
            return redirect(url_for('cashier_terminal'))

        if order.payment_method == 'CREDIT':
            if not order.customer_id:
                flash('Credit order has no registered customer and cannot be accepted.', 'error')
                return redirect(url_for('cashier_terminal'))
            cust = db.session.get(Customer, order.customer_id)
            issue = customer_access_issue(cust)
            if issue:
                flash(issue, 'error')
                return redirect(url_for('cashier_terminal'))
            if not cust.is_credit_eligible:
                flash('Customer is no longer authorized for credit.', 'error')
                return redirect(url_for('cashier_terminal'))
            available = max(0.0, parse_float(cust.credit_limit, 0.0) - parse_float(cust.outstanding_ar, 0.0))
            if order.total_amount > available + 1e-9:
                flash(f'Credit limit exceeded. Available credit is ₱{available:,.2f}.', 'error')
                return redirect(url_for('cashier_terminal'))
            cust.outstanding_ar = (cust.outstanding_ar or 0.0) + order.total_amount
            order.is_unpaid = True
            order.payment_verified = False
            order.status = 'COMPLETED'
        else:
            order.payment_verified = True
            order.status = 'COMPLETED'
            if order.customer_id:
                cust = db.session.get(Customer, order.customer_id)
                if cust and not customer_access_issue(cust):
                    cust.accumulated_spend = (cust.accumulated_spend or 0.0) + order.total_amount
                    earned = int(order.total_amount // 30)
                    if earned > 0:
                        cust.points_balance = (cust.points_balance or 0.0) + earned
                        db.session.add(RewardLedger(
                            customer_id=cust.id,
                            points_change=earned,
                            reason=f'Purchase Order #{order.id}',
                        ))
                    apply_member_marketing_rewards(cust, order)
        order.fulfillment_status = 'PREPARING'
        sync_craft_order_after_main_verification(order, accepted=True)
        sync_digital_order_after_main_verification(order, accepted=True)
        db.session.commit()
        if order.payment_method == 'CREDIT':
            flash(f'Order #{order.id} accepted as A/R Credit and added to Member Credit AR.', 'success')
        else:
            flash(f'Order #{order.id} accepted and completed. Details ready to print!', 'success')
    elif action == 'REJECT':
        if order.customer_id and parse_float(order.points_redeemed, 0.0) > 0:
            cust = db.session.get(Customer, order.customer_id)
            if cust:
                restored = round(parse_float(order.points_redeemed, 0.0), 2)
                cust.points_balance = round(parse_float(cust.points_balance, 0.0) + restored, 2)
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=restored,
                    reason=f'Cancelled Order #{order.id}: redeemed points restored',
                ))
        for item in order.items:
            if item.product_id:
                prod = db.session.get(Product, item.product_id)
                if prod:
                    prod.stock = parse_int(prod.stock, 0) + item.quantity
        order.status = 'CANCELLED'
        order.fulfillment_status = 'CANCELLED'
        sync_craft_order_after_main_verification(order, accepted=False)
        sync_digital_order_after_main_verification(order, accepted=False)
        db.session.commit()
        flash(f'Order #{order.id} cancelled and reserved stock restored.', 'info')
    else:
        flash('Invalid verification action.', 'error')

    return redirect(url_for('cashier_terminal'))


@app.route('/pos/order/<int:order_id>/fulfillment', methods=['POST'])
@require_cashier
def update_order_fulfillment(order_id):
    order = Order.query.get_or_404(order_id)
    new_status = request.form.get('fulfillment_status', '').strip().upper()
    allowed = {'PREPARING', 'READY', 'FULFILLED'}
    if order.status != 'COMPLETED' or new_status not in allowed:
        flash('Only accepted orders can move through Preparing, Ready, and Fulfilled.', 'error')
        return redirect(url_for('cashier_terminal'))
    current = (order.fulfillment_status or 'PREPARING').upper()
    ranks = {'SUBMITTED': 0, 'PREPARING': 1, 'READY': 2, 'FULFILLED': 3}
    if ranks.get(new_status, 0) < ranks.get(current, 0):
        flash('Fulfillment cannot be moved backward. This protects the customer tracker.', 'error')
        return redirect(url_for('cashier_terminal'))
    order.fulfillment_status = new_status
    db.session.commit()
    flash(f'Order #{order.id} is now {new_status.title()}.', 'success')
    return redirect(url_for('cashier_terminal'))

@app.route('/admin/settings/hours', methods=['POST'])
@require_cashier
def update_operating_hours():
    """Update store/delivery operating hours from the cashier settings modal."""
    time_fields = (
        'store_open_time',
        'store_close_time',
        'delivery_open_time',
        'delivery_close_time',
    )

    values = {}
    for key in time_fields:
        value = request.form.get(key, '').strip()
        try:
            datetime.strptime(value, '%H:%M')
        except (TypeError, ValueError):
            flash(f"Invalid time value for {key.replace('_', ' ')}.", "error")
            return redirect(url_for('cashier_terminal'))
        values[key] = value

    values['is_store_open'] = 'true' if request.form.get('is_store_open') else 'false'
    values['is_delivery_enabled'] = 'true' if request.form.get('is_delivery_enabled') else 'false'

    try:
        for key, value in values.items():
            setting = StoreSetting.query.filter_by(key=key).first()
            if setting:
                setting.value = value
            else:
                db.session.add(StoreSetting(key=key, value=value))
        db.session.commit()
        flash("Operating schedule updated.", "success")
    except Exception:
        db.session.rollback()
        app.logger.exception('Could not save operating schedule')
        flash("Could not save the operating schedule.", "error")

    if session.get('cashier_user'):
        return redirect(url_for('cashier_terminal'))
    return redirect(url_for('admin_dashboard'))


# ==================== INTEGRATED CRAFT SHOP ROUTES ====================

@app.route('/craft')
def craft_store():
    ip = get_client_ip()[:64]
    visitor = CraftSiteVisitor.query.filter_by(ip_address=ip).first()
    if visitor:
        visitor.visit_count = parse_int(visitor.visit_count, 0) + 1
        visitor.last_seen_at = utc_now()
    else:
        db.session.add(CraftSiteVisitor(ip_address=ip, visit_count=1))
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Could not update Craft Shop visitor tracking for %s', ip)

    selected_category = request.args.get('category', '').strip()
    query = CraftItem.query.filter_by(is_active=True)
    if selected_category:
        query = query.filter_by(category_name=selected_category)
    items = query.order_by(CraftItem.is_featured.desc(), CraftItem.is_top_seller.desc(), CraftItem.name.asc()).all()
    categories = CraftCategory.query.filter_by(is_active=True).order_by(CraftCategory.name.asc()).all()
    featured = CraftItem.query.filter_by(is_active=True, is_featured=True).order_by(CraftItem.name.asc()).limit(12).all()
    top_sellers = CraftItem.query.filter_by(is_active=True, is_top_seller=True).order_by(CraftItem.orders_count.desc(), CraftItem.name.asc()).limit(12).all()
    cust = db.session.get(Customer, session.get('customer_id')) if session.get('customer_id') else None
    liked_ids = {x.item_id for x in CraftItemLike.query.filter_by(ip_address=ip).all()}
    unique_visitors = CraftSiteVisitor.query.count()
    total_page_visits = db.session.query(db.func.coalesce(db.func.sum(CraftSiteVisitor.visit_count), 0)).scalar() or 0
    return render_template(
        'craft/index.html',
        items=items,
        categories=categories,
        selected_category=selected_category,
        featured=featured,
        top_sellers=top_sellers,
        cust=cust,
        liked_ids=liked_ids,
        site_visits=unique_visitors,
        total_page_visits=total_page_visits,
    )

@app.route('/craft/item/<int:item_id>')
def craft_item_detail(item_id):
    item = CraftItem.query.filter_by(id=item_id, is_active=True).first_or_404()
    ip = get_client_ip()[:64]
    viewed = CraftItemView.query.filter_by(item_id=item.id, ip_address=ip).first()
    if not viewed:
        try:
            db.session.add(CraftItemView(item_id=item.id, ip_address=ip))
            item.views = parse_int(item.views, 0) + 1
            db.session.commit()
        except Exception:
            # A concurrent duplicate request from the same IP must not inflate the counter.
            db.session.rollback()
    liked = bool(CraftItemLike.query.filter_by(item_id=item.id, ip_address=ip).first())
    return render_template('craft/item_detail.html', item=item, liked=liked)

@app.route('/craft/item/<int:item_id>/like', methods=['POST'])
def craft_like_item(item_id):
    item = CraftItem.query.filter_by(id=item_id, is_active=True).first_or_404()
    ip = get_client_ip()[:64]
    existing = CraftItemLike.query.filter_by(item_id=item.id, ip_address=ip).first()
    if existing:
        flash('You already liked this craft from this connection. One like per product per IP is allowed.', 'info')
        return redirect(url_for('craft_item_detail', item_id=item.id))
    try:
        db.session.add(CraftItemLike(item_id=item.id, ip_address=ip))
        item.likes = parse_int(item.likes, 0) + 1
        db.session.commit()
        flash('Thanks! Your like was counted.', 'success')
    except Exception:
        db.session.rollback()
        flash('This craft has already been liked from this connection.', 'info')
    return redirect(url_for('craft_item_detail', item_id=item.id))

@app.route('/craft/item/<int:item_id>/comment', methods=['POST'])
def craft_add_comment(item_id):
    item = CraftItem.query.filter_by(id=item_id, is_active=True).first_or_404()
    ip = get_client_ip()[:64]
    author = request.form.get('author', '').strip()[:80] or 'Customer'
    content = request.form.get('content', '').strip()[:1000]
    if content:
        # Comments remain open, but every comment is attributable to its source IP for moderation.
        db.session.add(CraftComment(item_id=item.id, author=author, content=content, ip_address=ip))
        db.session.commit()
    return redirect(url_for('craft_item_detail', item_id=item.id))

@app.route('/craft/order/<int:item_id>', methods=['GET', 'POST'])
def craft_order_item(item_id):
    item = CraftItem.query.filter_by(id=item_id, is_active=True).first_or_404()
    cust = db.session.get(Customer, session.get('customer_id')) if session.get('customer_id') else None
    error = None
    if request.method == 'POST':
        name = request.form.get('customer_name', '').strip()[:100]
        contact = request.form.get('contact_number', '').strip()[:50]
        email = request.form.get('email', '').strip()[:120] or None
        fb = request.form.get('fb_account', '').strip()[:150] or None
        qty = parse_int(request.form.get('quantity'), 1)
        payment_method = request.form.get('payment_method', 'CASH').strip().upper()
        gcash_ref = request.form.get('gcash_ref', '').strip()[:20] or None
        notes = request.form.get('notes', '').strip()[:1000] or None

        if not name or not contact:
            error = 'Name and contact number are required.'
        elif qty <= 0 or qty > 100:
            error = 'Quantity must be between 1 and 100.'
        elif payment_method not in CRAFT_PAYMENT_METHODS:
            error = 'Choose Cash or GCash.'
        elif payment_method == 'GCASH' and (not gcash_ref or len(gcash_ref) < 6):
            error = 'Please enter at least the last 6 digits of the GCash reference.'
        elif item.availability_type == 'IN_STOCK' and qty > parse_int(item.stock_quantity, 0):
            error = f'Sorry, only {item.stock_quantity} item(s) are currently in stock.'
        else:
            try:
                if item.availability_type == 'IN_STOCK':
                    item.stock_quantity = parse_int(item.stock_quantity, 0) - qty
                member = Customer.query.filter_by(contact=contact).first()
                craft_order = CraftOrder(
                    item_id=item.id,
                    customer_id=member.id if member else None,
                    customer_name=name,
                    contact_number=contact,
                    email=email,
                    fb_account=fb,
                    quantity=qty,
                    unit_price=max(0.0, parse_float(item.price, 0.0)),
                    unit_cost=max(0.0, parse_float(item.cost, 0.0)),
                    total_price=max(0.0, parse_float(item.price, 0.0)) * qty,
                    total_cost=max(0.0, parse_float(item.cost, 0.0)) * qty,
                    payment_method=payment_method,
                    payment_status='PENDING',
                    gcash_ref=gcash_ref,
                    status='PENDING',
                    notes=notes,
                )
                db.session.add(craft_order)
                db.session.flush()
                create_main_craft_order(craft_order)
                item.orders_count = parse_int(item.orders_count, 0) + qty
                db.session.commit()
                return render_template('craft/order_success.html', order=craft_order, item=item)
            except Exception:
                db.session.rollback()
                app.logger.exception('Craft order creation failed')
                error = 'We could not place the craft order. Please try again.'
    return render_template('craft/order_form.html', item=item, error=error, cust=cust)

@app.route('/admin/craft')
@require_admin
def craft_admin_dashboard():
    items = CraftItem.query.order_by(CraftItem.is_active.desc(), CraftItem.category_name.asc(), CraftItem.name.asc()).all()
    categories = CraftCategory.query.order_by(CraftCategory.name.asc()).all()
    orders = CraftOrder.query.order_by(CraftOrder.created_at.desc()).limit(200).all()
    ledger = CraftLedger.query.order_by(CraftLedger.created_at.desc()).limit(200).all()
    completed = [o for o in orders if o.status == 'COMPLETED']
    sale_revenue = sum(max(0.0, parse_float(o.total_price, 0.0)) for o in completed)
    sale_cost = sum(max(0.0, parse_float(o.total_cost, 0.0)) for o in completed)
    manual_income = sum(max(0.0, parse_float(x.amount, 0.0)) for x in ledger if x.event_type == 'OTHER_INCOME')
    expense_total = sum(max(0.0, parse_float(x.amount, 0.0)) for x in ledger if x.event_type in ('EXPENSE', 'REFUND'))
    metrics = {
        'site_unique_visitors': CraftSiteVisitor.query.count(),
        'site_total_visits': db.session.query(db.func.coalesce(db.func.sum(CraftSiteVisitor.visit_count), 0)).scalar() or 0,
        'product_unique_views': sum(parse_int(i.views, 0) for i in items),
        'total_likes': sum(parse_int(i.likes, 0) for i in items),
        'total_comments': CraftComment.query.count(),
        'total_orders': CraftOrder.query.count(),
        'completed_sales': sale_revenue,
        'recorded_cogs': sale_cost,
        'gross_profit': sale_revenue - sale_cost,
        'other_income': manual_income,
        'expenses_refunds': expense_total,
        'net_cash': sale_revenue + manual_income - expense_total,
        'pending_orders': sum(1 for o in orders if o.status in ('PENDING', 'READY')),
        'low_stock': sum(1 for i in items if i.is_active and i.availability_type == 'IN_STOCK' and parse_int(i.stock_quantity, 0) <= 3),
        'out_of_stock': sum(1 for i in items if i.is_active and i.availability_type == 'IN_STOCK' and parse_int(i.stock_quantity, 0) == 0),
    }
    return render_template('craft/admin.html', items=items, categories=categories, orders=orders, ledger=ledger, metrics=metrics)

@app.route('/admin/craft/category/add', methods=['POST'])
@require_admin
def craft_add_category():
    name = request.form.get('name', '').strip()[:80]
    if not name:
        flash('Craft category name is required.', 'error')
        return redirect(url_for('craft_admin_dashboard'))
    if CraftCategory.query.filter(db.func.lower(CraftCategory.name) == name.lower()).first():
        flash('That craft category already exists.', 'info')
        return redirect(url_for('craft_admin_dashboard'))
    try:
        image_url = craft_image_from_request(existing=CRAFT_DEFAULT_IMAGE)
        db.session.add(CraftCategory(name=name, image_url=image_url, is_active=True))
        db.session.commit()
        flash(f'Craft category {name} added.', 'success')
    except (OrderValidationError, Exception) as exc:
        db.session.rollback()
        if isinstance(exc, OrderValidationError):
            flash(str(exc), 'error')
        else:
            app.logger.exception('Craft category creation failed')
            flash('Could not add the craft category.', 'error')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/item/add', methods=['POST'])
@require_admin
def craft_add_item():
    name = request.form.get('name', '').strip()[:120]
    description = request.form.get('description', '').strip()[:3000]
    category_name = request.form.get('category_name', 'General').strip()[:80] or 'General'
    price = parse_float(request.form.get('price'), 0.0)
    cost = parse_float(request.form.get('cost'), 0.0)
    availability_type = request.form.get('availability_type', 'IN_STOCK').strip().upper()
    stock = parse_int(request.form.get('stock_quantity'), 0)
    if not name or price <= 0 or cost < 0 or stock < 0 or availability_type not in ('IN_STOCK', 'PREORDER'):
        flash('Enter a valid craft name, price, cost, availability, and stock.', 'error')
        return redirect(url_for('craft_admin_dashboard'))
    try:
        image_url = craft_image_from_request(existing=CRAFT_DEFAULT_IMAGE)
        db.session.add(CraftItem(
            name=name, description=description, category_name=category_name, price=price, cost=cost,
            image_url=image_url, availability_type=availability_type,
            stock_quantity=stock if availability_type == 'IN_STOCK' else 0,
            is_active=True, is_featured=bool(request.form.get('is_featured')),
            is_top_seller=bool(request.form.get('is_top_seller')),
        ))
        db.session.commit()
        flash(f'Craft item {name} added.', 'success')
    except OrderValidationError as exc:
        db.session.rollback(); flash(str(exc), 'error')
    except Exception:
        db.session.rollback(); app.logger.exception('Craft item creation failed'); flash('Could not add craft item.', 'error')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/item/<int:item_id>/update', methods=['POST'])
@require_admin
def craft_update_item(item_id):
    item = CraftItem.query.get_or_404(item_id)
    try:
        price = parse_float(request.form.get('price'), item.price)
        cost = parse_float(request.form.get('cost'), item.cost)
        stock = parse_int(request.form.get('stock_quantity'), item.stock_quantity)
        availability_type = request.form.get('availability_type', item.availability_type).strip().upper()
        if price <= 0 or cost < 0 or stock < 0 or availability_type not in ('IN_STOCK', 'PREORDER'):
            raise OrderValidationError('Invalid craft price, cost, stock, or availability.')
        item.name = request.form.get('name', item.name).strip()[:120] or item.name
        item.description = request.form.get('description', item.description or '').strip()[:3000]
        item.category_name = request.form.get('category_name', item.category_name).strip()[:80] or 'General'
        item.price = price
        item.cost = cost
        item.availability_type = availability_type
        item.stock_quantity = stock if availability_type == 'IN_STOCK' else 0
        item.is_active = bool(request.form.get('is_active'))
        item.is_featured = bool(request.form.get('is_featured'))
        item.is_top_seller = bool(request.form.get('is_top_seller'))
        item.image_url = craft_image_from_request(existing=item.image_url or CRAFT_DEFAULT_IMAGE)
        db.session.commit()
        flash(f'{item.name} updated.', 'success')
    except OrderValidationError as exc:
        db.session.rollback(); flash(str(exc), 'error')
    except Exception:
        db.session.rollback(); app.logger.exception('Craft item update failed'); flash('Could not update craft item.', 'error')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/item/<int:item_id>/duplicate', methods=['POST'])
@require_admin
def craft_duplicate_item(item_id):
    item = CraftItem.query.get_or_404(item_id)
    db.session.add(CraftItem(
        name=f'{item.name} (Copy)', description=item.description, category_name=item.category_name,
        price=item.price, cost=item.cost, image_url=item.image_url, availability_type=item.availability_type,
        stock_quantity=item.stock_quantity, is_active=False, is_featured=False, is_top_seller=False,
    ))
    db.session.commit()
    flash('Craft item duplicated as inactive.', 'success')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/item/<int:item_id>/delete', methods=['POST'])
@require_admin
def craft_delete_item(item_id):
    item = CraftItem.query.get_or_404(item_id)
    if CraftOrder.query.filter_by(item_id=item.id).first():
        item.is_active = False
        db.session.commit()
        flash('Item has order history, so it was archived instead of deleted.', 'info')
    else:
        db.session.delete(item); db.session.commit(); flash('Craft item deleted.', 'success')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/order/<int:order_id>/status', methods=['POST'])
@require_admin
def craft_update_order_status(order_id):
    craft_order = CraftOrder.query.get_or_404(order_id)
    new_status = request.form.get('status', '').strip().upper()
    if new_status not in ('PENDING', 'READY', 'CANCELLED'):
        flash('Use the Cashier POS to complete/accept a craft sale. Craft Admin can mark Pending, Ready, or Cancelled.', 'error')
        return redirect(url_for('craft_admin_dashboard'))
    if craft_order.status == 'COMPLETED':
        flash('Completed craft sales are locked. Use a Refund transaction instead of changing the sale.', 'error')
        return redirect(url_for('craft_admin_dashboard'))
    main_order = db.session.get(Order, craft_order.main_order_id) if craft_order.main_order_id else None
    if new_status == 'CANCELLED':
        craft_order.status = 'CANCELLED'
        craft_order.payment_status = 'CANCELLED'
        craft_restore_stock(craft_order)
        if main_order and main_order.status == 'VERIFICATION':
            main_order.status = 'CANCELLED'
    else:
        if craft_order.status == 'CANCELLED':
            item = db.session.get(CraftItem, craft_order.item_id)
            if item and item.availability_type == 'IN_STOCK':
                if parse_int(item.stock_quantity, 0) < parse_int(craft_order.quantity, 0):
                    flash('Not enough stock to reopen this cancelled craft order.', 'error')
                    return redirect(url_for('craft_admin_dashboard'))
                item.stock_quantity -= craft_order.quantity
            craft_order.stock_restored = False
            if main_order and main_order.status == 'CANCELLED':
                main_order.status = 'VERIFICATION'
        craft_order.status = new_status
    db.session.commit()
    flash(f'Craft Order #{craft_order.id} updated to {craft_order.status}.', 'success')
    return redirect(url_for('craft_admin_dashboard'))

@app.route('/admin/craft/transaction/add', methods=['POST'])
@require_admin
def craft_record_transaction():
    event_type = request.form.get('event_type', '').strip().upper()
    title = request.form.get('title', '').strip()[:150]
    amount = parse_float(request.form.get('amount'), 0.0)
    payment_method = request.form.get('payment_method', 'CASH').strip().upper()
    notes = request.form.get('notes', '').strip()[:255] or None
    if event_type not in ('EXPENSE', 'OTHER_INCOME', 'REFUND') or not title or amount <= 0:
        flash('Choose Expense, Other Income, or Refund and enter a valid title and amount.', 'error')
        return redirect(url_for('craft_admin_dashboard'))
    if payment_method not in CRAFT_PAYMENT_METHODS:
        payment_method = 'CASH'
    try:
        creator = session.get('admin_user') or 'admin'
        main_order_id = expense_id = None
        if event_type in ('EXPENSE', 'REFUND'):
            exp = Expense(
                title=f'[Craft {"Refund" if event_type == "REFUND" else "Expense"}] {title}',
                amount=amount,
                category='Craft Shop' if event_type == 'EXPENSE' else 'Craft Refund',
                created_by=creator,
            )
            db.session.add(exp); db.session.flush(); expense_id = exp.id
        else:
            main_order = Order(
                order_type='CRAFT/MISC', dining_option='TAKEOUT', customer_name='Craft Shop', contact_number='N/A',
                subtotal=amount, delivery_fee=0.0, total_amount=amount, payment_method=payment_method,
                payment_verified=True, status='COMPLETED', fulfillment_status='FULFILLED', notes=f'[CRAFT OTHER INCOME] {title}: {notes or ""}',
            )
            db.session.add(main_order); db.session.flush(); main_order_id = main_order.id
            db.session.add(OrderItem(
                order_id=main_order.id, product_id=None, product_name=f'[Craft Income] {title}',
                unit_price=amount, cost_price=0.0, quantity=1, subtotal=amount,
            ))
        db.session.add(CraftLedger(
            event_type=event_type, title=title, amount=amount, payment_method=payment_method,
            main_order_id=main_order_id, expense_id=expense_id, notes=notes, created_by=creator,
        ))
        db.session.commit()
        flash(f'Craft {event_type.replace("_", " ").title()} recorded and synced to the main system.', 'success')
    except Exception:
        db.session.rollback(); app.logger.exception('Craft transaction sync failed'); flash('Could not record craft transaction.', 'error')
    return redirect(url_for('craft_admin_dashboard'))

# ==================== DIGITAL BUSINESS PORTAL ====================

@app.route('/digital')
def digital_store():
    category = request.args.get('category', '').strip()
    query = DigitalItem.query.filter_by(is_active=True)
    if category:
        query = query.filter_by(category_name=category)
    return render_template('digital/index.html', items=query.order_by(DigitalItem.is_featured.desc(), DigitalItem.name.asc()).all(),
                           categories=DigitalCategory.query.filter_by(is_active=True).order_by(DigitalCategory.name).all(), selected_category=category)

@app.route('/digital/item/<int:item_id>', methods=['GET', 'POST'])
def digital_item_detail(item_id):
    item = DigitalItem.query.filter_by(id=item_id, is_active=True).first_or_404()
    if request.method == 'GET':
        item.views = parse_int(item.views, 0) + 1; db.session.commit()
        return render_template('digital/item.html', item=item)
    name = request.form.get('customer_name', '').strip()[:100]
    contact = request.form.get('contact_number', '').strip()[:50]
    email = request.form.get('email', '').strip()[:120]
    qty = max(1, min(100, parse_int(request.form.get('quantity'), 1)))
    method = request.form.get('payment_method', 'GCASH').upper()
    ref = request.form.get('gcash_ref', '').strip()[:30] or None
    if not name or not contact or '@' not in email or method not in CRAFT_PAYMENT_METHODS:
        flash('Name, contact number, a valid delivery email, and payment method are required.', 'error')
        return redirect(url_for('digital_item_detail', item_id=item.id))
    order = DigitalOrder(item_id=item.id, customer_name=name, contact_number=contact, email=email,
        quantity=qty, unit_price=item.price, unit_cost=item.cost or 0, total_price=item.price * qty,
        payment_method=method, gcash_ref=ref, requirements=request.form.get('requirements','').strip()[:3000] or None)
    db.session.add(order); db.session.flush(); create_main_digital_order(order)
    item.orders_count = parse_int(item.orders_count, 0) + qty; db.session.commit()
    return redirect(url_for('digital_order_status', token=order.tracking_token))

@app.route('/digital/order/<token>')
def digital_order_status(token):
    order = DigitalOrder.query.filter_by(tracking_token=token).first_or_404()
    can_download = order.payment_status == 'PAID' and order.status in ('READY', 'DELIVERED')
    return render_template('digital/order_status.html', order=order, can_download=can_download)

@app.route('/admin/digital')
@require_admin
def digital_admin():
    return render_template('digital/admin.html', items=DigitalItem.query.order_by(DigitalItem.is_active.desc(), DigitalItem.name).all(),
        categories=DigitalCategory.query.order_by(DigitalCategory.name).all(), orders=DigitalOrder.query.order_by(DigitalOrder.created_at.desc()).limit(200).all())

@app.route('/admin/digital/category/add', methods=['POST'])
@require_admin
def digital_category_add():
    name = request.form.get('name','').strip()[:80]
    if name and not DigitalCategory.query.filter(db.func.lower(DigitalCategory.name) == name.lower()).first():
        db.session.add(DigitalCategory(name=name)); db.session.commit(); flash('Digital category added.', 'success')
    return redirect(url_for('digital_admin'))

@app.route('/admin/digital/item/save', methods=['POST'])
@require_admin
def digital_item_save():
    item_id = parse_int(request.form.get('item_id'), 0)
    item = db.session.get(DigitalItem, item_id) if item_id else DigitalItem()
    price = parse_float(request.form.get('price'), 0)
    if not request.form.get('name','').strip() or price <= 0:
        flash('Enter a name and price greater than zero.', 'error'); return redirect(url_for('digital_admin'))
    item.name=request.form['name'].strip()[:120]; item.description=request.form.get('description','').strip()[:5000]
    item.category_name=request.form.get('category_name','General').strip()[:80] or 'General'
    item.product_type=request.form.get('product_type','DOWNLOAD').upper(); item.price=price
    item.cost=max(0,parse_float(request.form.get('cost'),0)); item.image_url=request.form.get('image_url','').strip() or CRAFT_DEFAULT_IMAGE
    item.sample_url=request.form.get('sample_url','').strip() or None; item.file_format=request.form.get('file_format','').strip()[:80] or None
    item.license_terms=request.form.get('license_terms','').strip()[:3000] or None
    item.turnaround_days=max(0,parse_int(request.form.get('turnaround_days'),0)); item.is_active=bool(request.form.get('is_active', '1'))
    item.is_featured=bool(request.form.get('is_featured'))
    if not item_id: db.session.add(item)
    db.session.commit(); flash('Digital offer saved.', 'success'); return redirect(url_for('digital_admin'))

@app.route('/admin/digital/order/<int:order_id>/update', methods=['POST'])
@require_admin
def digital_order_update(order_id):
    order = DigitalOrder.query.get_or_404(order_id)
    status = request.form.get('status','').upper()
    if status not in ('PENDING_PAYMENT','PAID','IN_PROGRESS','READY','DELIVERED','CANCELLED'):
        flash('Invalid digital fulfillment status.', 'error'); return redirect(url_for('digital_admin'))
    if order.payment_status != 'PAID' and status in ('READY','DELIVERED'):
        flash('Cashier must confirm payment before releasing a digital product.', 'error'); return redirect(url_for('digital_admin'))
    order.status=status; order.fulfillment_url=request.form.get('fulfillment_url','').strip() or None
    order.license_key=request.form.get('license_key','').strip()[:255] or None
    order.fulfillment_notes=request.form.get('fulfillment_notes','').strip()[:3000] or None
    if status == 'DELIVERED': order.completed_at=utc_now()
    db.session.commit(); flash('Digital order updated.', 'success'); return redirect(url_for('digital_admin'))

# ==================== EXPENSE PAYMENT CENTER ====================

@app.route('/admin/cashflow/payables')
@require_admin
def cashflow_payables():
    today = ph_today(); start = today - timedelta(days=60); end = today + timedelta(days=91)
    plans = CashFlowPlan.query.filter_by(entry_type='EXPENSE', is_active=True).all()
    saved = {(x.plan_id, x.occurrence_date): x for x in CashFlowExpensePayment.query.filter(CashFlowExpensePayment.occurrence_date >= start, CashFlowExpensePayment.occurrence_date < end).all()}
    rows=[]
    for plan in plans:
        if cashflow_plan_exclusion_reason(plan): continue
        for occurrence in cashflow_occurrence_dates(plan, start, end):
            payment=saved.get((plan.id, occurrence))
            rows.append({'plan':plan,'occurrence_date':occurrence,'due_date':payment.due_date if payment else occurrence,
                         'amount':payment.amount if payment else plan.amount,'status':payment.status if payment else 'PAYABLE','payment':payment})
    rows.sort(key=lambda x:(x['status']=='PAID', x['due_date'], x['plan'].title))
    return render_template('cash_flow_payables.html', rows=rows, today=today)

@app.route('/admin/cashflow/payables/update', methods=['POST'])
@require_admin
def cashflow_payable_update():
    plan = CashFlowPlan.query.get_or_404(parse_int(request.form.get('plan_id'),0))
    occurrence = date.fromisoformat(request.form.get('occurrence_date',''))
    payment = CashFlowExpensePayment.query.filter_by(plan_id=plan.id, occurrence_date=occurrence).first()
    if not payment:
        payment=CashFlowExpensePayment(plan_id=plan.id, occurrence_date=occurrence, due_date=occurrence, amount=plan.amount)
        db.session.add(payment)
    payment.due_date=date.fromisoformat(request.form.get('due_date') or occurrence.isoformat())
    payment.status='PAID' if request.form.get('status')=='PAID' else 'PAYABLE'
    payment.payment_method=request.form.get('payment_method','').strip()[:20] or None
    payment.reference=request.form.get('reference','').strip()[:100] or None
    payment.notes=request.form.get('notes','').strip()[:255] or None
    payment.paid_at=utc_now() if payment.status=='PAID' else None
    payment.paid_by=(session.get('admin_user') or 'admin') if payment.status=='PAID' else None
    db.session.commit(); flash(f'{plan.title} marked {payment.status.lower()}.', 'success')
    return redirect(url_for('cashflow_payables'))

# ==================== FLEXIBLE-HORIZON CASH FLOW PORTAL ====================

CASH_FLOW_EXCLUDED_EXPENSE_TERMS = (
    'tet',
    'joy',
    'delro',
    'motorcycle',
    'kevin',
    'investor',
)
CASH_FLOW_EXCLUDED_INCOME_TERMS = ('sample work',)


def cashflow_text_matches_term(text, term):
    pattern = r'(?<!\w)' + re.escape(term) + (r's?' if term == 'investor' else '') + r'(?!\w)'
    return bool(re.search(pattern, text, flags=re.IGNORECASE))


def cashflow_plan_exclusion_reason(plan):
    """Keep source schedules intact while excluding owner-requested items from analysis."""
    haystack = ' '.join(
        str(value or '')
        for value in (plan.title, plan.category, plan.notes, plan.created_by)
    ).casefold()
    if plan.entry_type == 'EXPENSE':
        match = next((term for term in CASH_FLOW_EXCLUDED_EXPENSE_TERMS if cashflow_text_matches_term(haystack, term)), None)
        return f'Excluded expense: {match}' if match else ''
    if plan.entry_type == 'INCOME':
        match = next((term for term in CASH_FLOW_EXCLUDED_INCOME_TERMS if cashflow_text_matches_term(haystack, term)), None)
        return f'Excluded income: {match}' if match else ''
    return ''


def cashflow_order_is_excluded(order):
    """Exclude Sample work from actual-sales history without deleting the order."""
    item_names = ' '.join(str(item.product_name or '') for item in order.items)
    haystack = ' '.join(
        str(value or '')
        for value in (order.order_type, order.customer_name, order.notes, item_names)
    ).casefold()
    return any(cashflow_text_matches_term(haystack, term) for term in CASH_FLOW_EXCLUDED_INCOME_TERMS)

@app.route('/admin/cash-flow')
@app.route('/admin/cashflow')
@require_admin
def cash_flow_portal():
    today = ph_today()
    start_raw = request.args.get('start', '').strip()
    try:
        if start_raw:
            parts = start_raw.split('-')
            period_start = date(int(parts[0]), int(parts[1]), 1)
        else:
            period_start = date(today.year, today.month, 1)
    except (ValueError, IndexError):
        period_start = date(today.year, today.month, 1)

    horizon_years = parse_int(request.args.get('years'), CASH_FLOW_DEFAULT_HORIZON_YEARS)
    horizon_years = max(1, min(CASH_FLOW_MAX_HORIZON_YEARS, horizon_years))
    horizon_months = horizon_years * 12
    period_end = cashflow_add_months(period_start, horizon_months)
    utc_start, utc_end = cashflow_utc_bounds(period_start, period_end)

    # Sales follow Macleen's existing accounting rule: completed POS orders plus Vault Drops.
    # This cash-flow portal intentionally uses a planning COGS rate of 60% of sales instead
    # of the recorded per-product costs used elsewhere in the accounting dashboard.
    completed_orders = Order.query.filter(
        Order.status == 'COMPLETED',
        Order.created_at >= utc_start,
        Order.created_at < utc_end,
    ).all()
    excluded_window_orders = [order for order in completed_orders if cashflow_order_is_excluded(order)]
    completed_orders = [order for order in completed_orders if not cashflow_order_is_excluded(order)]
    vault_drops = VaultDrop.query.filter(
        VaultDrop.created_at >= utc_start,
        VaultDrop.created_at < utc_end,
    ).all()

    daily_sales = {}
    daily_order_sales = {}
    daily_vault_sales = {}
    for order in completed_orders:
        day = utc_naive_to_ph(order.created_at).date()
        amount = max(0.0, parse_float(order.total_amount, 0.0))
        daily_order_sales[day] = daily_order_sales.get(day, 0.0) + amount
        daily_sales[day] = daily_sales.get(day, 0.0) + amount
    for drop in vault_drops:
        day = utc_naive_to_ph(drop.created_at).date()
        amount = max(0.0, parse_float(drop.amount, 0.0))
        daily_vault_sales[day] = daily_vault_sales.get(day, 0.0) + amount
        daily_sales[day] = daily_sales.get(day, 0.0) + amount

    # Build the forecast average from every recorded ACTUAL sales day up to today.
    # A day is counted in the average only when it has positive recorded sales; blank
    # days are intentionally excluded because those are the days this projection fills.
    # This keeps the projection based on real selling days instead of treating missing
    # records as zero sales. Vault Drops remain part of actual sales by project policy.
    _, history_end_utc = cashflow_utc_bounds(today, today + timedelta(days=1))
    historical_orders = Order.query.filter(
        Order.status == 'COMPLETED',
        Order.created_at < history_end_utc,
    ).all()
    excluded_historical_orders = [order for order in historical_orders if cashflow_order_is_excluded(order)]
    historical_orders = [order for order in historical_orders if not cashflow_order_is_excluded(order)]
    historical_vault_drops = VaultDrop.query.filter(
        VaultDrop.created_at < history_end_utc,
    ).all()
    historical_daily_sales = {}
    for order in historical_orders:
        day = utc_naive_to_ph(order.created_at).date()
        amount = max(0.0, parse_float(order.total_amount, 0.0))
        historical_daily_sales[day] = historical_daily_sales.get(day, 0.0) + amount
    for drop in historical_vault_drops:
        day = utc_naive_to_ph(drop.created_at).date()
        amount = max(0.0, parse_float(drop.amount, 0.0))
        historical_daily_sales[day] = historical_daily_sales.get(day, 0.0) + amount

    # Every recorded actual positive-sales day is used. Older manual settings, including
    # a stored ₱5,000 amount, remain in the database for audit history but are ignored.
    actual_sales_entries = sorted(
        ((day, amount) for day, amount in historical_daily_sales.items() if amount > 0),
        key=lambda item: item[0],
    )
    lifetime_actual_sales_day_count = len(actual_sales_entries)
    actual_sales_days = [amount for _, amount in actual_sales_entries]
    average_daily_sales = (sum(actual_sales_days) / len(actual_sales_days)) if actual_sales_days else 0.0
    plans = CashFlowPlan.query.order_by(CashFlowPlan.entry_type.asc(), CashFlowPlan.start_date.asc(), CashFlowPlan.id.asc()).all()
    included_plans = []
    excluded_plans = []
    daily_income = {}
    daily_expenses = {}
    for plan in plans:
        if not plan.is_active:
            continue
        exclusion_reason = cashflow_plan_exclusion_reason(plan)
        plan.analysis_exclusion_reason = exclusion_reason
        if exclusion_reason:
            excluded_plans.append(plan)
            continue
        included_plans.append(plan)
        amount = max(0.0, parse_float(plan.amount, 0.0))
        target = daily_income if plan.entry_type == 'INCOME' else daily_expenses
        for occurrence in cashflow_occurrence_dates(plan, period_start, period_end):
            target[occurrence] = target.get(occurrence, 0.0) + amount

    contribution_margin_rate = max(0.0001, 1.0 - CASH_FLOW_COGS_RATE)

    def cashflow_break_even_sales_for_month(month_start):
        month_end = cashflow_add_months(month_start, 1)
        month_days = max(1, (month_end - month_start).days)
        expense_total = income_total = 0.0
        cursor = month_start
        while cursor < month_end:
            expense_total += daily_expenses.get(cursor, 0.0)
            income_total += daily_income.get(cursor, 0.0)
            cursor += timedelta(days=1)
        required_contribution = max(0.0, expense_total - income_total)
        return required_contribution / contribution_margin_rate / month_days

    break_even_daily_by_month = {
        cashflow_add_months(period_start, month_index): cashflow_break_even_sales_for_month(
            cashflow_add_months(period_start, month_index)
        )
        for month_index in range(horizon_months)
    }
    break_even_daily_sales = max(break_even_daily_by_month.values(), default=0.0)
    projection_mode = 'BREAK_EVEN_ACTUALS'
    average_window = 'LIFETIME'
    average_window_label = 'All historical actual sales'
    projection_daily_sales = max(average_daily_sales, break_even_daily_sales)
    projection_source_label = 'Higher of lifetime actual-sales average and filtered break-even sales'
    projection_badge = 'BREAK-EVEN + ACTUALS'

    def cashflow_sales_for_day(day):
        actual = max(0.0, parse_float(daily_sales.get(day), 0.0))
        if actual > 0:
            return actual, False, 'ACTUAL'
        month_start = date(day.year, day.month, 1)
        break_even_target = break_even_daily_by_month.get(month_start, 0.0)
        return max(average_daily_sales, break_even_target), True, 'BREAK-EVEN + ACTUALS'

    monthly_rows = []
    running_net = 0.0
    totals = {
        'sales': 0.0,
        'cogs': 0.0,
        'gross_profit': 0.0,
        'income': 0.0,
        'expenses': 0.0,
        'net': 0.0,
    }

    for month_index in range(horizon_months):
        month_start = cashflow_add_months(period_start, month_index)
        next_month = cashflow_add_months(month_start, 1)
        cursor = month_start
        sales = extra_income = expenses = 0.0
        actual_sales_total = projected_sales_total = 0.0
        actual_days = projected_days = 0
        while cursor < next_month:
            day_sales, is_projected, _projection_kind = cashflow_sales_for_day(cursor)
            sales += day_sales
            if is_projected:
                projected_sales_total += day_sales
                projected_days += 1
            else:
                actual_sales_total += day_sales
                actual_days += 1
            extra_income += daily_income.get(cursor, 0.0)
            expenses += daily_expenses.get(cursor, 0.0)
            cursor += timedelta(days=1)

        cogs = sales * CASH_FLOW_COGS_RATE
        gross_profit = sales - cogs
        net = sales + extra_income - cogs - expenses
        running_net += net
        monthly_rows.append({
            'month_start': month_start,
            'month_key': month_start.strftime('%Y-%m'),
            'month_label': month_start.strftime('%B %Y'),
            'sales': sales,
            'actual_sales': actual_sales_total,
            'projected_sales': projected_sales_total,
            'actual_days': actual_days,
            'projected_days': projected_days,
            'cogs': cogs,
            'gross_profit': gross_profit,
            'income': extra_income,
            'expenses': expenses,
            'net': net,
            'running_net': running_net,
        })
        totals['sales'] += sales
        totals['cogs'] += cogs
        totals['gross_profit'] += gross_profit
        totals['income'] += extra_income
        totals['expenses'] += expenses
        totals['net'] += net

    detail_raw = request.args.get('month', '').strip()
    detail_month = None
    if detail_raw:
        try:
            y, m = [int(x) for x in detail_raw.split('-', 1)]
            candidate = date(y, m, 1)
            if period_start <= candidate < period_end:
                detail_month = candidate
        except (ValueError, TypeError):
            detail_month = None
    if detail_month is None:
        current_month = date(today.year, today.month, 1)
        detail_month = current_month if period_start <= current_month < period_end else period_start

    daily_rows = []
    cursor = detail_month
    detail_end = cashflow_add_months(detail_month, 1)
    while cursor < detail_end:
        sales, is_projected, projection_kind = cashflow_sales_for_day(cursor)
        cogs = sales * CASH_FLOW_COGS_RATE
        extra_income = daily_income.get(cursor, 0.0)
        expenses = daily_expenses.get(cursor, 0.0)
        daily_rows.append({
            'date': cursor,
            'order_sales': daily_order_sales.get(cursor, 0.0),
            'vault_sales': daily_vault_sales.get(cursor, 0.0),
            'sales': sales,
            'is_projected': is_projected,
            'projection_kind': projection_kind,
            'cogs': cogs,
            'income': extra_income,
            'expenses': expenses,
            'net': sales + extra_income - cogs - expenses,
        })
        cursor += timedelta(days=1)

    edit_plan = None
    edit_id = parse_int(request.args.get('edit'), 0)
    if edit_id:
        edit_plan = CashFlowPlan.query.get(edit_id)

    def plan_occurrences_in_view(plan):
        return sum(1 for _ in cashflow_occurrence_dates(plan, period_start, period_end))

    def plan_window_total(plan):
        return max(0.0, parse_float(plan.amount, 0.0)) * plan_occurrences_in_view(plan)

    active_expense_total = sum(
        plan_window_total(p) for p in included_plans if p.entry_type == 'EXPENSE'
    )
    active_income_total = sum(
        plan_window_total(p) for p in included_plans if p.entry_type == 'INCOME'
    )

    return render_template(
        'cash_flow_portal.html',
        period_start=period_start,
        period_end=period_end,
        period_last_day=period_end - timedelta(days=1),
        horizon_years=horizon_years,
        horizon_months=horizon_months,
        max_horizon_years=CASH_FLOW_MAX_HORIZON_YEARS,
        monthly_rows=monthly_rows,
        daily_rows=daily_rows,
        detail_month=detail_month,
        totals=totals,
        plans=plans,
        edit_plan=edit_plan,
        cogs_rate=CASH_FLOW_COGS_RATE,
        average_daily_sales=average_daily_sales,
        average_sales_day_count=len(actual_sales_days),
        lifetime_actual_sales_day_count=lifetime_actual_sales_day_count,
        average_window=average_window,
        average_window_label=average_window_label,
        projection_mode=projection_mode,
        break_even_daily_sales=break_even_daily_sales,
        projection_daily_sales=projection_daily_sales,
        projection_source_label=projection_source_label,
        projection_badge=projection_badge,
        max_duration=CASH_FLOW_MAX_DURATION,
        active_expense_total=active_expense_total,
        active_income_total=active_income_total,
        excluded_plans=excluded_plans,
        excluded_window_sales=sum(max(0.0, parse_float(o.total_amount, 0.0)) for o in excluded_window_orders),
        excluded_historical_sales=sum(max(0.0, parse_float(o.total_amount, 0.0)) for o in excluded_historical_orders),
        plan_end_date=cashflow_plan_end_date,
        plan_occurrences_in_view=plan_occurrences_in_view,
        plan_window_total=plan_window_total,
    )


@app.route('/admin/cash-flow/projection/save', methods=['POST'])
@require_admin
def cash_flow_projection_save():
    projection_mode = request.form.get('projection_mode', 'AVERAGE').strip().upper()
    average_window = request.form.get('average_window', 'LIFETIME').strip().upper()
    manual_daily_sales = parse_float(request.form.get('manual_daily_sales'), 0.0)
    manual_start_raw = request.form.get('manual_start_date', '').strip()
    view_start = request.form.get('view_start', '').strip()
    view_month = request.form.get('view_month', '').strip()
    view_years = request.form.get('view_years', '').strip()

    if projection_mode not in ('AVERAGE', 'MANUAL'):
        flash('Choose Actual Sales Average or Specific Daily Sales Amount.', 'error')
        return redirect(url_for('cash_flow_portal', start=view_start, month=view_month, years=view_years))
    if average_window not in ('3', '7', '15', '30', 'LIFETIME'):
        flash('Average range must be 3, 7, 15, 30 actual sales days, or Lifetime.', 'error')
        return redirect(url_for('cash_flow_portal', start=view_start, month=view_month, years=view_years))
    if manual_daily_sales < 0 or manual_daily_sales > 100000000:
        flash('Specific daily sales amount must be between ₱0.00 and ₱100,000,000.00.', 'error')
        return redirect(url_for('cash_flow_portal', start=view_start, month=view_month, years=view_years))

    existing_manual_start = StoreSetting.query.filter_by(key='cashflow_manual_start_date').first()
    if projection_mode == 'MANUAL':
        if not manual_start_raw:
            flash('Choose the starting date for the specific daily sales amount.', 'error')
            return redirect(url_for('cash_flow_portal', start=view_start, month=view_month, years=view_years))
        try:
            manual_start_date = datetime.strptime(manual_start_raw, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            flash('Specific sales starting date is invalid.', 'error')
            return redirect(url_for('cash_flow_portal', start=view_start, month=view_month, years=view_years))
    else:
        # Keep the last manual start date ready in case the user switches back later.
        manual_start_raw = (existing_manual_start.value if existing_manual_start else manual_start_raw).strip()

    values = {
        'cashflow_projection_mode': projection_mode,
        'cashflow_average_window': average_window,
        'cashflow_manual_daily_sales': f'{manual_daily_sales:.2f}',
        'cashflow_manual_start_date': manual_start_raw,
    }
    try:
        for key, value in values.items():
            setting = StoreSetting.query.filter_by(key=key).first()
            if setting:
                setting.value = value
            else:
                db.session.add(StoreSetting(key=key, value=value))
        db.session.commit()
        if projection_mode == 'MANUAL':
            flash(
                f'Cash-flow blank-day projection set to ₱{manual_daily_sales:,.2f} per day '
                f'starting {manual_start_date.strftime("%b %d, %Y")}.',
                'success',
            )
        else:
            label = 'Lifetime' if average_window == 'LIFETIME' else f'last {average_window} actual sales days'
            flash(f'Cash-flow blank-day projection switched back to the {label} average.', 'success')
    except Exception:
        db.session.rollback()
        app.logger.exception('Could not save cash-flow sales projection settings')
        flash('Could not save cash-flow projection settings.', 'error')

    redirect_kwargs = {}
    if view_start:
        redirect_kwargs['start'] = view_start
    if view_month:
        redirect_kwargs['month'] = view_month
    if view_years:
        redirect_kwargs['years'] = view_years
    return redirect(url_for('cash_flow_portal', **redirect_kwargs))


@app.route('/admin/cash-flow/plan/save', methods=['POST'])
@require_admin
def cash_flow_plan_save():
    plan_id = parse_int(request.form.get('plan_id'), 0)
    entry_type = request.form.get('entry_type', '').strip().upper()
    title = request.form.get('title', '').strip()
    amount = parse_float(request.form.get('amount'), 0.0)
    frequency = request.form.get('frequency', '').strip().upper()
    indefinite = request.form.get('indefinite') == 'on'
    duration_count = 0 if indefinite else parse_int(request.form.get('duration_count'), 0)
    category = request.form.get('category', '').strip()[:80]
    notes = request.form.get('notes', '').strip()[:255]
    start_raw = request.form.get('start_date', '').strip()
    view_start = request.form.get('view_start', '').strip()
    view_years = request.form.get('view_years', '').strip()
    return_kwargs = {}
    if view_start:
        return_kwargs['start'] = view_start
    if view_years:
        return_kwargs['years'] = view_years

    if entry_type not in CASH_FLOW_ENTRY_TYPES:
        flash('Choose Expense or Additional Income.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))
    if not title:
        flash('Please enter a title/description.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))
    if amount <= 0:
        flash('Amount must be greater than ₱0.00.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))
    if frequency not in CASH_FLOW_FREQUENCIES:
        flash('Frequency must be Daily, Weekly, Bi-weekly, or Monthly.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))
    max_count = CASH_FLOW_MAX_DURATION[frequency]
    if not indefinite and (duration_count < 1 or duration_count > max_count):
        unit = {'DAILY': 'days', 'WEEKLY': 'weeks', 'BIWEEKLY': 'bi-weekly periods', 'MONTHLY': 'months'}[frequency]
        flash(f'Duration must be between 1 and {max_count} {unit}, or choose Indefinite.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))
    try:
        start_date = date.fromisoformat(start_raw)
    except ValueError:
        flash('Please choose a valid start date.', 'error')
        return redirect(url_for('cash_flow_portal', **return_kwargs))

    if plan_id:
        plan = CashFlowPlan.query.get_or_404(plan_id)
    else:
        plan = CashFlowPlan(created_by=session.get('admin_user'))
        db.session.add(plan)

    plan.entry_type = entry_type
    plan.title = title[:150]
    plan.amount = amount
    plan.frequency = frequency
    plan.start_date = start_date
    plan.duration_count = duration_count
    plan.category = category or None
    plan.notes = notes or None
    plan.is_active = True if not plan_id else plan.is_active
    db.session.commit()

    action = 'updated' if plan_id else 'added'
    flash(f'Cash-flow {entry_type.lower()} schedule {action}.', 'success')
    return redirect(url_for('cash_flow_portal', **return_kwargs))


@app.route('/admin/cash-flow/plan/<int:plan_id>/toggle', methods=['POST'])
@require_admin
def cash_flow_plan_toggle(plan_id):
    plan = CashFlowPlan.query.get_or_404(plan_id)
    plan.is_active = not bool(plan.is_active)
    db.session.commit()
    flash(f"'{plan.title}' is now {'active' if plan.is_active else 'paused'}.", 'success')
    view_start = request.form.get('view_start', '').strip()
    view_years = request.form.get('view_years', '').strip()
    kwargs = {}
    if view_start:
        kwargs['start'] = view_start
    if view_years:
        kwargs['years'] = view_years
    return redirect(url_for('cash_flow_portal', **kwargs))


@app.route('/admin/cash-flow/plan/<int:plan_id>/delete', methods=['POST'])
@require_admin
def cash_flow_plan_delete(plan_id):
    plan = CashFlowPlan.query.get_or_404(plan_id)
    label = plan.title
    db.session.delete(plan)
    db.session.commit()
    flash(f"Cash-flow schedule '{label}' deleted.", 'info')
    view_start = request.form.get('view_start', '').strip()
    view_years = request.form.get('view_years', '').strip()
    kwargs = {}
    if view_start:
        kwargs['start'] = view_start
    if view_years:
        kwargs['years'] = view_years
    return redirect(url_for('cash_flow_portal', **kwargs))



# ==================== AI MARKETING ADMIN ====================

@app.route('/admin/marketing')
@require_admin
def marketing_admin():
    cfg = marketing_settings()
    groups = MarketingGroup.query.order_by(MarketingGroup.is_active.desc(), MarketingGroup.name.asc()).all()
    posts = MarketingPost.query.order_by(MarketingPost.created_at.desc()).limit(100).all()
    try:
        daily_menu_preview = build_daily_menu_caption()
        daily_menu_preview_error = ''
    except OrderValidationError as exc:
        daily_menu_preview = ''
        daily_menu_preview_error = str(exc)
    delivery_metrics = {
        'contacts': MessengerContact.query.count(),
        'sent': MessengerDelivery.query.filter(MessengerDelivery.status.in_(['SENT', 'DELIVERED', 'READ', 'CLICKED'])).count(),
        'delivered': MessengerDelivery.query.filter(MessengerDelivery.status.in_(['DELIVERED', 'READ', 'CLICKED'])).count(),
        'read': MessengerDelivery.query.filter(MessengerDelivery.status.in_(['READ', 'CLICKED'])).count(),
        'clicked': MessengerDelivery.query.filter_by(status='CLICKED').count(),
        'failed': MessengerDelivery.query.filter_by(status='FAILED').count(),
    }
    return render_template(
        'marketing_admin.html',
        settings=cfg,
        groups=groups,
        posts=posts,
        gemini_ready=gemini_configured(),
        openai_ready=openai_configured(),
        gemini_model=os.environ.get('GEMINI_MARKETING_MODEL', 'gemini-3.5-flash-lite'),
        openai_model=os.environ.get('OPENAI_MARKETING_MODEL', 'gpt-5.5'),
        cron_ready=bool(os.environ.get('MARKETING_CRON_TOKEN')),
        post_types=MARKETING_POST_TYPES,
        active_products=Product.query.filter_by(is_active=True).order_by(Product.name.asc()).all(),
        insight_imports=MarketingInsightImport.query.order_by(MarketingInsightImport.created_at.desc()).limit(10).all(),
        daily_menu_preview=daily_menu_preview,
        daily_menu_preview_error=daily_menu_preview_error,
        daily_menu_runs=FacebookMenuRun.query.order_by(FacebookMenuRun.created_at.desc()).limit(20).all(),
        messenger_deliveries=MessengerDelivery.query.order_by(MessengerDelivery.created_at.desc()).limit(20).all(),
        messenger_metrics=delivery_metrics,
        page_api_ready=facebook_page_api_ready(),
        webhook_ready=messenger_webhook_ready(),
        provider_ready=marketing_provider_ready(),
        messenger_start_url=messenger_menu_start_url(),
        messenger_webhook_url=url_for('meta_messenger_webhook', _external=True),
    )

@app.route('/admin/marketing/settings', methods=['POST'])
@require_admin
def marketing_save_settings():
    scope = request.form.get('business_scope', 'BOTH').upper()
    if scope not in ('FOODHOUSE', 'CRAFT', 'BOTH'):
        scope = 'BOTH'
    provider = request.form.get('ai_provider', 'GEMINI').upper()
    if provider not in ('GEMINI', 'AUTO', 'OPENAI', 'TEMPLATE'):
        provider = 'GEMINI'
    save_marketing_setting('marketing_enabled', 'true' if request.form.get('enabled') else 'false')
    save_marketing_setting('marketing_mode', 'MANUAL_POSTING')
    save_marketing_setting('marketing_ai_provider', provider)
    save_marketing_setting('marketing_business_scope', scope)
    save_marketing_setting('marketing_posts_per_week', max(1, min(7, parse_int(request.form.get('posts_per_week'), 4))))
    save_marketing_setting('marketing_start_hour', request.form.get('start_hour', '08:00'))
    save_marketing_setting('marketing_end_hour', request.form.get('end_hour', '19:00'))
    save_marketing_setting('marketing_repeat_cooldown_days', max(0, min(90, parse_int(request.form.get('repeat_cooldown_days'), 10))))
    save_marketing_setting('marketing_max_posts_per_day', max(1, min(3, parse_int(request.form.get('max_posts_per_day'), 1))))
    db.session.commit()
    flash('AI Marketing settings saved. Facebook publishing remains manual.', 'success')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/facebook-page', methods=['POST'])
@require_admin
def marketing_save_facebook_page():
    name = request.form.get('page_name', '').strip()[:150] or "Macleen's Facebook Page"
    page_url = request.form.get('page_url', '').strip()
    if page_url and not page_url.startswith(('https://facebook.com/', 'https://www.facebook.com/', 'https://m.facebook.com/')):
        flash('Enter a valid Facebook Page URL starting with https://www.facebook.com/.', 'error')
        return redirect(url_for('marketing_admin'))
    save_marketing_setting('marketing_facebook_page_name', name)
    save_marketing_setting('marketing_facebook_page_url', page_url)
    db.session.commit()
    flash('Facebook Page shortcut saved. No Meta API connection is required.', 'success')
    return redirect(url_for('marketing_admin'))


@app.route('/admin/marketing/daily-menu/settings', methods=['POST'])
@require_admin
def marketing_save_daily_menu_settings():
    send_time = request.form.get('daily_menu_time', '07:00').strip()
    try:
        datetime.strptime(send_time, '%H:%M')
    except ValueError:
        flash('Choose a valid daily menu time.', 'error')
        return redirect(url_for('marketing_admin') + '#daily-menu-automation')
    page_username = request.form.get('page_username', '').strip().lstrip('@')
    if page_username and not re.fullmatch(r'[A-Za-z0-9._-]{3,100}', page_username):
        flash('Facebook Page username may use letters, numbers, periods, underscores, and hyphens only.', 'error')
        return redirect(url_for('marketing_admin') + '#daily-menu-automation')

    save_marketing_setting('fb_daily_menu_enabled', 'true' if request.form.get('daily_menu_enabled') else 'false')
    save_marketing_setting('fb_daily_menu_time', send_time)
    save_marketing_setting('fb_daily_menu_require_approval', 'true' if request.form.get('require_approval') else 'false')
    save_marketing_setting('fb_daily_menu_auto_page_post', 'true' if request.form.get('auto_page_post') else 'false')
    save_marketing_setting('fb_daily_menu_messenger_reply', 'true' if request.form.get('messenger_reply') else 'false')
    save_marketing_setting('fb_daily_menu_provider_enabled', 'true' if request.form.get('provider_enabled') else 'false')
    save_marketing_setting('fb_daily_menu_intro', request.form.get('daily_menu_intro', '').strip()[:240])
    save_marketing_setting('fb_daily_menu_page_username', page_username)
    db.session.commit()

    missing = []
    if request.form.get('auto_page_post') and not facebook_page_api_ready():
        missing.append('Page API credentials')
    if request.form.get('messenger_reply') and not messenger_webhook_ready():
        missing.append('Messenger webhook credentials')
    if request.form.get('provider_enabled') and not marketing_provider_ready():
        missing.append('approved provider webhook')
    if missing:
        flash('Settings saved. Before those channels can send, configure: ' + ', '.join(missing) + '.', 'info')
    else:
        flash('Daily Facebook menu automation settings saved.', 'success')
    return redirect(url_for('marketing_admin') + '#daily-menu-automation')


@app.route('/admin/marketing/daily-menu/draft', methods=['POST'])
@require_admin
def marketing_create_daily_menu_draft():
    try:
        runs = create_daily_menu_drafts(session.get('admin_user') or 'admin')
        flash(f"Today's menu draft is ready for {len(runs)} configured channel(s).", 'success')
    except OrderValidationError as exc:
        flash(str(exc), 'error')
    return redirect(url_for('marketing_admin') + '#daily-menu-automation')


@app.route('/admin/marketing/daily-menu/publish', methods=['POST'])
@require_admin
def marketing_publish_daily_menu():
    try:
        result = publish_daily_menu_page(request.form.get('caption'), session.get('admin_user') or 'admin')
        flash(result['message'], 'success' if result.get('sent') else 'info')
    except OrderValidationError as exc:
        flash(str(exc), 'error')
    return redirect(url_for('marketing_admin') + '#daily-menu-automation')


@app.route('/admin/marketing/daily-menu/provider-send', methods=['POST'])
@require_admin
def marketing_send_daily_menu_provider():
    try:
        result = send_daily_menu_to_provider(request.form.get('caption'), session.get('admin_user') or 'admin')
        flash(result['message'], 'success' if result.get('sent') else 'info')
    except OrderValidationError as exc:
        flash(str(exc), 'error')
    return redirect(url_for('marketing_admin') + '#daily-menu-automation')


@app.route('/admin/marketing/daily-menu/test-reply', methods=['POST'])
@require_admin
def marketing_test_menu_reply():
    psid = request.form.get('psid', '').strip()
    if not re.fullmatch(r'[A-Za-z0-9_-]{5,180}', psid):
        flash('Enter a valid Page-scoped Messenger user ID from a received webhook.', 'error')
        return redirect(url_for('marketing_admin') + '#daily-menu-automation')
    try:
        delivery = send_messenger_menu_reply(psid)
        flash(f'Test menu reply sent. Delivery log #{delivery.id}.', 'success')
    except OrderValidationError as exc:
        flash(str(exc), 'error')
    return redirect(url_for('marketing_admin') + '#daily-menu-automation')

@app.route('/admin/marketing/run-agent', methods=['POST'])
@require_admin
def marketing_run_agent_now():
    try:
        result = run_marketing_agent_once()
        flash(result.get('message') or 'AI Marketing agent finished.', 'success' if result.get('ran') else 'info')
    except Exception as exc:
        app.logger.exception('Manual AI marketing agent run failed')
        flash(f'AI Marketing agent failed: {exc}', 'error')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/generate', methods=['POST'])
@require_admin
def marketing_generate():
    try:
        post = create_ai_marketing_post(
            business_hint=request.form.get('business', 'AUTO').upper(),
            post_type_hint=request.form.get('post_type', 'AUTO').upper(),
            product_id=request.form.get('product_id'),
        )
        if post.status == 'SKIPPED':
            flash(f'AI chose not to create a promotional post: {post.reason}', 'info')
        else:
            flash(f'AI draft #{post.id} created. Edit/copy it, then post it manually to Facebook.', 'success')
    except Exception as exc:
        app.logger.exception('AI marketing draft generation failed')
        flash(f'Could not generate AI post: {exc}', 'error')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/post/<int:post_id>/edit', methods=['POST'])
@require_admin
def marketing_edit_post(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.status == 'POSTED':
        flash('A post already marked as posted cannot be edited from this dashboard.', 'error')
        return redirect(url_for('marketing_admin'))
    caption = request.form.get('caption', '').strip()
    if not caption:
        flash('Caption cannot be blank.', 'error')
    else:
        post.caption = caption[:3000]
        post.status = 'DRAFT'
        post.error_message = None
        db.session.commit()
        flash(f'Draft #{post.id} updated.', 'success')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/post/<int:post_id>/mark-posted', methods=['POST'])
@require_admin
def marketing_page_mark_posted(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.target_type != 'FACEBOOK_PAGE':
        flash('This action is only for the manual Facebook Page queue.', 'error')
    elif post.status == 'SKIPPED':
        flash('A skipped AI decision cannot be marked as posted.', 'error')
    elif post.status == 'POSTED':
        flash('This post is already marked as posted.', 'info')
    else:
        post.status = 'POSTED'
        post.approved_at = post.approved_at or utc_now()
        post.published_at = utc_now()
        post.facebook_post_id = None
        post.error_message = None
        db.session.commit()
        flash(f'Post #{post.id} marked as manually posted to Facebook.', 'success')
    return redirect(url_for('marketing_admin'))

def _assign_marketing_insights_from_form(post):
    integer_fields = (
        'reach', 'impressions', 'reactions', 'comments', 'shares', 'saves',
        'link_clicks', 'new_followers',
    )
    for name in integer_fields:
        raw = request.form.get(name, '0').strip()
        try:
            value = int(raw or 0)
        except ValueError:
            raise OrderValidationError(f'{name.replace("_", " ").title()} must be a whole number.')
        setattr(post, f'insight_{name}', max(0, value))

    spend = parse_float(request.form.get('spend'), 0.0)
    if spend < 0:
        raise OrderValidationError('Ad spend cannot be negative.')
    post.insight_spend = round(spend, 2)
    post.insight_notes = request.form.get('insight_notes', '').strip()[:3000]
    post.insights_updated_at = utc_now()

@app.route('/admin/marketing/post/<int:post_id>/insights', methods=['POST'])
@require_admin
def marketing_save_insights(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.status != 'POSTED':
        flash('Mark the post as posted before adding Meta insights.', 'error')
        return redirect(url_for('marketing_admin') + f'#insights-{post.id}')
    try:
        _assign_marketing_insights_from_form(post)
    except OrderValidationError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('marketing_admin') + f'#insights-{post.id}')
    db.session.commit()
    flash(f'Meta insights saved for Post #{post.id}.', 'success')
    return redirect(url_for('marketing_admin') + f'#insights-{post.id}')

@app.route('/admin/marketing/insights/import', methods=['POST'])
@require_admin
def marketing_import_insights():
    upload = request.files.get('meta_export')
    if not upload:
        flash('Choose a Meta Business Suite CSV or XLSX export.', 'error')
        return redirect(url_for('marketing_admin') + '#meta-import')
    try:
        filename, summary, analysis = parse_meta_insight_upload(upload)
        db.session.add(MarketingInsightImport(filename=filename, row_count=summary['rows'],
            summary_json=json.dumps(summary), analysis=analysis, uploaded_by=session.get('admin_user') or 'admin'))
        db.session.commit()
        flash(f'Meta export analyzed: {summary["rows"]} post row(s).', 'success')
    except (OrderValidationError, Exception) as exc:
        db.session.rollback()
        if not isinstance(exc, OrderValidationError): app.logger.exception('Meta insight import failed')
        flash(str(exc) if isinstance(exc, OrderValidationError) else 'Could not analyze that Meta export.', 'error')
    return redirect(url_for('marketing_admin') + '#meta-import')

@app.route('/admin/marketing/post/<int:post_id>/analyze-insights', methods=['POST'])
@require_admin
def marketing_analyze_insights(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.status != 'POSTED':
        flash('Only posts marked as posted can be analyzed.', 'error')
        return redirect(url_for('marketing_admin') + f'#insights-{post.id}')
    try:
        _assign_marketing_insights_from_form(post)
    except OrderValidationError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('marketing_admin') + f'#insights-{post.id}')

    comparison_rows = MarketingPost.query.filter(
        MarketingPost.status == 'POSTED',
        MarketingPost.id != post.id,
        MarketingPost.insights_updated_at.isnot(None),
    ).order_by(MarketingPost.insights_updated_at.desc()).limit(10).all()
    payload = {
        'post': {
            'id': post.id,
            'business': post.business,
            'post_type': post.post_type,
            'caption': post.caption[:3000],
            'published_at': post.published_at.isoformat() if post.published_at else None,
            'insights_captured_at': post.insights_updated_at.isoformat() if post.insights_updated_at else None,
        },
        'metrics': {
            'reach': post.insight_reach or 0,
            'impressions': post.insight_impressions or 0,
            'reactions': post.insight_reactions or 0,
            'comments': post.insight_comments or 0,
            'shares': post.insight_shares or 0,
            'saves': post.insight_saves or 0,
            'link_clicks': post.insight_link_clicks or 0,
            'new_followers': post.insight_new_followers or 0,
            'spend_php': post.insight_spend or 0.0,
            'notes': post.insight_notes or '',
        },
        'recent_internal_comparisons': [
            {
                'post_type': row.post_type,
                'reach': row.insight_reach or 0,
                'reactions': row.insight_reactions or 0,
                'comments': row.insight_comments or 0,
                'shares': row.insight_shares or 0,
                'link_clicks': row.insight_link_clicks or 0,
            }
            for row in comparison_rows
        ],
    }
    try:
        result = analyze_marketing_insights(payload, provider=marketing_settings().get('ai_provider', 'GEMINI'))
        post.insight_analysis = str(result.get('analysis') or '')[:5000]
        post.insight_ai_model = str(result.get('model') or 'smart-template:insights')[:100]
        post.insights_analyzed_at = utc_now()
        db.session.commit()
        flash(f'AI analysis completed for Post #{post.id}.', 'success')
    except Exception as exc:
        db.session.rollback()
        app.logger.exception('Marketing insight analysis failed for post_id=%s', post.id)
        flash(f'Could not analyze Post #{post.id}: {exc}', 'error')
    return redirect(url_for('marketing_admin') + f'#insights-{post.id}')

@app.route('/admin/marketing/post/<int:post_id>/delete', methods=['POST'])
@require_admin
def marketing_delete_post(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.status == 'POSTED':
        flash('Posted history is retained for marketing memory and cannot be deleted here.', 'error')
    else:
        db.session.delete(post)
        db.session.commit()
        flash('Marketing draft removed.', 'info')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/group/add', methods=['POST'])
@require_admin
def marketing_group_add():
    name = request.form.get('name', '').strip()
    group_url = request.form.get('group_url', '').strip()
    if not name or not group_url.startswith(('https://facebook.com/', 'https://www.facebook.com/', 'https://m.facebook.com/')):
        flash('Enter a group name and a valid Facebook group URL.', 'error')
        return redirect(url_for('marketing_admin'))
    scope = request.form.get('business_scope', 'BOTH').upper()
    if scope not in ('FOODHOUSE', 'CRAFT', 'BOTH'):
        scope = 'BOTH'
    group = MarketingGroup(
        name=name[:150],
        group_url=group_url,
        business_scope=scope,
        post_types=request.form.get('post_types', '').strip()[:255],
        cooldown_days=max(0, min(90, parse_int(request.form.get('cooldown_days'), 7))),
        notes=request.form.get('notes', '').strip()[:3000],
        is_active=True,
    )
    db.session.add(group)
    db.session.commit()
    flash(f'Facebook group {name} added to the assisted queue.', 'success')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/group/<int:group_id>/toggle', methods=['POST'])
@require_admin
def marketing_group_toggle(group_id):
    group = MarketingGroup.query.get_or_404(group_id)
    group.is_active = not group.is_active
    db.session.commit()
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/group/<int:group_id>/delete', methods=['POST'])
@require_admin
def marketing_group_delete(group_id):
    group = MarketingGroup.query.get_or_404(group_id)
    db.session.delete(group)
    db.session.commit()
    flash('Group removed from the assisted queue.', 'info')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/group/<int:group_id>/generate', methods=['POST'])
@require_admin
def marketing_group_generate(group_id):
    group = MarketingGroup.query.get_or_404(group_id)
    if group.last_posted_at and group.cooldown_days and utc_now() - group.last_posted_at < timedelta(days=group.cooldown_days):
        flash(f'{group.name} is still inside its {group.cooldown_days}-day cooldown.', 'info')
        return redirect(url_for('marketing_admin'))
    try:
        post = create_ai_marketing_post(group=group)
        flash(f'Group-assisted draft #{post.id} created for {group.name}.', 'success' if post.status != 'SKIPPED' else 'info')
    except Exception as exc:
        app.logger.exception('Group-assisted marketing generation failed')
        flash(f'Could not generate group post: {exc}', 'error')
    return redirect(url_for('marketing_admin'))

@app.route('/admin/marketing/group-post/<int:post_id>/mark-posted', methods=['POST'])
@require_admin
def marketing_group_mark_posted(post_id):
    post = MarketingPost.query.get_or_404(post_id)
    if post.target_type != 'GROUP_ASSIST' or not post.group:
        flash('This is not a group-assisted post.', 'error')
    else:
        post.status = 'POSTED'
        post.published_at = utc_now()
        post.group.last_posted_at = utc_now()
        db.session.commit()
        flash('Group post marked as posted.', 'success')
    return redirect(url_for('marketing_admin'))

@app.route('/tasks/marketing/run', methods=['GET', 'POST'])
def marketing_cron_run():
    configured = os.environ.get('MARKETING_CRON_TOKEN', '').strip()
    supplied = request.headers.get('Authorization', '')
    if supplied.lower().startswith('bearer '):
        supplied = supplied[7:].strip()
    if not supplied:
        supplied = request.args.get('token', '').strip()
    if not configured or not secrets.compare_digest(supplied, configured):
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 401
    try:
        result = run_marketing_agent_once()
        try:
            daily_menu = run_daily_menu_automation_once(triggered_by='marketing-cron')
        except OrderValidationError as exc:
            daily_menu = {'ran': False, 'message': str(exc), 'channels': {}}
        return jsonify({'ok': True, **result, 'daily_menu': daily_menu})
    except Exception as exc:
        app.logger.exception('Scheduled AI marketing run failed')
        return jsonify({'ok': False, 'message': str(exc)}), 500


@app.route('/tasks/facebook-menu/run', methods=['GET', 'POST'])
def facebook_menu_cron_run():
    configured = os.environ.get('MARKETING_CRON_TOKEN', '').strip()
    supplied = request.headers.get('Authorization', '')
    if supplied.lower().startswith('bearer '):
        supplied = supplied[7:].strip()
    if not supplied:
        supplied = request.args.get('token', '').strip()
    if not configured or not secrets.compare_digest(supplied, configured):
        return jsonify({'ok': False, 'message': 'Unauthorized'}), 401
    try:
        result = run_daily_menu_automation_once(triggered_by='facebook-menu-cron')
        return jsonify({'ok': True, **result})
    except OrderValidationError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        app.logger.exception('Scheduled Facebook menu run failed')
        return jsonify({'ok': False, 'message': str(exc)}), 500


def _meta_webhook_signature_valid(raw_body):
    app_secret = os.environ.get('META_APP_SECRET', '').strip()
    signature = request.headers.get('X-Hub-Signature-256', '').strip()
    if not app_secret or not signature.startswith('sha256='):
        return False
    expected = hmac.new(app_secret.encode('utf-8'), raw_body, hashlib.sha256).hexdigest()
    return secrets.compare_digest(signature[7:], expected)


def _record_messenger_delivery_event(event):
    delivery_event = event.get('delivery') or {}
    mids = [str(mid) for mid in (delivery_event.get('mids') or []) if mid]
    if mids:
        for delivery in MessengerDelivery.query.filter(MessengerDelivery.message_id.in_(mids)).all():
            delivery.delivered_at = delivery.delivered_at or utc_now()
            if delivery.status == 'SENT':
                delivery.status = 'DELIVERED'

    read_event = event.get('read') or {}
    sender_id = str((event.get('sender') or {}).get('id') or '').strip()
    watermark = parse_int(read_event.get('watermark'), 0)
    if sender_id and watermark:
        contact = MessengerContact.query.filter_by(psid=sender_id).first()
        if contact:
            cutoff = datetime.fromtimestamp(watermark / 1000.0, tz=timezone.utc).replace(tzinfo=None)
            rows = MessengerDelivery.query.filter(
                MessengerDelivery.contact_id == contact.id,
                MessengerDelivery.sent_at.isnot(None),
                MessengerDelivery.sent_at <= cutoff,
            ).all()
            for delivery in rows:
                delivery.read_at = delivery.read_at or utc_now()
                if delivery.status in ('SENT', 'DELIVERED'):
                    delivery.status = 'READ'
    db.session.commit()


@app.route('/meta/messenger/webhook', methods=['GET', 'POST'])
def meta_messenger_webhook():
    if request.method == 'GET':
        configured = os.environ.get('META_WEBHOOK_VERIFY_TOKEN', '').strip()
        supplied = request.args.get('hub.verify_token', '')
        if request.args.get('hub.mode') == 'subscribe' and configured and secrets.compare_digest(supplied, configured):
            return Response(request.args.get('hub.challenge', ''), mimetype='text/plain')
        return Response('Webhook verification failed.', status=403, mimetype='text/plain')

    raw_body = request.get_data(cache=True)
    if not _meta_webhook_signature_valid(raw_body):
        return jsonify({'ok': False, 'message': 'Invalid Meta webhook signature.'}), 403
    payload = request.get_json(silent=True) or {}
    if payload.get('object') != 'page':
        return jsonify({'ok': True, 'ignored': True})

    handled = 0
    for entry in payload.get('entry') or []:
        for event in entry.get('messaging') or []:
            try:
                if event.get('delivery') or event.get('read'):
                    _record_messenger_delivery_event(event)
                    handled += 1
                    continue
                sender_id = str((event.get('sender') or {}).get('id') or '').strip()
                if not sender_id:
                    continue
                message = event.get('message') or {}
                if message.get('is_echo'):
                    continue
                contact = _messenger_contact(sender_id, 'MESSENGER')
                db.session.commit()
                text_value = str(message.get('text') or '').strip().casefold()
                postback = str((event.get('postback') or {}).get('payload') or '').strip().upper()
                referral = str((event.get('referral') or {}).get('ref') or '').strip().casefold()

                if text_value in {'stop', 'unsubscribe', 'cancel updates', 'stop updates'}:
                    contact.opted_out = True
                    db.session.commit()
                    send_messenger_text(sender_id, 'You will not receive marketing updates. You can still ask for today\'s menu anytime by sending MENU.', 'STOP_CONFIRM')
                    handled += 1
                elif (
                    marketing_settings()['daily_menu_messenger_reply']
                    and (text_value in {'menu', "today's menu", 'todays menu', 'ulam', 'price list'}
                         or postback == 'DAILY_MENU' or referral == 'daily_menu')
                ):
                    send_messenger_menu_reply(sender_id)
                    handled += 1
            except Exception:
                db.session.rollback()
                app.logger.exception('Could not process one Messenger webhook event')
    return jsonify({'ok': True, 'handled': handled})


@app.route('/m/menu/<string:tracking_token>')
def messenger_menu_click(tracking_token):
    delivery = MessengerDelivery.query.filter_by(tracking_token=tracking_token).first_or_404()
    delivery.clicked_at = delivery.clicked_at or utc_now()
    delivery.status = 'CLICKED'
    db.session.commit()
    return redirect(url_for('store_catalog'))


# ==================== BUSINESS EXPANSION & INVESTOR INTEREST ====================

INVESTOR_CONTACT_CHOICES = {
    'PHONE': 'Phone / SMS',
    'MESSENGER': 'Facebook Messenger',
    'EMAIL': 'Email',
}
INVESTOR_SUPPORT_AREAS = {
    'WHOLE_BUSINESS': 'Whole business expansion',
    'FOOD_HOUSE': 'Food House',
    'CRAFTS': 'Crafts and document services',
    'DIGITAL': 'Future digital assets',
    'EQUIPMENT': 'Equipment, seating, or supplier support',
}
INVESTOR_FUNDING_RANGES = {
    'DISCUSS_PRIVATELY': 'Prefer to discuss privately',
    'PHP_5K_20K': '₱5,000–₱20,000',
    'PHP_20K_50K': '₱20,001–₱50,000',
    'PHP_50K_100K': '₱50,001–₱100,000',
    'PHP_100K_PLUS': 'Above ₱100,000',
    'NON_CASH': 'Equipment, services, or another partnership',
}
INVESTOR_LEAD_STATUSES = {
    'NEW': 'New',
    'CONTACTED': 'Contacted',
    'REVIEWING': 'Private discussion / due diligence',
    'CLOSED': 'Closed',
}
INVESTOR_PAYOUT_OPTIONS = {
    'LUMP_SUM': 'Lump sum at maturity',
    'MONTHLY_INTEREST': 'Monthly interest + whole principal at maturity',
}
INVESTOR_OFFER_OPTIONS = {
    'PHP_50K_24M': {
        'label': '₱50,000 · 1.5% monthly · 2 years',
        'amount': 50000.0,
        'monthly_rate_percent': 1.5,
        'term_months': 24,
        'payout_options': ('LUMP_SUM',),
        'rate_note': 'Fixed proposed discussion rate',
    },
    'PHP_100K_24M': {
        'label': '₱100,000 · 1.5% monthly · 2 years',
        'amount': 100000.0,
        'monthly_rate_percent': 1.5,
        'term_months': 24,
        'payout_options': ('LUMP_SUM', 'MONTHLY_INTEREST'),
        'rate_note': 'Fixed proposed discussion rate',
    },
    'PHP_250K_24M': {
        'label': '₱250,000 · 2% monthly · 2 years',
        'amount': 250000.0,
        'monthly_rate_percent': 2.0,
        'term_months': 24,
        'payout_options': ('LUMP_SUM', 'MONTHLY_INTEREST'),
        'rate_note': 'Fixed proposed discussion rate',
    },
    'PHP_500K_36M': {
        'label': '₱500,000 · up to 2% monthly · 3 years',
        'amount': 500000.0,
        'monthly_rate_percent': 2.0,
        'term_months': 36,
        'payout_options': ('LUMP_SUM', 'MONTHLY_INTEREST'),
        'rate_note': 'Maximum proposed discussion rate; final rate may be lower',
    },
}

def calculate_investor_terms(amount, monthly_rate_percent, term_months, payout_option):
    """Return non-compounding/simple-interest figures for a proposal snapshot."""
    amount = round(max(0.0, parse_float(amount, 0.0)), 2)
    monthly_rate_percent = round(max(0.0, parse_float(monthly_rate_percent, 0.0)), 4)
    term_months = max(0, parse_int(term_months, 0))
    if payout_option not in INVESTOR_PAYOUT_OPTIONS:
        raise OrderValidationError('Please choose a valid payout option.')
    monthly_interest = round(amount * monthly_rate_percent / 100.0, 2)
    total_interest = round(monthly_interest * term_months, 2)
    total_contract = round(amount + total_interest, 2)
    maturity_payment = total_contract if payout_option == 'LUMP_SUM' else round(amount + monthly_interest, 2)
    return {
        'amount': amount,
        'monthly_rate_percent': monthly_rate_percent,
        'term_months': term_months,
        'monthly_interest_amount': monthly_interest,
        'total_interest_amount': total_interest,
        'maturity_payment_amount': maturity_payment,
        'total_contract_amount': total_contract,
    }

def investor_offer_view_options():
    rows = {}
    for code, offer in INVESTOR_OFFER_OPTIONS.items():
        payout_rows = {}
        for payout in offer['payout_options']:
            payout_rows[payout] = calculate_investor_terms(
                offer['amount'], offer['monthly_rate_percent'], offer['term_months'], payout
            )
        rows[code] = {**offer, 'calculations': payout_rows}
    return rows

def private_offer_access_valid():
    raw = session.get('investor_private_access_at')
    if not raw:
        return False
    try:
        granted_at = datetime.fromisoformat(raw)
    except (TypeError, ValueError):
        session.pop('investor_private_access_at', None)
        return False
    if utc_now() - granted_at > timedelta(hours=4):
        session.pop('investor_private_access_at', None)
        return False
    return True

def public_investor_settings():
    goal = max(0.0, parse_float(investor_setting('investor_funding_goal', '0'), 0.0))
    committed = max(0.0, parse_float(investor_setting('investor_funds_committed', '0'), 0.0))
    return {
        'headline': investor_setting(
            'investor_headline',
            "Help Macleen's grow from a neighborhood food house into a stronger physical and online marketplace.",
        ).strip(),
        'summary': investor_setting(
            'investor_summary',
            'We are strengthening customer choice, comfort, and convenience while keeping our offers affordable.',
        ).strip(),
        'contact_phone': investor_setting('investor_contact_phone', '0955 218 7601').strip(),
        'facebook_url': investor_setting('investor_facebook_url', '').strip(),
        'funding_goal': goal,
        'funds_committed': committed,
        'funding_progress': min(100.0, (committed / goal * 100.0) if goal else 0.0),
        'show_funding_numbers': investor_setting('investor_show_funding_numbers', 'false') == 'true',
        'private_offer_enabled': bool(investor_setting('investor_private_access_hash', '').strip()),
    }


def private_investor_business_snapshot(investor):
    """Return selected aggregate metrics only; never expose customer or schedule details."""
    today = ph_today()
    current_month = date(today.year, today.month, 1)
    next_month = cashflow_add_months(current_month, 1)

    completed_orders = Order.query.filter_by(status='COMPLETED').all()
    included_orders = [order for order in completed_orders if not cashflow_order_is_excluded(order)]
    excluded_sales_orders = [order for order in completed_orders if cashflow_order_is_excluded(order)]
    vault_drops = VaultDrop.query.all()

    daily_sales = {}
    for order in included_orders:
        if not order.created_at:
            continue
        day = utc_naive_to_ph(order.created_at).date()
        daily_sales[day] = daily_sales.get(day, 0.0) + max(0.0, parse_float(order.total_amount, 0.0))
    for drop in vault_drops:
        if not drop.created_at:
            continue
        day = utc_naive_to_ph(drop.created_at).date()
        daily_sales[day] = daily_sales.get(day, 0.0) + max(0.0, parse_float(drop.amount, 0.0))

    actual_days = sorted((day, amount) for day, amount in daily_sales.items() if amount > 0 and day <= today)
    all_time_sales = sum(amount for _, amount in actual_days)
    lifetime_daily_average = all_time_sales / len(actual_days) if actual_days else 0.0
    recent_start = today - timedelta(days=29)
    prior_start = today - timedelta(days=59)
    recent_30_sales = sum(amount for day, amount in actual_days if recent_start <= day <= today)
    prior_30_sales = sum(amount for day, amount in actual_days if prior_start <= day < recent_start)
    sales_trend_percent = ((recent_30_sales - prior_30_sales) / prior_30_sales * 100.0) if prior_30_sales > 0 else None

    included_month_expenses = 0.0
    included_month_income = 0.0
    excluded_schedule_count = 0
    active_plans = CashFlowPlan.query.filter_by(is_active=True).all()
    for plan in active_plans:
        if cashflow_plan_exclusion_reason(plan):
            excluded_schedule_count += 1
            continue
        occurrences = sum(1 for _ in cashflow_occurrence_dates(plan, current_month, next_month))
        amount = max(0.0, parse_float(plan.amount, 0.0)) * occurrences
        if plan.entry_type == 'EXPENSE':
            included_month_expenses += amount
        elif plan.entry_type == 'INCOME':
            included_month_income += amount

    contribution_margin_rate = max(0.0001, 1.0 - CASH_FLOW_COGS_RATE)
    required_contribution = max(0.0, included_month_expenses - included_month_income)
    monthly_break_even_sales = required_contribution / contribution_margin_rate
    days_in_month = max(1, (next_month - current_month).days)
    daily_break_even_sales = monthly_break_even_sales / days_in_month
    planning_daily_sales = max(lifetime_daily_average, daily_break_even_sales)
    break_even_gap = lifetime_daily_average - daily_break_even_sales
    break_even_attainment = (
        min(999.0, lifetime_daily_average / daily_break_even_sales * 100.0)
        if daily_break_even_sales > 0 else 100.0
    )

    inventory_value = sum((p.cost or 0.0) * max(0, p.stock or 0) for p in Product.query.all())
    inventory_value += sum((item.cost or 0.0) * max(0, item.stock_quantity or 0) for item in CraftItem.query.all())
    funding_basis = investor['funding_goal'] if investor['funding_goal'] > 0 else 500000.0
    use_of_funds = [
        ('Inventory & working capital', 28),
        ('Kitchen & productive equipment', 22),
        ('Tables, chairs & store improvements', 16),
        ('Crafts equipment & supplies', 12),
        ('Domain & website setup', 7),
        ('Branding & measured marketing', 7),
        ('Contingency & emergency reserve', 8),
    ]

    return {
        'as_of': today,
        'all_time_sales': all_time_sales,
        'recent_30_sales': recent_30_sales,
        'prior_30_sales': prior_30_sales,
        'sales_trend_percent': sales_trend_percent,
        'lifetime_daily_average': lifetime_daily_average,
        'actual_sales_days': len(actual_days),
        'first_sales_day': actual_days[0][0] if actual_days else None,
        'last_sales_day': actual_days[-1][0] if actual_days else None,
        'monthly_break_even_sales': monthly_break_even_sales,
        'daily_break_even_sales': daily_break_even_sales,
        'planning_daily_sales': planning_daily_sales,
        'break_even_gap': break_even_gap,
        'break_even_gap_abs': abs(break_even_gap),
        'break_even_attainment': break_even_attainment,
        'completed_orders': len(included_orders),
        'customer_accounts': Customer.query.count(),
        'active_food_offers': Product.query.filter_by(is_active=True).count(),
        'active_craft_offers': CraftItem.query.filter_by(is_active=True).count(),
        'inventory_value': inventory_value,
        'excluded_schedule_count': excluded_schedule_count,
        'excluded_sales_count': len(excluded_sales_orders),
        'funding_basis': funding_basis,
        'use_of_funds': [
            {'label': label, 'share': share, 'amount': funding_basis * share / 100.0}
            for label, share in use_of_funds
        ],
    }

@app.route('/investor-deck')
@app.route('/investor')
@app.route('/investors')
def investor_opportunity():
    flash('Investor discussions are private and available to selected invitees only.', 'info')
    return redirect(url_for('investor_private_offer'))
    session['investor_form_token'] = secrets.token_urlsafe(24)
    return render_template(
        'investor_deck.html',
        investor=public_investor_settings(),
        form_token=session['investor_form_token'],
        active_food_count=Product.query.filter_by(is_active=True).count(),
        active_craft_count=CraftItem.query.filter_by(is_active=True).count(),
        support_areas=INVESTOR_SUPPORT_AREAS,
        funding_ranges=INVESTOR_FUNDING_RANGES,
        contact_choices=INVESTOR_CONTACT_CHOICES,
        offers=INVESTOR_OFFER_OPTIONS,
        payout_options=INVESTOR_PAYOUT_OPTIONS,
    )

@app.route('/investors/interest', methods=['POST'])
def submit_investor_interest():
    flash('Public investor applications are closed. Private agreements are handled by invitation.', 'info')
    return redirect(url_for('investor_private_offer'))
    supplied_token = request.form.get('form_token', '')
    expected_token = session.pop('investor_form_token', '')
    if not supplied_token or not expected_token or not secrets.compare_digest(supplied_token, expected_token):
        flash('The form expired. Please review the page and submit again.', 'error')
        return redirect(url_for('investor_opportunity') + '#connect')

    # Quietly accept bot-filled honeypot submissions without saving them.
    if request.form.get('website', '').strip():
        return redirect(url_for('investor_opportunity') + '?submitted=1#connect')

    name = re.sub(r'\s+', ' ', request.form.get('name', '').strip())[:120]
    contact = re.sub(r'\s+', ' ', request.form.get('contact', '').strip())[:160]
    preferred_contact = request.form.get('preferred_contact', 'PHONE').strip().upper()
    business_area = request.form.get('business_area', 'WHOLE_BUSINESS').strip().upper()
    funding_range = request.form.get('funding_range', 'DISCUSS_PRIVATELY').strip().upper()
    message = request.form.get('message', '').strip()[:2000]

    if len(name) < 2 or len(contact) < 5:
        flash('Please enter your name and a valid phone, Messenger, or email contact.', 'error')
        return redirect(url_for('investor_opportunity') + '#connect')
    if preferred_contact not in INVESTOR_CONTACT_CHOICES:
        preferred_contact = 'PHONE'
    if business_area not in INVESTOR_SUPPORT_AREAS:
        business_area = 'WHOLE_BUSINESS'
    if funding_range not in INVESTOR_FUNDING_RANGES:
        funding_range = 'DISCUSS_PRIVATELY'
    if request.form.get('privacy_acknowledged') != 'yes':
        flash('Please confirm that Macleen\'s may contact you about a private business discussion.', 'error')
        return redirect(url_for('investor_opportunity') + '#connect')

    recent_duplicate = InvestorInterest.query.filter(
        db.func.lower(InvestorInterest.contact) == contact.lower(),
        InvestorInterest.created_at >= utc_now() - timedelta(minutes=10),
    ).first()
    if recent_duplicate:
        flash('Your interest was already received. Macleen\'s will contact you privately.', 'success')
        return redirect(url_for('investor_opportunity') + '?submitted=1#connect')

    db.session.add(InvestorInterest(
        name=name,
        contact=contact,
        preferred_contact=preferred_contact,
        business_area=business_area,
        funding_range=funding_range,
        message=message,
    ))
    db.session.commit()
    flash('Thank you. Your private introduction request has been received.', 'success')
    return redirect(url_for('investor_opportunity') + '?submitted=1#connect')

@app.route('/investors/private-offer', methods=['GET', 'POST'])
def investor_private_offer():
    investor = public_investor_settings()
    if request.method == 'POST':
        configured_hash = investor_setting('investor_private_access_hash', '').strip()
        lock_until_raw = session.get('investor_private_lock_until')
        try:
            lock_until = datetime.fromisoformat(lock_until_raw) if lock_until_raw else None
        except (TypeError, ValueError):
            lock_until = None
        if lock_until and utc_now() < lock_until:
            flash('Too many access attempts. Please wait 15 minutes or contact Macleen\'s.', 'error')
            return redirect(url_for('investor_private_offer'))

        access_code = request.form.get('access_code', '').strip()
        if configured_hash and access_code and check_password_hash(configured_hash, access_code):
            session['investor_private_access_at'] = utc_now().isoformat()
            session.pop('investor_private_attempts', None)
            session.pop('investor_private_lock_until', None)
            flash('Private proposal options unlocked for this session.', 'success')
            return redirect(url_for('investor_private_offer'))

        attempts = parse_int(session.get('investor_private_attempts'), 0) + 1
        if attempts >= 5:
            session['investor_private_attempts'] = 0
            session['investor_private_lock_until'] = (utc_now() + timedelta(minutes=15)).isoformat()
        else:
            session['investor_private_attempts'] = attempts
        flash('The private access code is not valid.', 'error')
        return redirect(url_for('investor_private_offer'))

    unlocked = private_offer_access_valid() and investor['private_offer_enabled']
    form_token = ''
    snapshot = None
    if unlocked:
        form_token = secrets.token_urlsafe(24)
        session['investor_proposal_form_token'] = form_token
        snapshot = private_investor_business_snapshot(investor)
    return render_template(
        'investor_private_offer.html',
        investor=investor,
        unlocked=unlocked,
        snapshot=snapshot,
        form_token=form_token,
        offers=investor_offer_view_options(),
        payout_options=INVESTOR_PAYOUT_OPTIONS,
        support_areas=INVESTOR_SUPPORT_AREAS,
        contact_choices=INVESTOR_CONTACT_CHOICES,
    ), 200, {'Cache-Control': 'no-store, private'}

@app.route('/investors/private-offer/submit', methods=['POST'])
def submit_private_investor_offer():
    if not private_offer_access_valid() or not investor_setting('investor_private_access_hash', '').strip():
        flash('Private proposal access expired. Enter the access code again.', 'error')
        return redirect(url_for('investor_private_offer'))
    supplied_token = request.form.get('form_token', '')
    expected_token = session.pop('investor_proposal_form_token', '')
    if not supplied_token or not expected_token or not secrets.compare_digest(supplied_token, expected_token):
        flash('The proposal form expired. Please review the terms and submit again.', 'error')
        return redirect(url_for('investor_private_offer'))
    if request.form.get('website', '').strip():
        return redirect(url_for('investor_private_offer') + '?submitted=1')

    name = re.sub(r'\s+', ' ', request.form.get('name', '').strip())[:120]
    contact = re.sub(r'\s+', ' ', request.form.get('contact', '').strip())[:160]
    preferred_contact = request.form.get('preferred_contact', 'PHONE').strip().upper()
    business_area = request.form.get('business_area', 'WHOLE_BUSINESS').strip().upper()
    offer_code = request.form.get('offer_code', '').strip().upper()
    payout_option = request.form.get('payout_option', '').strip().upper()
    message = request.form.get('message', '').strip()[:2000]

    if len(name) < 2 or len(contact) < 5:
        flash('Please enter your name and a valid contact detail.', 'error')
        return redirect(url_for('investor_private_offer'))
    if preferred_contact not in INVESTOR_CONTACT_CHOICES:
        preferred_contact = 'PHONE'
    if business_area not in INVESTOR_SUPPORT_AREAS:
        business_area = 'WHOLE_BUSINESS'
    if request.form.get('privacy_acknowledged') != 'yes' or request.form.get('proposal_acknowledged') != 'yes':
        flash('Please confirm the contact permission and non-binding proposal notice.', 'error')
        return redirect(url_for('investor_private_offer'))

    is_counter_offer = offer_code == 'COUNTER_OFFER'
    try:
        if is_counter_offer:
            amount = parse_float(request.form.get('counter_amount'), 0.0)
            monthly_rate = parse_float(request.form.get('counter_monthly_rate'), 0.0)
            term_months = parse_int(request.form.get('counter_term_months'), 0)
            if amount < 5000.0 or amount > 5000000.0:
                raise OrderValidationError('Counteroffer amount must be from ₱5,000 to ₱5,000,000.')
            if monthly_rate <= 0.0 or monthly_rate > 2.0:
                raise OrderValidationError('Counteroffer monthly rate must be above 0% and not more than 2%.')
            if term_months < 6 or term_months > 60:
                raise OrderValidationError('Counteroffer term must be from 6 to 60 months.')
        else:
            offer = INVESTOR_OFFER_OPTIONS.get(offer_code)
            if not offer:
                raise OrderValidationError('Please choose one proposal option or Counteroffer.')
            if payout_option not in offer['payout_options']:
                raise OrderValidationError('That payout method is not available for the selected proposal.')
            amount = offer['amount']
            monthly_rate = offer['monthly_rate_percent']
            term_months = offer['term_months']
        terms = calculate_investor_terms(amount, monthly_rate, term_months, payout_option)
    except OrderValidationError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('investor_private_offer'))

    recent_duplicate = InvestorInterest.query.filter(
        db.func.lower(InvestorInterest.contact) == contact.lower(),
        InvestorInterest.offer_code == offer_code,
        InvestorInterest.created_at >= utc_now() - timedelta(minutes=10),
    ).first()
    if recent_duplicate:
        flash('This proposal was already received and is waiting for cashier review.', 'success')
        return redirect(url_for('investor_private_offer') + '?submitted=1')

    db.session.add(InvestorInterest(
        name=name,
        contact=contact,
        preferred_contact=preferred_contact,
        business_area=business_area,
        funding_range=offer_code,
        offer_code=offer_code,
        payout_option=payout_option,
        proposed_amount=terms['amount'],
        monthly_rate_percent=terms['monthly_rate_percent'],
        term_months=terms['term_months'],
        monthly_interest_amount=terms['monthly_interest_amount'],
        total_interest_amount=terms['total_interest_amount'],
        maturity_payment_amount=terms['maturity_payment_amount'],
        total_contract_amount=terms['total_contract_amount'],
        is_counter_offer=is_counter_offer,
        message=message,
    ))
    db.session.commit()
    flash('Proposal received and sent to the cashier review queue. This is not yet an agreement.', 'success')
    return redirect(url_for('investor_private_offer') + '?submitted=1')

@app.route('/pos/investor-proposals')
@require_cashier
def cashier_investor_proposals():
    proposals = InvestorInterest.query.order_by(InvestorInterest.created_at.desc()).limit(200).all()
    return render_template(
        'cashier_investor_proposals.html',
        proposals=proposals,
        offers=INVESTOR_OFFER_OPTIONS,
        payout_options=INVESTOR_PAYOUT_OPTIONS,
        lead_statuses=INVESTOR_LEAD_STATUSES,
        support_areas=INVESTOR_SUPPORT_AREAS,
        contact_choices=INVESTOR_CONTACT_CHOICES,
        new_count=sum(1 for row in proposals if row.status == 'NEW'),
    )

@app.route('/pos/investor-proposals/<int:lead_id>/review', methods=['POST'])
@require_cashier
def cashier_review_investor_proposal(lead_id):
    lead = InvestorInterest.query.get_or_404(lead_id)
    status_value = request.form.get('status', '').strip().upper()
    notes = request.form.get('cashier_notes', '').strip()[:2000]
    if status_value not in INVESTOR_LEAD_STATUSES:
        flash('Invalid investor-proposal status.', 'error')
    else:
        lead.status = status_value
        lead.cashier_notes = notes
        lead.cashier_reviewed_by = session.get('cashier_user') or session.get('admin_user') or 'Staff'
        lead.cashier_reviewed_at = utc_now()
        db.session.commit()
        flash(f'Updated {lead.name}\'s proposal.', 'success')
    return redirect(url_for('cashier_investor_proposals') + f'#proposal-{lead.id}')

@app.route('/investor/dashboard')
@app.route('/admin/investors')
@require_admin
def investor_dashboard():
    month_ago = utc_now() - timedelta(days=30)
    all_completed = Order.query.filter_by(status='COMPLETED').all()
    month_completed = Order.query.filter(Order.status == 'COMPLETED', Order.created_at >= month_ago).all()
    all_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).scalar() or 0.0
    month_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).filter(VaultDrop.created_at >= month_ago).scalar() or 0.0
    month_expenses = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0.0)).filter(Expense.created_at >= month_ago).scalar() or 0.0

    total_sales = sum((o.total_amount or 0.0) for o in all_completed) + all_vault
    month_sales = sum((o.total_amount or 0.0) for o in month_completed) + month_vault
    month_cogs = sum(
        (item.cost_price or 0.0) * (item.quantity or 0)
        for order in month_completed for item in order.items
    )
    month_net_estimate = month_sales - month_cogs - month_expenses
    total_ar = db.session.query(db.func.coalesce(db.func.sum(Customer.outstanding_ar), 0.0)).scalar() or 0.0
    inventory_val = sum((p.cost or 0.0) * max(0, p.stock or 0) for p in Product.query.all())
    inventory_val += sum((item.cost or 0.0) * max(0, item.stock_quantity or 0) for item in CraftItem.query.all())
    leads = InvestorInterest.query.order_by(InvestorInterest.created_at.desc()).limit(150).all()

    return render_template(
        'investor_dashboard.html',
        investor=public_investor_settings(),
        total_sales=total_sales,
        total_orders=len(all_completed),
        month_sales=month_sales,
        month_net_estimate=month_net_estimate,
        total_ar=total_ar,
        inventory_val=inventory_val,
        total_customers=Customer.query.count(),
        active_products=Product.query.filter_by(is_active=True).count(),
        active_crafts=CraftItem.query.filter_by(is_active=True).count(),
        leads=leads,
        lead_statuses=INVESTOR_LEAD_STATUSES,
        support_areas=INVESTOR_SUPPORT_AREAS,
        funding_ranges=INVESTOR_FUNDING_RANGES,
        contact_choices=INVESTOR_CONTACT_CHOICES,
        offers=INVESTOR_OFFER_OPTIONS,
        payout_options=INVESTOR_PAYOUT_OPTIONS,
    )

@app.route('/admin/investors/settings', methods=['POST'])
@require_admin
def save_investor_settings():
    headline = re.sub(r'\s+', ' ', request.form.get('headline', '').strip())[:180]
    summary = re.sub(r'\s+', ' ', request.form.get('summary', '').strip())[:500]
    contact_phone = re.sub(r'\s+', ' ', request.form.get('contact_phone', '').strip())[:50]
    facebook_url = request.form.get('facebook_url', '').strip()[:300]
    funding_goal = max(0.0, parse_float(request.form.get('funding_goal'), 0.0))
    funds_committed = max(0.0, parse_float(request.form.get('funds_committed'), 0.0))
    private_access_code = request.form.get('private_access_code', '').strip()
    if not headline or not summary:
        flash('Investor headline and summary are required.', 'error')
        return redirect(url_for('investor_dashboard'))
    if facebook_url and not re.match(r'^https://(www\.)?facebook\.com/', facebook_url, flags=re.I):
        flash('Facebook URL must begin with https://facebook.com/ or https://www.facebook.com/.', 'error')
        return redirect(url_for('investor_dashboard'))
    if private_access_code and len(private_access_code) < 6:
        flash('Private proposal access code must contain at least 6 characters.', 'error')
        return redirect(url_for('investor_dashboard'))

    values = {
        'investor_headline': headline,
        'investor_summary': summary,
        'investor_contact_phone': contact_phone,
        'investor_facebook_url': facebook_url,
        'investor_funding_goal': f'{funding_goal:.2f}',
        'investor_funds_committed': f'{funds_committed:.2f}',
        'investor_show_funding_numbers': 'true' if request.form.get('show_funding_numbers') else 'false',
    }
    for key, value in values.items():
        save_investor_setting(key, value)
    if request.form.get('disable_private_access'):
        save_investor_setting('investor_private_access_hash', '')
    elif private_access_code:
        save_investor_setting('investor_private_access_hash', generate_password_hash(private_access_code))
    db.session.commit()
    flash('Investor page settings updated.', 'success')
    return redirect(url_for('investor_dashboard'))

@app.route('/admin/investors/interest/<int:lead_id>/status', methods=['POST'])
@require_admin
def update_investor_interest_status(lead_id):
    lead = InvestorInterest.query.get_or_404(lead_id)
    status_value = request.form.get('status', '').strip().upper()
    if status_value not in INVESTOR_LEAD_STATUSES:
        flash('Invalid investor-interest status.', 'error')
    else:
        lead.status = status_value
        db.session.commit()
        flash(f'Updated {lead.name}\'s inquiry.', 'success')
    return redirect(url_for('investor_dashboard') + '#inquiries')


# ==================== MASTER ADMIN, CONTROLS & SETTINGS ====================

@app.route('/admin')
@require_admin
def admin_dashboard():
    sort_by = request.args.get('sort', 'likes_desc')

    query = Product.query
    if sort_by == 'likes_desc': query = query.order_by(Product.total_likes.desc())
    elif sort_by == 'price_asc': query = query.order_by(Product.price.asc())
    elif sort_by == 'price_desc': query = query.order_by(Product.price.desc())
    elif sort_by == 'name_asc': query = query.order_by(Product.name.asc())
    elif sort_by == 'category': query = query.order_by(Product.category_name.asc(), Product.name.asc())
    elif sort_by == 'featured': query = query.order_by(Product.is_featured.desc(), Product.name.asc())
    elif sort_by == 'top_seller': query = query.order_by(Product.is_top_seller.desc(), Product.name.asc())

    products = query.all()
    categories = Category.query.all()
    staff_members = Staff.query.all()
    customers = Customer.query.order_by(Customer.id.desc()).all()
    delivery_zones = DeliveryZone.query.all()
    all_orders = Order.query.order_by(Order.created_at.desc()).limit(150).all()
    all_expenses = Expense.query.order_by(Expense.created_at.desc()).all()
    vault_drops = VaultDrop.query.order_by(VaultDrop.created_at.desc()).all()
    promotions = PromotionTracker.query.order_by(PromotionTracker.created_at.desc()).all()

    try:
        unique_visitors = SiteVisitor.query.count()
        total_accumulated_visits = db.session.query(db.func.coalesce(db.func.sum(SiteVisitor.visit_count), 0)).scalar() or unique_visitors
    except Exception:
        app.logger.exception('Admin visitor metrics query failed')
        unique_visitors = 0
        total_accumulated_visits = 0

    today = ph_today()
    day_start, next_day = ph_day_utc_bounds(today)
    week_ago = utc_now() - timedelta(days=7)
    month_ago = utc_now() - timedelta(days=30)

    daily_orders = Order.query.filter(Order.created_at >= day_start, Order.created_at < next_day, Order.status == 'COMPLETED').all()
    weekly_orders = Order.query.filter(Order.created_at >= week_ago, Order.status == 'COMPLETED').all()
    monthly_orders = Order.query.filter(Order.created_at >= month_ago, Order.status == 'COMPLETED').all()
    all_completed = Order.query.filter_by(status='COMPLETED').all()

    daily_exp = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0.0)).filter(Expense.created_at >= day_start, Expense.created_at < next_day).scalar() or 0.0
    weekly_exp = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0.0)).filter(Expense.created_at >= week_ago).scalar() or 0.0
    monthly_exp = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0.0)).filter(Expense.created_at >= month_ago).scalar() or 0.0
    total_exp_all = sum(e.amount for e in all_expenses)

    daily_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).filter(VaultDrop.created_at >= day_start, VaultDrop.created_at < next_day).scalar() or 0.0
    weekly_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).filter(VaultDrop.created_at >= week_ago).scalar() or 0.0
    monthly_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).filter(VaultDrop.created_at >= month_ago).scalar() or 0.0
    all_vault = db.session.query(db.func.coalesce(db.func.sum(VaultDrop.amount), 0.0)).scalar() or 0.0

    def calc_period(orders, exp, vault_drop_sales):
        order_rev = sum(o.total_amount for o in orders)
        total_rev = order_rev + vault_drop_sales
        cost = sum(sum((getattr(it, 'cost_price', 0.0) or 0.0) * it.quantity for it in o.items) for o in orders)
        gross_p = total_rev - cost
        net_p = gross_p - exp
        return {
            'order_rev': order_rev,
            'vault_sales': vault_drop_sales,
            'rev': total_rev,
            'cost': cost,
            'gross_p': gross_p,
            'exp': exp,
            'net_p': net_p
        }

    fin_daily = calc_period(daily_orders, daily_exp, daily_vault)
    fin_weekly = calc_period(weekly_orders, weekly_exp, weekly_vault)
    fin_monthly = calc_period(monthly_orders, monthly_exp, monthly_vault)
    fin_all = calc_period(all_completed, total_exp_all, all_vault)

    product_sales_stats = {}
    food_revenue_total = 0.0
    service_revenue_total = 0.0
    craft_revenue_total = 0.0

    for o in all_completed:
        for it in o.items:
            pname = it.product_name
            if pname not in product_sales_stats:
                product_sales_stats[pname] = {'qty': 0, 'revenue': 0.0, 'cost': 0.0}
            product_sales_stats[pname]['qty'] += (it.quantity or 0)
            product_sales_stats[pname]['revenue'] += (it.subtotal or 0.0)
            product_sales_stats[pname]['cost'] += (getattr(it, 'cost_price', 0.0) or 0.0) * (it.quantity or 0)

            if str(o.order_type or '').upper().startswith('CRAFT') or pname.startswith('[Craft]'):
                craft_revenue_total += (it.subtotal or 0.0)
            elif '[Service]' in pname or o.order_type == 'SERVICE/MISC' or 'Printing' in pname:
                service_revenue_total += (it.subtotal or 0.0)
            else:
                food_revenue_total += (it.subtotal or 0.0)

    food_revenue_total += all_vault
    total_ar = sum((c.outstanding_ar or 0.0) for c in customers)

    bonus_campaigns = BonusCampaign.query.order_by(BonusCampaign.created_at.desc()).all()
    product_suggestions = ProductSuggestion.query.order_by(ProductSuggestion.created_at.desc()).all()

    # Group near-identical requests (case/punctuation-insensitive) so repeated demand is obvious.
    demand_map = {}
    suggestion_repeat_counts = {}
    for suggestion in product_suggestions:
        normalized = re.sub(r'[^a-z0-9]+', ' ', (suggestion.suggestion_text or '').lower()).strip()
        normalized = re.sub(r'\s+', ' ', normalized)
        key = normalized or f'__suggestion_{suggestion.id}'
        bucket = demand_map.setdefault(key, {
            'label': suggestion.suggestion_text.strip(),
            'count': 0,
        })
        bucket['count'] += 1

    for suggestion in product_suggestions:
        normalized = re.sub(r'[^a-z0-9]+', ' ', (suggestion.suggestion_text or '').lower()).strip()
        normalized = re.sub(r'\s+', ' ', normalized)
        key = normalized or f'__suggestion_{suggestion.id}'
        suggestion_repeat_counts[suggestion.id] = demand_map[key]['count']

    suggestion_demand = sorted(
        demand_map.values(),
        key=lambda row: (-row['count'], row['label'].lower())
    )[:10]

    seven_days_ago = utc_now() - timedelta(days=7)

    marketing_metrics = {
        'qr_scans_7d': PortalEvent.query.filter(PortalEvent.event_type == 'QR_SCAN', PortalEvent.created_at >= seven_days_ago).count(),
        'portal_logins_7d': PortalEvent.query.filter(PortalEvent.event_type == 'LOGIN', PortalEvent.created_at >= seven_days_ago).count(),
        'registrations_7d': PortalEvent.query.filter(PortalEvent.event_type == 'REGISTER', PortalEvent.created_at >= seven_days_ago).count(),
        'counter_registrations_7d': PortalEvent.query.filter(PortalEvent.event_type == 'REGISTER', PortalEvent.source == 'counter', PortalEvent.created_at >= seven_days_ago).count(),
        'suggestions_7d': ProductSuggestion.query.filter(ProductSuggestion.created_at >= seven_days_ago).count(),
        'suggestions_total': ProductSuggestion.query.count(),
        'suggestions_open': ProductSuggestion.query.filter(ProductSuggestion.status != 'ARCHIVED').count(),
        'referral_rewards': ReferralReward.query.count(),
    }

    return render_template('admin.html', 
                           products=products, 
                           categories=categories, 
                           staff_members=staff_members, 
                           customers=customers, 
                           delivery_zones=delivery_zones, 
                           all_orders=all_orders, 
                           all_expenses=all_expenses, 
                           vault_drops=vault_drops, 
                           promotions=promotions, 
                           product_sales_stats=product_sales_stats, 
                           food_revenue_total=food_revenue_total, 
                           service_revenue_total=service_revenue_total, 
                           craft_revenue_total=craft_revenue_total, 
                           unique_visitors=unique_visitors, 
                           total_accumulated_visits=total_accumulated_visits, 
                           fin_daily=fin_daily, 
                           fin_weekly=fin_weekly, 
                           fin_monthly=fin_monthly, 
                           fin_all=fin_all, 
                           total_ar=total_ar, 
                           current_sort=sort_by,
                           bonus_campaigns=bonus_campaigns,
                           marketing_metrics=marketing_metrics,
                           product_suggestions=product_suggestions,
                           suggestion_repeat_counts=suggestion_repeat_counts,
                           suggestion_demand=suggestion_demand,
                           highlight_product_id=max(0, parse_int(request.args.get('added'), 0)),
                           suggestion_statuses=SUGGESTION_STATUSES)

@app.route('/admin/menu-vote/candidate/add', methods=['POST'])
@require_admin
def admin_add_menu_vote_candidate():
    name = re.sub(r'\s+', ' ', request.form.get('name', '').strip())
    category = re.sub(r'\s+', ' ', request.form.get('category_name', '').strip()) or 'Customer Requests'
    if len(name) < 2:
        flash('Enter a menu item or product name for voting.', 'error')
        return redirect(url_for('admin_dashboard') + '#menuVoteAdmin')
    candidate = ensure_menu_vote_candidate(name, category_name=category, commit=True)
    candidate.is_active = True
    db.session.commit()
    flash(f"Voting candidate '{candidate.name}' is active.", 'success')
    return redirect(url_for('admin_dashboard') + '#menuVoteAdmin')

@app.route('/admin/menu-vote/candidate/<int:candidate_id>/toggle', methods=['POST'])
@require_admin
def admin_toggle_menu_vote_candidate(candidate_id):
    candidate = MenuVoteCandidate.query.get_or_404(candidate_id)
    candidate.is_active = not bool(candidate.is_active)
    db.session.commit()
    flash(f"{candidate.name} voting is now {'active' if candidate.is_active else 'paused'}.", 'success')
    return redirect(url_for('admin_dashboard') + '#menuVoteAdmin')

@app.route('/admin/product-suggestion/<int:suggestion_id>/status', methods=['POST'])
@require_admin
def admin_update_product_suggestion_status(suggestion_id):
    suggestion = ProductSuggestion.query.get_or_404(suggestion_id)
    new_status = request.form.get('status', 'NEW').strip().upper()
    if new_status not in SUGGESTION_STATUSES:
        flash('Invalid suggestion status.', 'error')
        return redirect(url_for('admin_dashboard') + '#customerSuggestions')

    suggestion.status = new_status
    db.session.commit()
    flash(f"Suggestion #{suggestion.id} marked {new_status}.", 'success')
    return redirect(url_for('admin_dashboard') + '#customerSuggestions')

@app.route('/admin/allocate-vault-drop/<int:drop_id>', methods=['POST'])
@require_admin
def admin_allocate_vault_drop(drop_id):
    drop = VaultDrop.query.get_or_404(drop_id)
    cust_id = request.form.get('customer_id')
    product_id = request.form.get('product_id')
    allocated_amount = parse_float(request.form.get('amount'), 0.0)
    qty = parse_int(request.form.get('quantity'), 1)

    if allocated_amount <= 0 or allocated_amount > drop.amount:
        flash("Allocated amount must be between ₱0.01 and the remaining drop balance.", "error")
        return redirect(url_for('admin_dashboard'))

    cust = Customer.query.get(int(cust_id)) if cust_id else None
    prod = Product.query.get(int(product_id)) if product_id else None
    if cust:
        issue = customer_access_issue(cust)
        if issue:
            flash(issue, 'error')
            return redirect(url_for('admin_dashboard'))
    if qty < 1:
        flash('Quantity must be at least 1.', 'error')
        return redirect(url_for('admin_dashboard'))

    drop.amount = max(0.0, drop.amount - allocated_amount)

    order = Order(
        order_type='ALLOCATED_VAULT_SALE',
        dining_option='DINE-IN',
        customer_id=cust.id if cust else None,
        customer_name=cust.name if cust else 'Allocated Vault Customer',
        contact_number=cust.contact if cust else 'N/A',
        subtotal=allocated_amount,
        delivery_fee=0.0,
        total_amount=allocated_amount,
        payment_method='CASH',
        payment_verified=True,
        status='COMPLETED',
        fulfillment_status='FULFILLED',
        notes=f"Allocated from Vault Drop #{drop.drop_number}"
    )
    db.session.add(order)
    db.session.flush()

    db.session.add(OrderItem(
        order_id=order.id,
        product_id=prod.id if prod else None,
        product_name=prod.name if prod else f"Ulam Sale (Drop #{drop.drop_number})",
        unit_price=allocated_amount / qty,
        cost_price=max(0.0, parse_float(prod.cost, 0.0)) if prod else 0.0,
        quantity=qty,
        subtotal=allocated_amount
    ))

    if cust:
        cust.accumulated_spend = (cust.accumulated_spend or 0.0) + allocated_amount
        earned = int(allocated_amount // 30)
        if earned > 0:
            cust.points_balance = (cust.points_balance or 0.0) + earned
            db.session.add(RewardLedger(
                customer_id=cust.id,
                points_change=earned,
                reason=f"Allocated Sale from Drop #{drop.drop_number}"
            ))
        apply_member_marketing_rewards(cust, order)

    db.session.commit()
    flash(f"Allocated ₱{allocated_amount:,.2f} from Drop #{drop.drop_number} into a dedicated Order #{order.id}!", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reassign-order/<int:order_id>', methods=['POST'])
@require_admin
def admin_reassign_order(order_id):
    order = Order.query.get_or_404(order_id)
    new_cust_id = request.form.get('customer_id')

    if not new_cust_id:
        flash("Please select a target customer.", "error")
        return redirect(url_for('admin_dashboard'))

    new_cust = Customer.query.get_or_404(int(new_cust_id))
    issue = customer_access_issue(new_cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('admin_dashboard'))
    old_cust = Customer.query.get(order.customer_id) if order.customer_id else None
    redeemed = round(parse_float(order.points_redeemed, 0.0), 2)
    if redeemed > 0 and parse_float(new_cust.points_balance, 0.0) < redeemed - 1e-9:
        flash(f"'{new_cust.name}' needs at least {redeemed:,.2f} points to receive this discounted order.", 'error')
        return redirect(url_for('admin_dashboard'))

    if old_cust and order.status == 'COMPLETED' and order.payment_verified:
        earned = int(order.total_amount // 30)
        old_cust.points_balance = max(0.0, (old_cust.points_balance or 0.0) - earned)
        old_cust.accumulated_spend = max(0.0, (old_cust.accumulated_spend or 0.0) - order.total_amount)
        reverse_member_marketing_rewards_for_order(order)

    if old_cust and redeemed > 0:
        old_cust.points_balance = round(parse_float(old_cust.points_balance, 0.0) + redeemed, 2)
        db.session.add(RewardLedger(
            customer_id=old_cust.id,
            points_change=redeemed,
            reason=f"Order #{order.id} reassigned: redeemed points restored"
        ))

    order.customer_id = new_cust.id
    order.customer_name = new_cust.name
    order.contact_number = new_cust.contact
    if redeemed > 0:
        new_cust.points_balance = round(parse_float(new_cust.points_balance, 0.0) - redeemed, 2)
        db.session.add(RewardLedger(
            customer_id=new_cust.id,
            points_change=-redeemed,
            reason=f"Order #{order.id} reassigned: points discount transferred"
        ))

    if order.status == 'COMPLETED' and order.payment_verified:
        earned_new = int(order.total_amount // 30)
        new_cust.accumulated_spend = (new_cust.accumulated_spend or 0.0) + order.total_amount
        if earned_new > 0:
            new_cust.points_balance = (new_cust.points_balance or 0.0) + earned_new
            db.session.add(RewardLedger(
                customer_id=new_cust.id,
                points_change=earned_new,
                reason=f"Admin Reassigned Order #{order.id}"
            ))
        apply_member_marketing_rewards(new_cust, order)

    db.session.commit()
    flash(f"Transaction #{order.id} reassigned to '{new_cust.name}'.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/update-expense/<int:expense_id>', methods=['POST'])
@require_admin
def admin_update_expense(expense_id):
    exp = Expense.query.get_or_404(expense_id)
    new_title = request.form.get('title', exp.title).strip()
    new_amount = parse_float(request.form.get('amount'), exp.amount or 0.0)
    new_category = request.form.get('category', exp.category).strip()
    if not new_title or new_amount <= 0:
        flash('Expense title and amount must be valid.', 'error')
        return redirect(url_for('admin_dashboard'))
    exp.title = new_title
    exp.amount = new_amount
    exp.category = new_category
    db.session.commit()
    flash(f"Expense #{exp.id} updated successfully.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-expense/<int:expense_id>', methods=['POST'])
@require_admin
def admin_delete_expense(expense_id):
    exp = Expense.query.get_or_404(expense_id)
    db.session.delete(exp)
    db.session.commit()
    flash(f"Expense record removed.", "info")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-product/<int:product_id>', methods=['POST'])
@require_admin
def admin_delete_product(product_id):
    prod = Product.query.get_or_404(product_id)
    prod_name = prod.name

    OrderItem.query.filter_by(product_id=prod.id).update({'product_id': None})
    ProductLike.query.filter_by(product_id=prod.id).delete()
    ProductComment.query.filter_by(product_id=prod.id).delete()

    db.session.delete(prod)
    db.session.commit()
    flash(f"Product '{prod_name}' permanently deleted.", "info")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/revert-order/<int:order_id>', methods=['POST'])
@require_admin
def admin_revert_order(order_id):
    order = Order.query.get_or_404(order_id)
    
    for item in order.items:
        if item.product_id:
            p = Product.query.get(item.product_id)
            if p:
                p.stock += item.quantity

    if order.customer_id and order.status == 'COMPLETED' and order.payment_verified:
        cust = Customer.query.get(order.customer_id)
        if cust:
            earned = int(order.total_amount // 30)
            cust.points_balance = max(0.0, (cust.points_balance or 0.0) - earned)
            cust.accumulated_spend = max(0.0, (cust.accumulated_spend or 0.0) - order.total_amount)
            if earned > 0:
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=-earned,
                    reason=f"Admin Reverted Sale #{order.id}"
                ))
        reverse_member_marketing_rewards_for_order(order)

    if order.customer_id and parse_float(order.points_redeemed, 0.0) > 0:
        cust = Customer.query.get(order.customer_id)
        if cust:
            restored = round(parse_float(order.points_redeemed, 0.0), 2)
            cust.points_balance = round(parse_float(cust.points_balance, 0.0) + restored, 2)
            db.session.add(RewardLedger(
                customer_id=cust.id,
                points_change=restored,
                reason=f"Admin Reverted Sale #{order.id}: redeemed points restored"
            ))

    if order.is_unpaid and order.customer_id:
        cust = Customer.query.get(order.customer_id)
        if cust:
            cust.outstanding_ar = max(0.0, (cust.outstanding_ar or 0.0) - order.total_amount)

    db.session.delete(order)
    db.session.commit()
    flash(f"Order #{order_id} deleted & reverted.", "info")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/purge-dummy-orders/<int:cust_id>', methods=['POST'])
@require_admin
def admin_purge_dummy_orders(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    dummy_orders = Order.query.filter_by(customer_id=cust.id).all()
    for o in dummy_orders:
        db.session.delete(o)
    cust.points_balance = 0.0
    cust.accumulated_spend = 0.0
    cust.outstanding_ar = 0.0
    RewardLedger.query.filter_by(customer_id=cust.id).delete()
    db.session.commit()
    flash(f"Purged test orders & reset points for '{cust.name}'.", "info")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete-customer/<int:cust_id>', methods=['POST'])
@require_admin
def admin_delete_customer(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    Order.query.filter_by(customer_id=cust.id).delete()
    RewardLedger.query.filter_by(customer_id=cust.id).delete()
    ProductLike.query.filter_by(customer_id=cust.id).delete()
    ProductComment.query.filter_by(customer_id=cust.id).delete()
    db.session.delete(cust)
    db.session.commit()
    flash(f"Account '{cust.name}' permanently removed.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/update-logo', methods=['POST'])
@require_admin
def admin_update_logo():
    new_logo = request.form.get('logo_url', '').strip()
    if new_logo:
        setting = StoreSetting.query.filter_by(key='logo_url').first()
        if not setting:
            db.session.add(StoreSetting(key='logo_url', value=new_logo))
        else:
            setting.value = new_logo
        db.session.commit()
        flash("Company logo updated across all portals.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add-product', methods=['POST'])
@require_admin
def admin_add_product():
    name = request.form.get('name', '').strip()
    category_name = request.form.get('category_name', 'Meals').strip() or 'Meals'
    price = parse_float(request.form.get('price'), 0.0)
    cost = parse_float(request.form.get('cost'), 0.0)
    allow_custom_amount = bool(request.form.get('allow_custom_amount'))
    minimum_order_amount = parse_float(request.form.get('minimum_order_amount'), 0.0)
    option_schema_raw = request.form.get('option_schema', '').strip()
    size_schema_raw = request.form.get('size_schema', '').strip()
    stock = parse_int(request.form.get('stock'), 100)
    image_url = request.form.get('image_url', '').strip()
    start_t = request.form.get('available_start_time', '').strip() or None
    end_t = request.form.get('available_end_time', '').strip() or None
    prep_minutes = max(1, min(180, parse_int(request.form.get('prep_minutes'), 10)))
    is_featured = bool(request.form.get('is_featured'))
    is_top_seller = bool(request.form.get('is_top_seller'))

    if not name or price <= 0:
        flash('Please enter a valid product name and selling price.', 'error')
        return redirect(url_for('admin_dashboard') + '#add-dish')
    if cost < 0 or stock < 0:
        flash('Cost and stock cannot be negative.', 'error')
        return redirect(url_for('admin_dashboard') + '#add-dish')
    if allow_custom_amount:
        if minimum_order_amount <= 0:
            minimum_order_amount = price
        if minimum_order_amount > price:
            flash('Minimum order amount cannot be higher than the regular product price.', 'error')
            return redirect(url_for('admin_dashboard') + '#add-dish')
    else:
        minimum_order_amount = None

    try:
        option_schema, size_schema = normalize_product_configuration(option_schema_raw, size_schema_raw)
    except OrderValidationError as exc:
        flash(str(exc), 'error')
        return redirect(url_for('admin_dashboard') + '#add-dish')

    product = Product(
        name=name,
        category_name=category_name,
        price=price,
        cost=cost,
        allow_custom_amount=allow_custom_amount,
        minimum_order_amount=minimum_order_amount,
        option_schema=option_schema,
        size_schema=size_schema,
        stock=stock,
        image_url=image_url,
        available_start_time=start_t,
        available_end_time=end_t,
        prep_minutes=prep_minutes,
        is_featured=is_featured,
        is_top_seller=is_top_seller,
        is_active=True,
    )
    try:
        db.session.add(product)
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Admin could not add Food House product %r', name)
        flash('The product could not be added because the database rejected it. No partial product was saved.', 'error')
        return redirect(url_for('admin_dashboard') + '#add-dish')
    flash(f"Product '{name}' was added and is active. It is highlighted below.", 'success')
    return redirect(url_for('admin_dashboard', added=product.id) + f'#product-row-{product.id}')

@app.route('/admin/batch-update-products', methods=['POST'])
@require_admin
def admin_batch_update_products():
    live_request = request.headers.get('X-Macleens-Live') == '1'
    save_stage = 'reading the submitted catalog data'
    current_product_name = None

    def finish(success, message, status=200, updated_count=0, updated_ids=None, processed_ids=None):
        if live_request:
            return jsonify({
                'success': success,
                'action': 'bulk-products-updated',
                'message': message,
                'updated_count': updated_count,
                'updated_ids': updated_ids or [],
                'processed_ids': processed_ids or [],
                'release': APP_RELEASE,
            }), status
        flash(message, 'success' if success else 'error')
        return redirect(url_for('admin_dashboard') + '#catalog-editor')

    def required_float(field_name, label):
        raw = (request.form.get(field_name) or '').strip()
        try:
            value = float(raw)
        except (TypeError, ValueError):
            raise OrderValidationError(f'{label} must be a valid number.')
        if not math.isfinite(value):
            raise OrderValidationError(f'{label} must be a valid number.')
        return value

    def required_int(field_name, label):
        value = required_float(field_name, label)
        if not value.is_integer():
            raise OrderValidationError(f'{label} must be a whole number.')
        return int(value)

    try:
        save_stage = 'reading the submitted catalog data'
        dirty_tracking = request.form.get('bulk_dirty_tracking') == '1'
        submitted_product_ids = request.form.getlist('product_id')
        if dirty_tracking and not submitted_product_ids:
            return finish(
                True,
                'No new changes were detected. Your catalog is already up to date.',
                processed_ids=[],
            )
        if not submitted_product_ids:
            raise OrderValidationError('No products were received. Reload Admin and try saving again.')

        submitted_id_set = {
            parse_int(value, 0) for value in submitted_product_ids if parse_int(value, 0) > 0
        }
        if dirty_tracking:
            product_ids = [
                value for value in request.form.getlist('changed_product_id')
                if parse_int(value, 0) in submitted_id_set
            ]
        else:
            product_ids = submitted_product_ids

        if dirty_tracking and not product_ids:
            return finish(
                True,
                'No new changes were detected. Your catalog is already up to date.',
                processed_ids=[],
            )

        updated_ids = []
        seen_ids = set()
        for pid_raw in product_ids:
            pid = parse_int(pid_raw, 0)
            if pid <= 0 or pid in seen_ids:
                continue
            seen_ids.add(pid)
            save_stage = 'loading the selected product'
            prod = db.session.get(Product, pid)
            if not prod:
                continue
            current_product_name = prod.name

            save_stage = 'validating the edited fields'
            price = required_float(f'price_{pid}', f'Price for {prod.name}')
            cost = required_float(f'cost_{pid}', f'Cost for {prod.name}')
            stock = required_int(f'stock_{pid}', f'Stock for {prod.name}')
            prep_minutes = required_int(f'prep_minutes_{pid}', f'Preparation time for {prod.name}')
            allow_custom_amount = (f'allow_custom_amount_{pid}' in request.form)
            minimum_raw = (request.form.get(f'minimum_order_amount_{pid}') or '').strip()
            minimum_order_amount = required_float(
                f'minimum_order_amount_{pid}', f'Minimum amount for {prod.name}'
            ) if minimum_raw else None
            option_schema_raw = request.form.get(f'option_schema_{pid}', '').strip()
            size_schema_raw = request.form.get(f'size_schema_{pid}', '').strip()
            image_url = request.form.get(f'image_url_{pid}', '').strip()
            available_start_time = request.form.get(f'available_start_time_{pid}', '').strip() or None
            available_end_time = request.form.get(f'available_end_time_{pid}', '').strip() or None
            is_active = (f'is_active_{pid}' in request.form)
            is_featured = (f'is_featured_{pid}' in request.form)
            is_top_seller = (f'is_top_seller_{pid}' in request.form)

            if price <= 0 or cost < 0 or stock < 0:
                raise OrderValidationError(f'Invalid price, cost, or stock for {prod.name}.')
            if prep_minutes < 1 or prep_minutes > 180:
                raise OrderValidationError(f'Preparation time for {prod.name} must be from 1 to 180 minutes.')
            if allow_custom_amount:
                if minimum_order_amount is None or minimum_order_amount <= 0:
                    minimum_order_amount = price
                if minimum_order_amount > price:
                    raise OrderValidationError(f'Minimum order amount for {prod.name} cannot exceed its regular price.')
            else:
                minimum_order_amount = None

            save_stage = 'checking choices and priced sizes'
            option_schema, size_schema = normalize_product_configuration(
                option_schema_raw,
                size_schema_raw,
            )

            save_stage = 'preparing the database update'
            changed = any((
                abs(price - parse_float(prod.price, 0.0)) > 0.000001,
                abs(cost - parse_float(prod.cost, 0.0)) > 0.000001,
                allow_custom_amount != bool(prod.allow_custom_amount),
                minimum_order_amount != (
                    parse_float(prod.minimum_order_amount, 0.0)
                    if prod.minimum_order_amount is not None else None
                ),
                option_schema != (prod.option_schema or None),
                size_schema != (prod.size_schema or None),
                stock != (prod.stock or 0),
                prep_minutes != (prod.prep_minutes or 10),
                image_url != (prod.image_url or ''),
                available_start_time != (prod.available_start_time or None),
                available_end_time != (prod.available_end_time or None),
                is_active != bool(prod.is_active),
                is_featured != bool(prod.is_featured),
                is_top_seller != bool(prod.is_top_seller),
            ))
            if not changed:
                continue

            prod.price = price
            prod.cost = cost
            prod.allow_custom_amount = allow_custom_amount
            prod.minimum_order_amount = minimum_order_amount
            prod.option_schema = option_schema
            prod.size_schema = size_schema
            prod.stock = stock
            prod.prep_minutes = prep_minutes
            prod.image_url = image_url
            prod.available_start_time = available_start_time
            prod.available_end_time = available_end_time
            prod.is_active = is_active
            prod.is_featured = is_featured
            prod.is_top_seller = is_top_seller
            updated_ids.append(pid)
        save_stage = 'committing the database update'
        db.session.commit()
        count = len(updated_ids)
        message = (
            f'Saved changes to {count} product{"s" if count != 1 else ""}.'
            if count else 'No new changes were detected. Your catalog is already up to date.'
        )
        return finish(
            True,
            message,
            updated_count=count,
            updated_ids=updated_ids,
            processed_ids=list(seen_ids),
        )
    except OrderValidationError as exc:
        db.session.rollback()
        return finish(False, str(exc), status=400)
    except RequestEntityTooLarge:
        db.session.rollback()
        app.logger.warning('Bulk product update exceeded the request field limit')
        return finish(
            False,
            'The browser sent too much catalog data at once. Hard-refresh Admin to load the compact bulk editor, then change and save the product again.',
            status=413,
        )
    except Exception:
        db.session.rollback()
        error_reference = secrets.token_hex(4).upper()
        app.logger.exception(
            'Bulk product update failed [reference=%s, stage=%s, product=%r]',
            error_reference,
            save_stage,
            current_product_name,
        )
        target = f' for {current_product_name}' if current_product_name else ''
        return finish(
            False,
            f'The catalog could not be saved{target} while {save_stage}. Error reference: {error_reference}. No changes were saved.',
            status=500,
        )

@app.route('/admin/toggle-credit/<int:cust_id>', methods=['POST'])
@require_admin
def admin_toggle_credit(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    action = request.form.get('action', 'SET_LIMIT').upper()
    if action == 'TOGGLE':
        cust.is_credit_eligible = not cust.is_credit_eligible
    limit = request.form.get('credit_limit')
    if limit not in (None, ''):
        parsed_limit = parse_float(limit, -1)
        if parsed_limit < 0:
            flash('Credit limit cannot be negative.', 'error')
            return redirect(url_for('admin_dashboard'))
        cust.credit_limit = parsed_limit
    db.session.commit()
    state = 'enabled' if cust.is_credit_eligible else 'disabled'
    flash(f'Credit for {cust.name} is {state}; limit ₱{(cust.credit_limit or 0.0):,.2f}.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delivery-zones', methods=['POST'])
@require_admin
def admin_manage_delivery_zones():
    action = request.form.get('action')
    if action == 'ADD':
        place = request.form.get('place_name')
        brgy = request.form.get('barangay')
        rate = float(request.form.get('rate') or 40.0)
        dist = request.form.get('distance')
        note = request.form.get('note')
        db.session.add(DeliveryZone(place_name=place, barangay=brgy, rate=rate, distance=dist, note=note))
    elif action == 'DELETE':
        zid = request.form.get('zone_id')
        zone = DeliveryZone.query.get(zid)
        if zone:
            db.session.delete(zone)
    db.session.commit()
    flash("Delivery zones updated.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/renew-customer-card/<int:cust_id>', methods=['POST'])
@require_admin
def admin_renew_customer_card(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    cust.card_status = 'ACTIVE'
    cust.card_expires_at = ph_today() + timedelta(days=365)
    db.session.commit()
    flash(f"Rewards card for '{cust.name}' renewed for 1 year.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reset-customer-pin/<int:cust_id>', methods=['POST'])
@require_admin
def admin_reset_customer_pin(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    new_pin = request.form.get('new_pin', '').strip()
    if not is_valid_customer_pin(new_pin):
        flash('Customer PIN must be exactly 4 digits.', 'error')
    else:
        cust.pin_hash = generate_password_hash(new_pin)
        db.session.commit()
        flash(f"PIN for customer '{cust.name}' was reset successfully.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reset-staff-pin/<int:staff_id>', methods=['POST'])
@require_admin
def admin_reset_staff_pin(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    new_pin = request.form.get('new_pin', '').strip()
    if not new_pin.isdigit() or not (4 <= len(new_pin) <= 8):
        flash('Staff PIN must contain 4 to 8 digits.', 'error')
        return redirect(url_for('admin_dashboard'))
    staff.pin_hash = generate_password_hash(new_pin)
    db.session.commit()
    flash(f"PIN for staff account '{staff.username}' was changed.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/update-promo-financials/<int:promo_id>', methods=['POST'])
@require_admin
def admin_update_promo_financials(promo_id):
    promo = PromotionTracker.query.get_or_404(promo_id)
    price = parse_float(request.form.get('promo_price'), promo.promo_price)
    cost = parse_float(request.form.get('promo_cost'), promo.promo_cost or 0.0)
    if price <= 0 or cost < 0:
        flash('Promo price must be positive and promo cost cannot be negative.', 'error')
        return redirect(url_for('admin_dashboard'))
    promo.promo_price = price
    promo.promo_cost = cost
    db.session.commit()
    flash(f"Financials for '{promo.title}' updated.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/toggle-promo/<int:promo_id>', methods=['POST'])
@require_admin
def admin_toggle_promo(promo_id):
    promo = PromotionTracker.query.get_or_404(promo_id)
    promo.is_active = not promo.is_active
    if promo.is_active:
        promo.created_at = utc_now()
    db.session.commit()
    flash(f"Campaign '{promo.title}' status updated.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/toggle-promo-visibility/<int:promo_id>', methods=['POST'])
@require_admin
def admin_toggle_promo_visibility(promo_id):
    promo = PromotionTracker.query.get_or_404(promo_id)
    promo.is_visible = not getattr(promo, 'is_visible', True)
    db.session.commit()
    flash(f"Promo '{promo.title}' visibility updated.", "info")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/marketing/bonus-campaign/create', methods=['POST'])
@require_admin
def admin_create_bonus_campaign():
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    bonus_points = parse_float(request.form.get('bonus_points'), 0.0)
    points_multiplier = parse_float(request.form.get('points_multiplier'), 1.0)
    min_spend = parse_float(request.form.get('min_spend'), 0.0)
    start_date_raw = request.form.get('start_date', '').strip()
    end_date_raw = request.form.get('end_date', '').strip()
    start_time = request.form.get('start_time', '').strip() or None
    end_time = request.form.get('end_time', '').strip() or None
    weekdays = ','.join(request.form.getlist('weekdays')) or None

    if not title or (bonus_points <= 0 and points_multiplier <= 1.0) or min_spend < 0 or points_multiplier < 1.0:
        flash('Campaign needs a title and either fixed bonus points or a points multiplier above 1×.', 'error')
        return redirect(url_for('admin_dashboard'))
    try:
        start_date = datetime.strptime(start_date_raw, '%Y-%m-%d').date() if start_date_raw else None
        end_date = datetime.strptime(end_date_raw, '%Y-%m-%d').date() if end_date_raw else None
        if start_date and end_date and end_date < start_date:
            raise ValueError('End date must not be before start date.')
        if start_time:
            datetime.strptime(start_time, '%H:%M')
        if end_time:
            datetime.strptime(end_time, '%H:%M')
    except ValueError as exc:
        flash(f'Invalid campaign schedule: {exc}', 'error')
        return redirect(url_for('admin_dashboard'))

    db.session.add(BonusCampaign(
        title=title,
        description=description,
        bonus_points=bonus_points,
        points_multiplier=points_multiplier,
        min_spend=min_spend,
        start_date=start_date,
        end_date=end_date,
        start_time=start_time,
        end_time=end_time,
        weekdays=weekdays,
        is_active=True,
    ))
    db.session.commit()
    flash(f"Bonus campaign '{title}' created.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/marketing/bonus-campaign/<int:campaign_id>/toggle', methods=['POST'])
@require_admin
def admin_toggle_bonus_campaign(campaign_id):
    campaign = BonusCampaign.query.get_or_404(campaign_id)
    campaign.is_active = not campaign.is_active
    db.session.commit()
    flash(f"Bonus campaign '{campaign.title}' {'activated' if campaign.is_active else 'paused'}.", 'info')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/marketing/bonus-campaign/<int:campaign_id>/delete', methods=['POST'])
@require_admin
def admin_delete_bonus_campaign(campaign_id):
    campaign = BonusCampaign.query.get_or_404(campaign_id)
    BonusCampaignClaim.query.filter_by(campaign_id=campaign.id).delete(synchronize_session=False)
    db.session.delete(campaign)
    db.session.commit()
    flash(f"Bonus campaign '{campaign.title}' deleted.", 'info')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/marketing/promo/create', methods=['POST'])
@require_admin
def admin_create_member_promo():
    promo_code = re.sub(r'[^A-Z0-9_]+', '_', request.form.get('promo_code', '').strip().upper()).strip('_')
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    price = parse_float(request.form.get('promo_price'), 0.0)
    cost = parse_float(request.form.get('promo_cost'), 0.0)
    portal_only = request.form.get('portal_only') == 'on'
    if not promo_code or not title or price <= 0 or cost < 0:
        flash('Promo code, title, positive price, and valid cost are required.', 'error')
        return redirect(url_for('admin_dashboard'))
    if PromotionTracker.query.filter_by(promo_code=promo_code).first():
        flash('That promo code already exists.', 'error')
        return redirect(url_for('admin_dashboard'))
    db.session.add(PromotionTracker(
        promo_code=promo_code,
        title=title,
        description=description,
        promo_price=price,
        promo_cost=cost,
        is_active=True,
        is_visible=True,
        portal_only=portal_only,
    ))
    db.session.commit()
    flash(f"Promo '{title}' created and is live for its 3-day cycle.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/loyalty-cards')
@require_admin
def admin_loyalty_cards():
    customers = Customer.query.order_by(Customer.name.asc()).all()
    selected_id = parse_int(request.args.get('customer_id'), 0)
    selected = Customer.query.get(selected_id) if selected_id else (customers[0] if customers else None)
    theme = (request.args.get('theme') or (selected.card_theme if selected else None) or 'pink-classic').strip()
    if theme not in LOYALTY_CARD_THEMES:
        theme = 'pink-classic'
    qr_data = None
    qr_target = None
    if selected:
        base = _marketing_public_base_url()
        card_identifier = selected.card_number or selected.contact
        qr_target = f"{base}/portal/login?card={card_identifier}"
        qr_data = qr_svg_data_url(qr_target)
    return render_template(
        'loyalty_card_portal.html',
        customers=customers, selected=selected, theme=theme, themes=LOYALTY_CARD_THEMES,
        qr_data=qr_data, qr_target=qr_target,
    )

@app.route('/admin/loyalty-cards/<int:cust_id>/save', methods=['POST'])
@require_admin
def admin_save_loyalty_card(cust_id):
    cust = Customer.query.get_or_404(cust_id)
    theme = request.form.get('theme', cust.card_theme or 'pink-classic').strip()
    if theme not in LOYALTY_CARD_THEMES:
        theme = 'pink-classic'
    cust.card_theme = theme
    cust.card_logo_scale = loyalty_card_scale(request.form.get('card_logo_scale'), cust.card_logo_scale or 1.0)
    cust.card_photo_scale = loyalty_card_scale(request.form.get('card_photo_scale'), cust.card_photo_scale or 1.0)
    cust.card_qr_scale = loyalty_card_scale(request.form.get('card_qr_scale'), cust.card_qr_scale or 1.0)
    cust.card_text_scale = loyalty_card_scale(request.form.get('card_text_scale'), cust.card_text_scale or 1.0)
    cust.card_info_scale = loyalty_card_scale(request.form.get('card_info_scale'), cust.card_info_scale or 1.0)
    try:
        new_image = customer_profile_image_from_request(file_key='profile_photo', url_key='profile_image', existing=cust.profile_image)
        if new_image:
            cust.profile_image = new_image
        db.session.commit()
        flash(f'Card design saved for {cust.name}.', 'success')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    return redirect(url_for('admin_loyalty_cards', customer_id=cust.id, theme=theme))

@app.route('/admin/marketing/qr-kit')
@require_admin
def admin_qr_kit():
    qr_sources = [
        ('counter', 'Counter Registration', 'Scan to join rewards before or after a walk-in purchase.'),
        ('table', 'Table Product Suggestion', 'Scan while waiting and tell us what you would like Macleen\'s to offer.'),
        ('receipt', 'Receipt Rewards Check', 'Scan after purchase to check points and member benefits.'),
        ('facebook', 'Facebook / Social', 'Use this QR on printed social promotions and posters.'),
    ]
    return render_template('qr_kit.html', qr_sources=qr_sources)

@app.route('/marketing/qr/<string:source>.svg')
def marketing_qr_svg(source):
    allowed = {'counter', 'table', 'receipt', 'facebook'}
    if source not in allowed:
        return Response('Unknown QR source', status=404, mimetype='text/plain')
    target = url_for('portal_start', source=source, _external=True)
    image = qrcode.make(target, image_factory=qrcode.image.svg.SvgPathImage)
    buffer = io.BytesIO()
    image.save(buffer)
    return Response(buffer.getvalue(), mimetype='image/svg+xml', headers={'Cache-Control': 'no-store'})

# ==================== MACLEEN'S COMMUNITY ====================

def community_api_actor():
    cust = get_current_community_customer()
    if not cust:
        return None, None, (jsonify({'success': False, 'message': 'Please log in to your loyalty account again.'}), 401)
    profile = CommunityProfile.query.filter_by(customer_id=cust.id).first()
    if not profile:
        return cust, None, (jsonify({'success': False, 'message': 'Create your optional community profile first.'}), 403)
    return cust, profile, None

def community_profile_values(form, existing=None):
    handle = normalize_community_handle(form.get('handle'), allow_main_admin=bool(existing and existing.is_community_admin))
    if existing and existing.is_community_admin and handle != COMMUNITY_MAIN_ADMIN_HANDLE:
        raise OrderValidationError(f'The main community administrator must keep the reserved @{COMMUNITY_MAIN_ADMIN_HANDLE} handle.')
    duplicate = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == handle.lower()).first()
    if duplicate and (not existing or duplicate.id != existing.id):
        raise OrderValidationError('That community handle is already taken.')
    # A saved primary role is security-sensitive and cannot be switched from
    # profile settings. Residents use the reviewed student-application route.
    role = existing.role if existing else (form.get('role') or '').strip().upper()
    if role not in COMMUNITY_ROLES:
        raise OrderValidationError('Choose College Student or Binalbagan Resident.')
    current_year = ph_today().year
    values = {
        'handle': handle,
        'role': role,
        'campus_name': None,
        'department': None,
        'graduating_year': None,
        'vibe_status': None,
        'barangay': None,
        'resident_since_year': None,
        'verification_method': (existing.verification_method if existing else ('IN_PERSON_PENDING' if role == 'STUDENT' else 'SELF_DECLARED')),
        'public_bio': re.sub(r'\s+', ' ', (form.get('public_bio') or '').strip())[:160] or None,
        'is_profile_locked': str(form.get('is_profile_locked') or '').strip().lower() in {'1', 'true', 'on', 'yes'},
    }
    if role == 'STUDENT':
        campus = (form.get('campus_name') or '').strip()
        department = (form.get('department') or '').strip()
        graduating_year = parse_int(form.get('graduating_year'), 0)
        if campus not in COMMUNITY_CAMPUSES:
            raise OrderValidationError('Choose a campus from the provided list.')
        if department not in COMMUNITY_DEPARTMENTS:
            raise OrderValidationError('Choose your department or select Other.')
        if graduating_year < current_year or graduating_year > current_year + 10:
            raise OrderValidationError(f'Graduating year must be from {current_year} to {current_year + 10}.')
        values.update(
            campus_name=campus,
            department=department,
            graduating_year=graduating_year,
            vibe_status=(existing.vibe_status if existing and existing.role == 'STUDENT' and existing.vibe_status else 'Quiet study mode'),
        )
    else:
        barangay = (form.get('barangay') or '').strip()
        resident_since = parse_int(form.get('resident_since_year'), 0)
        if barangay not in BINALBAGAN_BARANGAYS:
            raise OrderValidationError('Choose one of Binalbagan’s 16 official barangays.')
        if resident_since < 1900 or resident_since > current_year:
            raise OrderValidationError(f'Resident-since year must be from 1900 to {current_year}.')
        values.update(barangay=barangay, resident_since_year=resident_since)
    return values

@app.route('/community')
def community_home():
    cust = get_current_community_customer()
    if not cust:
        return redirect(url_for('customer_login', next='community'))
    profile = CommunityProfile.query.filter_by(customer_id=cust.id).first()
    if not profile:
        if not COMMUNITY_REGISTRATION_OPEN:
            return render_template('community_setup.html', cust=cust, registration_closed=True, student_preapproved=False, campuses=COMMUNITY_CAMPUSES, departments=COMMUNITY_DEPARTMENTS, barangays=BINALBAGAN_BARANGAYS, today=ph_today()), 503
        return render_template(
            'community_setup.html',
            cust=cust,
            student_preapproved=bool(cust.community_student_preapproved),
            campuses=COMMUNITY_CAMPUSES,
            departments=COMMUNITY_DEPARTMENTS,
            barangays=BINALBAGAN_BARANGAYS,
            today=ph_today(),
        )

    checkin, new_drop = None, None
    now = utc_now()
    own_channel = community_channel_for_role(profile.role)
    visible_channels = list(COMMUNITY_CHANNELS) if profile.is_community_admin else [own_channel]
    posts = CommunityPost.query.filter(
        CommunityPost.status == 'PUBLISHED',
        CommunityPost.published_at.isnot(None),
        CommunityPost.published_at <= now,
        CommunityPost.channel.in_(visible_channels + ['GLOBAL']),
        or_(CommunityPost.expires_at.is_(None), CommunityPost.expires_at > now),
    ).order_by(CommunityPost.created_at.desc()).limit(50).all()
    channel_posts = {
        channel: [post for post in posts if post.channel in {'GLOBAL', channel}][:25]
        for channel in visible_channels
    }
    post_comments = {}
    reaction_map = {}
    vote_map = {}
    post_stats = {}
    poll_results = {}
    for post in posts:
        comments = CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED').order_by(CommunityComment.created_at.desc()).limit(3).all()
        post_comments[post.id] = list(reversed(comments))
        reaction = CommunityReaction.query.filter_by(post_id=post.id, customer_id=cust.id).first()
        reaction_map[post.id] = reaction.reaction_type if reaction else None
        vote = CommunityPollVote.query.filter_by(post_id=post.id, customer_id=cust.id).first()
        vote_map[post.id] = vote.option_id if vote else None
        post_stats[post.id] = {
            'reactions': CommunityReaction.query.filter_by(post_id=post.id).count(),
            'comments': CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED').count(),
            'reshares': CommunityPost.query.filter_by(
                reshared_post_id=(post.reshared_post_id or post.id), status='PUBLISHED',
            ).count(),
        }
        if post.post_type == 'POLL':
            poll_results[post.id] = community_poll_results(post)

    campus_leaders, barangay_leaders = [], []
    incoming_gifts, outgoing_gifts, drops, gift_products = [], [], [], []
    lifetime_punches = max(0, int(parse_float(cust.accumulated_spend, 0.0) // 30))
    friend_profile_ids = community_friend_profile_ids(profile.id)
    friends = CommunityProfile.query.filter(CommunityProfile.id.in_(friend_profile_ids)).order_by(CommunityProfile.handle.asc()).all() if friend_profile_ids else []
    incoming_connection_rows = CommunityConnection.query.filter(
        CommunityConnection.status == 'PENDING',
        or_(CommunityConnection.profile_a_id == profile.id, CommunityConnection.profile_b_id == profile.id),
        CommunityConnection.requested_by_profile_id != profile.id,
    ).order_by(CommunityConnection.created_at.desc()).all()
    requester_ids = {row.requested_by_profile_id for row in incoming_connection_rows}
    requester_profiles = {row.id: row for row in CommunityProfile.query.filter(CommunityProfile.id.in_(requester_ids)).all()} if requester_ids else {}
    main_admin_profile = CommunityProfile.query.filter_by(is_community_admin=True).first()
    follows_main_admin = bool(main_admin_profile and CommunityFollow.query.filter_by(
        follower_profile_id=profile.id, followed_profile_id=main_admin_profile.id,
    ).first())
    main_admin_followers = CommunityFollow.query.filter_by(followed_profile_id=main_admin_profile.id).count() if main_admin_profile else 0
    followed_ids = {
        row.followed_profile_id
        for row in CommunityFollow.query.filter_by(follower_profile_id=profile.id).all()
    }
    following_profiles = CommunityProfile.query.filter(
        CommunityProfile.id.in_(followed_ids)
    ).order_by(CommunityProfile.handle.asc()).all() if followed_ids else []
    suggested_profiles = community_suggested_profiles(profile, followed_ids=followed_ids)
    suggested_ids = [row.id for row in suggested_profiles]
    suggested_follower_counts = dict(
        db.session.query(CommunityFollow.followed_profile_id, db.func.count(CommunityFollow.id))
        .filter(CommunityFollow.followed_profile_id.in_(suggested_ids))
        .group_by(CommunityFollow.followed_profile_id).all()
    ) if suggested_ids else {}
    notifications = CommunityNotification.query.filter_by(recipient_profile_id=profile.id).order_by(CommunityNotification.created_at.desc()).limit(12).all()
    group_memberships = CommunityGroupMember.query.filter_by(
        profile_id=profile.id, status='ACTIVE',
    ).order_by(CommunityGroupMember.created_at.desc()).all()
    community_groups = [row for row in group_memberships if row.group and row.group.is_active]
    group_unread_counts = {}
    group_invites = CommunityGroupMember.query.filter_by(
        profile_id=profile.id, status='INVITED',
    ).order_by(CommunityGroupMember.created_at.desc()).all()
    mentionable_query = CommunityProfile.query.filter(
        CommunityProfile.id != profile.id,
        CommunityProfile.verification_status != 'REJECTED',
        or_(CommunityProfile.role != 'STUDENT', CommunityProfile.verification_status == 'VERIFIED', CommunityProfile.is_community_admin.is_(True)),
    )
    if not profile.is_community_admin:
        mentionable_query = mentionable_query.filter(or_(
            CommunityProfile.role == profile.role,
            CommunityProfile.is_community_admin.is_(True),
        ))
    mentionable_profiles = mentionable_query.order_by(CommunityProfile.handle.asc()).limit(300).all()
    return render_template(
        'community.html',
        cust=cust,
        profile=profile,
        own_channel=own_channel,
        visible_channels=visible_channels,
        channel_posts=channel_posts,
        post_comments=post_comments,
        reaction_map=reaction_map,
        vote_map=vote_map,
        post_stats=post_stats,
        poll_results=poll_results,
        modules=COMMUNITY_CHANNEL_MODULES,
        module_labels={value: label for rows in COMMUNITY_CHANNEL_MODULES.values() for value, label in rows} | {'FLASH_POLL': '8:00 AM Flash Poll'},
        vibes=COMMUNITY_VIBES,
        campuses=COMMUNITY_CAMPUSES,
        departments=COMMUNITY_DEPARTMENTS,
        barangays=BINALBAGAN_BARANGAYS,
        today=ph_today(),
        alerts=active_community_alerts(profile.role, include_all_roles=bool(profile.is_community_admin)),
        ads_by_channel={
            'CAMPUS': active_community_ads('STUDENT' if profile.is_community_admin else profile.role, 'CAMPUS'),
            'TOWN': active_community_ads('RESIDENT' if profile.is_community_admin else profile.role, 'TOWN'),
        },
        campus_leaders=campus_leaders,
        barangay_leaders=barangay_leaders,
        incoming_gifts=incoming_gifts,
        outgoing_gifts=outgoing_gifts,
        drops=drops,
        gift_products=gift_products,
        gift_daily_cap=COMMUNITY_GIFT_DAILY_CAP,
        privacy_minimum=COMMUNITY_GROUP_PRIVACY_MINIMUM,
        vapid_public_key=(os.environ.get('WEBPUSH_VAPID_PUBLIC_KEY') or '').strip(),
        checkin=checkin,
        new_drop=new_drop,
        lifetime_punches=lifetime_punches,
        punch_progress=lifetime_punches % 10,
        friend_profile_ids=friend_profile_ids,
        friends=friends,
        incoming_connections=incoming_connection_rows,
        requester_profiles=requester_profiles,
        can_interact=community_can_interact(profile, own_channel),
        main_admin_profile=main_admin_profile,
        follows_main_admin=follows_main_admin,
        main_admin_followers=main_admin_followers,
        followed_ids=followed_ids,
        following_profiles=following_profiles,
        suggested_profiles=suggested_profiles,
        suggested_follower_counts=suggested_follower_counts,
        notifications=notifications,
        community_groups=community_groups,
        group_unread_counts=group_unread_counts,
        group_invites=group_invites,
        community_group_max_members=COMMUNITY_GROUP_MAX_MEMBERS,
        community_max_owned_groups=COMMUNITY_MAX_OWNED_GROUPS,
        community_posting_open=COMMUNITY_POSTING_OPEN,
        community_group_workspaces_open=COMMUNITY_GROUP_WORKSPACES_OPEN,
        community_trusted_post_threshold=COMMUNITY_TRUSTED_POST_THRESHOLD,
        mentionable_profiles=mentionable_profiles,
    )

@app.route('/community/member/<string:handle>')
def community_member_profile(handle):
    cust = get_current_community_customer()
    if not cust:
        return redirect(url_for('customer_login', next='community'))
    viewer = CommunityProfile.query.filter_by(customer_id=cust.id).first()
    if not viewer:
        return redirect(url_for('community_home'))
    requested_handle = handle.strip().lower().lstrip('@')
    target = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == requested_handle).first()
    if not target or not community_can_view_profile(viewer, target):
        return render_template(
            'community_profile_unavailable.html',
            viewer_profile=viewer,
            requested_handle=requested_handle,
        ), 404

    friend_ids = community_friend_profile_ids(viewer.id)
    is_connected = target.id in friend_ids
    can_view_details = bool(
        viewer.id == target.id
        or viewer.is_community_admin
        or not target.is_profile_locked
        or is_connected
    )
    followed = bool(CommunityFollow.query.filter_by(
        follower_profile_id=viewer.id, followed_profile_id=target.id,
    ).first()) if viewer.id != target.id else False
    left, right = community_connection_pair(viewer.id, target.id)
    connection = CommunityConnection.query.filter_by(profile_a_id=left, profile_b_id=right).first() if viewer.id != target.id else None
    follower_count = CommunityFollow.query.filter_by(followed_profile_id=target.id).count()
    following_count = CommunityFollow.query.filter_by(follower_profile_id=target.id).count()

    posts = []
    post_comments = {}
    reaction_map = {}
    post_stats = {}
    if can_view_details:
        now = utc_now()
        visible_channels = list(COMMUNITY_CHANNELS) if viewer.is_community_admin else [community_channel_for_role(viewer.role)]
        posts = CommunityPost.query.filter(
            CommunityPost.author_profile_id == target.id,
            CommunityPost.status == 'PUBLISHED',
            CommunityPost.published_at.isnot(None),
            CommunityPost.published_at <= now,
            CommunityPost.channel.in_(visible_channels + ['GLOBAL']),
            or_(CommunityPost.expires_at.is_(None), CommunityPost.expires_at > now),
        ).order_by(CommunityPost.created_at.desc()).limit(40).all()
        for post in posts:
            post_comments[post.id] = list(reversed(
                CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED')
                .order_by(CommunityComment.created_at.desc()).limit(3).all()
            ))
            reaction = CommunityReaction.query.filter_by(post_id=post.id, customer_id=cust.id).first()
            reaction_map[post.id] = reaction.reaction_type if reaction else None
            post_stats[post.id] = {
                'reactions': CommunityReaction.query.filter_by(post_id=post.id).count(),
                'comments': CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED').count(),
                'reshares': CommunityPost.query.filter_by(
                    reshared_post_id=(post.reshared_post_id or post.id), status='PUBLISHED',
                ).count(),
            }

    return render_template(
        'community_profile.html',
        viewer=viewer,
        profile=viewer,
        member=target,
        can_view_details=can_view_details,
        is_connected=is_connected,
        connection=connection,
        is_following=followed,
        follower_count=follower_count,
        following_count=following_count,
        posts=posts,
        post_comments=post_comments,
        reaction_map=reaction_map,
        post_stats=post_stats,
        poll_results={},
        vote_map={},
        module_labels={value: label for rows in COMMUNITY_CHANNEL_MODULES.values() for value, label in rows} | {'FLASH_POLL': '8:00 AM Flash Poll'},
        can_interact=community_can_interact(viewer, community_channel_for_role(viewer.role)),
    )

@app.route('/community/profile', methods=['POST'])
def community_save_profile():
    cust = get_current_community_customer()
    if not cust:
        if request.headers.get('X-Macleens-Community') == '1':
            return jsonify({'success': False, 'message': 'Please log in again.'}), 401
        return redirect(url_for('customer_login', next='community'))
    profile = CommunityProfile.query.filter_by(customer_id=cust.id).first()
    is_new = profile is None
    try:
        if is_new and not COMMUNITY_REGISTRATION_OPEN:
            raise OrderValidationError('Community registration is temporarily paused by the administrator.')
        if is_new and not request.form.get('privacy_consent'):
            raise OrderValidationError('You must accept the community privacy notice and rules before joining.')
        values = community_profile_values(request.form, existing=profile)
        if not profile:
            profile = CommunityProfile(customer_id=cust.id, **values)
            if profile.role == 'STUDENT':
                if cust.community_student_preapproved:
                    profile.verification_status = 'VERIFIED'
                    profile.verification_method = 'ADMIN_TAG'
                    profile.verification_note = 'Pre-approved by authorized staff before community registration.'
                else:
                    profile.verification_status = 'PENDING'
                    profile.verification_method = 'IN_PERSON_PENDING'
                    profile.verification_note = 'Visit Macleen’s with a current student ID for private visual verification; no ID image is retained.'
            else:
                profile.verification_status = 'SELF_DECLARED'
            db.session.add(profile)
            db.session.flush()
            community_add_admin_notice(
                'NEW_MEMBER', f'profile:{profile.id}',
                f'@{profile.handle} joined as a {profile.role.lower()}. Review their role and verification status.',
                profile=profile,
            )
        else:
            for key, value in values.items():
                setattr(profile, key, value)
        db.session.commit()
        if request.headers.get('X-Macleens-Community') == '1':
            return jsonify({
                'success': True,
                'message': 'Community profile saved.',
                'handle': profile.handle,
                'role': profile.role,
                'profile_locked': bool(profile.is_profile_locked),
            })
        flash('Your optional community profile is ready. Welcome to Macleen’s Community!', 'success')
        return redirect(url_for('community_home'))
    except OrderValidationError as exc:
        db.session.rollback()
        if request.headers.get('X-Macleens-Community') == '1':
            return jsonify({'success': False, 'message': str(exc)}), 400
        flash(str(exc), 'error')
        return render_template(
            'community_setup.html',
            cust=cust,
            student_preapproved=bool(cust.community_student_preapproved),
            campuses=COMMUNITY_CAMPUSES,
            departments=COMMUNITY_DEPARTMENTS,
            barangays=BINALBAGAN_BARANGAYS,
            today=ph_today(),
        ), 400

@app.route('/community/api/student-application', methods=['POST'])
def community_student_application():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if profile.role != 'RESIDENT' or profile.is_community_admin:
        return jsonify({'success': False, 'message': 'Only a resident profile can submit this student application.'}), 403
    try:
        campus = (request.form.get('campus_name') or '').strip()
        department = (request.form.get('department') or '').strip()
        graduating_year = parse_int(request.form.get('graduating_year'), 0)
        current_year = ph_today().year
        if campus not in COMMUNITY_CAMPUSES:
            raise OrderValidationError('Choose a campus from the provided list.')
        if department not in COMMUNITY_DEPARTMENTS:
            raise OrderValidationError('Choose your department or select Other.')
        if graduating_year < current_year or graduating_year > current_year + 10:
            raise OrderValidationError(f'Graduating year must be from {current_year} to {current_year + 10}.')
        profile.student_application_campus = campus
        profile.student_application_department = department
        profile.student_application_graduating_year = graduating_year
        profile.student_application_status = 'PENDING'
        if cust.community_student_preapproved:
            profile.role = 'STUDENT'
            profile.campus_name = campus
            profile.department = department
            profile.graduating_year = graduating_year
            profile.vibe_status = 'Quiet study mode'
            profile.verification_status = 'VERIFIED'
            profile.verification_method = 'ADMIN_TAG'
            profile.student_application_status = 'APPROVED'
            profile.first_post_approved = False
            message = 'Student role approved from the staff pre-verification tag. Open Community again to enter Campus Hub.'
        else:
            profile.verification_method = 'IN_PERSON_PENDING'
            profile.verification_note = 'Awaiting private in-person student-ID check; no ID image is stored.'
            message = 'Student application sent. Bring your current student ID to Macleen’s for a private visual check; no ID photo is stored. You remain in Town Square until approval.'
        community_add_admin_notice(
            'STUDENT_APPLICATION', f'profile:{profile.id}:student-application:{utc_now().isoformat()}',
            f'@{profile.handle} submitted a student-access application for private staff review.',
            profile=profile,
        )
        db.session.commit()
        return jsonify({'success': True, 'message': message, 'status': profile.student_application_status})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/student-id-resubmit', methods=['POST'])
def community_student_id_resubmit():
    cust, profile, error = community_api_actor()
    if error:
        return error
    return jsonify({
        'success': False,
        'message': 'ID-image uploads are retired for privacy. Bring your current student ID to Macleen’s for a private visual check; no copy is retained.',
    }), 410

@app.route('/community/api/vibe', methods=['POST'])
def community_update_vibe():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if profile.role != 'STUDENT' or not community_can_interact(profile, 'CAMPUS'):
        return jsonify({'success': False, 'message': 'Vibe status is available after student verification.'}), 403
    data = request.get_json(silent=True) or request.form
    vibe = (data.get('vibe') or '').strip()
    if vibe not in COMMUNITY_VIBES:
        return jsonify({'success': False, 'message': 'Choose one of the available vibe statuses.'}), 400
    profile.vibe_status = vibe
    db.session.commit()
    return jsonify({'success': True, 'message': f'Vibe updated: {vibe}', 'vibe': vibe})

@app.route('/community/api/connections', methods=['POST'])
def community_connections():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if profile.role == 'STUDENT' and not community_can_interact(profile, 'CAMPUS'):
        return jsonify({'success': False, 'message': 'Student verification is required before connecting.'}), 403
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or 'REQUEST').strip().upper()
    try:
        if action == 'REQUEST':
            community_rate_limit(cust.id, CommunityConnection, 10080, 20, author_field='requested_by_profile_id', profile_id=profile.id)
            handle = normalize_community_handle(data.get('handle'))
            target = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == handle.lower()).first()
            if not target:
                raise OrderValidationError('No community member uses that handle.')
            if target.role == 'STUDENT' and target.verification_status != 'VERIFIED' and not target.is_community_admin:
                raise OrderValidationError('That student profile is still awaiting verification.')
            if target.id == profile.id:
                raise OrderValidationError('You cannot connect with your own profile.')
            same_feed = community_channel_for_role(target.role) == community_channel_for_role(profile.role)
            if not profile.is_community_admin and not target.is_community_admin and not same_feed:
                raise OrderValidationError('Connections are limited to members in your role-locked community.')
            left, right = community_connection_pair(profile.id, target.id)
            connection = CommunityConnection.query.filter_by(profile_a_id=left, profile_b_id=right).first()
            if connection and connection.status == 'ACCEPTED':
                raise OrderValidationError(f'@{target.handle} is already your connection.')
            if connection and connection.status == 'PENDING':
                if connection.requested_by_profile_id == target.id:
                    connection.status = 'ACCEPTED'
                    connection.responded_at = utc_now()
                    message = f'You and @{target.handle} are now connected.'
                else:
                    raise OrderValidationError(f'A request to @{target.handle} is already pending.')
            else:
                if connection:
                    connection.requested_by_profile_id = profile.id
                    connection.status = 'PENDING'
                    connection.created_at = utc_now()
                    connection.responded_at = None
                else:
                    db.session.add(CommunityConnection(
                        profile_a_id=left, profile_b_id=right,
                        requested_by_profile_id=profile.id, status='PENDING',
                    ))
                message = f'Connection request sent to @{target.handle}.'
        else:
            connection = db.session.get(CommunityConnection, parse_int(data.get('connection_id'), 0))
            if not connection or profile.id not in {connection.profile_a_id, connection.profile_b_id}:
                raise OrderValidationError('That connection request is unavailable.')
            if action == 'ACCEPT':
                if connection.status != 'PENDING' or connection.requested_by_profile_id == profile.id:
                    raise OrderValidationError('Only the receiving member can accept this request.')
                connection.status = 'ACCEPTED'
                connection.responded_at = utc_now()
                message = 'Connection accepted. Vibe status is now shared between you.'
            elif action in {'DECLINE', 'REMOVE'}:
                db.session.delete(connection)
                message = 'Connection removed.'
            else:
                raise OrderValidationError('Choose a valid connection action.')
        db.session.commit()
        return jsonify({'success': True, 'message': message})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/follows/<string:handle>', methods=['POST'])
def community_toggle_follow(handle):
    cust, profile, error = community_api_actor()
    if error:
        return error
    if profile.role == 'STUDENT' and not community_can_interact(profile, 'CAMPUS'):
        return jsonify({'success': False, 'message': 'Student verification is required before following members.'}), 403
    target = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == handle.strip().lower().lstrip('@')).first()
    if not target:
        return jsonify({'success': False, 'message': 'That community member was not found.'}), 404
    if target.role == 'STUDENT' and target.verification_status != 'VERIFIED' and not target.is_community_admin:
        return jsonify({'success': False, 'message': 'That student profile is still awaiting verification.'}), 403
    if target.id == profile.id:
        return jsonify({'success': False, 'message': 'You cannot follow your own profile.'}), 400
    same_feed = community_channel_for_role(target.role) == community_channel_for_role(profile.role)
    if not profile.is_community_admin and not target.is_community_admin and not same_feed:
        return jsonify({'success': False, 'message': 'That member is outside your role-locked community.'}), 403
    existing = CommunityFollow.query.filter_by(follower_profile_id=profile.id, followed_profile_id=target.id).first()
    awarded = 0.0
    if existing:
        db.session.delete(existing)
        following = False
        message = f'You unfollowed @{target.handle}.'
    else:
        db.session.add(CommunityFollow(follower_profile_id=profile.id, followed_profile_id=target.id))
        following = True
        message = f'You are now following @{target.handle}.'
        notification_key = f'follower:{profile.id}'
        if not CommunityNotification.query.filter_by(recipient_profile_id=target.id, kind='FOLLOW', target_key=notification_key).first():
            db.session.add(CommunityNotification(
                recipient_profile_id=target.id, actor_profile_id=profile.id, kind='FOLLOW',
                target_key=notification_key, message=f'@{profile.handle} followed you.',
            ))
    db.session.commit()
    if awarded:
        message += f' +{awarded:g} loyalty points (one-time reward).'
    return jsonify({
        'success': True, 'following': following, 'count': CommunityFollow.query.filter_by(followed_profile_id=target.id).count(),
        'points_awarded': awarded, 'points_balance': round(parse_float(cust.points_balance, 0.0), 2), 'message': message,
    })

@app.route('/community/api/notifications/read', methods=['POST'])
def community_read_notifications():
    cust, profile, error = community_api_actor()
    if error:
        return error
    CommunityNotification.query.filter_by(recipient_profile_id=profile.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'success': True, 'message': 'Notifications marked as read.'})

@app.route('/community/api/groups', methods=['POST'])
def community_create_group():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not COMMUNITY_GROUP_WORKSPACES_OPEN:
        return jsonify({'success': False, 'message': 'New Community workspaces are temporarily paused by the administrator.'}), 503
    if not community_can_interact(profile, community_channel_for_role(profile.role)):
        return jsonify({'success': False, 'message': 'Student verification is required before creating group workspaces.'}), 403
    try:
        active_owned = CommunityGroup.query.filter_by(created_by_profile_id=profile.id, is_active=True).count()
        if active_owned >= COMMUNITY_MAX_OWNED_GROUPS and not profile.is_community_admin:
            raise OrderValidationError(f'Each member can own up to {COMMUNITY_MAX_OWNED_GROUPS} active workspaces. Reuse an existing workspace or ask staff for help.')
        community_rate_limit(cust.id, CommunityGroup, 1440, COMMUNITY_MAX_OWNED_GROUPS, author_field='created_by_profile_id', profile_id=profile.id)
        data = request.get_json(silent=True) or request.form
        name = re.sub(r'\s+', ' ', (data.get('name') or '').strip())
        if len(name) < 3 or len(name) > 80:
            raise OrderValidationError('Group name must contain 3–80 characters.')
        channel = (data.get('channel') or '').strip().upper() if profile.is_community_admin else community_channel_for_role(profile.role)
        if channel not in COMMUNITY_CHANNELS or not community_can_interact(profile, channel):
            raise OrderValidationError('Choose a group channel you are allowed to manage.')
        external_chat_url = community_safe_link(data.get('external_chat_url'))
        raw_handles = str(data.get('invite_handles') or '')
        handle_values = []
        for raw in re.split(r'[,\s]+', raw_handles):
            candidate = raw.strip().lower().lstrip('@')
            if candidate and candidate not in handle_values:
                handle_values.append(candidate)
        if len(handle_values) > COMMUNITY_GROUP_MAX_MEMBERS - 1:
            raise OrderValidationError(f'A group can contain at most {COMMUNITY_GROUP_MAX_MEMBERS} members including you.')
        invitees = []
        for handle_value in handle_values:
            target = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == handle_value).first()
            if not target:
                raise OrderValidationError(f'No community member uses @{handle_value}.')
            if target.id == profile.id:
                continue
            provisional = CommunityGroup(channel=channel)
            if not community_group_target_allowed(provisional, target):
                raise OrderValidationError(f'@{target.handle} is outside this role-locked group or still awaiting verification.')
            invitees.append(target)
        group = CommunityGroup(name=name, channel=channel, created_by_profile_id=profile.id, external_chat_url=external_chat_url, is_active=True)
        db.session.add(group)
        db.session.flush()
        db.session.add(CommunityGroupMember(
            group_id=group.id, profile_id=profile.id, invited_by_profile_id=profile.id,
            member_role='OWNER', status='ACTIVE', joined_at=utc_now(), last_read_at=utc_now(),
        ))
        for target in invitees:
            membership = CommunityGroupMember(
                group_id=group.id, profile_id=target.id, invited_by_profile_id=profile.id,
                member_role='MEMBER', status='INVITED',
            )
            db.session.add(membership)
            db.session.flush()
            db.session.add(CommunityNotification(
                recipient_profile_id=target.id, actor_profile_id=profile.id,
                kind='GROUP_INVITE', target_key=f'group-member:{membership.id}',
                message=f'@{profile.handle} invited you to the group “{group.name}”.',
            ))
        db.session.commit()
        return jsonify({
            'success': True, 'message': f'Group “{group.name}” created with {len(invitees)} invitation(s).',
            'group_id': group.id, 'url': url_for('community_group_chat', group_id=group.id),
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/group-invites/<int:membership_id>', methods=['POST'])
def community_group_invite_response(membership_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    membership = db.session.get(CommunityGroupMember, membership_id)
    if not membership or membership.profile_id != profile.id or membership.status != 'INVITED' or not membership.group or not membership.group.is_active:
        return jsonify({'success': False, 'message': 'That group invitation is unavailable.'}), 404
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or '').strip().upper()
    if action == 'ACCEPT':
        if not community_group_target_allowed(membership.group, profile):
            return jsonify({'success': False, 'message': 'Your current verified role cannot enter this group.'}), 403
        membership.status = 'ACTIVE'
        membership.joined_at = utc_now()
        membership.last_read_at = utc_now()
        message = f'You joined “{membership.group.name}”.'
        group_url = url_for('community_group_chat', group_id=membership.group_id)
    elif action == 'DECLINE':
        membership.status = 'DECLINED'
        message = f'Invitation to “{membership.group.name}” declined.'
        group_url = None
    else:
        return jsonify({'success': False, 'message': 'Choose Accept or Decline.'}), 400
    db.session.commit()
    return jsonify({'success': True, 'message': message, 'url': group_url})

@app.route('/community/groups/<int:group_id>')
def community_group_chat(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        flash('That private group workspace is unavailable or requires an accepted invitation.', 'error')
        return redirect(url_for('community_home'))
    active_memberships = CommunityGroupMember.query.filter_by(group_id=group.id, status='ACTIVE').order_by(CommunityGroupMember.joined_at.asc()).all()
    existing_profile_ids = {
        row.profile_id for row in CommunityGroupMember.query.filter_by(group_id=group.id).all()
    }
    messages = []
    tasks = CommunityGroupTask.query.filter_by(group_id=group.id).order_by(
        CommunityGroupTask.status.asc(), CommunityGroupTask.due_date.asc(), CommunityGroupTask.created_at.desc(),
    ).limit(100).all()
    notes = CommunityGroupNote.query.filter_by(group_id=group.id, status='ACTIVE').order_by(
        CommunityGroupNote.is_pinned.desc(), CommunityGroupNote.updated_at.desc(),
    ).limit(60).all()
    polls = CommunityGroupPoll.query.filter_by(group_id=group.id).order_by(CommunityGroupPoll.created_at.desc()).limit(30).all()
    poll_payloads = {row.id: community_group_poll_payload(row, profile.id) for row in polls}
    candidate_query = CommunityProfile.query.filter(
        CommunityProfile.id.notin_(existing_profile_ids),
        CommunityProfile.verification_status != 'REJECTED',
    )
    invite_candidates = [row for row in candidate_query.order_by(CommunityProfile.handle.asc()).limit(300).all() if community_group_target_allowed(group, row)]
    return render_template(
        'community_group.html', cust=cust, profile=profile, group=group, membership=membership,
        active_memberships=active_memberships, messages=messages, tasks=tasks, notes=notes,
        polls=polls, poll_payloads=poll_payloads, invite_candidates=invite_candidates,
        is_group_owner=membership.member_role == 'OWNER', task_statuses=COMMUNITY_GROUP_TASK_STATUSES,
        task_priorities=COMMUNITY_GROUP_TASK_PRIORITIES, report_reasons=COMMUNITY_REPORT_REASONS,
        today=ph_today(), group_max_members=COMMUNITY_GROUP_MAX_MEMBERS,
        internal_chat_open=COMMUNITY_INTERNAL_CHAT_OPEN,
    )

@app.route('/community/api/groups/<int:group_id>/messages', methods=['GET', 'POST'])
def community_group_messages(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        return error
    if not COMMUNITY_INTERNAL_CHAT_OPEN:
        return jsonify({
            'success': False,
            'message': 'Internal live chat is paused in Community Lite. Use the workspace tasks, notes, polls, or the owner’s optional external Messenger link.',
        }), 410
    if request.method == 'GET':
        after_id = max(0, parse_int(request.args.get('after_id'), 0))
        messages = CommunityGroupMessage.query.filter(
            CommunityGroupMessage.group_id == group.id,
            CommunityGroupMessage.status == 'PUBLISHED',
            CommunityGroupMessage.id > after_id,
        ).order_by(CommunityGroupMessage.id.asc()).limit(100).all()
        membership.last_read_at = utc_now()
        db.session.commit()
        return jsonify({'success': True, 'messages': [community_group_message_payload(row) for row in messages]})
    try:
        community_rate_limit(cust.id, CommunityGroupMessage, 60, 30, profile_id=profile.id)
        data = request.get_json(silent=True) or request.form
        body = re.sub(r'\s+', ' ', (data.get('body') or '').strip())
        if not body or len(body) > 1000:
            raise OrderValidationError('Message must contain 1–1,000 characters.')
        hits = community_moderation_hits(body)
        message = CommunityGroupMessage(
            group_id=group.id, author_profile_id=profile.id, body=body,
            status='PENDING' if hits else 'PUBLISHED',
            moderation_hits=json.dumps(hits, ensure_ascii=False) if hits else None,
        )
        db.session.add(message)
        db.session.flush()
        group.updated_at = utc_now()
        membership.last_read_at = utc_now()
        if hits:
            community_add_admin_notice(
                'GROUP_MESSAGE_REVIEW', f'group-message:{message.id}',
                f'A safety phrase held a message in group “{group.name}” from @{profile.handle}.',
                profile=profile,
            )
        db.session.commit()
        return jsonify({
            'success': True, 'pending': bool(hits),
            'message': 'Message held for a private safety review.' if hits else 'Message sent.',
            'chat_message': None if hits else community_group_message_payload(message),
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/groups/<int:group_id>/members', methods=['POST'])
def community_group_members(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        return error
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or 'INVITE').strip().upper()
    if action == 'LEAVE':
        if membership.member_role == 'OWNER':
            return jsonify({'success': False, 'message': 'The group owner cannot leave. Remove other members or keep the group for now.'}), 400
        membership.status = 'LEFT'
        membership.joined_at = None
        group.updated_at = utc_now()
        db.session.commit()
        return jsonify({'success': True, 'message': f'You left “{group.name}”.'})
    if membership.member_role != 'OWNER' and not profile.is_community_admin:
        return jsonify({'success': False, 'message': 'Only the group owner can manage invitations.'}), 403
    try:
        if action == 'INVITE':
            total_members = CommunityGroupMember.query.filter(
                CommunityGroupMember.group_id == group.id,
                CommunityGroupMember.status.in_(('ACTIVE', 'INVITED')),
            ).count()
            if total_members >= COMMUNITY_GROUP_MAX_MEMBERS:
                raise OrderValidationError(f'This group has reached its {COMMUNITY_GROUP_MAX_MEMBERS}-member limit.')
            handle = str(data.get('handle') or '').strip().lower().lstrip('@')
            target = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == handle).first()
            if not community_group_target_allowed(group, target):
                raise OrderValidationError('That handle is outside this role-locked group or still awaiting verification.')
            if target.id == profile.id:
                raise OrderValidationError('You are already in this group.')
            target_membership = CommunityGroupMember.query.filter_by(group_id=group.id, profile_id=target.id).first()
            if target_membership and target_membership.status in {'ACTIVE', 'INVITED'}:
                raise OrderValidationError(f'@{target.handle} is already a member or has a pending invitation.')
            if target_membership:
                target_membership.status = 'INVITED'
                target_membership.invited_by_profile_id = profile.id
                target_membership.created_at = utc_now()
                target_membership.joined_at = None
            else:
                target_membership = CommunityGroupMember(
                    group_id=group.id, profile_id=target.id, invited_by_profile_id=profile.id,
                    member_role='MEMBER', status='INVITED',
                )
                db.session.add(target_membership)
            db.session.flush()
            if not CommunityNotification.query.filter_by(
                recipient_profile_id=target.id, kind='GROUP_INVITE', target_key=f'group-member:{target_membership.id}',
            ).first():
                db.session.add(CommunityNotification(
                    recipient_profile_id=target.id, actor_profile_id=profile.id, kind='GROUP_INVITE',
                    target_key=f'group-member:{target_membership.id}', message=f'@{profile.handle} invited you to “{group.name}”.',
                ))
            message = f'Invitation sent to @{target.handle}.'
        elif action == 'REMOVE':
            target_membership = db.session.get(CommunityGroupMember, parse_int(data.get('membership_id'), 0))
            if not target_membership or target_membership.group_id != group.id or target_membership.member_role == 'OWNER':
                raise OrderValidationError('That group member cannot be removed.')
            target_membership.status = 'REMOVED'
            message = 'Member removed from the group.'
        else:
            raise OrderValidationError('Choose a valid membership action.')
        group.updated_at = utc_now()
        db.session.commit()
        return jsonify({'success': True, 'message': message})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/groups/<int:group_id>/tasks', methods=['POST'])
def community_group_tasks(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        return error
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or 'CREATE').strip().upper()
    try:
        if action == 'CREATE':
            community_rate_limit(cust.id, CommunityGroupTask, 1440, 30, author_field='created_by_profile_id', profile_id=profile.id)
            title = re.sub(r'\s+', ' ', (data.get('title') or '').strip())
            details = re.sub(r'\s+', ' ', (data.get('details') or '').strip())[:500] or None
            priority = (data.get('priority') or 'NORMAL').strip().upper()
            assignee_id = parse_int(data.get('assigned_to_profile_id'), 0)
            assignee_membership = CommunityGroupMember.query.filter_by(group_id=group.id, profile_id=assignee_id, status='ACTIVE').first()
            if not title or len(title) > 120:
                raise OrderValidationError('Task title must contain 1–120 characters.')
            if priority not in COMMUNITY_GROUP_TASK_PRIORITIES:
                raise OrderValidationError('Choose Low, Normal, or High priority.')
            if not assignee_membership:
                raise OrderValidationError('Assign the task to an active group member.')
            due_raw = str(data.get('due_date') or '').strip()
            due_date = date.fromisoformat(due_raw) if due_raw else None
            if due_date and (due_date < ph_today() or due_date > ph_today() + timedelta(days=730)):
                raise OrderValidationError('Task due date must be today or within the next two years.')
            task = CommunityGroupTask(
                group_id=group.id, created_by_profile_id=profile.id,
                assigned_to_profile_id=assignee_id, title=title, details=details,
                priority=priority, status='TODO', due_date=due_date,
            )
            db.session.add(task)
            db.session.flush()
            if assignee_id != profile.id:
                db.session.add(CommunityNotification(
                    recipient_profile_id=assignee_id, actor_profile_id=profile.id, kind='GROUP_TASK',
                    target_key=f'group-task:{task.id}', message=f'@{profile.handle} assigned you “{task.title}” in {group.name}.',
                ))
            message = 'Task created and assigned.'
        elif action == 'STATUS':
            task = db.session.get(CommunityGroupTask, parse_int(data.get('task_id'), 0))
            status = (data.get('status') or '').strip().upper()
            if not task or task.group_id != group.id or status not in COMMUNITY_GROUP_TASK_STATUSES:
                raise OrderValidationError('That task or status is unavailable.')
            allowed = profile.is_community_admin or membership.member_role == 'OWNER' or profile.id in {task.created_by_profile_id, task.assigned_to_profile_id}
            if not allowed:
                raise OrderValidationError('Only the assignee, task creator, or group owner can update this task.')
            task.status = status
            task.completed_at = utc_now() if status == 'DONE' else None
            message = f'Task moved to {status.title()}.'
        else:
            raise OrderValidationError('Choose a valid task action.')
        group.updated_at = utc_now()
        db.session.commit()
        return jsonify({
            'success': True, 'message': message,
            'task': {
                'id': task.id, 'title': task.title, 'details': task.details,
                'priority': task.priority, 'status': task.status,
                'due_date': task.due_date.isoformat() if task.due_date else None,
                'assignee': f'@{task.assignee.handle}' if task.assignee else 'Unassigned',
            },
        })
    except (OrderValidationError, ValueError) as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc) if str(exc) else 'Enter a valid task due date.'}), 400

@app.route('/community/api/groups/<int:group_id>/notes', methods=['POST'])
def community_group_notes(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        return error
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or 'CREATE').strip().upper()
    try:
        if action == 'CREATE':
            community_rate_limit(cust.id, CommunityGroupNote, 1440, 20, profile_id=profile.id)
            note = CommunityGroupNote(group_id=group.id, author_profile_id=profile.id)
        else:
            note = db.session.get(CommunityGroupNote, parse_int(data.get('note_id'), 0))
            if not note or note.group_id != group.id or note.status != 'ACTIVE':
                raise OrderValidationError('That shared note is unavailable.')
            if not (profile.is_community_admin or membership.member_role == 'OWNER' or note.author_profile_id == profile.id):
                raise OrderValidationError('Only the note author or group owner can change this note.')
        if action in {'CREATE', 'UPDATE'}:
            title = re.sub(r'\s+', ' ', (data.get('title') or '').strip())
            body = (data.get('body') or '').strip()
            if not title or len(title) > 120 or not body or len(body) > 2000:
                raise OrderValidationError('A note needs a 1–120 character title and 1–2,000 characters of content.')
            note.title = title
            note.body = body
            if action == 'CREATE':
                db.session.add(note)
            message = 'Shared note created.' if action == 'CREATE' else 'Shared note updated.'
        elif action == 'TOGGLE_PIN':
            note.is_pinned = not note.is_pinned
            message = 'Note pin updated.'
        elif action == 'ARCHIVE':
            note.status = 'ARCHIVED'
            message = 'Note archived from the group workspace.'
        else:
            raise OrderValidationError('Choose a valid note action.')
        group.updated_at = utc_now()
        db.session.commit()
        return jsonify({
            'success': True, 'message': message,
            'note': {'id': note.id, 'title': note.title, 'body': note.body, 'is_pinned': bool(note.is_pinned), 'status': note.status},
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/groups/<int:group_id>/polls', methods=['POST'])
def community_group_polls(group_id):
    cust, profile, group, membership, error = community_group_actor(group_id)
    if error:
        return error
    data = request.get_json(silent=True) or request.form
    action = (data.get('action') or 'CREATE').strip().upper()
    try:
        if action == 'CREATE':
            community_rate_limit(cust.id, CommunityGroupPoll, 1440, 10, profile_id=profile.id)
            question = re.sub(r'\s+', ' ', (data.get('question') or '').strip())
            raw_options = data.get('options') if isinstance(data, dict) else request.form.getlist('option')
            if isinstance(raw_options, str):
                raw_options = raw_options.splitlines()
            options = []
            seen = set()
            for raw in raw_options or []:
                option = re.sub(r'\s+', ' ', str(raw).strip())[:100]
                if option and option.casefold() not in seen:
                    seen.add(option.casefold())
                    options.append(option)
            if not question or len(question) > 240 or not (2 <= len(options) <= 6):
                raise OrderValidationError('A poll needs a 1–240 character question and 2–6 different choices.')
            hours = parse_int(data.get('duration_hours'), 24)
            if hours not in {1, 6, 12, 24, 48, 72, 168}:
                raise OrderValidationError('Choose a supported poll duration.')
            poll = CommunityGroupPoll(
                group_id=group.id, author_profile_id=profile.id, question=question,
                status='OPEN', closes_at=utc_now() + timedelta(hours=hours),
            )
            db.session.add(poll)
            db.session.flush()
            for index, option in enumerate(options):
                db.session.add(CommunityGroupPollOption(poll_id=poll.id, option_text=option, sort_order=index))
            db.session.flush()
            message = 'Group poll opened.'
        elif action == 'VOTE':
            poll = db.session.get(CommunityGroupPoll, parse_int(data.get('poll_id'), 0))
            option = db.session.get(CommunityGroupPollOption, parse_int(data.get('option_id'), 0))
            if not poll or poll.group_id != group.id or not option or option.poll_id != poll.id:
                raise OrderValidationError('That poll choice is unavailable.')
            if poll.status != 'OPEN' or (poll.closes_at and poll.closes_at <= utc_now()):
                raise OrderValidationError('This poll is closed.')
            vote = CommunityGroupPollVote.query.filter_by(poll_id=poll.id, profile_id=profile.id).first()
            if vote:
                vote.option_id = option.id
            else:
                db.session.add(CommunityGroupPollVote(poll_id=poll.id, option_id=option.id, profile_id=profile.id))
            message = 'Vote recorded. You may change it while the poll is open.'
        elif action == 'CLOSE':
            poll = db.session.get(CommunityGroupPoll, parse_int(data.get('poll_id'), 0))
            if not poll or poll.group_id != group.id:
                raise OrderValidationError('That poll is unavailable.')
            if not (profile.is_community_admin or membership.member_role == 'OWNER' or poll.author_profile_id == profile.id):
                raise OrderValidationError('Only the poll creator or group owner can close it.')
            poll.status = 'CLOSED'
            message = 'Poll closed.'
        else:
            raise OrderValidationError('Choose a valid poll action.')
        group.updated_at = utc_now()
        db.session.commit()
        return jsonify({'success': True, 'message': message, 'poll': community_group_poll_payload(poll, profile.id)})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/group-messages/<int:message_id>/report', methods=['POST'])
def community_group_message_report(message_id):
    message = CommunityGroupMessage.query.get_or_404(message_id)
    cust, profile, group, membership, error = community_group_actor(message.group_id)
    if error:
        return error
    if message.author_profile_id == profile.id:
        return jsonify({'success': False, 'message': 'You cannot report your own message.'}), 400
    data = request.get_json(silent=True) or request.form
    reason = (data.get('reason') or '').strip().upper()
    details = re.sub(r'\s+', ' ', (data.get('details') or '').strip())[:240] or None
    if reason not in COMMUNITY_REPORT_REASONS:
        return jsonify({'success': False, 'message': 'Choose a valid report reason.'}), 400
    if CommunityGroupMessageReport.query.filter_by(message_id=message.id, reporter_profile_id=profile.id).first():
        return jsonify({'success': False, 'message': 'You already reported this message.'}), 400
    db.session.add(CommunityGroupMessageReport(
        message_id=message.id, reporter_profile_id=profile.id, reason=reason, details=details,
    ))
    db.session.flush()
    count = CommunityGroupMessageReport.query.filter_by(message_id=message.id, status='OPEN').count()
    message.flags_count = count
    quarantined = count >= 3
    if quarantined:
        message.status = 'QUARANTINED'
    community_add_admin_notice(
        'GROUP_MESSAGE_REPORT', f'group-message:{message.id}',
        f'A member reported a message in private group “{group.name}”.', profile=profile,
    )
    db.session.commit()
    return jsonify({'success': True, 'quarantined': quarantined, 'message': 'Report sent privately to Community Admin.'})

@app.route('/community/api/posts', methods=['POST'])
def community_create_post():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not COMMUNITY_POSTING_OPEN:
        return jsonify({'success': False, 'message': 'Community posting is temporarily paused by the administrator.'}), 503
    try:
        community_rate_limit(cust.id, CommunityPost, 60, 5, profile_id=profile.id)
        channel = (request.form.get('channel') or '').strip().upper()
        if channel not in COMMUNITY_CHANNELS or not community_can_interact(profile, channel):
            raise OrderValidationError('Your role can publish only in its own community channel.')
        allowed_modules = {value for value, _label in COMMUNITY_CHANNEL_MODULES[channel]}
        module = (request.form.get('module') or '').strip().upper()
        if module not in allowed_modules:
            raise OrderValidationError('Choose a valid post category.')
        body = re.sub(r'\s+', ' ', (request.form.get('body') or '').strip())
        if not body or len(body) > 280:
            raise OrderValidationError('Post text must contain 1–280 characters.')
        post_type = (request.form.get('post_type') or 'TEXT').strip().upper()
        if post_type != 'TEXT' or request.files.get('image') or request.form.get('link_url') or request.form.getlist('poll_option'):
            raise OrderValidationError('Community member posts are word-only. Images, links, and member polls are disabled.')
        mentioned_profiles = community_validate_mentions(profile, channel, body)
        hits = community_moderation_hits(body)
        published_count = CommunityPost.query.filter_by(author_profile_id=profile.id, status='PUBLISHED').count()
        requires_review = (not profile.is_community_admin and published_count < COMMUNITY_TRUSTED_POST_THRESHOLD) or bool(hits)
        post = CommunityPost(
            author_profile_id=profile.id,
            channel=channel,
            module=module,
            post_type=post_type,
            body=body,
            image_data=None,
            link_url=None,
            status='PENDING' if requires_review else 'PUBLISHED',
            moderation_hits=json.dumps(hits, ensure_ascii=False) if hits else None,
            published_at=None if requires_review else utc_now(),
            score_awarded=False,
        )
        db.session.add(post)
        db.session.flush()
        community_sync_post_mentions(post, mentioned_profiles, notify=not requires_review)
        db.session.commit()
        message = ('Post held for a quick safety review.' if hits else f'Your first {COMMUNITY_TRUSTED_POST_THRESHOLD} posts require staff approval.') if requires_review else 'Post published.'
        return jsonify({
            'success': True,
            'pending': requires_review,
            'message': message,
            'post': serialize_community_post(post, cust.id) if not requires_review else None,
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Community post creation failed for customer_id=%s', cust.id)
        return jsonify({'success': False, 'message': 'The post could not be saved. No partial post was published.'}), 500

@app.route('/community/api/posts/<int:post_id>/react', methods=['POST'])
def community_react(post_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    post = CommunityPost.query.get_or_404(post_id)
    if post.status != 'PUBLISHED' or not community_can_interact(profile, post.channel):
        return jsonify({'success': False, 'message': 'This channel is read-only for your role.'}), 403
    data = request.get_json(silent=True) or request.form
    reaction_type = (data.get('reaction_type') or 'LIKE').strip().upper()
    if reaction_type not in COMMUNITY_REACTIONS:
        return jsonify({'success': False, 'message': 'Invalid reaction.'}), 400
    reaction = CommunityReaction.query.filter_by(post_id=post.id, customer_id=cust.id).first()
    active = True
    if reaction and reaction.reaction_type == reaction_type:
        db.session.delete(reaction)
        active = False
    elif reaction:
        reaction.reaction_type = reaction_type
    else:
        db.session.add(CommunityReaction(post_id=post.id, customer_id=cust.id, reaction_type=reaction_type))
    awarded = 0.0
    if active and post.author_profile_id and post.author_profile_id != profile.id:
        key = f'post:{post.id}:actor:{profile.id}'
        if not CommunityNotification.query.filter_by(recipient_profile_id=post.author_profile_id, kind='REACTION', target_key=key).first():
            db.session.add(CommunityNotification(
                recipient_profile_id=post.author_profile_id, actor_profile_id=profile.id,
                kind='REACTION', target_key=key, message=f'@{profile.handle} liked your post.',
            ))
    db.session.commit()
    message = f'+{awarded:g} loyalty points earned (one time for this post).' if awarded else ('Post liked.' if active else 'Like removed.')
    return jsonify({'success': True, 'active': active, 'reaction_type': reaction_type, 'count': CommunityReaction.query.filter_by(post_id=post.id).count(), 'points_awarded': awarded, 'points_balance': round(parse_float(cust.points_balance, 0.0), 2), 'message': message})

@app.route('/community/api/posts/<int:post_id>/comments', methods=['POST'])
def community_comment(post_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not COMMUNITY_POSTING_OPEN:
        return jsonify({'success': False, 'message': 'Community comments are temporarily paused by the administrator.'}), 503
    post = CommunityPost.query.get_or_404(post_id)
    if post.status != 'PUBLISHED' or not community_can_interact(profile, post.channel):
        return jsonify({'success': False, 'message': 'This channel is read-only for your role.'}), 403
    try:
        community_rate_limit(cust.id, CommunityComment, 60, 20, profile_id=profile.id)
        data = request.get_json(silent=True) or request.form
        body = re.sub(r'\s+', ' ', (data.get('body') or '').strip())
        if not body or len(body) > 280:
            raise OrderValidationError('Comment must contain 1–280 characters.')
        mentioned_profiles = community_validate_mentions(profile, post.channel, body)
        hits = community_moderation_hits(body)
        comment = CommunityComment(
            post_id=post.id,
            author_profile_id=profile.id,
            body=body,
            status='PENDING' if hits else 'PUBLISHED',
            moderation_hits=json.dumps(hits, ensure_ascii=False) if hits else None,
            score_awarded=False,
        )
        db.session.add(comment)
        db.session.flush()
        if not hits:
            for target in mentioned_profiles:
                if not CommunityNotification.query.filter_by(recipient_profile_id=target.id, kind='COMMENT_MENTION', target_key=f'comment:{comment.id}').first():
                    db.session.add(CommunityNotification(
                        recipient_profile_id=target.id, actor_profile_id=profile.id, kind='COMMENT_MENTION',
                        target_key=f'comment:{comment.id}', message=f'@{profile.handle} tagged you in a comment.',
                    ))
        db.session.commit()
        return jsonify({
            'success': True,
            'pending': bool(hits),
            'message': 'Comment held for safety review.' if hits else 'Comment added.',
            'comment': None if hits else {
                'id': comment.id,
                'body': comment.body,
                'handle': f'@{profile.handle}',
                'created_at': ph_datetime_filter(comment.created_at),
            },
            'count': CommunityComment.query.filter_by(post_id=post.id, status='PUBLISHED').count(),
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/posts/<int:post_id>/reshare', methods=['POST'])
def community_reshare_post(post_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not COMMUNITY_RESHARING_OPEN:
        return jsonify({'success': False, 'message': 'Public resharing is paused in Community Lite. Share the original post link outside the app if appropriate.'}), 410
    source = CommunityPost.query.get_or_404(post_id)
    root = source.reshared_post if source.reshared_post_id and source.reshared_post else source
    if source.status != 'PUBLISHED' or root.status != 'PUBLISHED' or not community_can_interact(profile, source.channel):
        return jsonify({'success': False, 'message': 'This post is unavailable in your role-locked community.'}), 403
    try:
        community_rate_limit(cust.id, CommunityPost, 60, 5, profile_id=profile.id)
        existing = CommunityPost.query.filter(
            CommunityPost.author_profile_id == profile.id,
            CommunityPost.reshared_post_id == root.id,
            CommunityPost.status.in_(('PENDING', 'PUBLISHED')),
        ).first()
        if existing:
            raise OrderValidationError('You already reshared this post. It remains on your wall.')
        data = request.get_json(silent=True) or request.form
        if profile.is_community_admin:
            channel = (data.get('channel') or '').strip().upper()
            if channel not in COMMUNITY_CHANNELS:
                channel = root.channel if root.channel in COMMUNITY_CHANNELS else community_channel_for_role(profile.role)
        else:
            channel = community_channel_for_role(profile.role)
        if not community_can_interact(profile, channel):
            raise OrderValidationError('Choose a community channel you are allowed to use.')
        original_handle = f'@{root.author.handle}' if root.author else 'Macleen’s'
        requires_review = not profile.first_post_approved
        reshare = CommunityPost(
            author_profile_id=profile.id,
            reshared_post_id=root.id,
            channel=channel,
            module=root.module,
            post_type='TEXT',
            body=f'Shared {original_handle}’s post.',
            image_data=None,
            link_url=None,
            status='PENDING' if requires_review else 'PUBLISHED',
            published_at=None if requires_review else utc_now(),
            score_awarded=not requires_review,
        )
        db.session.add(reshare)
        db.session.flush()
        if not requires_review:
            profile.community_score = round(parse_float(profile.community_score, 0.0) + 1.0, 2)
            if root.author_profile_id and root.author_profile_id != profile.id:
                db.session.add(CommunityNotification(
                    recipient_profile_id=root.author_profile_id,
                    actor_profile_id=profile.id,
                    kind='RESHARE',
                    target_key=f'reshare:{reshare.id}',
                    message=f'@{profile.handle} reshared your post.',
                ))
        db.session.commit()
        return jsonify({
            'success': True,
            'pending': requires_review,
            'message': 'Your reshare was sent for first-post approval.' if requires_review else 'Post reshared to your wall.',
            'post': None if requires_review else serialize_community_post(reshare, cust.id),
            'count': CommunityPost.query.filter_by(reshared_post_id=root.id, status='PUBLISHED').count(),
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Community reshare failed for customer_id=%s post_id=%s', cust.id, post_id)
        return jsonify({'success': False, 'message': 'The post could not be reshared. Nothing was published.'}), 500

@app.route('/community/api/posts/<int:post_id>/report', methods=['POST'])
def community_report_post(post_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    post = CommunityPost.query.get_or_404(post_id)
    if not community_can_interact(profile, post.channel):
        return jsonify({'success': False, 'message': 'This post is outside your role-locked community or your student verification is pending.'}), 403
    if post.author_profile_id == profile.id:
        return jsonify({'success': False, 'message': 'You cannot report your own post. Ask staff if you need it removed.'}), 400
    if post.status != 'PUBLISHED':
        return jsonify({'success': False, 'message': 'This post is no longer public.'}), 400
    try:
        community_rate_limit(cust.id, CommunityReport, 1440, 10, author_field='reporter_customer_id')
        if CommunityReport.query.filter_by(post_id=post.id, reporter_customer_id=cust.id).first():
            raise OrderValidationError('You already reported this post. Staff will review it.')
        data = request.get_json(silent=True) or request.form
        reason = (data.get('reason') or '').strip().upper()
        details = re.sub(r'\s+', ' ', (data.get('details') or '').strip())[:240] or None
        if reason not in COMMUNITY_REPORT_REASONS:
            raise OrderValidationError('Choose a valid report reason.')
        db.session.add(CommunityReport(post_id=post.id, reporter_customer_id=cust.id, reason=reason, details=details))
        db.session.flush()
        report_count = CommunityReport.query.filter_by(post_id=post.id, status='OPEN').count()
        post.flags_count = report_count
        quarantined = report_count >= 3
        if quarantined:
            post.status = 'QUARANTINED'
            db.session.add(CommunityModerationAction(
                post_id=post.id,
                profile_id=post.author_profile_id,
                admin_username='system',
                action='AUTO_QUARANTINE',
                note='Temporarily hidden after 3 unique open reports; requires staff review.',
            ))
        db.session.commit()
        return jsonify({'success': True, 'quarantined': quarantined, 'message': 'Report received privately. The post is temporarily hidden for review.' if quarantined else 'Report received privately. Staff will review it.'})
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400

@app.route('/community/api/polls/<int:post_id>/vote', methods=['POST'])
def community_poll_vote(post_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    post = CommunityPost.query.get_or_404(post_id)
    if post.status != 'PUBLISHED' or post.post_type != 'POLL' or not community_can_interact(profile, post.channel):
        return jsonify({'success': False, 'message': 'This poll is not available for your role.'}), 403
    if post.expires_at and post.expires_at <= utc_now():
        return jsonify({'success': False, 'message': 'This poll has ended.'}), 400
    data = request.get_json(silent=True) or request.form
    option_id = parse_int(data.get('option_id'), 0)
    option = CommunityPollOption.query.filter_by(id=option_id, post_id=post.id).first()
    if not option:
        return jsonify({'success': False, 'message': 'Choose a valid poll option.'}), 400
    if CommunityPollVote.query.filter_by(post_id=post.id, customer_id=cust.id).first():
        return jsonify({'success': False, 'message': 'You already voted in this poll.'}), 400
    score = 0.0
    db.session.add(CommunityPollVote(
        post_id=post.id,
        option_id=option.id,
        customer_id=cust.id,
        role_snapshot=profile.role,
        score_awarded=False,
    ))
    db.session.commit()
    return jsonify({'success': True, 'message': 'Vote recorded.', 'score': profile.community_score, 'results': community_poll_results(post)})

@app.route('/community/api/gifts', methods=['POST'])
def community_send_gift():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not COMMUNITY_GIFTING_OPEN:
        return jsonify({'success': False, 'message': 'New peer gifts are paused in Community Lite to protect loyalty balances. Existing issued vouchers remain valid.'}), 410
    if profile.role == 'STUDENT' and not community_can_interact(profile, 'CAMPUS'):
        return jsonify({'success': False, 'message': 'Student verification is required before sending gifts.'}), 403
    try:
        data = request.get_json(silent=True) or request.form
        if not check_password_hash(cust.pin_hash, str(data.get('pin') or '').strip()):
            raise OrderValidationError('Incorrect 4-digit loyalty PIN. No points were transferred.')
        recipient_handle = normalize_community_handle(data.get('recipient_handle'))
        recipient_profile = CommunityProfile.query.filter(db.func.lower(CommunityProfile.handle) == recipient_handle.lower()).first()
        if not recipient_profile:
            raise OrderValidationError('No community member uses that handle.')
        if recipient_profile.customer_id == cust.id:
            raise OrderValidationError('You cannot send a gift to your own account.')
        gift_type = (data.get('gift_type') or 'POINTS').strip().upper()
        note = re.sub(r'\s+', ' ', (data.get('note') or '').strip())[:120] or None
        product = None
        if gift_type == 'POINTS':
            points_amount = round(parse_float(data.get('points_amount'), 0.0), 2)
            if points_amount < 1 or points_amount > 20:
                raise OrderValidationError('A point gift must be from 1 to 20 points.')
            status = 'RECEIVED'
        elif gift_type == 'PRODUCT':
            product = Product.query.filter_by(id=parse_int(data.get('product_id'), 0)).with_for_update().first()
            if not product or not product.is_active or parse_int(product.stock, 0) <= 0:
                raise OrderValidationError('Choose an active in-stock product voucher.')
            points_amount = round(product_starting_price(product), 2)
            if points_amount <= 0 or points_amount > COMMUNITY_GIFT_DAILY_CAP:
                raise OrderValidationError(f'Gift vouchers must cost no more than {COMMUNITY_GIFT_DAILY_CAP:g} points.')
            status = 'AVAILABLE'
        else:
            raise OrderValidationError('Choose a points gift or product voucher.')
        start_today, next_day = ph_day_utc_bounds()
        spent_today = db.session.query(db.func.coalesce(db.func.sum(CommunityGift.points_amount), 0.0)).filter(
            CommunityGift.sender_customer_id == cust.id,
            CommunityGift.created_at >= start_today,
            CommunityGift.created_at < next_day,
            CommunityGift.status.notin_(('CANCELLED', 'REVERSED')),
        ).scalar() or 0.0
        if parse_float(spent_today, 0.0) + points_amount > COMMUNITY_GIFT_DAILY_CAP + 1e-9:
            raise OrderValidationError(f'Daily gifting is capped at {COMMUNITY_GIFT_DAILY_CAP:g} points for account safety.')
        locked_customers = Customer.query.filter(Customer.id.in_([cust.id, recipient_profile.customer_id])).order_by(Customer.id.asc()).with_for_update().all()
        locked_by_id = {row.id: row for row in locked_customers}
        cust = locked_by_id.get(cust.id)
        recipient = locked_by_id.get(recipient_profile.customer_id)
        if not cust or not recipient:
            raise OrderValidationError('A member account became unavailable. No gift was sent.')
        # Recheck the daily total after the sender row lock. This closes the
        # double-click/concurrent-request window that could otherwise exceed the cap.
        locked_spent_today = db.session.query(db.func.coalesce(db.func.sum(CommunityGift.points_amount), 0.0)).filter(
            CommunityGift.sender_customer_id == cust.id,
            CommunityGift.created_at >= start_today,
            CommunityGift.created_at < next_day,
            CommunityGift.status.notin_(('CANCELLED', 'REVERSED')),
        ).scalar() or 0.0
        if parse_float(locked_spent_today, 0.0) + points_amount > COMMUNITY_GIFT_DAILY_CAP + 1e-9:
            raise OrderValidationError(f'Daily gifting is capped at {COMMUNITY_GIFT_DAILY_CAP:g} points for account safety.')
        if parse_float(cust.points_balance, 0.0) < points_amount - 1e-9:
            raise OrderValidationError(f'You need {points_amount:g} points for this gift. Available: {parse_float(cust.points_balance, 0.0):g}.')
        claim_code = f'MFH-{secrets.token_hex(4).upper()}'
        gift = CommunityGift(
            sender_customer_id=cust.id,
            recipient_customer_id=recipient.id,
            gift_type=gift_type,
            points_amount=points_amount,
            product_id=product.id if product else None,
            product_name=product.name if product else None,
            note=note,
            claim_code=claim_code,
            status=status,
        )
        cust.points_balance = round(parse_float(cust.points_balance, 0.0) - points_amount, 2)
        if product:
            # Reserve inventory now so a valid paid voucher can be honored later.
            product.stock = parse_int(product.stock, 0) - 1
        db.session.add(RewardLedger(customer_id=cust.id, points_change=-points_amount, reason=f'Community gift sent to @{recipient_profile.handle}'))
        if gift_type == 'POINTS':
            recipient.points_balance = round(parse_float(recipient.points_balance, 0.0) + points_amount, 2)
            db.session.add(RewardLedger(customer_id=recipient.id, points_change=points_amount, reason=f'Community gift from @{profile.handle}'))
        db.session.add(gift)
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'{product.name} voucher sent to @{recipient_profile.handle}.' if product else f'{points_amount:g} points sent to @{recipient_profile.handle}.',
            'points_balance': cust.points_balance,
            'claim_code': claim_code if product else None,
        })
    except OrderValidationError as exc:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(exc)}), 400
    except Exception:
        db.session.rollback()
        app.logger.exception('Community gifting failed for customer_id=%s', cust.id)
        return jsonify({'success': False, 'message': 'The gift could not be completed. No points were transferred.'}), 500

@app.route('/community/api/ads/<int:ad_id>/impression', methods=['POST'])
def community_ad_impression(ad_id):
    cust, profile, error = community_api_actor()
    if error:
        return error
    CommunityAd.query.get_or_404(ad_id)
    # Do not write one database row/update per screen view. Clicks remain the
    # useful low-volume conversion signal until batched analytics is added.
    return jsonify({'success': True, 'recorded': False, 'mode': 'clicks_only'})

@app.route('/community/ad/<int:ad_id>/click')
def community_ad_click(ad_id):
    cust = get_current_community_customer()
    if not cust:
        return redirect(url_for('customer_login', next='community'))
    ad = CommunityAd.query.get_or_404(ad_id)
    if not ad.is_active:
        return redirect(url_for('community_home'))
    ad.click_count = (ad.click_count or 0) + 1
    db.session.commit()
    return redirect(community_safe_cta(ad.cta_url))

@app.route('/community/api/push/subscribe', methods=['POST'])
def community_push_subscribe():
    cust, profile, error = community_api_actor()
    if error:
        return error
    if not (os.environ.get('WEBPUSH_VAPID_PUBLIC_KEY') and os.environ.get('WEBPUSH_VAPID_PRIVATE_KEY')):
        return jsonify({'success': False, 'message': 'Background push is not configured yet. In-app Flash Perch alerts remain active.'}), 503
    data = request.get_json(silent=True) or {}
    endpoint = str(data.get('endpoint') or '').strip()
    keys = data.get('keys') if isinstance(data.get('keys'), dict) else {}
    p256dh = str(keys.get('p256dh') or '').strip()
    auth = str(keys.get('auth') or '').strip()
    if not endpoint.startswith('https://') or len(endpoint) > 3000 or not p256dh or not auth:
        return jsonify({'success': False, 'message': 'The browser returned an invalid push subscription.'}), 400
    subscription = CommunityPushSubscription.query.filter_by(endpoint=endpoint).first()
    if subscription:
        subscription.customer_id = cust.id
        subscription.p256dh = p256dh
        subscription.auth = auth
        subscription.is_active = True
    else:
        db.session.add(CommunityPushSubscription(customer_id=cust.id, endpoint=endpoint, p256dh=p256dh, auth=auth, is_active=True))
    profile.push_opt_in = True
    db.session.commit()
    return jsonify({'success': True, 'message': 'Flash Perch browser alerts enabled.'})

@app.route('/community/api/push/unsubscribe', methods=['POST'])
def community_push_unsubscribe():
    cust, profile, error = community_api_actor()
    if error:
        return error
    CommunityPushSubscription.query.filter_by(customer_id=cust.id).update({'is_active': False})
    profile.push_opt_in = False
    db.session.commit()
    return jsonify({'success': True, 'message': 'Background alerts disabled. In-app alerts remain visible.'})

@app.route('/admin/community')
@require_admin
def community_admin():
    profiles = CommunityProfile.query.order_by(CommunityProfile.created_at.desc()).all()
    admin_notices = CommunityAdminNotice.query.order_by(CommunityAdminNotice.created_at.desc()).limit(60).all()
    review_posts = CommunityPost.query.filter(CommunityPost.status.in_(('PENDING', 'QUARANTINED', 'HIDDEN'))).order_by(CommunityPost.created_at.asc()).all()
    pending_comments = CommunityComment.query.filter_by(status='PENDING').order_by(CommunityComment.created_at.asc()).all()
    reports = CommunityReport.query.filter_by(status='OPEN').order_by(CommunityReport.created_at.asc()).all()
    group_review_messages = [] if not COMMUNITY_INTERNAL_CHAT_OPEN else CommunityGroupMessage.query.filter(or_(
        CommunityGroupMessage.status.in_(('PENDING', 'QUARANTINED')),
        CommunityGroupMessage.flags_count > 0,
    )).order_by(CommunityGroupMessage.created_at.asc()).all()
    ads = CommunityAd.query.order_by(CommunityAd.created_at.desc()).all()
    alerts = CommunityAlert.query.order_by(CommunityAlert.created_at.desc()).limit(30).all()
    flash_polls = CommunityPost.query.filter_by(is_flash_poll=True).order_by(CommunityPost.publish_date.desc()).limit(30).all()
    keywords = CommunityKeyword.query.order_by(CommunityKeyword.category.asc(), CommunityKeyword.phrase.asc()).all()
    pending_drops = []
    recent_actions = CommunityModerationAction.query.order_by(CommunityModerationAction.created_at.desc()).limit(40).all()
    active_products = Product.query.filter(Product.is_active.is_(True), Product.stock > 0).order_by(Product.name.asc()).all()
    main_admin_profile = CommunityProfile.query.filter_by(is_community_admin=True).first()
    today = ph_today()
    start_today, next_day = ph_day_utc_bounds(today)
    new_members_today = CommunityProfile.query.filter(
        CommunityProfile.created_at >= start_today,
        CommunityProfile.created_at < next_day,
    ).count()
    pending_students = CommunityProfile.query.filter(
        CommunityProfile.role == 'STUDENT',
        CommunityProfile.verification_status == 'PENDING',
    ).count() + CommunityProfile.query.filter_by(student_application_status='PENDING').count()
    stats = {
        'profiles': len(profiles),
        'students': sum(1 for row in profiles if row.role == 'STUDENT'),
        'residents': sum(1 for row in profiles if row.role == 'RESIDENT'),
        'published_posts': CommunityPost.query.filter_by(status='PUBLISHED').count(),
        'review_queue': len(review_posts) + len(pending_comments) + len(group_review_messages),
        'open_reports': len(reports),
        'today_checkins': CommunityCheckin.query.filter_by(checkin_date=today).count(),
        'today_store_checkins': CommunityStoreCheckin.query.filter_by(checkin_date=today).count(),
        'new_members_today': new_members_today,
        'pending_students': pending_students,
        'unread_join_notices': sum(1 for row in admin_notices if not row.is_read),
    }
    return render_template(
        'community_admin.html',
        profiles=profiles,
        admin_notices=admin_notices,
        review_posts=review_posts,
        pending_comments=pending_comments,
        reports=reports,
        group_review_messages=group_review_messages,
        ads=ads,
        alerts=alerts,
        flash_polls=flash_polls,
        keywords=keywords,
        pending_drops=pending_drops,
        recent_actions=recent_actions,
        active_products=active_products,
        main_admin_profile=main_admin_profile,
        stats=stats,
        report_reasons=COMMUNITY_REPORT_REASONS,
        today=today,
        webpush_configured=bool(os.environ.get('WEBPUSH_VAPID_PUBLIC_KEY') and os.environ.get('WEBPUSH_VAPID_PRIVATE_KEY') and os.environ.get('WEBPUSH_VAPID_SUBJECT')),
        production_sqlite_warning=bool(IS_PRODUCTION and db.engine.dialect.name == 'sqlite'),
        community_controls={
            'registration': COMMUNITY_REGISTRATION_OPEN,
            'posting': COMMUNITY_POSTING_OPEN,
            'workspaces': COMMUNITY_GROUP_WORKSPACES_OPEN,
            'internal_chat': COMMUNITY_INTERNAL_CHAT_OPEN,
            'social_rewards': COMMUNITY_SOCIAL_REWARDS_OPEN,
            'gifting': COMMUNITY_GIFTING_OPEN,
            'resharing': COMMUNITY_RESHARING_OPEN,
        },
    )

@app.route('/admin/community/notices/read', methods=['POST'])
@require_admin
def community_admin_read_notices():
    notice_id = parse_int(request.form.get('notice_id'), 0)
    query = CommunityAdminNotice.query.filter_by(is_read=False)
    if notice_id:
        query = query.filter_by(id=notice_id)
    updated = query.update({'is_read': True}, synchronize_session=False)
    db.session.commit()
    flash(f'{updated} Community Admin alert(s) marked as read.', 'success')
    return redirect(url_for('community_admin') + '#join-alerts')

@app.route('/admin/community/group-message/<int:message_id>/status', methods=['POST'])
@require_admin
def community_admin_group_message_status(message_id):
    message = CommunityGroupMessage.query.get_or_404(message_id)
    action = (request.form.get('action') or '').strip().upper()
    if action not in {'PUBLISH', 'REMOVE'}:
        flash('Invalid group-message moderation action.', 'error')
        return redirect(url_for('community_admin') + '#group-message-review')
    message.status = 'PUBLISHED' if action == 'PUBLISH' else 'REMOVED'
    message.flags_count = 0
    CommunityGroupMessageReport.query.filter_by(message_id=message.id, status='OPEN').update({'status': 'CLOSED'})
    db.session.add(CommunityModerationAction(
        profile_id=message.author_profile_id,
        admin_username=session.get('admin_user') or 'admin',
        action=f'GROUP_MESSAGE_{action}',
        note=f'Private group message #{message.id} in group #{message.group_id}.',
    ))
    db.session.commit()
    flash(f'Group message #{message.id} is now {message.status.lower()}.', 'success')
    return redirect(url_for('community_admin') + '#group-message-review')

@app.route('/admin/community/profile/<int:profile_id>/verification', methods=['POST'])
@require_admin
def community_admin_verify_profile(profile_id):
    profile = CommunityProfile.query.get_or_404(profile_id)
    action = (request.form.get('action') or '').strip().upper()
    if action not in {'VERIFY', 'REJECT', 'RESET'}:
        flash('Invalid community verification action.', 'error')
        return redirect(url_for('community_admin'))
    reviewing_application = profile.role == 'RESIDENT' and profile.student_application_status == 'PENDING'
    if action == 'VERIFY' and reviewing_application:
        profile.role = 'STUDENT'
        profile.campus_name = profile.student_application_campus
        profile.department = profile.student_application_department
        profile.graduating_year = profile.student_application_graduating_year
        profile.vibe_status = profile.vibe_status or 'Quiet study mode'
        profile.student_application_status = 'APPROVED'
        profile.first_post_approved = False
    elif action == 'REJECT' and reviewing_application:
        profile.student_application_status = 'REJECTED'
    profile.verification_status = {'VERIFY': 'VERIFIED', 'REJECT': ('SELF_DECLARED' if reviewing_application else 'REJECTED'), 'RESET': 'PENDING'}[action]
    profile.verification_note = re.sub(r'\s+', ' ', (request.form.get('note') or '').strip())[:255] or None
    if action in {'VERIFY', 'REJECT'} and profile.student_id_image_data:
        profile.student_id_image_data = None
        profile.student_id_deleted_at = utc_now()
    db.session.add(CommunityModerationAction(
        profile_id=profile.id,
        admin_username=session.get('admin_user') or 'admin',
        action=f'PROFILE_{action}',
        note=profile.verification_note,
    ))
    db.session.commit()
    flash(f'@{profile.handle} verification set to {profile.verification_status}.', 'success')
    return redirect(url_for('community_admin') + '#profiles')

@app.route('/admin/community/student-tag', methods=['POST'])
@require_admin
def community_admin_student_tag():
    identifier = (request.form.get('customer_identifier') or '').strip()
    cust = get_customer_by_identifier(identifier)
    if not cust:
        flash('Customer not found. Enter the exact mobile number or loyalty card number.', 'error')
        return redirect(url_for('community_admin') + '#profiles')
    cust.community_student_preapproved = True
    cust.community_student_preapproved_at = utc_now()
    cust.community_student_preapproved_by = session.get('admin_user') or 'admin'
    profile = CommunityProfile.query.filter_by(customer_id=cust.id).first()
    if profile and profile.role == 'STUDENT':
        profile.verification_status = 'VERIFIED'
        profile.verification_method = 'ADMIN_TAG'
        if profile.student_id_image_data:
            profile.student_id_image_data = None
            profile.student_id_deleted_at = utc_now()
    db.session.add(CommunityModerationAction(
        profile_id=profile.id if profile else None,
        admin_username=session.get('admin_user') or 'admin', action='STUDENT_PREAPPROVAL',
        note=f'Private loyalty customer #{cust.id} marked as known student by staff.',
    ))
    db.session.commit()
    flash(f'{cust.name} is pre-approved as a student. Their mobile number remains private.', 'success')
    return redirect(url_for('community_admin') + '#profiles')

@app.route('/admin/community/main-admin', methods=['POST'])
@require_admin
def community_admin_assign_main_admin():
    target = db.session.get(CommunityProfile, parse_int(request.form.get('profile_id'), 0))
    if not target:
        flash('Choose an existing community profile.', 'error')
        return redirect(url_for('community_admin') + '#profiles')
    conflict = CommunityProfile.query.filter(
        db.func.lower(CommunityProfile.handle) == COMMUNITY_MAIN_ADMIN_HANDLE,
        CommunityProfile.id != target.id,
    ).first()
    if conflict:
        flash(f'@{COMMUNITY_MAIN_ADMIN_HANDLE} already belongs to another profile. Select that profile or resolve the duplicate first.', 'error')
        return redirect(url_for('community_admin') + '#profiles')
    CommunityProfile.query.filter(CommunityProfile.id != target.id).update({'is_community_admin': False})
    target.handle = COMMUNITY_MAIN_ADMIN_HANDLE
    target.is_community_admin = True
    target.verification_status = 'VERIFIED'
    target.first_post_approved = True
    db.session.add(CommunityModerationAction(
        profile_id=target.id, admin_username=session.get('admin_user') or 'admin',
        action='ASSIGN_MAIN_COMMUNITY_ADMIN', note=f'Assigned reserved @{COMMUNITY_MAIN_ADMIN_HANDLE} access to both role feeds.',
    ))
    db.session.commit()
    flash(f'@{COMMUNITY_MAIN_ADMIN_HANDLE} is now the main administrator for Campus Hub and Town Square.', 'success')
    return redirect(url_for('community_admin') + '#profiles')

@app.route('/admin/community/post/<int:post_id>/status', methods=['POST'])
@require_admin
def community_admin_post_status(post_id):
    post = CommunityPost.query.get_or_404(post_id)
    action = (request.form.get('action') or '').strip().upper()
    if action not in {'PUBLISH', 'HIDE', 'REMOVE'}:
        flash('Invalid community post action.', 'error')
        return redirect(url_for('community_admin'))
    if action == 'PUBLISH' and post.author and post.author.role == 'STUDENT' and post.author.verification_status != 'VERIFIED':
        flash(f'Approve @{post.author.handle} in the private student verification queue before publishing a Campus Hub post.', 'error')
        return redirect(url_for('community_admin') + '#profiles')
    old_status = post.status
    post.status = {'PUBLISH': 'PUBLISHED', 'HIDE': 'HIDDEN', 'REMOVE': 'REMOVED'}[action]
    if action == 'PUBLISH':
        post.published_at = post.published_at or utc_now()
        if post.author:
            approved_count = CommunityPost.query.filter_by(author_profile_id=post.author.id, status='PUBLISHED').count()
            post.author.first_post_approved = approved_count >= COMMUNITY_TRUSTED_POST_THRESHOLD
            post.score_awarded = False
        if post.author:
            community_sync_post_mentions(post, notify=True)
        CommunityReport.query.filter_by(post_id=post.id, status='OPEN').update({'status': 'RESOLVED'})
        post.flags_count = 0
    note = re.sub(r'\s+', ' ', (request.form.get('note') or '').strip())[:240] or None
    db.session.add(CommunityModerationAction(
        post_id=post.id,
        profile_id=post.author_profile_id,
        admin_username=session.get('admin_user') or 'admin',
        action=f'POST_{action}',
        note=note or f'{old_status} → {post.status}',
    ))
    db.session.commit()
    flash(f'Community post #{post.id} is now {post.status.lower()}.', 'success')
    return redirect(url_for('community_admin') + '#review')

@app.route('/admin/community/comment/<int:comment_id>/status', methods=['POST'])
@require_admin
def community_admin_comment_status(comment_id):
    comment = CommunityComment.query.get_or_404(comment_id)
    action = (request.form.get('action') or '').strip().upper()
    if action not in {'PUBLISH', 'REMOVE'}:
        flash('Invalid comment action.', 'error')
        return redirect(url_for('community_admin'))
    comment.status = 'PUBLISHED' if action == 'PUBLISH' else 'REMOVED'
    if action == 'PUBLISH':
        comment.score_awarded = False
    db.session.add(CommunityModerationAction(
        post_id=comment.post_id,
        profile_id=comment.author_profile_id,
        admin_username=session.get('admin_user') or 'admin',
        action=f'COMMENT_{action}',
    ))
    db.session.commit()
    flash(f'Comment #{comment.id} is now {comment.status.lower()}.', 'success')
    return redirect(url_for('community_admin') + '#review')

@app.route('/admin/community/ad', methods=['POST'])
@require_admin
def community_admin_create_ad():
    try:
        title = re.sub(r'\s+', ' ', (request.form.get('title') or '').strip())
        body = re.sub(r'\s+', ' ', (request.form.get('body') or '').strip())
        target_role = (request.form.get('target_role') or 'ALL').strip().upper()
        channel = (request.form.get('channel') or 'ALL').strip().upper()
        cta_label = re.sub(r'\s+', ' ', (request.form.get('cta_label') or 'View menu').strip())[:40]
        cta_url = community_safe_cta(request.form.get('cta_url'), default='/')
        image_url = (request.form.get('image_url') or '').strip() or None
        if not title or len(title) > 120 or not body or len(body) > 240:
            raise OrderValidationError('Ad title and message are required and must fit their limits.')
        if target_role not in {'ALL', *COMMUNITY_ROLES} or channel not in {'ALL', *COMMUNITY_CHANNELS}:
            raise OrderValidationError('Choose a valid ad audience and channel.')
        if image_url:
            image_url = community_safe_link(image_url, required=True)
        start_at = community_local_datetime(request.form.get('start_at'))
        end_at = community_local_datetime(request.form.get('end_at'))
        if start_at and end_at and end_at <= start_at:
            raise OrderValidationError('Ad end time must be after its start time.')
        db.session.add(CommunityAd(
            title=title, body=body, target_role=target_role, channel=channel,
            cta_label=cta_label or 'View menu', cta_url=cta_url, image_url=image_url,
            start_at=start_at, end_at=end_at, is_active=True,
        ))
        db.session.commit()
        flash('Native community ad created. It will be inserted after every fifth post for its audience.', 'success')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    return redirect(url_for('community_admin') + '#ads')

@app.route('/admin/community/ad/<int:ad_id>/toggle', methods=['POST'])
@require_admin
def community_admin_toggle_ad(ad_id):
    ad = CommunityAd.query.get_or_404(ad_id)
    ad.is_active = not ad.is_active
    db.session.commit()
    flash(f'Community ad {"activated" if ad.is_active else "paused"}.', 'success')
    return redirect(url_for('community_admin') + '#ads')

@app.route('/admin/community/alert', methods=['POST'])
@require_admin
def community_admin_create_alert():
    try:
        title = re.sub(r'\s+', ' ', (request.form.get('title') or '').strip())
        body = re.sub(r'\s+', ' ', (request.form.get('body') or '').strip())
        target_role = (request.form.get('target_role') or 'ALL').strip().upper()
        duration_minutes = parse_int(request.form.get('duration_minutes'), 60)
        if not title or len(title) > 100 or not body or len(body) > 240:
            raise OrderValidationError('Alert title and message are required and must fit their limits.')
        if target_role not in {'ALL', *COMMUNITY_ROLES}:
            raise OrderValidationError('Choose a valid alert audience.')
        if duration_minutes < 5 or duration_minutes > 720:
            raise OrderValidationError('Flash Perch duration must be from 5 minutes to 12 hours.')
        cta_url = community_safe_cta(request.form.get('cta_url'), default='/community')
        now = utc_now()
        alert = CommunityAlert(
            title=title, body=body, target_role=target_role, cta_url=cta_url,
            starts_at=now, ends_at=now + timedelta(minutes=duration_minutes),
            created_by=session.get('admin_user') or 'admin', is_active=True,
        )
        db.session.add(alert)
        db.session.commit()
        delivery = send_community_alert_push(alert)
        if delivery['configured']:
            flash(f'Flash Perch alert is live. Browser push sent: {delivery["sent"]}; failed: {delivery["failed"]}.', 'success')
        else:
            flash('Flash Perch alert is live in the app. Background Web Push will activate after VAPID keys are configured.', 'success')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    return redirect(url_for('community_admin') + '#alerts')

@app.route('/admin/community/alert/<int:alert_id>/toggle', methods=['POST'])
@require_admin
def community_admin_toggle_alert(alert_id):
    alert = CommunityAlert.query.get_or_404(alert_id)
    alert.is_active = not alert.is_active
    db.session.commit()
    flash(f'Flash Perch alert {"activated" if alert.is_active else "ended"}.', 'success')
    return redirect(url_for('community_admin') + '#alerts')

@app.route('/admin/community/flash-poll', methods=['POST'])
@require_admin
def community_admin_create_flash_poll():
    try:
        question = re.sub(r'\s+', ' ', (request.form.get('question') or '').strip())
        publish_date_raw = (request.form.get('publish_date') or '').strip()
        try:
            publish_date = date.fromisoformat(publish_date_raw)
        except ValueError:
            raise OrderValidationError('Choose a valid Flash Poll date.')
        if publish_date < ph_today() or publish_date > ph_today() + timedelta(days=60):
            raise OrderValidationError('Flash Poll date must be today or within the next 60 days.')
        if not question or len(question) > 280:
            raise OrderValidationError('Flash Poll question must contain 1–280 characters.')
        if CommunityPost.query.filter_by(is_flash_poll=True, publish_date=publish_date).first():
            raise OrderValidationError('Only one 8:00 AM Flash Poll may be scheduled for that date.')
        raw_options = [re.sub(r'\s+', ' ', value.strip()) for value in request.form.getlist('poll_option') if value.strip()]
        unique_options = []
        seen = set()
        for option in raw_options:
            key = option.casefold()
            if key not in seen:
                seen.add(key)
                unique_options.append(option[:80])
        if not (2 <= len(unique_options) <= 4):
            raise OrderValidationError('A Flash Poll needs 2–4 different choices.')
        start_local = datetime.combine(publish_date, time(hour=8), tzinfo=MANILA_TZ)
        end_local = datetime.combine(publish_date, time(hour=23, minute=59, second=59), tzinfo=MANILA_TZ)
        post = CommunityPost(
            author_profile_id=None,
            channel='GLOBAL',
            module='FLASH_POLL',
            post_type='POLL',
            body=question,
            status='PUBLISHED',
            is_flash_poll=True,
            publish_date=publish_date,
            published_at=start_local.astimezone(timezone.utc).replace(tzinfo=None),
            expires_at=end_local.astimezone(timezone.utc).replace(tzinfo=None),
            score_awarded=True,
        )
        db.session.add(post)
        db.session.flush()
        for index, option in enumerate(unique_options):
            db.session.add(CommunityPollOption(post_id=post.id, option_text=option, sort_order=index))
        db.session.commit()
        flash(f'8:00 AM Flash Poll scheduled for {publish_date.strftime("%b %d, %Y")}.', 'success')
    except OrderValidationError as exc:
        db.session.rollback()
        flash(str(exc), 'error')
    return redirect(url_for('community_admin') + '#polls')

@app.route('/admin/community/keyword', methods=['POST'])
@require_admin
def community_admin_add_keyword():
    phrase = re.sub(r'\s+', ' ', (request.form.get('phrase') or '').strip().casefold())
    category = re.sub(r'[^A-Z0-9_]+', '_', (request.form.get('category') or 'ABUSE').strip().upper())[:40] or 'ABUSE'
    if len(phrase) < 2 or len(phrase) > 100:
        flash('Moderation phrase must contain 2–100 characters.', 'error')
        return redirect(url_for('community_admin') + '#keywords')
    if CommunityKeyword.query.filter(db.func.lower(CommunityKeyword.phrase) == phrase).first():
        flash('That moderation phrase already exists.', 'info')
        return redirect(url_for('community_admin') + '#keywords')
    db.session.add(CommunityKeyword(phrase=phrase, category=category, is_active=True))
    db.session.commit()
    flash('Moderation phrase added. Matches will be held for review, not automatically deleted.', 'success')
    return redirect(url_for('community_admin') + '#keywords')

@app.route('/admin/community/keyword/<int:keyword_id>/toggle', methods=['POST'])
@require_admin
def community_admin_toggle_keyword(keyword_id):
    keyword = CommunityKeyword.query.get_or_404(keyword_id)
    keyword.is_active = not keyword.is_active
    db.session.commit()
    flash(f'Moderation phrase {"enabled" if keyword.is_active else "paused"}.', 'success')
    return redirect(url_for('community_admin') + '#keywords')

@app.route('/admin/community/drop/<int:drop_id>/approve', methods=['POST'])
@require_admin
def community_admin_approve_drop(drop_id):
    drop = CommunityDrop.query.get_or_404(drop_id)
    if drop.reward_type != 'STAFF_FREEBIE' or drop.status not in {'PENDING_APPROVAL', 'ACTIVE'}:
        flash('That Mystery Drop cannot be changed.', 'error')
        return redirect(url_for('community_admin') + '#drops')
    product = db.session.get(Product, parse_int(request.form.get('product_id'), 0))
    if not product or not product.is_active or parse_int(product.stock, 0) <= 0:
        flash('Choose an active in-stock product for the freebie.', 'error')
        return redirect(url_for('community_admin') + '#drops')
    drop.product_id = product.id
    drop.reward_title = f'Free {product.name}'
    drop.status = 'ACTIVE'
    drop.approved_by = session.get('admin_user') or 'admin'
    drop.approved_at = utc_now()
    db.session.commit()
    flash(f'Mystery Drop approved: Free {product.name}.', 'success')
    return redirect(url_for('community_admin') + '#drops')

@app.route('/pos/community-gift/<int:gift_id>/redeem', methods=['POST'])
@require_cashier
def cashier_redeem_community_gift(gift_id):
    gift = CommunityGift.query.get_or_404(gift_id)
    if gift.gift_type != 'PRODUCT' or gift.status != 'AVAILABLE':
        flash('That community gift voucher is no longer available.', 'error')
        return redirect(url_for('cashier_terminal'))
    product = gift.product
    if not product:
        flash('The gifted product record is unavailable. Do not mark the voucher claimed.', 'error')
        return redirect(url_for('cashier_terminal'))
    recipient = gift.recipient
    order = Order(
        order_type='COMMUNITY_GIFT',
        dining_option='TAKEOUT',
        customer_id=recipient.id,
        customer_name=f'{recipient.name} (Gift Voucher)',
        contact_number=recipient.contact,
        subtotal=0.0,
        total_amount=0.0,
        payment_method='REWARD',
        payment_verified=True,
        status='COMPLETED',
        fulfillment_status='FULFILLED',
        notes=f'Community gift {gift.claim_code} from {gift.sender.name}',
    )
    db.session.add(order)
    db.session.flush()
    db.session.add(OrderItem(
        order_id=order.id,
        product_id=product.id,
        product_name=f'[Community Gift] {product.name}',
        unit_price=0.0,
        cost_price=max(0.0, parse_float(product.cost, 0.0)),
        quantity=1,
        subtotal=0.0,
    ))
    gift.status = 'CLAIMED'
    gift.claimed_by = session.get('cashier_user') or session.get('admin_user') or 'staff'
    gift.claimed_at = utc_now()
    db.session.commit()
    flash(f'Gift voucher {gift.claim_code} redeemed for {recipient.name}: {product.name}.', 'success')
    return redirect(url_for('cashier_terminal'))

@app.route('/pos/community-drop/<int:drop_id>/redeem', methods=['POST'])
@require_cashier
def cashier_redeem_community_drop(drop_id):
    drop = CommunityDrop.query.get_or_404(drop_id)
    product = drop.product
    cust = db.session.get(Customer, drop.customer_id)
    if drop.reward_type != 'STAFF_FREEBIE' or drop.status != 'ACTIVE' or not product or not cust:
        flash('That Mystery Drop is not ready to redeem.', 'error')
        return redirect(url_for('cashier_terminal'))
    if not product.is_active or parse_int(product.stock, 0) <= 0:
        flash('The selected Mystery Drop product is out of stock. Ask Admin to assign another product.', 'error')
        return redirect(url_for('cashier_terminal'))
    order = Order(
        order_type='MYSTERY_DROP', dining_option='TAKEOUT', customer_id=cust.id,
        customer_name=f'{cust.name} (Mystery Drop)', contact_number=cust.contact,
        subtotal=0.0, total_amount=0.0, payment_method='REWARD', payment_verified=True,
        status='COMPLETED', fulfillment_status='FULFILLED',
        notes=f'30-day Community Mystery Drop #{drop.id}',
    )
    db.session.add(order)
    db.session.flush()
    db.session.add(OrderItem(
        order_id=order.id, product_id=product.id, product_name=f'[Mystery Drop] {product.name}',
        unit_price=0.0, cost_price=max(0.0, parse_float(product.cost, 0.0)), quantity=1, subtotal=0.0,
    ))
    product.stock = parse_int(product.stock, 0) - 1
    drop.status = 'REDEEMED'
    drop.redeemed_order_id = order.id
    drop.redeemed_at = utc_now()
    db.session.commit()
    flash(f'Mystery Drop redeemed for {cust.name}: Free {product.name}.', 'success')
    return redirect(url_for('cashier_terminal'))

# ==================== CUSTOMER PORTAL ====================

@app.route('/portal/start/<string:source>')
def portal_start(source):
    allowed = {'counter', 'table', 'receipt', 'facebook'}
    if source not in allowed:
        source = 'direct'
    session['portal_source'] = source
    track_portal_event('QR_SCAN', source=source, customer_id=session.get('customer_id'))
    db.session.commit()
    section = {'table': 'suggest', 'counter': 'rewards', 'receipt': 'rewards', 'facebook': 'deals'}.get(source, 'rewards')
    if session.get('customer_id'):
        return redirect(url_for('customer_dashboard') + f'#{section}')
    return redirect(url_for('customer_login', src=source, next=section))

@app.route('/portal/login', methods=['GET', 'POST'])
def customer_login():
    source = request.args.get('src', '').strip() or session.get('portal_source') or 'direct'
    next_section = request.args.get('next', '').strip()
    card_hint = request.args.get('card', '').strip()
    if source:
        session['portal_source'] = source[:30]
    if request.method == 'POST':
        contact = request.form.get('contact', '').strip()
        pin = request.form.get('pin', '').strip()
        next_section = request.form.get('next', '').strip() or next_section
        cust = get_customer_by_identifier(contact)
        if cust and check_password_hash(cust.pin_hash, pin):
            issue = customer_access_issue(cust)
            if issue:
                if cust.card_expires_at and ph_today() > cust.card_expires_at and (cust.card_status or 'ACTIVE').upper() == 'ACTIVE':
                    cust.card_status = 'EXPIRED'
                    db.session.commit()
                flash(issue, 'error')
                return render_template('customer_login.html', source=source, next_section=next_section, card_hint=contact or card_hint)

            portal_source = session.get('portal_source', source)
            session.clear()
            session['customer_id'] = cust.id
            session['portal_source'] = portal_source
            session.permanent = True
            today = ph_today()
            if cust.last_daily_login != today:
                yesterday = today - timedelta(days=1)
                cust.login_streak = (cust.login_streak or 0) + 1 if cust.last_daily_login == yesterday else 1
                cust.points_balance = (cust.points_balance or 0.0) + 0.5
                cust.last_daily_login = today
                db.session.add(RewardLedger(
                    customer_id=cust.id,
                    points_change=0.5,
                    reason=f'Daily Login Reward (Day {cust.login_streak})',
                ))
            cust.last_active_at = utc_now()
            track_portal_event('LOGIN', source=portal_source, customer_id=cust.id)
            db.session.commit()
            if next_section == 'community':
                return redirect(url_for('community_home'))
            target = url_for('customer_dashboard')
            return redirect(target + (f'#{next_section}' if next_section else ''))
        flash('Invalid Contact or PIN.', 'error')
    return render_template('customer_login.html', source=source, next_section=next_section, card_hint=card_hint)

@app.route('/portal/register', methods=['GET', 'POST'])
def customer_register():
    ref_code = request.args.get('ref', '').strip()
    source = request.args.get('src', '').strip() or session.get('portal_source') or 'direct'
    next_section = request.args.get('next', '').strip()
    if source:
        session['portal_source'] = source[:30]
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        contact = request.form.get('contact', '').strip()
        messenger = request.form.get('fb_messenger', '').strip()
        pin = request.form.get('pin', '').strip()
        address = request.form.get('default_address', '').strip()
        landmark = request.form.get('default_landmark', '').strip()
        ref = request.form.get('referral_code', '').strip() or ref_code
        source = request.form.get('source', '').strip() or source
        next_section = request.form.get('next', '').strip() or next_section

        if not name or not contact:
            flash('Name and mobile number are required.', 'error')
            return redirect(url_for('customer_register', ref=ref_code or None, src=source, next=next_section or None))
        if not is_valid_customer_pin(pin):
            flash('Security PIN must be exactly 4 digits.', 'error')
            return redirect(url_for('customer_register', ref=ref_code or None, src=source, next=next_section or None))
        if Customer.query.filter_by(contact=contact).first():
            flash('Contact number already registered.', 'error')
            return redirect(url_for('customer_login', src=source, next=next_section or None))

        today = ph_today()
        try:
            new_cust = Customer(
                name=name,
                email=email,
                contact=contact,
                fb_messenger=messenger,
                default_address=address,
                default_landmark=landmark,
                points_balance=0.5,
                pin_hash=generate_password_hash(pin),
                card_number=None,
                card_status='ACTIVE',
                card_expires_at=today + timedelta(days=365),
                referred_by=ref if ref else None,
                last_daily_login=today,
                login_streak=1,
                last_active_at=utc_now(),
            )
            db.session.add(new_cust)
            db.session.flush()
            new_cust.card_number = generate_unique_card_number(new_cust.id)

            db.session.add(RewardLedger(
                customer_id=new_cust.id,
                points_change=0.5,
                reason='Welcome Login Bonus',
            ))
            # Referral rewards are intentionally held until the new member completes a paid purchase.
            # This prevents fake registrations and rewards both sides for an actual new customer.
            track_portal_event('REGISTER', source=source, customer_id=new_cust.id)
            db.session.commit()
            portal_source = source
            session.clear()
            session['customer_id'] = new_cust.id
            session['portal_source'] = portal_source
            session.permanent = True
            if ref:
                flash('🎉 Welcome! +0.5 point added. Complete your first paid purchase to unlock the referral bonus for both of you!', 'success')
            else:
                flash('🎉 Welcome! You earned 0.5 points!', 'success')
            if next_section == 'community':
                return redirect(url_for('community_home'))
            target = url_for('customer_dashboard')
            return redirect(target + (f'#{next_section}' if next_section else ''))
        except Exception:
            db.session.rollback()
            app.logger.exception('Customer registration failed for contact=%s', contact)
            flash('Registration could not be completed. Please try again or ask staff for help.', 'error')
            return redirect(url_for('customer_register', ref=ref_code or None, src=source, next=next_section or None))

    return render_template('customer_register.html', ref_code=ref_code, source=source, next_section=next_section)

@app.route('/portal/dashboard')
def customer_dashboard():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login'))

    cust = Customer.query.get(session['customer_id'])
    if not cust:
        session.pop('customer_id', None)
        flash('Account session expired. Please log in again.', 'info')
        return redirect(url_for('customer_login'))

    issue = customer_access_issue(cust)
    if issue:
        session.pop('customer_id', None)
        flash(issue, 'error')
        return redirect(url_for('customer_login'))

    if getattr(cust, 'login_streak', None) is None:
        cust.login_streak = 1
        db.session.commit()

    my_orders = Order.query.filter_by(customer_id=cust.id).order_by(Order.created_at.desc()).limit(30).all()
    active_promos = PromotionTracker.query.filter_by(is_active=True, is_visible=True).order_by(PromotionTracker.created_at.desc()).all()
    # Preserve the existing 3-day promo cycle automatically.
    active_promos = [p for p in active_promos if (utc_now() - p.created_at).days <= 3]
    bonus_campaigns = get_active_bonus_campaigns()

    reward_target = 20.0
    balance = float(cust.points_balance or 0.0)
    points_to_reward = max(0.0, reward_target - balance)
    reward_progress_pct = min(100.0, (balance / reward_target * 100.0) if reward_target else 100.0)
    recent_rewards = RewardLedger.query.filter_by(customer_id=cust.id).order_by(RewardLedger.created_at.desc()).limit(8).all()
    referral_rewards_count = ReferralReward.query.filter_by(referrer_customer_id=cust.id).count()
    favorite_items = Product.query.join(
        CustomerWishlist, CustomerWishlist.product_id == Product.id
    ).filter(
        CustomerWishlist.customer_id == cust.id,
        Product.is_active.is_(True),
    ).order_by(Product.name).all()
    open_group_orders = GroupOrder.query.filter_by(
        organizer_customer_id=cust.id, status='OPEN'
    ).order_by(GroupOrder.created_at.desc()).limit(5).all()
    base = _marketing_public_base_url()
    card_identifier = cust.card_number or cust.contact
    qr_target = f"{base}/portal/login?card={card_identifier}"
    qr_data = qr_svg_data_url(qr_target)

    return render_template(
        'customer_dashboard.html',
        cust=cust,
        orders=my_orders,
        active_promos=active_promos,
        bonus_campaigns=bonus_campaigns,
        reward_target=reward_target,
        points_to_reward=points_to_reward,
        reward_progress_pct=reward_progress_pct,
        recent_rewards=recent_rewards,
        referral_rewards_count=referral_rewards_count,
        favorite_items=favorite_items,
        open_group_orders=open_group_orders,
        loyalty_card_themes=LOYALTY_CARD_THEMES,
        qr_data=qr_data,
        qr_target=qr_target,
        today=ph_today(),
    )

@app.route('/portal/menu-vote/<int:candidate_id>', methods=['POST'])
def customer_menu_vote(candidate_id):
    # Legacy compatibility endpoint: voting was retired in favor of free-text product requests.
    if 'customer_id' not in session:
        return redirect(url_for('customer_login', next='suggest'))
    flash('Menu voting has been replaced by “What would you like to buy from our shop?”. Please send your request there.', 'info')
    return redirect(url_for('customer_dashboard') + '#suggest')

@app.route('/portal/suggest-product', methods=['POST'])
def customer_submit_product_suggestion():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login', next='suggest'))

    cust = Customer.query.get_or_404(session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('customer_login'))

    suggestion_text = re.sub(r'\s+', ' ', request.form.get('suggestion', '').strip())
    if len(suggestion_text) < 2:
        flash('Please tell us what you would like Macleen\'s to offer.', 'error')
        return redirect(url_for('customer_dashboard') + '#suggest')
    if len(suggestion_text) > 500:
        flash('Please keep your suggestion to 500 characters or less.', 'error')
        return redirect(url_for('customer_dashboard') + '#suggest')

    db.session.add(ProductSuggestion(
        customer_id=cust.id,
        customer_name=cust.name,
        suggestion_text=suggestion_text,
        status='NEW',
    ))
    # Keep customer requests as literal free-text words in Admin.
    # Do not convert suggestions into products, vote candidates, or catalog records.
    track_portal_event('PRODUCT_SUGGESTION', customer_id=cust.id)
    db.session.commit()
    flash('💡 Thank you! Your product suggestion was sent to Macleen\'s.', 'success')
    return redirect(url_for('customer_dashboard') + '#suggest')

@app.route('/portal/logout')
def customer_logout():
    if 'customer_id' in session:
        try:
            cust = Customer.query.get(session['customer_id'])
            if cust and hasattr(cust, 'last_active_at'):
                cust.last_active_at = utc_now() - timedelta(hours=1)
                db.session.commit()
        except Exception:
            db.session.rollback()
            app.logger.exception('Could not mark customer offline during logout')
    session.pop('customer_id', None)
    flash('Successfully logged out.', 'info')
    return redirect(url_for('store_catalog'))

@app.route('/portal/reserve-promo/<string:promo_code>', methods=['POST'])
def customer_reserve_promo_by_code(promo_code):
    if 'customer_id' not in session:
        return redirect(url_for('customer_login', next='deals'))

    cust = Customer.query.get_or_404(session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('customer_login'))
    promo = PromotionTracker.query.filter_by(promo_code=promo_code).first()

    if not promo or not promo.is_active or not promo.is_visible:
        flash('This promotion campaign is currently archived.', 'info')
        return redirect(url_for('customer_dashboard') + '#deals')

    days_active = (utc_now() - promo.created_at).days
    if days_active > 3:
        promo.is_active = False
        db.session.commit()
        flash('This 3-day promotion campaign has ended and is now archived.', 'info')
        return redirect(url_for('customer_dashboard') + '#deals')

    order = Order(
        order_type='PICKUP',
        dining_option='TAKEOUT',
        customer_id=cust.id,
        customer_name=f'{cust.name} (Member Promo)',
        contact_number=cust.contact,
        pickup_time='Today / In-Store Claim',
        target_time='In-Store Counter Claim',
        subtotal=promo.promo_price,
        delivery_fee=0.0,
        total_amount=promo.promo_price,
        payment_method='CASH',
        payment_verified=False,
        status='VERIFICATION',
        notes=f'[3-DAY PROMO CLAIM] {promo.title} (₱{promo.promo_price:,.2f} Deal)'
    )
    db.session.add(order)
    db.session.flush()

    db.session.add(OrderItem(
        order_id=order.id,
        product_id=None,
        product_name=f'[PROMO 3-DAY] {promo.title}',
        unit_price=promo.promo_price,
        cost_price=max(0.0, parse_float(promo.promo_cost, 0.0)),
        quantity=1,
        subtotal=promo.promo_price
    ))

    promo.claims_count = (promo.claims_count or 0) + 1
    promo.total_revenue = (promo.total_revenue or 0.0) + promo.promo_price
    track_portal_event('PROMO_RESERVE', customer_id=cust.id)
    db.session.commit()
    flash(f'🎉 Deal reserved! Order #{order.id} queued at Cashier counter.', 'success')
    return redirect(url_for('customer_dashboard') + '#deals')

@app.route('/portal/card-theme', methods=['POST'])
def customer_update_card_theme():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login', next='cardDesign'))
    cust = Customer.query.get_or_404(session['customer_id'])
    issue = customer_access_issue(cust)
    if issue:
        flash(issue, 'error')
        return redirect(url_for('customer_login'))

    theme = (request.form.get('theme') or '').strip()
    if theme not in LOYALTY_CARD_THEMES:
        flash('Please choose one of the available loyalty card themes.', 'error')
        return redirect(url_for('customer_dashboard') + '#cardDesign')

    cust.card_theme = theme
    db.session.commit()
    flash(f'🎫 Your loyalty card theme is now {LOYALTY_CARD_THEMES[theme]}.', 'success')
    return redirect(url_for('customer_dashboard') + '#cardDesign')

@app.route('/portal/update-profile-pic', methods=['POST'])
def update_profile_pic():
    if 'customer_id' not in session:
        return redirect(url_for('customer_login'))
    cust = Customer.query.get_or_404(session['customer_id'])
    try:
        new_image = customer_profile_image_from_request(existing=cust.profile_image)
        if new_image:
            cust.profile_image = new_image
            db.session.commit()
            flash('Profile picture updated!', 'success')
        else:
            flash('Choose a photo from your device or enter a photo URL.', 'info')
    except OrderValidationError as exc:
        flash(str(exc), 'error')
    return redirect(url_for('customer_dashboard') + '#account')

# ==================== INDEPENDENT STAFF AUTH ====================

@app.route('/staff/login', methods=['GET', 'POST'])
def staff_login():
    target = request.args.get('target', '').lower()
    if request.method == 'POST':
        user = request.form.get('username', '').strip().lower()
        pin = request.form.get('pin', '').strip()
        staff = Staff.query.filter(db.func.lower(Staff.username) == user).first()

        if staff and staff.active and check_password_hash(staff.pin_hash, pin):
            session.clear()
            session.permanent = False
            session['_staff_last_activity'] = utc_now().isoformat()
            if staff.role == 'ADMIN':
                session['admin_id'] = staff.id
                session['admin_user'] = staff.username
                return redirect(url_for('admin_dashboard'))
            if staff.role == 'CASHIER':
                session['cashier_id'] = staff.id
                session['cashier_user'] = staff.username
                return redirect(url_for('cashier_terminal'))
        flash('Invalid Username or PIN.', 'error')
    return render_template('staff_login.html', target=target)

@app.route('/staff/logout')
def staff_logout():
    clear_staff_session()
    flash('Logged out.', 'info')
    return redirect(url_for('staff_login'))

@app.route('/healthz')
def healthz():
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({
            'status': 'ok', 'release': APP_RELEASE, 'time_ph': ph_now().isoformat(),
            'community_mode': 'lite',
            'database': db.engine.dialect.name,
            'production_database_warning': bool(IS_PRODUCTION and db.engine.dialect.name == 'sqlite'),
        })
    except Exception:
        app.logger.exception('Health check failed')
        return jsonify({'status': 'error'}), 500

@app.errorhandler(500)
def internal_server_error(error):
    original = getattr(error, 'original_exception', None)
    app.logger.error('Unhandled server error on %s %s', request.method, request.path, exc_info=True)
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'message': 'Server error. Please try again.'}), 500
    return 'Macleen\'s Food House encountered a server error. Please try again.', 500

# Run schema/setup during application import so Render/Gunicorn fails visibly at startup if the DB is incompatible.
with app.app_context():
    run_db_setup()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
